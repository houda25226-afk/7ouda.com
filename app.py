"""
تطبيق Streamlit لتصنيف إفادات المكالمات + حساب الوقت المهدر + داشبورد
باستخدام الموديل: Mahmoud252002/7oudaModel

التشغيل:
    pip install -r requirements.txt
    streamlit run app.py

الصفحات:
  - تصنيف المكالمات (حسب الشركة والفترة)
  - الوعود (قائمة + مكسورة)
  - الإهمال ومتابعة الإهمال
  - تحليل نشاط المحصّلين (Dashboard + تصدير HTML)

أسماء الأعمدة قابلة للتعديل من قسم الإعدادات أعلى الملف.
"""

import io
import hashlib
import re
import zipfile
from io import BytesIO
from datetime import time as dt_time
from datetime import datetime

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import plotly.io as pio
import streamlit as st
import torch
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
from transformers import AutoTokenizer, AutoModelForSequenceClassification

# ==========================================================
# إعدادات وأسماء الأعمدة
# ==========================================================

MODEL_REPO = "Mahmoud252002/7oudaModel"
MAX_LENGTH = 256

ORIGINAL_TEXT_COL = "Notes"         # اسم العمود الأصلي في الملف
MODEL_TEXT_COL = "الافادة"          # الاسم اللي بيتحول له مؤقتًا لـ الموديل
CLASSIFICATION_COL = "التصنيف"      # عمود النتيجة: 1 = ناجحة / 0 = غير ناجحة
WASTED_TIME_COL = "الوقت_المهدر_دقيقة"

ID_CANDIDATES = ["Account ID", "account id", "AccountID", "ID", "id", "رقم الحساب", "الرقم التعريفي", "Account No", "account no", "Account Number"]
SALES_PERSON_CANDIDATES = ["Create By", "create by", "CreateBy", "Created By", "created by", "Sales Person", "sales person", "المحصّل", "Salesperson", "salesperson", "SalesPerson"]
COLLECTED_BY_CANDIDATES = ["Collected by", "collected by", "Collected By", "COLLECTED BY", "Created by", "created by", "Created By", "CREATED BY", "المحصل", "المحصّل", "Collector", "collector"]
ACCOUNT_NUMBER_CANDIDATES = ["Customer Account number", "Customer Account Number", "customer account number", "Customer Account No", "Account Number", "account number", "Account No", "رقم حساب العميل", "رقم الحساب"]
CREATED_ON_CANDIDATES = ["Created On", "created on", "CreatedOn", "تاريخ الافادة"]
CLAIM_CANDIDATES = ["Claim", "claim", "CLAIM", "رقم المطالبة", "رقم المطالبه"]
DUPLICATE_WINDOW_MINUTES = 20
DURATION_CANDIDATES = ["Call Duration", "call duration", "CallDuration", "Duration", "duration", "مدة المكالمة", "Call Time", "call time", "Talk Time", "talk time", "Duration (min)", "مدة"]

# ========================================================== 
# إعدادات تويب الوعود القائمة (المحفظة)
# ==========================================================
# أسماء الأعمدة المتوقعة في ملف المحفظة — لو اتغيرت غيّرها من هنا.
PROMISE_SUB_STATE_CANDIDATES = ["Sub State", "sub state", "SubState", "الحالة الفرعية"]
PROMISE_DUE_DATE_CANDIDATES = ["Follow up Due Date", "follow up due date", "FollowUpDueDate", "تاريخ المتابعة", "Due Date"]
PROMISE_NET_AMOUNT_CANDIDATES = ["Net Amount", "net amount", "NetAmount", "صافي المبلغ", "مبلغ المديونية"]

# قيمة Sub State اللي بتمثل وعد قائم
PROMISE_SUB_STATE_VALUE = "واعد بالسداد"

# المحصّلين اللي بنستبعدهم من الوعود القائمة
PROMISE_EXCLUDED_SALES = [
    "Archive Companies  II Anas",
    "Closed payments  II Anas",
    "Hold Companies  II Anas",
    "Op II Ibrahim Qassem",
    "قانونى -الوطنية",
]


# ==========================================================
# إعدادات تويب الإهمال
# ==========================================================
NEGLECT_SUB_STATES_DEFAULT = [
    "تم ابلاغ العميل - اتصال",
    "لايرد مع التكرار",
    "جدولة",
    "واعد بالسداد",
    "تم ابلاغ العميل - واتسب",
    "لا يرد",
    "إعفاء || بإنتظار المستند",
    "متوفي",
    "مغلق مع التكرار"
]

NEGLECT_LAST_DATE_CANDIDATES = ["Follow up Last Date", "follow up last date", "Last Follow Up", "تاريخ آخر متابعة"]
NEGLECT_RESULT_KEY = "neglect_result"
APP_DATA_CACHE_KEY = "app_uploaded_data_cache"
DASHBOARD_SOURCE_HASH_KEY = "dashboard_source_hash"


def uploaded_file_hash(uploaded_file):
    if uploaded_file is None:
        return None
    return hashlib.sha256(uploaded_file.getvalue()).hexdigest()


def _clear_cached_results(result_keys):
    for key in result_keys:
        if key.startswith("period_results:"):
            period_key = key.split(":", 1)[1]
            st.session_state.setdefault("period_results", {}).pop(period_key, None)
        elif key == "dashboard_source":
            st.session_state.pop(DASHBOARD_SOURCE_KEY, None)
            st.session_state.pop(DASHBOARD_SOURCE_HASH_KEY, None)
        else:
            st.session_state.pop(key, None)


def sync_file_cache(widget_key, cache_scope, result_keys):
    """يحافظ على النتائج عبر rerun ويمسحها فقط عند إزالة الملف أو تغييره."""
    uploaded = st.session_state.get(widget_key)
    cache = st.session_state.setdefault(APP_DATA_CACHE_KEY, {})
    previous = cache.get(cache_scope)
    if uploaded is None:
        _clear_cached_results(result_keys)
        cache.pop(cache_scope, None)
        return
    current_hash = uploaded_file_hash(uploaded)
    if previous and previous.get("file_hash") != current_hash:
        _clear_cached_results(result_keys)
    cache[cache_scope] = {"file_hash": current_hash, "filename": uploaded.name}

def init_neglect_state():
    if "neglect_sub_states" not in st.session_state:
        st.session_state["neglect_sub_states"] = NEGLECT_SUB_STATES_DEFAULT.copy()
    if "neglect_available_states" not in st.session_state:
        st.session_state["neglect_available_states"] = []
    if "neglect_mode" not in st.session_state:
        st.session_state["neglect_mode"] = "neglect"  # 'neglect' or 'followup'

# التاريخ المستهدف: تاريخ اليوم (بيتم تحديده مرة واحدة عند أول عرض للصفحة)
TODAY_KEY = "promises_today"


def _init_promises_today():
    """بنحدد تاريخ اليوم مرة واحدة في أول رن للصفحة لـ ميترفرش مع كل إعادة تشغيل."""
    if TODAY_KEY not in st.session_state:
        st.session_state[TODAY_KEY] = datetime.now().date()


def parse_date_cell(val):
    """بيحول خلية التاريخ لـ date مهما كان شكلها (datetime / date / نص)."""
    if pd.isna(val):
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, pd.Timestamp):
        return val.date()
    if hasattr(val, "date"):
        return val.date()
    txt = str(val).strip()
    if not txt:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d", "%d-%m-%Y", "%d %B %Y"):
        try:
            return datetime.strptime(txt, fmt).date()
        except (ValueError, TypeError):
            continue
    try:
        return pd.to_datetime(txt, errors="coerce").date()
    except Exception:
        return None


PROMISES_RESULT_KEY = "promises_result"  # كاش نتائج الوعود القائمة (فلتر: اليوم)
BROKEN_RESULT_KEY = "promises_broken_result"  # كاش نتائج الوعود المكسورة (فلتر: قبل اليوم)


def _run_promises_pipeline(
    uploaded,
    result_key,
    due_mode,
    count_label,
):
    """معالجة ملف المحفظة (فلترة + تجميع) وحفظ النتيجة في كاش الصفحة.

    - ``due_mode='today'``: Follow up Due Date == تاريخ اليوم (وعود قائمة).
    - ``due_mode='before'``: Follow up Due Date < تاريخ اليوم (وعود مكسورة).
    - ``count_label``: نص ملصق التجميع (مثلًا \"الوعود القائمة\" أو \"الوعود المكسورة\").

    بترجع ``True`` لو تم الحفظ في الكاش، و``False`` لو اتعرضت من الكاش مباشرة.
    """
    _file_hash = hashlib.sha256(uploaded.getvalue()).hexdigest()
    cached = st.session_state.get(result_key)
    if cached and cached.get("file_hash") == _file_hash:
        return False  # النتيجة موجودة في الكاش من رفع الملف ده — مش يلزم وجود إعادة معالجة

    try:
        raw_df = read_uploaded_dataframe(uploaded)
    except Exception as e:
        st.error(f"تعذر قراءة الملف: {e}")
        return False

    total_in_file = len(raw_df) - 1  # قبل حذف أول صف
    df = raw_df

    # 1) حذف أول صف بعد العناوين (زي قاعدة باقي الملفات في التطبيق)
    if len(df) > 0:
        df = df.iloc[1:].reset_index(drop=True)

    sales_col = find_column(df, SALES_PERSON_CANDIDATES)
    substate_col = find_column(df, PROMISE_SUB_STATE_CANDIDATES)
    duedate_col = find_column(df, PROMISE_DUE_DATE_CANDIDATES)
    net_col = find_column(df, PROMISE_NET_AMOUNT_CANDIDATES)

    missing = [n for n, c in [
        ("المحصّل (Salesperson)", sales_col),
        ("الحالة الفرعية (Sub State)", substate_col),
        ("تاريخ المتابعة (Follow up Due Date)", duedate_col),
    ] if not c]
    if missing:
        st.error(
            "تعذر العثور على أعمدة مهمة في الملف. الأعمدة المطلوبة: "
            f"{', '.join(missing)}\n\nالأعمدة الموجودة في الملف: {', '.join(df.columns.astype(str))}"
        )
        return False

    # 2) فلترة Salesperson — نستبعد المحصّلين المحددين
    sales_vals = df[sales_col].astype(str).str.strip()
    keep_sales = ~sales_vals.isin(PROMISE_EXCLUDED_SALES)
    dropped_sales = int((~keep_sales).sum())
    df = df[keep_sales].copy()

    # 3) فلترة Sub State = واعد بالسداد
    if substate_col:
        sub_vals = df[substate_col].astype(str).str.strip()
        keep_sub = sub_vals == PROMISE_SUB_STATE_VALUE
        dropped_sub = int((~keep_sub).sum())
        df = df[keep_sub].copy()
    else:
        dropped_sub = 0

    # 4) فلترة Follow up Due Date حسب وضع التبويبة
    target_date = st.session_state[TODAY_KEY]
    due_vals = pd.Series([parse_date_cell(v) for v in df[duedate_col]], index=df.index)
    if due_mode == "before":
        keep_due = due_vals.apply(lambda d: d is not None and d < target_date)
        due_desc = f"التاريخ قبل اليوم (< {target_date.strftime('%Y-%m-%d')})"
    else:
        keep_due = due_vals == target_date
        due_desc = f"التاريخ يساوي اليوم ({target_date.strftime('%Y-%m-%d')})"
    dropped_due = int((~keep_due).sum())
    df = df[keep_due].copy()

    # 5) الجدول التجميعي لكل محصّل
    if net_col and net_col in df.columns:
        summary_df = df.groupby(sales_col).agg(
            **{
                count_label: (duedate_col, "count"),
                "صافي المديونية (Net Amount)": (net_col, "sum"),
            }
        )
    else:
        summary_df = df.groupby(sales_col).agg(
            **{count_label: (duedate_col, "count")}
        )
    summary_df = summary_df.sort_values(count_label, ascending=False).reset_index()
    summary_df.columns = ["المحصّل " + str(sales_col), count_label] + (
        ["صافي المديونية (Net Amount)"] if net_col and net_col in df.columns else []
    )

    # 💾 حفظ النتائج في الكاش — تفضل موجودة لحد ما نعمل reload أو نشيل الملف
    st.session_state[result_key] = {
        "df": df,
        "summary_df": summary_df,
        "target_date": target_date,
        "filename": uploaded.name,
        "file_hash": _file_hash,
        "sales_col": sales_col,
        "substate_col": substate_col,
        "duedate_col": duedate_col,
        "net_col": net_col,
        "total_in_file": total_in_file,
        "dropped_sales": dropped_sales,
        "dropped_sub": dropped_sub,
        "dropped_due": dropped_due,
        "due_mode": due_mode,
        "due_desc": due_desc,
    }
    return True


PROMISES_COMPANY_KEY = "promises_selected_company"
PROMISES_COMPANY_OPTIONS = ["الوطنية للتأمين", "تري للتأمين"]
PROMISE_COMPANY_CANDIDATES = ["Company", "company", "Company Name", "اسم الشركة", "الشركة"]


def _build_promises_agent_summary(df, sales_col, net_col):
    if not sales_col or sales_col not in df.columns:
        return pd.DataFrame()
    summary = df.groupby(sales_col).size().reset_index(name="عدد الوعود")
    if net_col and net_col in df.columns:
        amounts = pd.to_numeric(df[net_col], errors="coerce").fillna(0)
        amount_summary = df.assign(_promise_amount=amounts).groupby(sales_col)["_promise_amount"].sum().reset_index(name="إجمالي المديونية")
        summary = summary.merge(amount_summary, on=sales_col, how="left")
    return summary.sort_values("عدد الوعود", ascending=False).reset_index(drop=True)


def render_promises_dashboard(df, summary, mode_label):
    sales_col = summary.get("sales_col")
    net_col = summary.get("net_col")
    render_promises_filter_notice()
    df = get_promises_view(df, sales_col)
    total = len(df)
    total_amount = pd.to_numeric(df[net_col], errors="coerce").fillna(0).sum() if net_col and net_col in df.columns else 0
    agent_summary = _build_promises_agent_summary(df, sales_col, net_col)
    agent_count = len(agent_summary)
    avg_amount = total_amount / total if total else 0

    st.subheader(f"📊 ملخص الوعود — {mode_label}")
    k1, k2, k3, k4 = st.columns(4)
    k1.metric("🤝 إجمالي الوعود", f"{total:,}")
    k2.metric("👥 عدد المحصّلين", f"{agent_count:,}")
    k3.metric("💰 إجمالي المديونية", f"{total_amount:,.0f}" if net_col else "—")
    k4.metric("📈 متوسط المديونية", f"{avg_amount:,.0f}" if net_col else "—")

    if total and sales_col and not agent_summary.empty:
        st.markdown("#### 📈 تحليلات الوعود التفاعلية")
        left, right = st.columns(2)
        with left:
            count_fig = px.bar(
                agent_summary.head(15).sort_values("عدد الوعود"),
                x="عدد الوعود",
                y=sales_col,
                orientation="h",
                text="عدد الوعود",
                color="عدد الوعود",
                color_continuous_scale=[THEME["surface_2"], COLOR_ACCENT],
                template=PLOTLY_TEMPLATE,
            )
            count_fig.update_layout(**PLOTLY_LAYOUT, title="عدد الوعود حسب المحصّل", xaxis_title="عدد الوعود", yaxis_title="", coloraxis_showscale=False, height=430)
            count_fig.update_traces(customdata=agent_summary.head(15).sort_values("عدد الوعود")[sales_col], hovertemplate="<b>%{y}</b><br>عدد الوعود: %{x:,}<extra></extra>")
            render_selectable_chart(count_fig, f"promises_count_{mode_label}", filter_key=PROMISES_AGENT_FILTER_KEY)
        with right:
            if net_col and "إجمالي المديونية" in agent_summary.columns:
                amount_fig = px.bar(
                    agent_summary.head(15).sort_values("إجمالي المديونية"),
                    x="إجمالي المديونية",
                    y=sales_col,
                    orientation="h",
                    text="إجمالي المديونية",
                    color="إجمالي المديونية",
                    color_continuous_scale=[COLOR_ACCENT, COLOR_WARN],
                    template=PLOTLY_TEMPLATE,
                )
                amount_fig.update_layout(**PLOTLY_LAYOUT, title="إجمالي المديونية حسب المحصّل", xaxis_title="إجمالي المديونية", yaxis_title="", coloraxis_showscale=False, height=430)
                amount_fig.update_traces(customdata=agent_summary.head(15).sort_values("إجمالي المديونية")[sales_col], hovertemplate="<b>%{y}</b><br>إجمالي المديونية: %{x:,.0f}<extra></extra>")
                render_selectable_chart(amount_fig, f"promises_amount_{mode_label}", filter_key=PROMISES_AGENT_FILTER_KEY)
            else:
                st.info("لا يوجد عمود صافي المديونية لعرض الرسم المالي.")

    display_cols = [c for c in [summary.get("sales_col"), summary.get("substate_col"), summary.get("duedate_col"), summary.get("net_col")] if c and c in df.columns]
    st.subheader(f"📋 بيانات الوعود — {mode_label}")
    if display_cols:
        st.dataframe(df[display_cols], use_container_width=True, hide_index=True)
    if not df.empty:
        out_excel = io.BytesIO()
        with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="الوعود")
        st.download_button(
            f"⬇️ تحميل بيانات الوعود — {mode_label}",
            data=out_excel.getvalue(),
            file_name=f"الوعود_{mode_label}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"promises_unified_download_{mode_label}",
            type="primary",
        )


def _filter_promises_by_company(df, company_label):
    """يفلتر عمود الشركة إن وُجد؛ وإذا لم يوجد فكل الملف يُعامل كملف الشركة المختارة."""
    company_col = find_column(df, PROMISE_COMPANY_CANDIDATES)
    if not company_col or company_col not in df.columns:
        return df.copy()
    short_name = "الوطنية" if "الوطنية" in company_label else "تري"
    values = df[company_col].astype(str).str.strip()
    mask = values.str.contains(company_label, case=False, na=False) | values.str.contains(short_name, case=False, na=False)
    return df.loc[mask].copy() if mask.any() else df.iloc[0:0].copy()


def _combine_promises_cached_results(company_label, result_keys=None):
    parts = []
    result_keys = result_keys or (PROMISES_RESULT_KEY, BROKEN_RESULT_KEY)
    for key, promise_type in zip(result_keys, ("الوعود القائمة", "الوعود المكسورة")):
        cached = st.session_state.get(key)
        if not cached or cached.get("df") is None:
            continue
        part = _filter_promises_by_company(cached["df"], company_label)
        if part.empty:
            continue
        part = part.copy()
        part["نوع الوعد"] = promise_type
        parts.append(part)
    if not parts:
        return pd.DataFrame(), None
    combined = pd.concat(parts, ignore_index=True, sort=False)
    meta = st.session_state.get(result_keys[0]) or st.session_state.get(result_keys[1])
    return combined, meta


def render_promises_kpi_dashboard(total, standing_count, broken_count, agent_count, total_amount):
    cards = [
        ("🤝<br>إجمالي الوعود", total, {"valueformat": ",d"}, THEME["text"]),
        ("📗<br>الوعود القائمة", standing_count, {"valueformat": ",d"}, COLOR_SUCCESS),
        ("📕<br>الوعود المكسورة", broken_count, {"valueformat": ",d"}, COLOR_FAIL),
        ("👥<br>عدد المحصّلين", agent_count, {"valueformat": ",d"}, THEME["text"]),
        ("💰<br>إجمالي المديونية", total_amount, {"valueformat": ",.0f"}, COLOR_WARN),
    ]
    figure = go.Figure()
    count = len(cards)
    gap = 0.018
    width = (1 - gap * (count + 1)) / count
    for index, (label, value, number_format, number_color) in enumerate(cards):
        x0 = gap + index * (width + gap)
        x1 = x0 + width
        figure.add_shape(
            type="path",
            xref="paper",
            yref="paper",
            path=_rounded_rect_path(x0, x1, 0.06, 0.94, radius=0.022),
            line={"color": THEME["border"], "width": 1},
            fillcolor=THEME["surface"],
            layer="below",
        )
        figure.add_trace(
            go.Indicator(
                mode="number",
                value=float(value or 0),
                domain={"x": [x0 + 0.012, x1 - 0.012], "y": [0.12, 0.88]},
                title={"text": label, "font": {"size": 18, "color": THEME["text_dim"]}, "align": "center"},
                number={"font": {"size": 32, "color": number_color}, **number_format},
            )
        )
    figure.update_layout(
        height=200,
        template=PLOTLY_TEMPLATE,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font={"family": "Tajawal, sans-serif", "color": THEME["text"]},
        margin={"t": 8, "b": 8, "l": 8, "r": 8},
    )
    st.plotly_chart(figure, use_container_width=True, config=PLOTLY_CONFIG, key="promises_kpi_dashboard")


def render_combined_promises_dashboard(df, meta, company_label):
    sales_col = meta.get("sales_col") if meta else None
    net_col = meta.get("net_col") if meta else None
    render_promises_filter_notice()
    df = get_promises_view(df, sales_col)
    total = len(df)
    standing_count = int((df["نوع الوعد"] == "الوعود القائمة").sum()) if "نوع الوعد" in df.columns else 0
    broken_count = int((df["نوع الوعد"] == "الوعود المكسورة").sum()) if "نوع الوعد" in df.columns else 0
    total_amount = pd.to_numeric(df[net_col], errors="coerce").fillna(0).sum() if net_col and net_col in df.columns else 0
    agent_count = int(df[sales_col].nunique()) if sales_col and sales_col in df.columns else 0

    st.subheader(f"📊 ملخص الوعود — {company_label}")
    render_promises_kpi_dashboard(total, standing_count, broken_count, agent_count, total_amount if net_col else 0)

    if total and sales_col and sales_col in df.columns:
        st.markdown("#### 📈 تحليلات الوعود القائمة والمكسورة")
        agent_type = df.groupby([sales_col, "نوع الوعد"]).size().reset_index(name="عدد الوعود")
        agent_order = agent_type.groupby(sales_col)["عدد الوعود"].sum().sort_values(ascending=False).head(15).index.tolist()
        agent_type = agent_type[agent_type[sales_col].isin(agent_order)]
        agent_type["_sort"] = agent_type[sales_col].map({name: i for i, name in enumerate(agent_order)})
        agent_type = agent_type.sort_values(["_sort", "نوع الوعد"], ascending=[True, True])
        left, right = st.columns(2)
        with left:
            fig = px.bar(
                agent_type,
                x="عدد الوعود",
                y=sales_col,
                color="نوع الوعد",
                barmode="group",
                orientation="h",
                text="عدد الوعود",
                color_discrete_map={"الوعود القائمة": COLOR_SUCCESS, "الوعود المكسورة": COLOR_FAIL},
                template=PLOTLY_TEMPLATE,
            )
            fig.update_layout(**{
                **PLOTLY_LAYOUT,
                "title": "القائمة والمكسورة حسب المحصّل",
                "xaxis_title": "عدد الوعود",
                "yaxis_title": "",
                "height": 500,
                "legend_title_text": "",
                "margin": dict(t=78, b=62, l=170, r=70),
                "xaxis": dict(tickformat=",.0f", automargin=True),
                "uniformtext_minsize": 12,
                "uniformtext_mode": "hide",
            })
            fig.update_traces(
                texttemplate="%{x:,.0f}",
                textposition="outside",
                textfont=dict(size=15, color=THEME["text"]),
                cliponaxis=False,
                customdata=agent_type[sales_col],
                hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:,.0f} وعد<extra></extra>",
            )
            render_selectable_chart(fig, "promises_combined_by_agent", filter_key=PROMISES_AGENT_FILTER_KEY)
        with right:
            if net_col and net_col in df.columns:
                amount_work = df.assign(_amount=pd.to_numeric(df[net_col], errors="coerce").fillna(0))
                agent_amount = amount_work.groupby([sales_col, "نوع الوعد"])["_amount"].sum().reset_index(name="إجمالي المديونية")
                agent_amount = agent_amount[agent_amount[sales_col].isin(agent_order)]
                fig = px.bar(
                    agent_amount,
                    x="إجمالي المديونية",
                    y=sales_col,
                    color="نوع الوعد",
                    barmode="group",
                    orientation="h",
                    text="إجمالي المديونية",
                    color_discrete_map={"الوعود القائمة": COLOR_SUCCESS, "الوعود المكسورة": COLOR_FAIL},
                    template=PLOTLY_TEMPLATE,
                )
                fig.update_layout(**{
                    **PLOTLY_LAYOUT,
                    "title": "إجمالي المديونية حسب المحصّل",
                    "xaxis_title": "إجمالي المديونية",
                    "yaxis_title": "",
                    "height": 500,
                    "legend_title_text": "",
                    "margin": dict(t=78, b=62, l=170, r=105),
                    "xaxis": dict(tickformat=",.0f", separatethousands=True, automargin=True),
                    "uniformtext_minsize": 11,
                    "uniformtext_mode": "hide",
                })
                fig.update_traces(
                    texttemplate="%{x:,.0f}",
                    textposition="outside",
                    textfont=dict(size=14, color=THEME["text"]),
                    cliponaxis=False,
                    customdata=agent_amount[sales_col],
                    hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:,.0f} جنيه<extra></extra>",
                )
                render_selectable_chart(fig, "promises_combined_amount_by_agent", filter_key=PROMISES_AGENT_FILTER_KEY)
            else:
                st.info("لا يوجد عمود صافي المديونية لعرض الرسم المالي.")

        type_counts = df["نوع الوعد"].value_counts().rename_axis("نوع الوعد").reset_index(name="عدد الوعود")
        fig = px.pie(
            type_counts,
            values="عدد الوعود",
            names="نوع الوعد",
            hole=0.55,
            color="نوع الوعد",
            color_discrete_map={"الوعود القائمة": COLOR_SUCCESS, "الوعود المكسورة": COLOR_FAIL},
            template=PLOTLY_TEMPLATE,
        )
        fig.update_layout(**{
            **PLOTLY_LAYOUT,
            "title": "توزيع الوعود القائمة والمكسورة",
            "height": 420,
            "legend_title_text": "",
            "margin": dict(t=78, b=45, l=35, r=35),
        })
        fig.update_traces(
            texttemplate="%{label}<br>%{value:,.0f} (%{percent:.1%})",
            textfont=dict(size=16, color=THEME["text"]),
            textinfo="text",
            hovertemplate="<b>%{label}</b><br>عدد الوعود: %{value:,.0f}<br>النسبة: %{percent:.1%}<extra></extra>",
        )
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG, key="promises_combined_type_share")

    if sales_col and sales_col in df.columns and "نوع الوعد" in df.columns:
        agent_promise_table = (
            df.groupby([sales_col, "نوع الوعد"]).size()
            .unstack(fill_value=0)
            .reset_index()
        )
        if "الوعود القائمة" not in agent_promise_table.columns:
            agent_promise_table["الوعود القائمة"] = 0
        if "الوعود المكسورة" not in agent_promise_table.columns:
            agent_promise_table["الوعود المكسورة"] = 0
        agent_promise_table["الإجمالي"] = agent_promise_table["الوعود القائمة"] + agent_promise_table["الوعود المكسورة"]
        agent_promise_table = agent_promise_table.rename(columns={sales_col: "المحصّل"})
        agent_promise_table = agent_promise_table[["المحصّل", "الوعود القائمة", "الوعود المكسورة", "الإجمالي"]].sort_values("الإجمالي", ascending=False)
        st.subheader("📊 ملخص الوعود حسب المحصّل")
        st.dataframe(agent_promise_table, use_container_width=True, hide_index=True)

    display_cols = [c for c in [sales_col, "نوع الوعد", meta.get("substate_col") if meta else None, meta.get("duedate_col") if meta else None, net_col] if c and c in df.columns]
    st.subheader("📋 تفاصيل الوعود القائمة والمكسورة")
    if display_cols:
        st.dataframe(df[display_cols], use_container_width=True, hide_index=True)

    report_date = datetime.now().strftime("%Y-%m-%d")
    company_file_name = company_label.replace(" ", "_")
    standing_df = df[df["نوع الوعد"] == "الوعود القائمة"].copy() if "نوع الوعد" in df.columns else pd.DataFrame()
    broken_df = df[df["نوع الوعد"] == "الوعود المكسورة"].copy() if "نوع الوعد" in df.columns else pd.DataFrame()

    def _excel_bytes(report_df, sheet_name):
        buffer = io.BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            report_df.to_excel(writer, index=False, sheet_name=sheet_name)
        return buffer.getvalue()

    download_left, download_right = st.columns(2)
    with download_left:
        st.download_button(
            "⬇️ تحميل تقرير الوعود القائمة",
            data=_excel_bytes(standing_df, "الوعود القائمة"),
            file_name=f"تقرير_الوعود_القائمة_{company_file_name}_{report_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="promises_standing_download",
            type="primary",
            disabled=standing_df.empty,
        )
    with download_right:
        st.download_button(
            "⬇️ تحميل تقرير الوعود المكسورة",
            data=_excel_bytes(broken_df, "الوعود المكسورة"),
            file_name=f"تقرير_الوعود_المكسورة_{company_file_name}_{report_date}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key="promises_broken_download",
            type="primary",
            disabled=broken_df.empty,
        )


def page_promises():
    """صفحة واحدة تجمع الوعود القائمة والمكسورة داخل محفظة الشركة المختارة فقط."""
    _init_promises_today()
    page_header(
        "PROMISES",
        "📚 الوعود",
        "لكل شركة محفظة مستقلة؛ اختر الشركة وارفع ملفها لعرض الوعود القائمة والمكسورة معًا",
    )
    company_label = st.radio("اختر الشركة", PROMISES_COMPANY_OPTIONS, horizontal=True, key=PROMISES_COMPANY_KEY)
    company_slug = "wataniya" if company_label == "الوطنية للتأمين" else "tary"
    standing_key = f"promises_{company_slug}_standing"
    broken_key = f"promises_{company_slug}_broken"
    upload_key = f"promises_upload_{company_slug}"
    cache_scope = f"promises_upload:{company_slug}"
    result_keys = (standing_key, broken_key)

    uploaded = st.file_uploader(
        f"📂 ارفع محفظة {company_label} (Excel أو CSV)",
        type=["xlsx", "xls", "csv"],
        key=upload_key,
        on_change=sync_file_cache,
        args=(upload_key, cache_scope, result_keys),
    )
    if uploaded is not None:
        st.caption(f"المحفظة المختارة: {company_label} · الملف: {uploaded.name}")
        _run_promises_pipeline(uploaded, standing_key, due_mode="today", count_label="عدد الوعود القائمة")
        _run_promises_pipeline(uploaded, broken_key, due_mode="before", count_label="عدد الوعود المكسورة")
    else:
        cached_file = st.session_state.get(APP_DATA_CACHE_KEY, {}).get(cache_scope, {})
        cached_result = st.session_state.get(standing_key) or st.session_state.get(broken_key)
        if cached_result or cached_file:
            saved_name = (cached_result or cached_file).get("filename", "المحفظة المحفوظة")
            st.success(f"✅ محفظة {company_label} محفوظة: {saved_name}. لن تُحذف عند التنقل بين التبويبات.")

    combined, meta = _combine_promises_cached_results(company_label, result_keys=result_keys)
    if combined.empty or meta is None:
        st.info(f"📂 ارفع محفظة {company_label} لعرض الوعود القائمة والمكسورة معًا.")
        return
    render_combined_promises_dashboard(combined, meta, company_label)


st.set_page_config(
    page_title="لوحة تحليل المكالمات",
    page_icon="🎙️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ==========================================================
# الهوية البصرية (Theme)
# ==========================================================
# ملحوظة مهمة: الـ direction: rtl متطبق بس على محتوى النص
# (الـ block-container والـ sidebar content) مش على هيكل الصفحة كله،
# لـ الـ Sidebar يفضل ثابت فعليًا على الشمال زي ما اتطلب،
# بدل ما ينقلب يمين بسبب انعكاس اتجاه الـ flex layout.


# ==========================================================
# إعدادات المظهر (Light / Dark Mode)
# ==========================================================
THEMES = {
    "dark": {
        "bg": "#0E1420",
        "bg_glow": "#10192C",
        "surface": "#151F30",
        "surface_2": "#1B2A42",
        "surface_3": "#111B2C",
        "surface_hover": "#1D304C",
        "sidebar_bg": "#0B111C",
        "accent_surface": "#122B2A",
        "accent": "#4F8C8D",
        "accent_strong": "#3D7375",
        "on_accent": "#FFFFFF",
        "success": "#4F927D",
        "danger": "#B96578",
        "warn": "#B39A5A",
        "text": "#F3F6FA",
        "text_dim": "#B9C6D6",
        "text_muted": "#9FB0C6",
        "placeholder": "#8FA3B8",
        "border": "rgba(15, 157, 138, 0.20)",
        "border_soft": "rgba(148, 163, 184, 0.18)",
        "input_bg": "#1B2A42",
        "chart_marker": "#0E1420",
        "chart_text": "#FFFFFF",
        "danger_soft": "#3B1420", "danger_text": "#FECDD3",
        "warn_soft": "#3A2A08", "warn_text": "#FDE68A",
    },
    "light": {
        "bg": "#F4F7FB",
        "bg_glow": "#EAF1F8",
        "surface": "#FFFFFF",
        "surface_2": "#EEF4F8",
        "surface_3": "#F7FAFC",
        "surface_hover": "#E2ECF3",
        "sidebar_bg": "#FFFFFF",
        "accent_surface": "#E6F7F4",
        "accent": "#397B7D",
        "accent_strong": "#2F6668",
        "on_accent": "#FFFFFF",
        "success": "#397D69",
        "danger": "#A65367",
        "warn": "#977D42",
        "text": "#172033",
        "text_dim": "#526174",
        "text_muted": "#6B7A8C",
        "placeholder": "#718096",
        "border": "rgba(8, 127, 112, 0.22)",
        "border_soft": "rgba(71, 85, 105, 0.18)",
        "input_bg": "#FFFFFF",
        "chart_marker": "#CBD5E1",
        "chart_text": "#172033",
        "danger_soft": "#FDE2E7", "danger_text": "#7F1D35",
        "warn_soft": "#FEF3C7", "warn_text": "#78350F",
    },
}

def _detect_native_streamlit_theme() -> str:
    """
    بيقرأ الوضع (Light/Dark) اللي المستخدم مختاره فعليًا من قائمة
    إعدادات Streamlit نفسها ("⋮" ← Settings ← Choose app theme)
    عن طريق st.context.theme.type (متاحة من Streamlit 1.46+).
    لو مش متاحة لأي سبب (نسخة قديمة من Streamlit)، بيرجع لآخر قيمة
    محفوظة في session_state، ولو لا يوجد هيستخدم "dark" كافتراضي.
    """
    try:
        ctx_theme = st.context.theme
        theme_type = getattr(ctx_theme, "type", None)
        if theme_type is None and hasattr(ctx_theme, "get"):
            theme_type = ctx_theme.get("type")
        if theme_type in ("light", "dark"):
            return theme_type
    except Exception:
        pass
    return st.session_state.get("theme_mode", "dark")


THEME_NAME = _detect_native_streamlit_theme()
st.session_state["theme_mode"] = THEME_NAME
THEME = THEMES.get(THEME_NAME, THEMES["dark"])


# Streamlit native theme is the source of truth for the app UI.
# Plotly receives the matching palette below; no custom CSS is injected.

def page_header(eyebrow: str, title: str, subtitle: str, centered: bool = False):
    if eyebrow:
        st.caption(eyebrow.upper())
    st.title(title)
    st.caption(subtitle)
    st.divider()


def find_column(df: pd.DataFrame, candidates: list):
    cols_lower = {c.lower().strip(): c for c in df.columns}
    for cand in candidates:
        if cand.lower().strip() in cols_lower:
            return cols_lower[cand.lower().strip()]
    return None


# ==========================================================
# منطق الموديل
# ==========================================================

@st.cache_resource(show_spinner="جارٍ تحميل النموذج من Hugging Face...")
def load_model():
    tokenizer = AutoTokenizer.from_pretrained(MODEL_REPO)
    model = AutoModelForSequenceClassification.from_pretrained(MODEL_REPO)
    device = "cuda" if torch.cuda.is_available() else "cpu"
    model.to(device)
    model.eval()
    return tokenizer, model, device


def predict_batch(texts, tokenizer, model, device, batch_size=16):
    all_preds, all_confidences = [], []
    progress_bar = st.progress(0, text="جارٍ التصنيف...")
    total = len(texts)

    for i in range(0, total, batch_size):
        batch = texts[i : i + batch_size]
        batch = [str(t) if pd.notna(t) and str(t).strip() != "" else "" for t in batch]

        inputs = tokenizer(
            batch, return_tensors="pt", truncation=True, padding=True, max_length=MAX_LENGTH
        ).to(device)

        with torch.no_grad():
            logits = model(**inputs).logits
            probs = torch.softmax(logits, dim=1)
            preds = torch.argmax(probs, dim=1)
            confidences = torch.max(probs, dim=1).values

        all_preds.extend(preds.cpu().tolist())
        all_confidences.extend(confidences.cpu().tolist())

        done = min(i + batch_size, total)
        pct = int(done / total * 100) if total else 100
        progress_bar.progress(
            done / total if total else 1.0,
            text=f"جارٍ التصنيف... {pct}%  ({done}/{total})",
        )

    progress_bar.empty()
    return all_preds, all_confidences


def remove_claim_duplicates(df, claim_col, time_col, classification_col=CLASSIFICATION_COL):
    """حذف تكرارات Claim وفق اليوم والتصنيف والوقت، مع الاحتفاظ بالصف الأحدث."""
    stats = {
        "input_rows": len(df),
        "output_rows": len(df),
        "removed_rows": 0,
        "duplicate_groups": 0,
        "success_priority_groups": 0,
        "non_success_window_groups": 0,
        "skipped_rows": 0,
    }
    if not claim_col or claim_col not in df.columns or not time_col or time_col not in df.columns:
        stats["skipped_rows"] = len(df)
        return df.copy(), stats
    if classification_col not in df.columns:
        stats["skipped_rows"] = len(df)
        return df.copy(), stats

    work = df.copy()
    work["_dedup_time"] = pd.to_datetime(work[time_col], errors="coerce")
    work["_dedup_claim"] = work[claim_col].astype("string").str.strip()
    work["_dedup_day"] = work["_dedup_time"].dt.date
    work["_dedup_order"] = range(len(work))
    valid = (
        work["_dedup_time"].notna()
        & work["_dedup_day"].notna()
        & work["_dedup_claim"].notna()
        & work["_dedup_claim"].ne("")
    )
    stats["skipped_rows"] = int((~valid).sum())
    keep_indices = set(work.index[~valid])
    valid_work = work.loc[valid]

    for (_, _), group in valid_work.groupby(["_dedup_claim", "_dedup_day"], sort=False):
        group = group.sort_values(["_dedup_time", "_dedup_order"])
        if len(group) == 1:
            keep_indices.add(group.index[0])
            continue

        stats["duplicate_groups"] += 1
        successful = pd.to_numeric(group[classification_col], errors="coerce").eq(1)
        if successful.any():
            # وجود ناجحة يلغي غير الناجحة، ونحتفظ بآخر إفادة ناجحة.
            keep_indices.add(group.loc[successful].index[-1])
            stats["success_priority_groups"] += 1
            continue

        # عند كون كل التكرارات غير ناجحة: نكوّن مجموعات متجاورة بفارق أقل من 20 دقيقة.
        cluster = [group.index[0]]
        previous_time = group.iloc[0]["_dedup_time"]
        for row_index, row in group.iloc[1:].iterrows():
            current_time = row["_dedup_time"]
            if current_time - previous_time < pd.Timedelta(minutes=DUPLICATE_WINDOW_MINUTES):
                cluster.append(row_index)
            else:
                keep_indices.add(cluster[-1])
                cluster = [row_index]
            previous_time = current_time
        keep_indices.add(cluster[-1])
        stats["non_success_window_groups"] += 1

    result = df.loc[sorted(keep_indices, key=lambda index: work.loc[index, "_dedup_order"])].copy()
    stats["output_rows"] = len(result)
    stats["removed_rows"] = stats["input_rows"] - stats["output_rows"]
    return result.reset_index(drop=True), stats


def render_duplicate_summary(stats):
    """عرض نتيجة تنظيف التكرارات بعد التصنيف."""
    if not stats:
        return
    if stats.get("removed_rows", 0) > 0:
        st.success(
            f"🧹 تم حذف {stats['removed_rows']:,} تكرار من أصل {stats['input_rows']:,} صف "
            f"والاحتفاظ بـ {stats['output_rows']:,} صف."
        )
        st.caption(
            f"مجموعات بأولوية ناجحة: {stats.get('success_priority_groups', 0)} · "
            f"مجموعات غير ناجحة ضمن نافذة {DUPLICATE_WINDOW_MINUTES} دقيقة: "
            f"{stats.get('non_success_window_groups', 0)}"
        )
    elif stats.get("skipped_rows", 0) == stats.get("input_rows", 0):
        st.warning("لم تُطبَّق إزالة التكرارات: يلزم وجود عمود Claim وعمود التاريخ والوقت.")
    else:
        st.info("لم يتم العثور على تكرارات مطابقة وفق قواعد Claim واليوم والتصنيف.")


# ==========================================================
# حساب الوقت المهدر بين المكالمات لكل محصّل
# ==========================================================

def subtract_break_overlap(prev_time, curr_time, break_start, break_end, gap_minutes):
    """بيخصم من الفجوة أي جزء واقع جوه وقت الاستراحة المحدد."""
    if break_start is None or break_end is None or pd.isna(prev_time) or pd.isna(curr_time):
        return gap_minutes
    day = prev_time.date()
    break_start_dt = pd.Timestamp.combine(day, break_start)
    break_end_dt = pd.Timestamp.combine(day, break_end)
    overlap_start = max(prev_time, break_start_dt)
    overlap_end = min(curr_time, break_end_dt)
    overlap_minutes = max((overlap_end - overlap_start).total_seconds() / 60, 0)
    return max(gap_minutes - overlap_minutes, 0)


def calculate_wasted_time(df, sales_col, time_col, break_start, break_end):
    """
    بتحسب الوقت المهدر (بالدقايق) بين كل مكالمة واللي قبلها لنفس المحصّل،
    بعد استبعاد وقت الاستراحة. أول مكالمة لكل محصّل = صفر (لا يوجد مكالمة قبلها نقيس منها).
    """
    work = df.copy()
    work["_orig_idx"] = work.index
    work[time_col] = pd.to_datetime(work[time_col], errors="coerce")
    work = work.sort_values([sales_col, time_col])
    work["_prev_time"] = work.groupby(sales_col)[time_col].shift(1)

    def compute_row(row):
        if pd.isna(row["_prev_time"]) or pd.isna(row[time_col]):
            return 0.0
        gap_min = (row[time_col] - row["_prev_time"]).total_seconds() / 60
        gap_min = subtract_break_overlap(row["_prev_time"], row[time_col], break_start, break_end, gap_min)
        return round(max(gap_min, 0.0), 1)

    work[WASTED_TIME_COL] = work.apply(compute_row, axis=1)
    work = work.set_index("_orig_idx").sort_index()
    df[WASTED_TIME_COL] = work[WASTED_TIME_COL]
    return df


# Tables use Streamlit's native rendering; no cell-level CSS is injected.

# ==========================================================
# داشبورد مشترك (يُستخدم بعد التصنيف مباشرة، وكمان في تويب الداشبورد)
# ==========================================================

# لوحة ألوان موحّدة للداشبورد كله — تتبدل حسب المود المختار
COLOR_SUCCESS = THEMES[THEME_NAME]["success"]
COLOR_FAIL = THEMES[THEME_NAME]["danger"]
COLOR_ACCENT = THEMES[THEME_NAME]["accent"]
COLOR_WARN = THEMES[THEME_NAME]["warn"]
CHART_COLORS = {"ناجحة": COLOR_SUCCESS, "غير ناجحة": COLOR_FAIL}
# Palette النشاط: ثلاث عائلات لونية هادئة فقط بدرجات متقاربة.
ACTIVITY_AGENT_PALETTE = [
    "#2F6F73", "#477F82", "#628B8E", "#7D9A9D", "#98AEB2", "#B2C1C3",
    "#5F7D8C", "#8095A2", "#A6B4B9",
]
ACTIVITY_STATE_PALETTE = ["#2F6F73", "#628B8E", "#8095A2", "#A6B4B9"]
ACTIVITY_OUTCOME_COLORS = {"ناجحة": "#2F6F73", "غير ناجحة": "#8095A2"}


def _activity_agent_color_map(values):
    names = sorted({str(value) for value in values if pd.notna(value)})
    return {name: ACTIVITY_AGENT_PALETTE[index % len(ACTIVITY_AGENT_PALETTE)] for index, name in enumerate(names)}


PLOTLY_TEMPLATE = "plotly_dark" if THEME_NAME == "dark" else "plotly_white"
PLOTLY_LAYOUT = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(0,0,0,0)",
    font_color=THEME["text_dim"],
    font_family="Tajawal, sans-serif",
    font_size=13,
    margin=dict(t=60, b=50, l=50, r=20),
    title_font_size=18,
    legend_font_size=12,
    hovermode="closest",
    hoverlabel=dict(
        bgcolor=THEME["surface"],
        bordercolor=THEME["border"],
        font=dict(family="Tajawal, sans-serif", size=13, color=THEME["text"]),
    ),
)
PLOTLY_CONFIG = {
    "displayModeBar": True,
    "displaylogo": False,
    "responsive": True,
    "scrollZoom": True,
    "doubleClick": "reset+autosize",
    "toImageButtonOptions": {
        "format": "png",
        "filename": "classification_chart",
        "height": 900,
        "width": 1500,
        "scale": 2,
    },
}


CLASSIFICATION_AGENT_FILTER_KEY = "classification_selected_agent"
PROMISES_AGENT_FILTER_KEY = "promises_selected_agent"


def _event_value(item, key, default=None):
    if isinstance(item, dict):
        return item.get(key, default)
    return getattr(item, key, default)


def _extract_selected_agent(event, fig):
    selection = _event_value(event, "selection")
    points = _event_value(selection, "points", []) if selection is not None else []
    if not points:
        return None
    point = points[0]
    selected = _event_value(point, "customdata")
    if isinstance(selected, (list, tuple)):
        selected = selected[0] if selected else None
    if selected is None:
        curve_number = _event_value(point, "curve_number", _event_value(point, "curveNumber", 0))
        point_index = _event_value(point, "point_index", _event_value(point, "pointNumber"))
        if point_index is None or curve_number >= len(fig.data):
            return None
        trace = fig.data[curve_number]
        orientation = getattr(trace, "orientation", None)
        values = trace.y if orientation == "h" else trace.x
        if values is not None and point_index < len(values):
            selected = values[point_index]
    return str(selected).strip() if selected is not None else None


def render_selectable_chart(fig, key, filter_key=CLASSIFICATION_AGENT_FILTER_KEY):
    """عرض رسم Plotly مع التقاط اختيار محصّل وإعادة تشغيل الصفحة لتطبيق الفلتر."""
    try:
        event = st.plotly_chart(
            fig,
            use_container_width=True,
            config=PLOTLY_CONFIG,
            key=key,
            on_select="rerun",
            selection_mode=("points",),
        )
    except TypeError:
        # توافق مع إصدارات Streamlit القديمة التي لا تدعم on_select.
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG, key=key)
        return
    selected = _extract_selected_agent(event, fig)
    if selected:
        current = st.session_state.get(filter_key)
        if selected != current:
            st.session_state[filter_key] = selected
            st.rerun()


def get_promises_view(df, sales_col):
    if not sales_col or sales_col not in df.columns:
        return df
    selected = st.session_state.get(PROMISES_AGENT_FILTER_KEY)
    if not selected:
        return df
    mask = df[sales_col].astype(str).str.strip().eq(str(selected).strip())
    if not mask.any():
        st.session_state.pop(PROMISES_AGENT_FILTER_KEY, None)
        return df
    return df.loc[mask].copy()


def render_promises_filter_notice():
    selected = st.session_state.get(PROMISES_AGENT_FILTER_KEY)
    if not selected:
        return
    c1, c2 = st.columns([4, 1])
    with c1:
        st.info(f"🎯 الفلتر النشط: عرض كل ملخصات الوعود للمحصّل «{selected}»")
    with c2:
        if st.button("إظهار الكل", key="clear_promises_agent_filter", use_container_width=True):
            st.session_state.pop(PROMISES_AGENT_FILTER_KEY, None)
            st.rerun()


def get_classification_view(df, sales_col):
    """تطبيق المحصّل المختار على بيانات التصنيف مع إبقاء العرض كاملًا افتراضيًا."""
    if not sales_col or sales_col not in df.columns:
        return df
    selected = st.session_state.get(CLASSIFICATION_AGENT_FILTER_KEY)
    if not selected:
        return df
    mask = df[sales_col].astype(str).str.strip().eq(str(selected).strip())
    if not mask.any():
        st.session_state.pop(CLASSIFICATION_AGENT_FILTER_KEY, None)
        return df
    return df.loc[mask].copy()


def render_classification_filter_notice(df, sales_col):
    selected = st.session_state.get(CLASSIFICATION_AGENT_FILTER_KEY)
    if not selected:
        return
    c1, c2 = st.columns([4, 1])
    with c1:
        st.info(f"🎯 الفلتر النشط: عرض كل مؤشرات ورسوم المحصّل «{selected}»")
    with c2:
        if st.button("إظهار الكل", key="clear_classification_agent_filter", use_container_width=True):
            st.session_state.pop(CLASSIFICATION_AGENT_FILTER_KEY, None)
            st.rerun()


DASHBOARD_AGENT_FILTER_KEY = "dashboard_selected_agent"

ACTIVITY_NO_ANSWER_STATES = [
    "لا يرد",
    "لا يرد مع التكرار",
    "مغلق",
    "مغلق مع التكرار",
]
ACTIVITY_PROMISE_STATES = ["واعد بالسداد"]
ACTIVITY_PAYMENT_STATES = [
    "سدد كامل المديونية",
    "جدولة",
    "جدولة مقفلة",
    "سدد كامل المديونية بخصم",
]


def _state_key(value):
    """توحيد Sub State للتعامل مع اختلاف المسافات والهمزات والصياغة."""
    text = "" if pd.isna(value) else str(value).strip().casefold()
    text = (
        text.replace("أ", "ا")
        .replace("إ", "ا")
        .replace("آ", "ا")
        .replace("ى", "ي")
        .replace("ة", "ه")
    )
    return re.sub(r"[\s_-]+", "", text)


def _classify_activity_sub_state(value):
    """إرجاع مجموعة موحدة للحالة الفرعية المطلوبة في Dashboard النشاط."""
    key = _state_key(value)
    if not key:
        return "غير محدد"
    if "لايرد" in key and "تكرار" in key:
        return "لا يرد مع التكرار"
    if "مغلق" in key and "تكرار" in key:
        return "مغلق مع التكرار"
    if "لايرد" in key:
        return "لا يرد"
    if key == "مغلق" or ("مغلق" in key and "جدوله" not in key):
        return "مغلق"
    if "واعد" in key and "سداد" in key:
        return "واعد بالسداد"
    if "خصم" in key and "كامل" in key and "مديون" in key:
        return "سدد كامل المديونية بخصم"
    if ("سدد" in key or "سداد" in key) and "كامل" in key and "مديون" in key:
        return "سدد كامل المديونية"
    if "جدوله" in key and ("مغلق" in key or "مقفل" in key):
        return "جدولة مقفلة"
    if "جدوله" in key:
        return "جدولة"
    return "أخرى"


def _activity_success_mask(df, class_col):
    if not class_col or class_col not in df.columns:
        return pd.Series(False, index=df.index)
    numeric = pd.to_numeric(df[class_col], errors="coerce")
    text = df[class_col].astype(str).str.strip().str.casefold()
    text_mask = text.isin({"1", "true", "yes", "ناجحة", "ناجحه", "successful"})
    return numeric.eq(1) | text_mask


def _calculate_wasted_time_by_day(df, sales_col, time_col, break_start=None, break_end=None):
    """حساب الفجوات بين مكالمات المحصل داخل كل يوم فقط، مع خصم البريك."""
    if not sales_col or sales_col not in df.columns or not time_col or time_col not in df.columns:
        return pd.Series(0.0, index=df.index)
    work = df[[sales_col, time_col]].copy()
    work["_calc_time"] = pd.to_datetime(work[time_col], errors="coerce")
    work["_calc_day"] = work["_calc_time"].dt.date
    work["_orig_idx"] = work.index
    work = work.sort_values([sales_col, "_calc_day", "_calc_time"])
    work["_prev_time"] = work.groupby([sales_col, "_calc_day"])["_calc_time"].shift(1)

    def gap_for_row(row):
        if pd.isna(row["_prev_time"]) or pd.isna(row["_calc_time"]):
            return 0.0
        gap = max((row["_calc_time"] - row["_prev_time"]).total_seconds() / 60, 0.0)
        return round(subtract_break_overlap(row["_prev_time"], row["_calc_time"], break_start, break_end, gap), 1)

    work[WASTED_TIME_COL] = work.apply(gap_for_row, axis=1)
    return work.set_index("_orig_idx")[WASTED_TIME_COL].reindex(df.index).fillna(0.0)


def _calculate_daily_work_hours(df, sales_col, time_col, break_start=None, break_end=None):
    """حساب فترة نشاط كل محصل في كل يوم ثم خصم الجزء المتداخل مع البريك."""
    empty = pd.DataFrame(columns=["المحصّل", "أيام النشاط", "متوسط ساعات العمل/اليوم", "إجمالي ساعات العمل"])
    if not sales_col or sales_col not in df.columns or not time_col or time_col not in df.columns:
        return empty
    work = pd.DataFrame({
        "المحصّل": df[sales_col].fillna("غير محدد").astype(str).str.strip(),
        "_activity_time": pd.to_datetime(df[time_col], errors="coerce"),
    }).dropna(subset=["_activity_time"])
    if work.empty:
        return empty
    work["_activity_day"] = work["_activity_time"].dt.date
    daily = work.groupby(["المحصّل", "_activity_day"], as_index=False)["_activity_time"].agg(
        بداية="min", نهاية="max"
    )

    def net_minutes(row):
        minutes = max((row["نهاية"] - row["بداية"]).total_seconds() / 60, 0.0)
        if break_start is not None and break_end is not None:
            minutes = subtract_break_overlap(row["بداية"], row["نهاية"], break_start, break_end, minutes)
        return max(minutes, 0.0)

    daily["دقائق العمل"] = daily.apply(net_minutes, axis=1)
    summary = daily.groupby("المحصّل")["دقائق العمل"].agg(["count", "mean", "sum"]).reset_index()
    summary = summary.rename(columns={"count": "أيام النشاط"})
    summary["متوسط ساعات العمل/اليوم"] = (summary["mean"] / 60).round(2)
    summary["إجمالي ساعات العمل"] = (summary["sum"] / 60).round(2)
    return summary[["المحصّل", "أيام النشاط", "متوسط ساعات العمل/اليوم", "إجمالي ساعات العمل"]]


def _build_activity_summary(df, class_col, sales_col, time_col, break_start=None, break_end=None):
    """بناء جدول مؤشرات المحصلين ومجموعة بيانات الرسوم من ملف النشاط."""
    if not sales_col or sales_col not in df.columns:
        return pd.DataFrame(), df.copy(), None
    work = df.copy()
    work["_agent_display"] = work[sales_col].fillna("غير محدد").astype(str).str.strip()
    work.loc[work["_agent_display"] == "", "_agent_display"] = "غير محدد"
    work["_success_bool"] = _activity_success_mask(work, class_col)
    sub_col = find_column(work, PROMISE_SUB_STATE_CANDIDATES)
    if sub_col:
        work["_activity_state"] = work[sub_col].map(_classify_activity_sub_state)
    else:
        work["_activity_state"] = "غير محدد"

    if time_col and time_col in work.columns and WASTED_TIME_COL not in work.columns:
        work[WASTED_TIME_COL] = _calculate_wasted_time_by_day(
            work, "_agent_display", time_col, break_start, break_end
        )
    if WASTED_TIME_COL in work.columns:
        work[WASTED_TIME_COL] = pd.to_numeric(work[WASTED_TIME_COL], errors="coerce").fillna(0)

    agent = work.groupby("_agent_display", dropna=False).size().rename("إجمالي المكالمات").to_frame()
    agent["المكالمات الناجحة"] = work[work["_success_bool"]].groupby("_agent_display").size()
    agent["المكالمات الناجحة"] = agent["المكالمات الناجحة"].fillna(0).astype(int)
    agent["المكالمات غير الناجحة"] = agent["إجمالي المكالمات"] - agent["المكالمات الناجحة"]
    agent["نسبة النجاح (%)"] = (
        agent["المكالمات الناجحة"] / agent["إجمالي المكالمات"].replace(0, pd.NA) * 100
    ).fillna(0).round(1)

    state_columns = ACTIVITY_NO_ANSWER_STATES + ACTIVITY_PROMISE_STATES + ACTIVITY_PAYMENT_STATES
    state_table = pd.crosstab(work["_agent_display"], work["_activity_state"])
    for state in state_columns:
        if state not in state_table.columns:
            state_table[state] = 0
    state_table = state_table.reindex(columns=state_columns, fill_value=0)
    agent = agent.join(state_table, how="left").fillna(0)
    agent["إجمالي لا يرد"] = agent[ACTIVITY_NO_ANSWER_STATES].sum(axis=1).astype(int)
    agent["نسبة من إجمالي المكالمات (%)"] = (
        agent["إجمالي المكالمات"] / max(len(work), 1) * 100
    ).round(1)

    if WASTED_TIME_COL in work.columns:
        agent["إجمالي الوقت المهدر (دقيقة)"] = work.groupby("_agent_display")[WASTED_TIME_COL].sum().round(1)
    else:
        agent["إجمالي الوقت المهدر (دقيقة)"] = 0.0

    hours = _calculate_daily_work_hours(work, "_agent_display", time_col, break_start, break_end)
    if not hours.empty:
        agent = agent.reset_index().rename(columns={"_agent_display": "المحصّل"}).merge(hours, on="المحصّل", how="left").set_index("المحصّل")
    else:
        agent["أيام النشاط"] = 0
        agent["متوسط ساعات العمل/اليوم"] = 0.0
        agent["إجمالي ساعات العمل"] = 0.0

    agent = agent.reset_index().rename(columns={"_agent_display": "المحصّل"})
    return agent.fillna(0), work, sub_col


def _activity_layout(**overrides):
    return {**PLOTLY_LAYOUT, **overrides}


def render_activity_kpi_cards(total, success, agent_count, success_rate, wasted_minutes):
    cards = [
        ("👥<br>عدد المحصّلين", agent_count, {"valueformat": ",d"}, THEME["text"]),
        ("📞<br>إجمالي المكالمات", total, {"valueformat": ",d"}, THEME["text"]),
        ("✅<br>المكالمات الناجحة", success, {"valueformat": ",d"}, ACTIVITY_OUTCOME_COLORS["ناجحة"]),
        ("📈<br>نسبة النجاح", success_rate, {"valueformat": ".1f", "suffix": "%"}, ACTIVITY_AGENT_PALETTE[1]),
        ("⏱️<br>إجمالي الوقت المهدر", wasted_minutes, {"valueformat": ".1f", "suffix": " دقيقة"}, ACTIVITY_AGENT_PALETTE[3]),
    ]
    figure = go.Figure()
    gap = 0.014
    width = (1 - gap * (len(cards) + 1)) / len(cards)
    for index, (label, value, number_format, color) in enumerate(cards):
        x0 = gap + index * (width + gap)
        x1 = x0 + width
        figure.add_shape(
            type="path",
            path=_rounded_rect_path(x0, x1, 0.04, 0.96, radius=0.022),
            xref="paper", yref="paper", layer="below",
            fillcolor=THEME["surface"], line={"color": THEME["border"], "width": 1},
        )
        figure.add_trace(go.Indicator(
            mode="number", value=float(value or 0),
            domain={"x": [x0 + 0.008, x1 - 0.008], "y": [0.13, 0.87]},
            title={"text": label, "font": {"size": 16, "color": THEME["text_dim"]}, "align": "center"},
            number={"font": {"size": 28, "color": color}, **number_format},
        ))
    figure.update_layout(
        height=205, template=PLOTLY_TEMPLATE, paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)", font={"family": "Tajawal, sans-serif", "color": THEME["text"]},
        margin={"t": 8, "b": 8, "l": 8, "r": 8},
    )
    st.plotly_chart(figure, use_container_width=True, config=PLOTLY_CONFIG, key="activity_kpi_cards")


def _render_dashboard_agent_filter_notice():
    selected = st.session_state.get(DASHBOARD_AGENT_FILTER_KEY)
    if not selected:
        return
    c1, c2 = st.columns([4, 1])
    with c1:
        st.info(f"🎯 الفلتر التفاعلي النشط: كل المؤشرات للمحصّل «{selected}»")
    with c2:
        if st.button("إظهار الكل", key="clear_dashboard_agent_filter", use_container_width=True):
            st.session_state.pop(DASHBOARD_AGENT_FILTER_KEY, None)
            st.rerun()


def _dashboard_activity_view(df, sales_col):
    selected = st.session_state.get(DASHBOARD_AGENT_FILTER_KEY)
    if not selected or not sales_col or sales_col not in df.columns:
        return df
    mask = df[sales_col].fillna("غير محدد").astype(str).str.strip().eq(str(selected).strip())
    if not mask.any():
        st.session_state.pop(DASHBOARD_AGENT_FILTER_KEY, None)
        return df
    return df.loc[mask].copy()


def _render_activity_daily_chart(work, time_col, class_col=None):
    if not time_col or time_col not in work.columns:
        st.info("يلزم وجود عمود Created On لعرض النشاط على مدار الأيام.")
        return
    trend = work.copy()
    trend["_activity_time"] = pd.to_datetime(trend[time_col], errors="coerce")
    trend = trend.dropna(subset=["_activity_time"])
    if trend.empty:
        st.info("لا توجد تواريخ صالحة لعرض النشاط اليومي.")
        return
    trend["اليوم"] = trend["_activity_time"].dt.strftime("%Y-%m-%d")
    trend["_success_for_day"] = _activity_success_mask(trend, class_col)
    daily = trend.groupby(["اليوم", "_agent_display"], as_index=False).agg(
        **{"عدد المكالمات": ("_agent_display", "size"), "المكالمات الناجحة": ("_success_for_day", "sum")}
    )
    daily_totals = daily.groupby("اليوم", as_index=False).agg(
        **{"إجمالي المكالمات": ("عدد المكالمات", "sum"), "إجمالي الناجحة": ("المكالمات الناجحة", "sum")}
    )
    daily_totals["نسبة النجاح (%)"] = (
        daily_totals["إجمالي الناجحة"] / daily_totals["إجمالي المكالمات"].replace(0, pd.NA) * 100
    ).fillna(0).round(1)
    sort_options = {
        "التاريخ تصاعديًا": "date",
        "إجمالي المكالمات تنازليًا": "calls",
        "نسبة النجاح تنازليًا": "success_rate",
    }
    sort_label = st.selectbox(
        "ترتيب الـ Histogram اليومي",
        list(sort_options.keys()),
        key="activity_daily_sort_v1",
        help="الترتيب يغيّر ترتيب الأيام على المحور الأفقي فقط.",
    )
    sort_mode = sort_options[sort_label]
    if sort_mode == "calls":
        ordered_days = daily_totals.sort_values(["إجمالي المكالمات", "اليوم"], ascending=[False, True])["اليوم"].tolist()
    elif sort_mode == "success_rate":
        ordered_days = daily_totals.sort_values(["نسبة النجاح (%)", "إجمالي المكالمات", "اليوم"], ascending=[False, False, True])["اليوم"].tolist()
    else:
        ordered_days = sorted(daily_totals["اليوم"].tolist())
    daily["اليوم"] = pd.Categorical(daily["اليوم"], categories=ordered_days, ordered=True)
    daily = daily.sort_values(["اليوم", "_agent_display"])
    daily_totals["اليوم"] = pd.Categorical(daily_totals["اليوم"], categories=ordered_days, ordered=True)
    daily_totals = daily_totals.sort_values("اليوم")

    fig = px.bar(
        daily, x="اليوم", y="عدد المكالمات", color="_agent_display", barmode="group",
        text_auto=True, custom_data=["_agent_display"], template=PLOTLY_TEMPLATE,
        labels={"_agent_display": "المحصّل"}, color_discrete_sequence=ACTIVITY_AGENT_PALETTE,
    )
    fig.add_trace(go.Scatter(
        x=daily_totals["اليوم"].astype(str), y=daily_totals["نسبة النجاح (%)"],
        name="نسبة النجاح", mode="lines+markers+text", text=daily_totals["نسبة النجاح (%)"].map(lambda value: f"{value:.1f}%"),
        textposition="top center", line={"color": ACTIVITY_AGENT_PALETTE[0], "width": 3},
        marker={"color": ACTIVITY_AGENT_PALETTE[0], "size": 9, "line": {"color": THEME["surface"], "width": 2}},
        yaxis="y2", customdata=[[""] for _ in range(len(daily_totals))],
        hovertemplate="<b>%{x}</b><br>نسبة النجاح: %{y:.1f}%<extra></extra>",
    ))
    fig.update_layout(**_activity_layout(
        title="📊 Combo Chart يومي: المكالمات ونسبة النجاح", title_x=0.5,
        xaxis_title="اليوم", yaxis_title="عدد المكالمات", height=400, bargap=0.14,
        legend_title_text="", hovermode="x unified",
        legend={"orientation": "h", "yanchor": "top", "y": -0.16, "x": 0.5, "xanchor": "center"},
        margin={"t": 62, "b": 78, "l": 50, "r": 55},
        xaxis={"type": "category", "categoryorder": "array", "categoryarray": ordered_days, "tickangle": -25},
        yaxis={"title": "عدد المكالمات", "rangemode": "tozero"},
        yaxis2={"title": "نسبة النجاح (%)", "overlaying": "y", "side": "right", "range": [0, 100], "ticksuffix": "%", "showgrid": False},
    ))
    fig.update_traces(selector={"type": "bar"}, marker_line_width=0, hovertemplate="<b>%{x}</b><br>%{fullData.name}: %{y:,} مكالمة<extra></extra>")
    render_selectable_chart(fig, "dashboard_activity_daily", filter_key=DASHBOARD_AGENT_FILTER_KEY)


def _render_activity_hourly_chart(work, time_col):
    if not time_col or time_col not in work.columns:
        st.info("يلزم وجود عمود Created On لعرض النشاط حسب الساعة.")
        return
    trend = work.copy()
    trend["_activity_time"] = pd.to_datetime(trend[time_col], errors="coerce")
    trend = trend.dropna(subset=["_activity_time"])
    if trend.empty:
        st.info("لا توجد أوقات صالحة لعرض النشاط الساعي.")
        return
    trend["الساعة"] = trend["_activity_time"].dt.hour
    hour_min = int(trend["الساعة"].min())
    hour_max = int(trend["الساعة"].max())
    hourly = trend.groupby(["الساعة", "_agent_display"], as_index=False).size().rename(columns={"size": "عدد المكالمات"})
    fig = px.bar(
        hourly, x="الساعة", y="عدد المكالمات", color="_agent_display", barmode="stack",
        text_auto=True, custom_data=["_agent_display"], template=PLOTLY_TEMPLATE,
        labels={"_agent_display": "المحصّل"}, color_discrete_sequence=ACTIVITY_AGENT_PALETTE,
    )
    fig.update_layout(**_activity_layout(
        title="🕒 Histogram ساعي لنشاط المحصلين", title_x=0.5, xaxis_title="ساعة اليوم", yaxis_title="عدد المكالمات",
        xaxis={"dtick": 1, "tickvals": list(range(hour_min, hour_max + 1)), "range": [max(-0.5, hour_min - 0.5), min(23.5, hour_max + 0.5)]}, height=400, bargap=0.06, legend_title_text="",
        legend={"orientation": "h", "yanchor": "top", "y": -0.16, "x": 0.5, "xanchor": "center"},
        margin={"t": 62, "b": 78, "l": 50, "r": 16},
    ))
    fig.update_traces(
        marker_line_width=0,
        customdata=trend["_agent_display"],
        hovertemplate="<b>الساعة %{x}:00</b><br>%{fullData.name}: %{y:,} مكالمة<extra></extra>",
    )
    render_selectable_chart(fig, "dashboard_activity_hourly", filter_key=DASHBOARD_AGENT_FILTER_KEY)


def _render_activity_outcome_donut(work, class_col):
    if not class_col or class_col not in work.columns:
        st.info("يلزم وجود عمود التصنيف لعرض الناجحة مقابل غير الناجحة.")
        return
    success = int(work["_success_bool"].sum())
    failed = int(len(work) - success)
    donut_df = pd.DataFrame({"النتيجة": ["ناجحة", "غير ناجحة"], "العدد": [success, failed]})
    rate = success / len(work) * 100 if len(work) else 0
    fig = px.pie(
        donut_df, names="النتيجة", values="العدد", hole=0.62,
        color="النتيجة", color_discrete_map=ACTIVITY_OUTCOME_COLORS, template=PLOTLY_TEMPLATE,
    )
    fig.update_traces(
        textinfo="percent", textfont_size=15,
        marker={"line": {"color": THEME["surface"], "width": 3}},
        hovertemplate="<b>%{label}</b><br>العدد: %{value:,}<br>النسبة: %{percent}<extra></extra>",
    )
    fig.update_layout(**_activity_layout(
        title="🎯 الناجحة مقابل غير الناجحة", title_x=0.5, height=400,
        legend={"orientation": "h", "yanchor": "top", "y": -0.12, "x": 0.5, "xanchor": "center"},
        margin={"t": 62, "b": 62, "l": 16, "r": 16},
        annotations=[{"text": f"{rate:.1f}%<br>نجاح", "x": 0.5, "y": 0.5, "font": {"size": 22, "color": COLOR_SUCCESS}, "showarrow": False}],
    ))
    st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG, key="dashboard_outcome_donut")


def _render_activity_no_answer_chart(agent):
    available = [state for state in ACTIVITY_NO_ANSWER_STATES if state in agent.columns]
    if not available:
        st.info("يلزم وجود عمود Sub State لعرض حالات لا يرد ومغلق.")
        return
    plot = agent[["المحصّل"] + available].copy()
    plot["إجمالي لا يرد"] = plot[available].sum(axis=1)
    plot = plot.sort_values("إجمالي لا يرد", ascending=True)
    long = plot.melt(id_vars=["المحصّل"], value_vars=available, var_name="الحالة", value_name="العدد")
    fig = px.bar(
        long, x="العدد", y="المحصّل", orientation="h", color="الحالة", barmode="stack", text_auto=True,
        template=PLOTLY_TEMPLATE, category_orders={"الحالة": ACTIVITY_NO_ANSWER_STATES},
        color_discrete_sequence=ACTIVITY_STATE_PALETTE,
    )
    fig.update_layout(**_activity_layout(
        title="📵 حالات لا يرد لكل محصل (تشمل مغلق والتكرار)", title_x=0.5, xaxis_title="عدد الحالات", yaxis_title="",
        height=400, legend_title_text="", legend={"orientation": "h", "yanchor": "top", "y": -0.16, "x": 0.5, "xanchor": "center"},
        margin={"t": 62, "b": 78, "l": 100, "r": 16}, yaxis={"categoryorder": "total ascending"},
    ))
    fig.update_traces(
        marker_line_width=0,
        customdata=long["المحصّل"],
        hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:,}<extra></extra>",
    )
    render_selectable_chart(fig, "dashboard_no_answer_states", filter_key=DASHBOARD_AGENT_FILTER_KEY)


def _render_activity_table(agent):
    table_columns = [
        "المحصّل", "إجمالي المكالمات", "المكالمات الناجحة", "نسبة النجاح (%)", "نسبة من إجمالي المكالمات (%)",
        "واعد بالسداد", "سدد كامل المديونية", "جدولة", "جدولة مقفلة", "سدد كامل المديونية بخصم",
        "لا يرد", "لا يرد مع التكرار", "مغلق", "مغلق مع التكرار", "إجمالي لا يرد",
        "أيام النشاط", "متوسط ساعات العمل/اليوم", "إجمالي ساعات العمل", "إجمالي الوقت المهدر (دقيقة)",
    ]
    table = agent[[c for c in table_columns if c in agent.columns]].copy()
    for col in ["نسبة النجاح (%)", "نسبة من إجمالي المكالمات (%)", "متوسط ساعات العمل/اليوم", "إجمالي ساعات العمل", "إجمالي الوقت المهدر (دقيقة)"]:
        if col in table.columns:
            table[col] = pd.to_numeric(table[col], errors="coerce").fillna(0).round(2)
    st.dataframe(table.sort_values("إجمالي المكالمات", ascending=False), use_container_width=True, hide_index=True)


def render_activity_dashboard(df, class_col=None, sales_col=None, time_col=None, break_start=None, break_end=None):
    """Dashboard تحليل نشاط المحصلين: KPI، اتجاهات زمنية، حالات Sub State، ساعات العمل، وجدول تفصيلي."""
    if not sales_col or sales_col not in df.columns:
        st.error("لا يوجد عمود واضح للمحصّل (Create By / Sales Person) في الملف.")
        return
    _render_dashboard_agent_filter_notice()
    view = _dashboard_activity_view(df, sales_col)
    agent, work, sub_col = _build_activity_summary(view, class_col, sales_col, time_col, break_start, break_end)
    if agent.empty:
        st.info("لا توجد مكالمات قابلة للعرض بعد تطبيق الفلاتر.")
        return
    total = len(work)
    success = int(work["_success_bool"].sum())
    success_rate = success / total * 100 if total else 0
    wasted = float(pd.to_numeric(work.get(WASTED_TIME_COL, pd.Series(dtype=float)), errors="coerce").fillna(0).sum())
    st.subheader("📌 مؤشرات الأداء الرئيسية")
    render_activity_kpi_cards(total, success, int(agent["المحصّل"].nunique()), success_rate, wasted)
    st.caption("اضغط على اسم أي محصل داخل الرسوم التفاعلية لتطبيق فلتر موحد على الكروت والرسوم والجدول.")

    daily_col, hourly_col = st.columns(2)
    with daily_col:
        with st.container(border=True):
            _render_activity_daily_chart(work, time_col, class_col)
    with hourly_col:
        with st.container(border=True):
            _render_activity_hourly_chart(work, time_col)

    outcome_col, no_answer_col = st.columns(2)
    with outcome_col:
        with st.container(border=True):
            _render_activity_outcome_donut(work, class_col)
    with no_answer_col:
        with st.container(border=True):
            _render_activity_no_answer_chart(agent)

    with st.container(border=True):
        st.subheader("📋 جدول أداء كل محصل وحالات Sub State وساعات العمل")
        if not sub_col:
            st.warning("لم يتم العثور على عمود Sub State؛ ستظهر أعمدة الحالات بصفر حتى يتم رفع ملف يحتوي على العمود.")
        _render_activity_table(agent)


def render_full_dashboard(df, class_col=None, sales_col=None, time_col=None, break_start=None, break_end=None):
    render_activity_dashboard(df, class_col, sales_col, time_col, break_start, break_end)


# ==========================================================
# تصدير الداشبورد كصفحة ويب (HTML) مستقلة — نفس الشكل والألوان
# ==========================================================


def build_dashboard_html(df, class_col, sales_col, time_col, source_name="", filter_hint="", filter_summary=None) -> str:
    """إنشاء نسخة HTML مستقلة من Dashboard النشاط بنفس التسلسل والألوان والرسوم الأساسية."""
    from html import escape
    import json

    # التقرير المصدّر له Light Mode مستقل حتى يظل واضحًا عند فتحه في أي متصفح.
    background = "#F5F7FB"
    surface = "#FFFFFF"
    border = "#D9E2EC"
    text = "#1F2937"
    text_dim = "#526174"
    export_success = "#2F6F73"
    export_fail = "#8095A2"
    export_accent = "#477F82"
    export_warn = "#628B8E"
    export_template = "plotly_white"
    work = df.copy()
    if sales_col and sales_col in work.columns:
        work["_agent_display"] = work[sales_col].fillna("غير محدد").astype(str).str.strip()
    else:
        work["_agent_display"] = "غير محدد"
    work["_success_bool"] = _activity_success_mask(work, class_col)
    sub_col_for_export = find_column(work, PROMISE_SUB_STATE_CANDIDATES)
    if sub_col_for_export:
        work["_activity_state"] = work[sub_col_for_export].map(_classify_activity_sub_state)
    total = len(work)
    success = int(work["_success_bool"].sum())
    rate = success / total * 100 if total else 0
    wasted = pd.to_numeric(work.get(WASTED_TIME_COL, pd.Series(dtype=float)), errors="coerce").fillna(0).sum()
    agent_count = int(work["_agent_display"].nunique()) if total else 0
    filter_summary = filter_summary or {}
    filter_items = [
        ("👤 المحصل", filter_summary.get("المحصل", "كل المحصلين")),
        ("📊 الحالة الفرعية", filter_summary.get("الحالة الفرعية", "كل الحالات")),
        ("📅 التاريخ", filter_summary.get("التاريخ", "كل التواريخ")),
        ("🏷️ التصنيف", filter_summary.get("التصنيف", "الكل")),
    ]
    timed_source = pd.to_datetime(work[time_col], errors="coerce") if time_col and time_col in work.columns else pd.Series(pd.NaT, index=work.index)
    raw_records = []
    for row_index, row in work.iterrows():
        timestamp = timed_source.loc[row_index] if row_index in timed_source.index else pd.NaT
        raw_records.append({
            "agent": str(row.get("_agent_display", "غير محدد")),
            "time": timestamp.isoformat() if pd.notna(timestamp) else "",
            "success": bool(row.get("_success_bool", False)),
            "wasted": float(pd.to_numeric(row.get(WASTED_TIME_COL, 0), errors="coerce") or 0),
            "state": str(_classify_activity_sub_state(row.get(find_column(work, PROMISE_SUB_STATE_CANDIDATES), ""))) if find_column(work, PROMISE_SUB_STATE_CANDIDATES) else "",
        })
    raw_records_json = json.dumps(raw_records, ensure_ascii=False).replace("</", "<\\/")
    date_values = sorted({record["time"][:10] for record in raw_records if record.get("time")})
    export_date_min = date_values[0] if date_values else ""
    export_date_max = date_values[-1] if date_values else ""
    agent_color_json = json.dumps(_activity_agent_color_map(work["_agent_display"]), ensure_ascii=False)
    state_color_json = json.dumps(dict(zip(ACTIVITY_NO_ANSWER_STATES, ACTIVITY_STATE_PALETTE)), ensure_ascii=False)

    def metric_card(card_id, label, value, color):
        return (
            f'<div id="{card_id}" style="background:{surface};border:1px solid {border};border-radius:14px;'
            f'padding:22px 14px;text-align:center;min-height:112px;box-sizing:border-box">'
            f'<div style="color:{text_dim};font-size:15px;margin-bottom:12px">{label}</div>'
            f'<div data-role="value" style="color:{color};font-size:28px;font-weight:700;line-height:1.2">{value}</div></div>'
        )

    parts = [
        '<!doctype html><html lang="ar" dir="rtl"><head><meta charset="utf-8">'
        '<meta name="viewport" content="width=device-width, initial-scale=1">'
        '<title>داشبورد تحليل نشاط المحصلين</title></head>',
        f'<body style="margin:0;background:{background};color:{text};font-family:Tahoma,Arial,sans-serif;line-height:1.6">',
        '<main style="max-width:1500px;margin:0 auto;padding:28px 30px">',
        f'<header style="background:{surface};border:1px solid {border};border-radius:16px;padding:24px 28px;margin-bottom:22px">'
        '<div style="font-size:13px;color:' + COLOR_ACCENT + ';letter-spacing:1px">ACTIVITY DASHBOARD</div>'
        '<h1 style="margin:4px 0 2px;font-size:30px">📊 تحليل نشاط المحصلين</h1>'
        f'<div style="color:{text_dim};font-size:14px">مصدر البيانات: {escape(source_name or "ملف النشاط")}</div>',
    ]
    if filter_hint:
        parts.append(f'<div style="margin-top:12px;color:{text_dim};font-size:13px">الفلاتر النشطة: {escape(filter_hint)}</div>')
    agent_options = sorted(work["_agent_display"].dropna().astype(str).unique().tolist())
    state_options = sorted(work["_activity_state"].dropna().astype(str).unique().tolist()) if "_activity_state" in work.columns else []
    parts.extend([
        '</header>',
        f'<section style="background:{surface};border:0;border-radius:12px;padding:8px 0;margin-bottom:14px">',
        f'<h2 style="margin:0 0 8px;text-align:center;font-size:18px;color:{text}">🎚️ فلاتر التقرير التفاعلية</h2>',
        '<section id="interactive-filters" style="display:grid;grid-template-columns:repeat(4,minmax(0,1fr));gap:9px;align-items:end">',
        f'<div style="position:relative;display:flex;flex-direction:column;gap:7px;color:{text_dim};font-size:13px"><span>👤 المحصلون</span><button type="button" class="multi-trigger" data-target="agent-menu" style="background:{background};color:{text};border:1px solid {border};border-radius:9px;padding:8px;font-size:13px;text-align:right;cursor:pointer"><span id="agent-label">كل المحصلين</span>⌄</button><div id="agent-menu" class="multi-menu" style="display:none;position:absolute;z-index:20;top:74px;right:0;left:0;background:#FFFFFF;color:{text};border:1px solid {border};border-radius:10px;padding:8px;box-shadow:0 10px 24px rgba(15,23,42,.16);max-height:230px;overflow-y:auto"><label style="display:block;padding:8px;border-bottom:1px solid {border};font-weight:700"><input type="checkbox" class="select-all-agent"> كل المحصلين</label>',
    ])
    for value in agent_options:
        parts.append(f'<label style="display:block;padding:8px 6px;border-radius:7px;cursor:pointer"><input type="checkbox" class="agent-option" value="{escape(value, quote=True)}"> {escape(value)}</label>')
    parts.extend([
        f'</div></div><div style="position:relative;display:flex;flex-direction:column;gap:7px;color:{text_dim};font-size:13px"><span>📊 الحالات الفرعية</span><button type="button" class="multi-trigger" data-target="state-menu" style="background:{background};color:{text};border:1px solid {border};border-radius:9px;padding:8px;font-size:13px;text-align:right;cursor:pointer"><span id="state-label">كل الحالات</span>⌄</button><div id="state-menu" class="multi-menu" style="display:none;position:absolute;z-index:20;top:74px;right:0;left:0;background:#FFFFFF;color:{text};border:1px solid {border};border-radius:10px;padding:8px;box-shadow:0 10px 24px rgba(15,23,42,.16);max-height:230px;overflow-y:auto"><label style="display:block;padding:8px;border-bottom:1px solid {border};font-weight:700"><input type="checkbox" class="select-all-state"> كل الحالات</label>',
    ])
    for value in state_options:
        parts.append(f'<label style="display:block;padding:8px 6px;border-radius:7px;cursor:pointer"><input type="checkbox" class="state-option" value="{escape(value, quote=True)}"> {escape(value)}</label>')
    parts.extend([
        f'</div></div><label style="display:flex;flex-direction:column;gap:7px;color:{text_dim};font-size:13px"><span>🏷️ التصنيف</span><select id="filter-class" style="background:{background};color:{text};border:1px solid {border};border-radius:9px;padding:8px;font-size:13px"><option value="">الكل</option><option value="success">ناجحة</option><option value="failure">غير ناجحة</option></select></label>',
        f'<label style="display:flex;flex-direction:column;gap:7px;color:{text_dim};font-size:13px"><span>📅 من تاريخ</span><input id="filter-date-from" type="date" value="{export_date_min}" min="{export_date_min}" max="{export_date_max}" style="background:{background};color:{text};border:1px solid {border};border-radius:9px;padding:8px;font-size:13px"></label>',
        f'<label style="display:flex;flex-direction:column;gap:7px;color:{text_dim};font-size:13px"><span>📅 إلى تاريخ</span><input id="filter-date-to" type="date" value="{export_date_max}" min="{export_date_min}" max="{export_date_max}" style="background:{background};color:{text};border:1px solid {border};border-radius:9px;padding:8px;font-size:13px"></label>',
        f'<div style="display:flex;gap:8px;align-items:end"><button id="reset-filters" type="button" style="flex:1;background:{export_accent};color:#fff;border:0;border-radius:9px;padding:8px;font-size:13px;cursor:pointer">↺ إعادة ضبط</button></div>',
        '</section><div id="filter-status" style="text-align:center;color:' + text_dim + ';font-size:12px;margin-top:12px">عرض كل البيانات</div></section>',
        '<section id="kpi-grid" style="display:grid;grid-template-columns:repeat(auto-fit,minmax(180px,1fr));gap:14px;margin-bottom:18px">',
        metric_card("kpi-agents", "👥 عدد المحصلين", f"{agent_count:,}", text),
        metric_card("kpi-total", "📞 إجمالي المكالمات", f"{total:,}", text),
        metric_card("kpi-success", "✅ المكالمات الناجحة", f"{success:,}", export_success),
        metric_card("kpi-rate", "📈 نسبة النجاح", f"{rate:.1f}%", export_accent),
        metric_card("kpi-wasted", "⏱️ إجمالي الوقت المهدر", f"{wasted:,.1f} دقيقة", export_warn),
        '</section>',
    ])

    figs = []
    if class_col and class_col in work.columns:
        donut_df = pd.DataFrame({"النتيجة": ["ناجحة", "غير ناجحة"], "العدد": [success, total - success]})
        fig = px.pie(donut_df, names="النتيجة", values="العدد", hole=0.62, color="النتيجة", color_discrete_map={"ناجحة": export_success, "غير ناجحة": export_fail}, template=export_template)
        fig.update_traces(textinfo="percent", textfont_size=15, marker={"line": {"color": surface, "width": 3}}, hovertemplate="<b>%{label}</b><br>العدد: %{value:,}<br>النسبة: %{percent}<extra></extra>")
        fig.update_layout(**_activity_layout(title="🎯 الناجحة مقابل غير الناجحة", title_x=0.5, height=410, margin={"t":68,"b":65,"l":20,"r":20}, legend={"orientation":"h","y":-0.1,"x":0.5,"xanchor":"center"}, annotations=[{"text":f"{rate:.1f}%<br>نجاح","x":0.5,"y":0.5,"font":{"size":22,"color":export_success},"showarrow":False}]))
        figs.append(("🎯 توزيع نتائج المكالمات", fig))

    if time_col and time_col in work.columns:
        work["_activity_time"] = pd.to_datetime(work[time_col], errors="coerce")
        timed = work.dropna(subset=["_activity_time"]).copy()
        if not timed.empty:
            timed["اليوم"] = timed["_activity_time"].dt.strftime("%Y-%m-%d")
            timed["_success_for_day"] = _activity_success_mask(timed, class_col)
            daily = timed.groupby(["اليوم", "_agent_display"], as_index=False).agg(**{"عدد المكالمات": ("_agent_display", "size"), "المكالمات الناجحة": ("_success_for_day", "sum")})
            daily_totals = daily.groupby("اليوم", as_index=False).agg(**{"إجمالي المكالمات": ("عدد المكالمات", "sum"), "إجمالي الناجحة": ("المكالمات الناجحة", "sum")})
            daily_totals["نسبة النجاح (%)"] = (daily_totals["إجمالي الناجحة"] / daily_totals["إجمالي المكالمات"].replace(0, pd.NA) * 100).fillna(0).round(1)
            ordered_days = sorted(daily_totals["اليوم"].tolist())
            daily["اليوم"] = pd.Categorical(daily["اليوم"], categories=ordered_days, ordered=True)
            daily = daily.sort_values(["اليوم", "_agent_display"])
            daily_totals["اليوم"] = pd.Categorical(daily_totals["اليوم"], categories=ordered_days, ordered=True)
            daily_totals = daily_totals.sort_values("اليوم")
            day_fig = px.bar(daily, x="اليوم", y="عدد المكالمات", color="_agent_display", barmode="group", text_auto=True, custom_data=["_agent_display"], template=export_template, labels={"_agent_display":"المحصل"}, color_discrete_sequence=ACTIVITY_AGENT_PALETTE)
            day_fig.add_trace(go.Scatter(x=daily_totals["اليوم"].astype(str), y=daily_totals["نسبة النجاح (%)"], name="نسبة النجاح", mode="lines+markers+text", text=daily_totals["نسبة النجاح (%)"].map(lambda value: f"{value:.1f}%"), textposition="top center", line={"color": export_accent, "width": 3}, marker={"color": export_accent, "size": 9}, yaxis="y2", hovertemplate="<b>%{x}</b><br>نسبة النجاح: %{y:.1f}%<extra></extra>"))
            day_fig.update_layout(**_activity_layout(title="📊 Combo Chart يومي: المكالمات ونسبة النجاح", title_x=0.5, xaxis_title="اليوم", yaxis_title="عدد المكالمات", height=430, bargap=0.18, margin={"t":68,"b":95,"l":55,"r":65}, xaxis={"type":"category","categoryorder":"array","categoryarray":ordered_days,"tickangle":-25}, yaxis={"rangemode":"tozero"}, yaxis2={"title":"نسبة النجاح (%)","overlaying":"y","side":"right","range":[0,100],"ticksuffix":"%","showgrid":False}, legend={"orientation":"h","y":-0.2,"x":0.5,"xanchor":"center"}))
            day_fig.update_traces(selector={"type":"bar"}, marker_line_width=0, hovertemplate="<b>%{x}</b><br>%{fullData.name}: %{y:,} مكالمة<extra></extra>")
            figs.append(("📊 Combo Chart النشاط اليومي", day_fig))

            timed["الساعة"] = timed["_activity_time"].dt.hour
            hour_min = int(timed["الساعة"].min())
            hour_max = int(timed["الساعة"].max())
            hourly = timed.groupby(["الساعة", "_agent_display"], as_index=False).size().rename(columns={"size":"عدد المكالمات"})
            hour_fig = px.bar(hourly, x="الساعة", y="عدد المكالمات", color="_agent_display", barmode="stack", text_auto=True, custom_data=["_agent_display"], template=export_template, labels={"_agent_display":"المحصل"}, color_discrete_sequence=ACTIVITY_AGENT_PALETTE)
            hour_fig.update_layout(**_activity_layout(title="🕒 Histogram ساعي لنشاط المحصلين", title_x=0.5, xaxis_title="ساعة اليوم", yaxis_title="عدد المكالمات", height=430, bargap=0.08, margin={"t":68,"b":95,"l":55,"r":20}, xaxis={"dtick":1,"tickvals":list(range(hour_min, hour_max + 1)),"range":[max(-0.5, hour_min - 0.5), min(23.5, hour_max + 0.5)]}, legend={"orientation":"h","y":-0.2,"x":0.5,"xanchor":"center"}))
            hour_fig.update_traces(marker_line_width=0, hovertemplate="<b>الساعة %{x}:00</b><br>%{fullData.name}: %{y:,} مكالمة<extra></extra>")
            figs.append(("🕒 Histogram النشاط الساعي", hour_fig))

    sub_col = find_column(work, PROMISE_SUB_STATE_CANDIDATES)
    if sub_col:
        work["_activity_state"] = work[sub_col].map(_classify_activity_sub_state)
        state_counts = work.pivot_table(index="_agent_display", columns="_activity_state", aggfunc="size", fill_value=0)
        for state_name in ACTIVITY_NO_ANSWER_STATES:
            if state_name not in state_counts.columns:
                state_counts[state_name] = 0
        state_counts = state_counts.reindex(columns=ACTIVITY_NO_ANSWER_STATES, fill_value=0).reset_index().rename(columns={"_agent_display":"المحصّل"})
        state_long = state_counts.melt(id_vars=["المحصّل"], var_name="الحالة", value_name="العدد")
        no_fig = px.bar(state_long, x="العدد", y="المحصّل", orientation="h", color="الحالة", barmode="stack", text_auto=True, template=export_template, category_orders={"الحالة":ACTIVITY_NO_ANSWER_STATES}, color_discrete_sequence=[export_success, export_accent, export_warn, "#A6B4B9"])
        no_fig.update_layout(**_activity_layout(title="📵 حالات لا يرد لكل محصل", title_x=0.5, xaxis_title="عدد الحالات", yaxis_title="", height=430, margin={"t":68,"b":80,"l":105,"r":20}, legend={"orientation":"h","y":-0.18,"x":0.5,"xanchor":"center"}, yaxis={"categoryorder":"total ascending"}))
        no_fig.update_traces(hovertemplate="<b>%{y}</b><br>%{fullData.name}: %{x:,}<extra></extra>")
        figs.append(("📵 تحليل حالات Sub State", no_fig))

    parts.append('<section style="display:grid;grid-template-columns:repeat(auto-fit,minmax(460px,1fr));gap:18px">')
    include_js = True
    for index, (heading, fig) in enumerate(figs):
        parts.append(f'<article id="chart-card-{index}" style="background:{surface};border:1px solid {border};border-radius:16px;padding:10px 14px 4px;min-width:0"><h2 style="font-size:17px;margin:8px 10px;color:{text};text-align:center">{heading}</h2>')
        parts.append(pio.to_html(fig, full_html=False, include_plotlyjs=include_js, config=PLOTLY_CONFIG, div_id=f"activity_plot_{index}", default_width="100%", default_height=f"{max(fig.layout.height or 430, 450)}px"))
        parts.append('</article>')
        include_js = False
    parts.append('</section>')

    if sales_col and sales_col in df.columns:
        agent_table, _, _ = _build_activity_summary(df, class_col, sales_col, time_col)
        columns = ["المحصّل", "إجمالي المكالمات", "المكالمات الناجحة", "نسبة النجاح (%)", "نسبة من إجمالي المكالمات (%)", "واعد بالسداد", "إجمالي لا يرد", "أيام النشاط", "متوسط ساعات العمل/اليوم", "إجمالي ساعات العمل", "إجمالي الوقت المهدر (دقيقة)"]
        columns = [column for column in columns if column in agent_table.columns]
        parts.append(f'<section id="activity-summary-table" style="background:{surface};border:1px solid {border};border-radius:16px;padding:18px;margin-top:18px"><h2 style="font-size:19px;margin:0 0 12px;text-align:center">📋 ملخص أداء كل محصل</h2><div style="overflow-x:auto"><table style="width:100%;border-collapse:collapse;font-size:13px"><thead><tr>')
        for column in columns:
            parts.append(f'<th style="padding:10px;border-bottom:1px solid {border};color:{text_dim};white-space:nowrap;text-align:right">{escape(column)}</th>')
        parts.append('</tr></thead><tbody>')
        for _, row in agent_table.sort_values("إجمالي المكالمات", ascending=False).iterrows():
            parts.append('<tr>')
            for column in columns:
                value = row[column]
                if isinstance(value, float):
                    value = f"{value:,.2f}"
                parts.append(f'<td style="padding:9px;border-bottom:1px solid rgba(128,145,170,.18);white-space:nowrap">{escape(str(value))}</td>')
            parts.append('</tr>')
        parts.append('</tbody></table></div></section>')

    interactive_js = """
<script>
const activityData = __ACTIVITY_DATA__;
const agentColors = __AGENT_COLORS__;
const stateColors = __STATE_COLORS__;
const dailyPlot = document.getElementById('activity_plot_1');
const hourlyPlot = document.getElementById('activity_plot_2');
const donutPlot = document.getElementById('activity_plot_0');
const statePlot = document.getElementById('activity_plot_3');
const fmt = n => Number(n || 0).toLocaleString('en-US');
function updateMultiLabels() {
  const agent = [...document.querySelectorAll('.agent-option:checked')].map(o => o.value);
  const state = [...document.querySelectorAll('.state-option:checked')].map(o => o.value);
  document.getElementById('agent-label').textContent = agent.length ? `${agent.length} محصل محدد` : 'كل المحصلين';
  document.getElementById('state-label').textContent = state.length ? `${state.length} حالة محددة` : 'كل الحالات';
}
function selectedRows() {
  const values = cls => [...document.querySelectorAll('.' + cls + ':checked')].map(option => option.value).filter(Boolean);
  const agent = values('agent-option');
  const state = values('state-option');
  const cls = document.getElementById('filter-class').value;
  const from = document.getElementById('filter-date-from').value;
  const to = document.getElementById('filter-date-to').value;
  return activityData.filter(row => ((!agent.length) || agent.includes(row.agent)) && ((!state.length) || state.includes(row.state)) && (!cls || (cls === 'success' ? row.success : !row.success)) && (!from || !row.time || row.time.slice(0,10) >= from) && (!to || !row.time || row.time.slice(0,10) <= to));
}
function setKpi(id, value) { const el = document.querySelector('#' + id + ' [data-role=value]'); if (el) el.textContent = value; }
function refreshDashboard() {
  const rows = selectedRows();
  const agents = [...new Set(rows.map(r => r.agent))].sort();
  const success = rows.filter(r => r.success).length;
  const rate = rows.length ? success / rows.length * 100 : 0;
  setKpi('kpi-agents', fmt(agents.length)); setKpi('kpi-total', fmt(rows.length)); setKpi('kpi-success', fmt(success)); setKpi('kpi-rate', rate.toFixed(1) + '%'); setKpi('kpi-wasted', Number(rows.reduce((s,r) => s + (r.wasted || 0), 0)).toLocaleString('en-US', {maximumFractionDigits:1}) + ' دقيقة');
  document.getElementById('filter-status').textContent = `عرض ${fmt(rows.length)} مكالمة من أصل ${fmt(activityData.length)} — ${fmt(agents.length)} محصل`;
  const days = [...new Set(rows.filter(r => r.time).map(r => r.time.slice(0,10)))].sort();
  const dailyTraces = agents.map(agent => ({type:'bar', name:agent, x:days, y:days.map(day => rows.filter(r => r.agent===agent && r.time.slice(0,10)===day).length), marker:{color:agentColors[agent] || '#6F9FB5'}, texttemplate:'%{y}', textposition:'inside', hovertemplate:'<b>%{x}</b><br>%{fullData.name}: %{y} مكالمة<extra></extra>'}));
  const dailySuccess = days.map(day => { const d=rows.filter(r => r.time && r.time.slice(0,10)===day); return d.length ? d.filter(r=>r.success).length/d.length*100 : 0; });
  dailyTraces.push({type:'scatter', mode:'lines+markers+text', name:'نسبة النجاح', x:days, y:dailySuccess, text:dailySuccess.map(v=>v.toFixed(1)+'%'), textposition:'top center', line:{color:'__ACCENT__',width:3}, marker:{color:'__ACCENT__',size:8}, yaxis:'y2', hovertemplate:'<b>%{x}</b><br>نسبة النجاح: %{y:.1f}%<extra></extra>'});
  if (dailyPlot) Plotly.react(dailyPlot, dailyTraces, {...dailyPlot.layout, xaxis:{...(dailyPlot.layout?.xaxis||{}), type:'category', categoryarray:days}, yaxis2:{...(dailyPlot.layout?.yaxis2||{}), range:[0,100], ticksuffix:'%'}});
  const hours = Array.from({length:24},(_,i)=>i); const hourlyTraces = agents.map(agent=>({type:'bar',name:agent,x:hours,y:hours.map(h=>rows.filter(r=>r.agent===agent && r.time && new Date(r.time).getHours()===h).length),marker:{color:agentColors[agent]||'#6F9FB5'},texttemplate:'%{y}',textposition:'inside'}));
  if (hourlyPlot) Plotly.react(hourlyPlot, hourlyTraces, {...hourlyPlot.layout, barmode:'stack', xaxis:{...(hourlyPlot.layout?.xaxis||{}), dtick:1, range:[-0.5,23.5]}});
  if (donutPlot) Plotly.react(donutPlot, [{type:'pie',labels:['ناجحة','غير ناجحة'],values:[success, rows.length-success],hole:.62,marker:{colors:['__SUCCESS__','__FAIL__']},textinfo:'percent'}], donutPlot.layout);
  const stateNames = Object.keys(stateColors); const stateTraces = stateNames.map(state=>({type:'bar',name:state,x:agents,y:agents.map(a=>rows.filter(r=>r.agent===a && r.state===state).length),marker:{color:stateColors[state]},texttemplate:'%{y}',textposition:'inside'}));
  if (statePlot) Plotly.react(statePlot, stateTraces, {...statePlot.layout, barmode:'stack'});
}
document.querySelectorAll('.multi-trigger').forEach(trigger => trigger.addEventListener('click', event => { event.stopPropagation(); const menu = document.getElementById(trigger.dataset.target); document.querySelectorAll('.multi-menu').forEach(other => { if (other !== menu) other.style.display = 'none'; }); menu.style.display = menu.style.display === 'block' ? 'none' : 'block'; }));
document.addEventListener('click', () => document.querySelectorAll('.multi-menu').forEach(menu => menu.style.display = 'none'));
document.querySelectorAll('.agent-option,.state-option').forEach(option => option.addEventListener('change', () => { updateMultiLabels(); refreshDashboard(); }));
document.querySelector('.select-all-agent')?.addEventListener('change', event => { document.querySelectorAll('.agent-option').forEach(option => option.checked = event.target.checked); updateMultiLabels(); refreshDashboard(); });
document.querySelector('.select-all-state')?.addEventListener('change', event => { document.querySelectorAll('.state-option').forEach(option => option.checked = event.target.checked); updateMultiLabels(); refreshDashboard(); });
['filter-class','filter-date-from','filter-date-to'].forEach(id => document.getElementById(id)?.addEventListener('change', refreshDashboard));
document.getElementById('reset-filters')?.addEventListener('click', () => { document.querySelectorAll('.agent-option,.state-option,.select-all-agent,.select-all-state').forEach(option => option.checked=false); document.getElementById('filter-class').value=''; document.getElementById('filter-date-from').value='__DATE_MIN__'; document.getElementById('filter-date-to').value='__DATE_MAX__'; updateMultiLabels(); refreshDashboard(); });
updateMultiLabels();
refreshDashboard();
</script>
""".replace('__ACTIVITY_DATA__', raw_records_json).replace('__AGENT_COLORS__', agent_color_json).replace('__STATE_COLORS__', state_color_json).replace('__ACCENT__', json.dumps(export_accent)).replace('__SUCCESS__', json.dumps(export_success)).replace('__FAIL__', json.dumps(export_fail)).replace('__DATE_MIN__', export_date_min).replace('__DATE_MAX__', export_date_max)
    parts.append(interactive_js)
    parts.extend(['<footer style="color:' + text_dim + ';font-size:12px;text-align:center;margin-top:24px">تم إنشاء التقرير من لوحة تحليل نشاط المحصلين</footer></main></body></html>'])
    return "".join(parts)


# ==========================================================
# تويب 1: التصنيف
# ==========================================================

# ==========================================================
# أدوات اختيار الشركة / الفترة / التجميع اليومي
# ==========================================================

COMPANIES = ["الوطنية للتأمين", "تري للتأمين"]
STATUS_CANDIDATES = [
    "Main State", "Final State", "Status", "State", "Call Status",
    "الحالة الرئيسية", "الحالة النهائية", "الحالة", "حالة المكالمة",
]

CLOSED_WORDS = ["مغلق", "مغلقه", "مغلقه", "closed", "غير متاح"]
def init_activity_state():
    defaults = {
        "selected_company": None,
        "selected_period": None,
        "period_1_start": dt_time(9, 0),
        "period_1_end": dt_time(12, 30),
        "period_2_start": dt_time(13, 0),
        "period_2_end": dt_time(17, 0),
        "period_1_has_break": False,
        "period_2_has_break": False,
        "period_1_break_start": dt_time(11, 0),
        "period_1_break_end": dt_time(11, 15),
        "period_2_break_start": dt_time(15, 0),
        "period_2_break_end": dt_time(15, 15),
        "period_1_break_duration": 15,
        "period_2_break_duration": 15,
        "daily_has_break": False,
        "daily_break_start": dt_time(13, 0),
        "daily_break_end": dt_time(13, 15),
        "daily_break_duration": 15,
        "monthly_has_break": False,
        "monthly_break_start": dt_time(13, 0),
        "monthly_break_end": dt_time(13, 15),
        "monthly_break_duration": 15,
        "selected_week": "week_1",
        "period_results": {},
        "weekly_week_1_has_break": False,
        "weekly_week_1_break_start": dt_time(13, 0),
        "weekly_week_1_break_end": dt_time(13, 15),
        "weekly_week_1_break_duration": 15,
        "weekly_week_1_result": None,
        "weekly_week_2_has_break": False,
        "weekly_week_2_break_start": dt_time(13, 0),
        "weekly_week_2_break_end": dt_time(13, 15),
        "weekly_week_2_break_duration": 15,
        "weekly_week_2_result": None,
        "weekly_week_3_has_break": False,
        "weekly_week_3_break_start": dt_time(13, 0),
        "weekly_week_3_break_end": dt_time(13, 15),
        "weekly_week_3_break_duration": 15,
        "weekly_week_3_result": None,
        "weekly_week_4_has_break": False,
        "weekly_week_4_break_start": dt_time(13, 0),
        "weekly_week_4_break_end": dt_time(13, 15),
        "weekly_week_4_break_duration": 15,
        "weekly_week_4_result": None,
        "daily_result": None,
        "monthly_result": None,
        "dashboard_result": None,
    }
    for key, value in defaults.items():
        st.session_state.setdefault(key, value)


def render_company_selector():
    st.subheader("🏢 اختر شركة التصنيف")
    c1, c2 = st.columns(2)
    for col, company_name in zip((c1, c2), COMPANIES):
        with col:
            selected = st.session_state.get("selected_company") == company_name
            button_label = f"✓ {company_name}" if selected else company_name
            if st.button(
                button_label,
                key=f"company_{company_name}",
                use_container_width=True,
                type="primary" if selected else "secondary",
            ):
                st.session_state["selected_company"] = company_name
                st.session_state["selected_period"] = None
                st.rerun()
    company = st.session_state.get("selected_company")
    if company:
        st.success(f"الشركة المختارة: {company}")


def render_period_selector():
    if not st.session_state.get("selected_company"):
        return
    st.subheader("🕒 اختر فترة النشاط")
    cols = st.columns(5)
    buttons = [("الفترة الأولى", "🟢", "period_1"), ("الفترة الثانية", "🔵", "period_2"), ("التجميع اليومي", "📊", "daily"), ("التجميع الأسبوعي", "📅", "weekly"), ("التجميع الشهري", "🗓️", "monthly")]
    for col, (title, icon, key) in zip(cols, buttons):
        with col:
            selected = st.session_state.get("selected_period") == key
            if st.button(f"{icon} {title}", key=f"period_btn_{key}", use_container_width=True, type="primary" if selected else "secondary"):
                st.session_state["selected_period"] = key
                st.rerun()
    selected_period = st.session_state.get("selected_period")
    if not selected_period:
        return
    company = st.session_state["selected_company"]
    if selected_period == "period_1":
        render_period_settings("period_1", "الفترة الأولى")
    elif selected_period == "period_2":
        render_period_settings("period_2", "الفترة الثانية")
    elif selected_period == "daily":
        st.subheader("📊 التجميع اليومي")
        st.caption(f"{company} · من بداية اليوم إلى نهايته")
        render_aggregate_tab("daily", "التجميع اليومي")
    elif selected_period == "weekly":
        st.subheader("📅 التجميع الأسبوعي")
        st.caption(f"{company} · اختر الأسبوع لرفع الملف")
        weeks = [("الأسبوع الأول", "week_1"), ("الأسبوع الثاني", "week_2"), ("الأسبوع الثالث", "week_3"), ("الأسبوع الرابع", "week_4")]
        week_cols = st.columns(4)
        for col, (title, key) in zip(week_cols, weeks):
            with col:
                if st.button(title, key=f"btn_{key}", use_container_width=True, type="primary" if st.session_state.get("selected_week") == key else "secondary"):
                    st.session_state["selected_week"] = key
                    st.rerun()
        current = st.session_state.get("selected_week", "week_1")
        title = next(label for label, key in weeks if key == current)
        render_aggregate_tab(f"weekly_{current}", f"التجميع الأسبوعي ({title})")
    elif selected_period == "monthly":
        st.subheader("🗓️ التجميع الشهري")
        st.caption(f"{company} · نشاط الشهر بالكامل")
        render_aggregate_tab("monthly", "التجميع الشهري")


def render_period_upload_and_classify(period_key: str, period_title: str):
    uploaded_file = st.file_uploader(
        f"📂 ارفع ملف {period_title} (CSV أو Excel)",
        type=["csv", "xlsx", "xls"],
        key=f"upload_{period_key}",
        on_change=sync_file_cache,
        args=(f"upload_{period_key}", f"period_upload:{period_key}", (f"period_results:{period_key}",)),
    )
    if uploaded_file is not None:
        st.caption(f"الملف المختار: {uploaded_file.name}")
        classify_period_file(uploaded_file, period_key)
    else:
        _show_period_results_from_cache(period_key)


def _render_break_switch(label: str, has_break_key: str):
    st.session_state[has_break_key] = st.toggle(label, value=st.session_state[has_break_key], key=f"{has_break_key}_switch")
    st.caption("✅ يوجد استراحة" if st.session_state[has_break_key] else "لا يوجد استراحة")


def render_period_settings(period_key: str, period_title: str):
    company = st.session_state["selected_company"]
    start_key, end_key = f"{period_key}_start", f"{period_key}_end"
    has_break_key = f"{period_key}_has_break"
    break_start_key, break_end_key = f"{period_key}_break_start", f"{period_key}_break_end"
    st.subheader(f"⏱️ {period_title}")
    st.caption(f"نشاط {company} · حدّد الميعاد قبل رفع ملف الفترة")
    c1, c2 = st.columns(2)
    with c1:
        st.session_state[start_key] = st.time_input(f"بداية {period_title}", value=st.session_state[start_key], key=f"{start_key}_input")
    with c2:
        st.session_state[end_key] = st.time_input(f"نهاية {period_title}", value=st.session_state[end_key], key=f"{end_key}_input")
    if st.session_state[start_key] >= st.session_state[end_key]:
        st.warning("يجب أن تسبق بداية الفترة نهايتها.")
    _render_break_switch(f"هل يوجد استراحة في {period_title}؟", has_break_key)
    if st.session_state[has_break_key]:
        b1, b2 = st.columns(2)
        with b1:
            st.session_state[break_start_key] = st.time_input("🕐 بداية الاستراحة", value=st.session_state[break_start_key], key=f"{break_start_key}_input")
        with b2:
            duration_key = f"{period_key}_break_duration"
            current_duration = max(5, int(st.session_state[duration_key]))
            st.session_state[duration_key] = st.slider("⏳ مدة الاستراحة (دقيقة)", 5, 120, current_duration, 5, key=f"{duration_key}_slider")
            start_minutes = st.session_state[break_start_key].hour * 60 + st.session_state[break_start_key].minute
            end_minutes = start_minutes + st.session_state[duration_key]
            st.session_state[break_end_key] = dt_time(end_minutes // 60, end_minutes % 60)
        st.info(f"☕ الاستراحة من {st.session_state[break_start_key]:%H:%M} إلى {st.session_state[break_end_key]:%H:%M} ({st.session_state[duration_key]} دقيقة)")
    st.info(f"النشاط: {period_title} {company} من {st.session_state[start_key]:%H:%M} إلى {st.session_state[end_key]:%H:%M}")
    render_period_upload_and_classify(period_key, period_title)


@st.cache_data(show_spinner=False)
def read_uploaded_dataframe(uploaded_file):
    """قراءة الملف مع كاش على محتوى البايتات — مش بيتعاد إلا لو محتوى الملف اتغير.

    Streamlit بيعمل rerun كامل مع أي أكشن، والدالة دي كانت بتتدعي مع كل rerun
    وكانت بتقرأ الإكسيل من الصفر كل مرة. الـ @st.cache_data بيخزن النتيجة
    ويعيدها فوراً طالما محتوى الملف (البايتات + الاسم) نفسهما.
    """
    data = uploaded_file.getvalue()
    name = uploaded_file.name.lower()
    if name.endswith(".csv"):
        return pd.read_csv(io.BytesIO(data))
    return pd.read_excel(io.BytesIO(data))


def classify_period_file(uploaded_file, period_key):
    period_title = {"period_1": "الفترة الأولى", "period_2": "الفترة الثانية"}[period_key]

    # 💾 لو الفترة دي اتصنّفت قبل كده والنتيجة لسه في الذاكرة — نعرضها من الكاش من غير إعادة قراءة أو تصنيف
    stored = st.session_state["period_results"].get(period_key)
    current_file_hash = uploaded_file_hash(uploaded_file)
    if stored and stored.get("uploaded_hash") == current_file_hash:
        st.success(f"تم تصنيف {period_title} بنجاح ✅ — {len(stored['df']):,} مكالمة")
        _render_period_results(stored, period_key)
        return

    try:
        df = read_uploaded_dataframe(uploaded_file)
    except Exception as e:
        st.error(f"تعذر قراءة الملف: {e}")
        return

    # نفس قاعدة الملف الحالية: حذف أول صف بعد العناوين.
    if len(df) > 0:
        df = df.iloc[1:].reset_index(drop=True)

    if ORIGINAL_TEXT_COL in df.columns:
        df = df.rename(columns={ORIGINAL_TEXT_COL: MODEL_TEXT_COL})

    if MODEL_TEXT_COL not in df.columns:
        st.error(
            f"عمود النص ('{ORIGINAL_TEXT_COL}' أو '{MODEL_TEXT_COL}') غير موجود. "
            f"الأعمدة الموجودة: {', '.join(df.columns.astype(str))}"
        )
        return

    if st.button(
        f"🚀 بدء تصنيف {period_title}",
        key=f"classify_{period_key}",
        type="primary",
        use_container_width=True,
    ):
        with st.spinner("جارٍ تجهيز النموذج وتصنيف الملف..."):
            tokenizer, model, device = load_model()
            texts = df[MODEL_TEXT_COL].tolist()
            preds, confidences = predict_batch(texts, tokenizer, model, device)

        result_df = df.copy()
        result_df[CLASSIFICATION_COL] = preds
        result_df["نسبة_الثقة"] = [round(c * 100, 1) for c in confidences]

        sales_col = find_column(result_df, SALES_PERSON_CANDIDATES)
        time_col = find_column(result_df, CREATED_ON_CANDIDATES)
        claim_col = find_column(result_df, CLAIM_CANDIDATES)
        result_df, duplicate_stats = remove_claim_duplicates(result_df, claim_col, time_col)

        break_start = (
            st.session_state[f"{period_key}_break_start"]
            if st.session_state[f"{period_key}_has_break"]
            else None
        )
        break_end = (
            st.session_state[f"{period_key}_break_end"]
            if st.session_state[f"{period_key}_has_break"]
            else None
        )

        if sales_col and time_col:
            result_df = calculate_wasted_time(
                result_df, sales_col, time_col, break_start, break_end
            )
        else:
            st.warning(
                "تعذر العثور على عمود المحصّل أو عمود التاريخ والوقت، فلن يتم حساب الوقت المهدر. "
                "تأكد من أسماء الأعمدة."
            )

        result_df = result_df.rename(columns={MODEL_TEXT_COL: ORIGINAL_TEXT_COL})

        # معلومات الفترة والشركة تبقى مع النتيجة بدون التأثير على الموديل.
        result_df["الشركة"] = st.session_state["selected_company"]
        result_df["الفترة"] = period_title
        result_df["بداية الفترة"] = st.session_state[f"{period_key}_start"].strftime("%H:%M")
        result_df["نهاية الفترة"] = st.session_state[f"{period_key}_end"].strftime("%H:%M")

        st.session_state["period_results"][period_key] = {
            "df": result_df,
            "sales_col": sales_col,
            "time_col": time_col,
            "claim_col": claim_col,
            "duplicate_stats": duplicate_stats,
            "company": st.session_state["selected_company"],
            "period_title": period_title,
            "uploaded_filename": uploaded_file.name,
            "uploaded_hash": current_file_hash,
        }
        st.session_state["last_result_df"] = result_df
        st.session_state["last_sales_col"] = sales_col
        st.session_state["last_time_col"] = time_col
        st.rerun()

    stored = st.session_state["period_results"].get(period_key)
    if stored:
        st.success(f"تم تصنيف {period_title} بنجاح ✅ — {len(stored['df']):,} مكالمة")
        _render_period_results(stored, period_key)


def _xml_escape(value):
    return (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
        .replace("'", "&apos;")
    )


def _sanitize_table_name(name: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9_]", "_", name)
    if not cleaned or not cleaned[0].isalpha():
        cleaned = "T_" + cleaned
    return cleaned[:60]


def _build_base_workbook_bytes(df: pd.DataFrame, data_sheet_name: str, table_name: str, pivot_sheet_name: str = "Pivot Table") -> bytes:
    """بيبني ملف إكسيل بـ openpyxl فيه شيت البيانات كـ Excel Table (ListObject) رسمي + شيت فاضي للـ Pivot."""
    wb = Workbook()
    ws = wb.active
    ws.title = data_sheet_name
    ws.append([str(c) for c in df.columns])
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    last_row = max(ws.max_row, 2)
    last_col_letter = get_column_letter(len(df.columns))
    tab = Table(displayName=table_name, ref=f"A1:{last_col_letter}{last_row}")
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True)
    ws.add_table(tab)
    wb.create_sheet(pivot_sheet_name)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _inject_native_pivot_table(xlsx_bytes, df_columns, table_name, pivot_sheet_name, row_field, data_field, data_field_label, col_field=None, subtotal="count"):
    """بيحقن Pivot Table حقيقي (native، قابل للتحديث والسحب والإفلات) جوه ملف الإكسيل — نفس اللي بتعمله
    يدوي في إكسيل بـ Insert > PivotTable. بيتحدث تلقائي من الـ Excel Table لما تفتح الملف."""
    df_columns = list(df_columns)
    row_idx = df_columns.index(row_field)
    col_idx = df_columns.index(col_field) if col_field else None
    data_idx = df_columns.index(data_field)
    n_fields = len(df_columns)

    zin = zipfile.ZipFile(BytesIO(xlsx_bytes), "r")
    data = {n: zin.read(n) for n in zin.namelist()}
    zin.close()

    wb_xml = data["xl/workbook.xml"].decode("utf-8")
    wb_rels = data["xl/_rels/workbook.xml.rels"].decode("utf-8")

    name_to_rid = {}
    for tag in re.findall(r"<sheet [^>]*/>", wb_xml):
        m_name = re.search(r'name="([^"]+)"', tag)
        m_rid = re.search(r'r:id="([^"]+)"', tag)
        if m_name and m_rid:
            name_to_rid[m_name.group(1)] = m_rid.group(1)
    pivot_rid = name_to_rid[pivot_sheet_name]

    rid_to_target = {}
    for tag in re.findall(r"<Relationship [^>]*/>", wb_rels):
        m_id = re.search(r'Id="([^"]+)"', tag)
        m_target = re.search(r'Target="([^"]+)"', tag)
        if m_id and m_target:
            rid_to_target[m_id.group(1)] = m_target.group(1)
    pivot_sheet_target = rid_to_target[pivot_rid]
    pivot_sheet_path = pivot_sheet_target.lstrip("/")
    if not pivot_sheet_path.startswith("xl/"):
        pivot_sheet_path = "xl/" + pivot_sheet_path
    sheet_file = pivot_sheet_path.split("/")[-1]

    existing_ids = [int(re.sub(r"\D", "", rid)) for rid in re.findall(r'Id="(rId\d+)"', wb_rels)]
    cache_rid = f"rId{max(existing_ids) + 1}"

    wb_xml = wb_xml.replace(
        "</workbook>",
        f'<pivotCaches><pivotCache cacheId="1" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="{cache_rid}"/></pivotCaches></workbook>',
    )
    data["xl/workbook.xml"] = wb_xml.encode("utf-8")

    new_rel = f'<Relationship Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="pivotCache/pivotCacheDefinition1.xml" Id="{cache_rid}"/>'
    wb_rels = wb_rels.replace("</Relationships>", new_rel + "</Relationships>")
    data["xl/_rels/workbook.xml.rels"] = wb_rels.encode("utf-8")

    data[f"xl/worksheets/_rels/{sheet_file}.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotTable" Target="../pivotTables/pivotTable1.xml"/>'
        "</Relationships>"
    ).encode("utf-8")

    ct = data["[Content_Types].xml"].decode("utf-8")
    overrides = (
        '<Override PartName="/xl/pivotCache/pivotCacheDefinition1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml"/>'
        '<Override PartName="/xl/pivotCache/pivotCacheRecords1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml"/>'
        '<Override PartName="/xl/pivotTables/pivotTable1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.pivotTable+xml"/>'
    )
    data["[Content_Types].xml"] = ct.replace("</Types>", overrides + "</Types>").encode("utf-8")

    cache_fields_xml = "".join(
        f'<cacheField name="{_xml_escape(col)}" numFmtId="0"><sharedItems/></cacheField>' for col in df_columns
    )
    data["xl/pivotCache/pivotCacheDefinition1.xml"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotCacheDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" r:id="rId1" '
        'refreshOnLoad="1" refreshedBy="Claude" refreshedDate="45900" createdVersion="6" refreshedVersion="6" '
        f'minRefreshableVersion="3" recordCount="0">'
        f'<cacheSource type="worksheet"><worksheetSource name="{table_name}"/></cacheSource>'
        f'<cacheFields count="{n_fields}">{cache_fields_xml}</cacheFields>'
        "</pivotCacheDefinition>"
    ).encode("utf-8")
    data["xl/pivotCache/pivotCacheRecords1.xml"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotCacheRecords xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships" count="0"/>'
    ).encode("utf-8")
    data["xl/pivotCache/_rels/pivotCacheDefinition1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" Target="pivotCacheRecords1.xml"/>'
        "</Relationships>"
    ).encode("utf-8")

    pivot_fields_parts = []
    for i in range(n_fields):
        if i == row_idx:
            pivot_fields_parts.append('<pivotField axis="axisRow" showAll="0"><items count="1"><item t="default"/></items></pivotField>')
        elif col_idx is not None and i == col_idx:
            pivot_fields_parts.append('<pivotField axis="axisCol" showAll="0"><items count="1"><item t="default"/></items></pivotField>')
        elif i == data_idx:
            pivot_fields_parts.append('<pivotField dataField="1" showAll="0"/>')
        else:
            pivot_fields_parts.append('<pivotField showAll="0"/>')
    pivot_fields_xml = "".join(pivot_fields_parts)

    col_fields_xml = f'<colFields count="1"><field x="{col_idx}"/></colFields><colItems count="1"><i><x/></i></colItems>' if col_idx is not None else ""

    data["xl/pivotTables/pivotTable1.xml"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<pivotTableDefinition xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
        'name="PivotTable1" cacheId="1" applyNumberFormats="0" applyBorderFormats="0" applyFontFormats="0" '
        'applyPatternFormats="0" applyAlignmentFormats="0" applyWidthHeightFormats="1" dataCaption="Values" '
        'updatedVersion="6" minRefreshableVersion="3" useAutoFormatting="1" itemPrintTitles="1" createdVersion="6" '
        'indent="0" outline="1" outlineData="1" multipleFieldFilters="0">'
        '<location ref="A3:C10" firstHeaderRow="1" firstDataRow="2" firstDataCol="1"/>'
        f'<pivotFields count="{n_fields}">{pivot_fields_xml}</pivotFields>'
        f'<rowFields count="1"><field x="{row_idx}"/></rowFields>'
        "<rowItems count=\"1\"><i><x/></i></rowItems>"
        f'{col_fields_xml}'
        "<dataFields count=\"1\">"
        f'<dataField name="{_xml_escape(data_field_label)}" fld="{data_idx}" subtotal="{subtotal}" baseField="0" baseItem="0"/>'
        "</dataFields>"
        '<pivotTableStyleInfo name="PivotStyleMedium9" showRowHeaders="1" showColHeaders="1" showRowStripes="0" showColStripes="0" showLastColumn="1"/>'
        "</pivotTableDefinition>"
    ).encode("utf-8")
    data["xl/pivotTables/_rels/pivotTable1.xml.rels"] = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
        '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition" Target="../pivotCache/pivotCacheDefinition1.xml"/>'
        "</Relationships>"
    ).encode("utf-8")

    out_buf = BytesIO()
    zout = zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED)
    for n, content in data.items():
        zout.writestr(n, content)
    zout.close()
    return out_buf.getvalue()


def build_excel_with_native_pivot(result_df: pd.DataFrame, data_sheet_name: str, key_prefix: str):
    """بيبني ملف إكسيل فيه شيت البيانات (كـ Excel Table) + شيت Pivot Table حقيقي (native) —
    المحصل في الصفوف، ومجموع عمود التصنيف (SUM) في القيم.
    بيرجع (bytes, تم إضافة بيفوت ولا لأ)."""
    collected_col = find_column(result_df, COLLECTED_BY_CANDIDATES)
    class_col = CLASSIFICATION_COL if CLASSIFICATION_COL in result_df.columns else None

    table_name = _sanitize_table_name(f"tbl_{key_prefix}")
    base_bytes = _build_base_workbook_bytes(result_df, data_sheet_name, table_name)

    if not (collected_col and class_col):
        return base_bytes, False

    final_bytes = _inject_native_pivot_table(
        base_bytes,
        result_df.columns,
        table_name,
        "Pivot Table",
        row_field=collected_col,
        data_field=class_col,
        data_field_label="مجموع التصنيف",
        subtotal="sum",
    )
    return final_bytes, True


def render_pivot_section(df: pd.DataFrame, key_prefix: str):
    """معاينة سريعة جوه السيستم بس (نفس منطق الـ Pivot اللي هيتضاف حقيقي في ملف الإكسيل):
    الصفوف = المحصل، القيم = مجموع (SUM) عمود التصنيف."""
    if df is None or df.empty:
        return

    collected_col = find_column(df, COLLECTED_BY_CANDIDATES)
    class_col = CLASSIFICATION_COL if CLASSIFICATION_COL in df.columns else None

    with st.expander("📊 معاينة الـ Pivot Table (هتلاقي النسخة الحقيقية القابلة للتعديل جوه ملف الإكسيل بعد التحميل)", expanded=False):
        missing = [
            label for label, col in [
                ("المحصل (Created by)", collected_col),
                ("التصنيف", class_col),
            ] if col is None
        ]
        if missing:
            st.warning("تعذر إنشاء الـ Pivot — الأعمدة دي مش موجودة في الملف: " + "، ".join(missing))
            return

        pivot_df = pd.pivot_table(
            df,
            index=collected_col,
            values=class_col,
            aggfunc="sum",
            fill_value=0,
            margins=True,
            margins_name="الإجمالي",
        )
        pivot_df = pivot_df.rename_axis(index="المحصل").rename(columns={class_col: "مجموع التصنيف"})
        st.dataframe(pivot_df, use_container_width=True)


def _render_period_results(stored, period_key):
    """عرض نتائج الفترة المحفوظة (جدول + كروت + شارتات + تنزيل) — تُستخدم بعد الضغط على زر التصنيف وبعد شيل الملف."""
    period_title = stored["period_title"]
    result_df = stored["df"]
    sales_col = stored["sales_col"]
    time_col = stored["time_col"]

    render_duplicate_summary(stored.get("duplicate_stats"))
    st.dataframe(result_df, use_container_width=True, hide_index=True)

    render_period_charts(result_df, sales_col, time_col, period_title)

    render_pivot_section(result_df, f"period_{period_key}")

    if result_df is not None:
        excel_bytes, pivot_added = build_excel_with_native_pivot(result_df, "النتائج", f"period_{period_key}")
        if not pivot_added:
            st.caption("⚠️ اتنزل الملف من غير Pivot Table لأن عمود المحصل (Collected by) أو رقم حساب العميل (Customer Account number) مش موجود في الملف.")
        st.download_button(
            f"⬇️ تحميل نتائج {period_title}",
            data=excel_bytes,
            file_name=f"نتائج_{period_title}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            key=f"download_{period_key}",
        )


def _show_period_results_from_cache(period_key):
    """عرض النتائج المحفوظة بصمت لما يكون لا يوجد ملف مرفوع — بدون أي رسائل أو أزرار."""
    stored = st.session_state["period_results"].get(period_key)
    if stored:
        _render_period_results(stored, period_key)


def normalize_status(value):
    text = str(value).strip().lower()
    if any(word in text for word in CLOSED_WORDS):
        return "مغلق"
    return None


def get_status_series(df):
    status_col = find_column(df, STATUS_CANDIDATES)
    if not status_col:
        return None, None
    return status_col, df[status_col].fillna("").map(normalize_status)


def build_agent_activity(df, sales_col):
    if not sales_col or sales_col not in df.columns:
        return pd.DataFrame()

    work = df.copy()
    work["_status_norm"] = None
    status_col, status_series = get_status_series(work)
    if status_col:
        work["_status_norm"] = status_series

    grouped = work.groupby(sales_col).size().rename("إجمالي المكالمات").to_frame()
    grouped["ناجحة"] = work[work[CLASSIFICATION_COL] == 1].groupby(sales_col).size()
    grouped["ناجحة"] = grouped["ناجحة"].fillna(0).astype(int)

    counts = work[work["_status_norm"] == "مغلق"].groupby(sales_col).size()
    grouped["مغلق"] = counts.fillna(0).astype(int)

    # ===== عدد إفادات "لا يرد" في عمود Notes لكل محصّل =====
    notes_col = ORIGINAL_TEXT_COL if ORIGINAL_TEXT_COL in work.columns else None
    lrd_counts = None
    if notes_col:
        lrd = work[notes_col].astype(str).str.contains("لا يرد|لايرد|لا\\s*يرد", regex=True, na=False)
        lrd_counts = work[lrd].groupby(sales_col).size()
    grouped["إفادات لا يرد"] = lrd_counts.reindex(grouped.index, fill_value=0).astype(int) if lrd_counts is not None else 0
    total_calls = work.groupby(sales_col).size()
    lrd_pct = []
    for _, r in grouped.iterrows():
        lrd_n = int(r["إفادات لا يرد"])
        total_n = int(r["إجمالي المكالمات"])
        lrd_pct.append(round(lrd_n / total_n * 100, 1) if total_n else 0.0)
    grouped["نسبة لا يرد (%)"] = lrd_pct

    # ===== متوسط الفجوة/المدة لكل محصّل (دقيقة) =====
    # لو الملف فيه عمود "مدة المكالمة" نستخدمه، وإلا نستخدم متوسط الوقت المهدر بين المكالمات.
    dur_col = find_column(work, DURATION_CANDIDATES)
    avg_key = "متوسط مدة المكالمة (دقيقة)"
    if dur_col and dur_col in work.columns:
        dur = pd.to_numeric(work[dur_col], errors="coerce")
        avg = dur.groupby(work[sales_col]).mean()
        grouped[avg_key] = avg.round(1)
        grouped[avg_key] = grouped[avg_key].fillna(0)
    else:
        wt_col = WASTED_TIME_COL if WASTED_TIME_COL in work.columns else None
        if wt_col:
            wt = pd.to_numeric(work[wt_col], errors="coerce")
            avg = wt.groupby(work[sales_col]).mean()
            grouped[avg_key] = avg.round(1)
            grouped[avg_key] = grouped[avg_key].fillna(0)
        else:
            grouped[avg_key] = 0.0

    return grouped.fillna(0).reset_index().rename(columns={sales_col: "المحصّل"})


def _rounded_rect_path(x0, x1, y0, y1, radius=0.018):
    """إنشاء مسار SVG مستدير الزوايا داخل إحداثيات Plotly الورقية."""
    radius = min(radius, (x1 - x0) / 3, (y1 - y0) / 3)
    return (
        f"M {x0 + radius},{y0} "
        f"L {x1 - radius},{y0} Q {x1},{y0} {x1},{y0 + radius} "
        f"L {x1},{y1 - radius} Q {x1},{y1} {x1 - radius},{y1} "
        f"L {x0 + radius},{y1} Q {x0},{y1} {x0},{y1 - radius} "
        f"L {x0},{y0 + radius} Q {x0},{y0} {x0 + radius},{y0} Z"
    )


def render_kpi_dashboard(total, success, agent_count, success_rate, avg_wasted=None):
    """لوحة KPI مركزية مبنية بـ Plotly لضمان محاذاة موحدة داخل كل كارت."""
    cards = [
        ("📞<br>إجمالي المكالمات", total, {"valueformat": ",d"}, THEME["text"]),
        ("✅<br>المكالمات الناجحة", success, {"valueformat": ",d"}, COLOR_SUCCESS),
        ("👥<br>عدد المحصّلين", agent_count, {"valueformat": ",d"}, THEME["text"]),
        ("📈<br>نسبة النجاح", success_rate, {"valueformat": ".1f", "suffix": "%"}, COLOR_ACCENT),
    ]
    if avg_wasted is not None:
        cards.append(("⏱️<br>متوسط الوقت المهدر", avg_wasted, {"valueformat": ".1f", "suffix": " دقيقة"}, COLOR_WARN))

    figure = go.Figure()
    count = len(cards)
    gap = 0.018
    width = (1 - gap * (count + 1)) / count
    for index, (label, value, number_format, number_color) in enumerate(cards):
        x0 = gap + index * (width + gap)
        x1 = x0 + width
        figure.add_shape(
            type="path",
            xref="paper",
            yref="paper",
            path=_rounded_rect_path(x0, x1, 0.06, 0.94, radius=0.022),
            line={"color": THEME["border"], "width": 1},
            fillcolor=THEME["surface"],
            layer="below",
        )
        figure.add_trace(
            go.Indicator(
                mode="number",
                value=value,
                domain={"x": [x0 + 0.012, x1 - 0.012], "y": [0.12, 0.88]},
                title={"text": label, "font": {"size": 18, "color": THEME["text_dim"]}, "align": "center"},
                number={"font": {"size": 32, "color": number_color}, **number_format},
            )
        )
    figure.update_layout(
        height=200,
        template=PLOTLY_TEMPLATE,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font={"family": "Tajawal, sans-serif", "color": THEME["text"]},
        margin={"t": 8, "b": 8, "l": 8, "r": 8},
    )
    st.plotly_chart(figure, use_container_width=True, config=PLOTLY_CONFIG)


def render_period_charts(df, sales_col, time_col, period_title):
    if not sales_col or sales_col not in df.columns:
        st.info("لا يوجد عمود واضح للمحصّل لعرض نشاط المحصّلين.")
        return
    render_classification_filter_notice(df, sales_col)
    df = get_classification_view(df, sales_col)
    agent = build_agent_activity(df, sales_col)
    if agent.empty:
        return
    total = len(df)
    success = int((df[CLASSIFICATION_COL] == 1).sum()) if CLASSIFICATION_COL in df.columns else 0
    success_rate = round(success / total * 100, 1) if total else 0
    avg_wasted = None
    if WASTED_TIME_COL in df.columns:
        avg_wasted = float(pd.to_numeric(df[WASTED_TIME_COL], errors="coerce").mean())
        if pd.isna(avg_wasted):
            avg_wasted = 0.0
    st.subheader("📌 ملخص نتائج التصنيف")
    render_kpi_dashboard(total, success, len(agent), success_rate, avg_wasted)
    st.divider()
    render_agent_activity_charts(agent, df, sales_col, period_title)
    with st.expander("📋 عرض جدول تفاصيل كل محصّل"):
        st.dataframe(agent, use_container_width=True, hide_index=True)


def render_agent_activity_charts(agent, df, sales_col, period_title):
    agent_sorted = agent.sort_values("ناجحة", ascending=False).reset_index(drop=True)
    names = [str(n) for n in agent_sorted["المحصّل"]]

    def total_vs_success_chart():
        fig = go.Figure([
            go.Bar(name="إجمالي المكالمات", x=names, y=agent_sorted["إجمالي المكالمات"], marker_color=THEME["text_dim"], customdata=names),
            go.Bar(name="المكالمات الناجحة", x=names, y=agent_sorted["ناجحة"], marker_color=COLOR_SUCCESS, customdata=names),
        ])
        fig.update_layout(
            **PLOTLY_LAYOUT,
            title=f"📞 إجمالي المكالمات والناجحة — {period_title}",
            barmode="group",
            xaxis_title="",
            yaxis_title="عدد المكالمات",
            height=430,
            legend=dict(orientation="h", yanchor="bottom", y=-0.28, x=0.5, xanchor="center"),
        )
        fig.update_traces(
            marker_line_width=0,
            hovertemplate="<b>%{x}</b><br>%{fullData.name}: %{y:,} مكالمة<extra></extra>",
        )
        render_selectable_chart(fig, "classification_total_success_chart")

    rates = [
        (int(row["ناجحة"]) / int(row["إجمالي المكالمات"]) * 100)
        if int(row["إجمالي المكالمات"]) else 0
        for _, row in agent_sorted.iterrows()
    ]

    def success_rate_chart():
        rate_df = agent_sorted.assign(**{"نسبة النجاح": rates}).sort_values("نسبة النجاح")
        fig = px.bar(
            rate_df,
            x="نسبة النجاح",
            y="المحصّل",
            orientation="h",
            text="نسبة النجاح",
            color="نسبة النجاح",
            color_continuous_scale=[COLOR_FAIL, COLOR_WARN, COLOR_SUCCESS],
            template=PLOTLY_TEMPLATE,
        )
        fig.update_layout(
            **PLOTLY_LAYOUT,
            title="📈 نسبة نجاح كل محصّل",
            xaxis_range=[0, 100],
            xaxis_title="نسبة النجاح (%)",
            yaxis_title="",
            height=430,
            coloraxis_showscale=False,
        )
        fig.update_traces(
            texttemplate="%{text:.1f}%",
            textposition="outside",
            cliponaxis=False,
            marker_line_width=0,
            customdata=rate_df["المحصّل"],
            hovertemplate="<b>%{y}</b><br>نسبة النجاح: %{x:.1f}%<extra></extra>",
        )
        render_selectable_chart(fig, "classification_success_rate_chart")

    st.subheader(f"📈 تحليلات الأداء — {period_title}")
    first, second = st.columns(2)
    with first:
        with st.container(border=True):
            total_vs_success_chart()
    with second:
        with st.container(border=True):
            success_rate_chart()

    with st.container(border=True):
        render_success_fail_chart(agent, period_title)

    duration_col, no_answer_col = st.columns(2)
    with duration_col:
        with st.container(border=True):
            render_avg_duration_chart(agent, period_title)
    with no_answer_col:
        with st.container(border=True):
            render_no_answer_chart(df, sales_col, period_title)


def render_success_fail_chart(agent, period_title):
    ordered = agent.sort_values("ناجحة", ascending=False).reset_index(drop=True)
    names = [str(n) for n in ordered["المحصّل"]]
    failed = ordered["إجمالي المكالمات"] - ordered["ناجحة"]
    fig = go.Figure([
        go.Bar(name="المكالمات الناجحة", x=names, y=ordered["ناجحة"], marker_color=COLOR_SUCCESS, customdata=names),
        go.Bar(name="المكالمات غير الناجحة", x=names, y=failed, marker_color=COLOR_FAIL, customdata=names),
    ])
    fig.update_layout(**PLOTLY_LAYOUT, title=f"✅ الناجحة مقابل غير الناجحة ({period_title})", barmode="group", xaxis_title="", yaxis_title="عدد المكالمات")
    fig.update_traces(hovertemplate="<b>%{x}</b><br>%{fullData.name}: %{y:,} مكالمة<extra></extra>")
    render_selectable_chart(fig, f"classification_success_fail_{period_title}")


def render_avg_duration_chart(agent, period_title):
    key = "متوسط مدة المكالمة (دقيقة)"
    if key not in agent.columns:
        return
    ordered = agent.sort_values(key, ascending=True)
    fig = px.bar(ordered, x=key, y="المحصّل", orientation="h", text=key, color=key, color_continuous_scale=[COLOR_ACCENT, COLOR_WARN], template=PLOTLY_TEMPLATE)
    fig.update_layout(**PLOTLY_LAYOUT, title=f"⏱️ متوسط مدة المكالمات ({period_title})", xaxis_title="دقيقة", yaxis_title="")
    fig.update_traces(customdata=ordered["المحصّل"], hovertemplate="<b>%{y}</b><br>متوسط المدة: %{x:.1f} دقيقة<extra></extra>")
    render_selectable_chart(fig, f"classification_avg_duration_{period_title}")


def render_no_answer_chart(df, sales_col, period_title):
    notes_col = ORIGINAL_TEXT_COL if ORIGINAL_TEXT_COL in df.columns else None
    if not notes_col or sales_col not in df.columns:
        return
    work = df.copy()
    work["_lrd"] = work[notes_col].astype(str).str.contains(r"لا يرد|لايرد|لا\s*يرد", regex=True, na=False)
    counts = work[work["_lrd"]].groupby(sales_col).size().reset_index(name="عدد إفادات لا يرد")
    if counts.empty:
        return
    fig = px.bar(counts, x=sales_col, y="عدد إفادات لا يرد", color_discrete_sequence=[COLOR_FAIL], template=PLOTLY_TEMPLATE)
    fig.update_layout(**PLOTLY_LAYOUT, title=f"📝 إفادات لا يرد لكل محصّل ({period_title})", xaxis_title="", yaxis_title="العدد")
    fig.update_traces(customdata=counts[sales_col], hovertemplate="<b>%{x}</b><br>عدد إفادات لا يرد: %{y}<extra></extra>")
    render_selectable_chart(fig, f"classification_no_answer_{period_title}")


def _render_aggregate_break_settings(period_key, period_title):
    has_break_key = f"{period_key}_has_break"
    break_start_key = f"{period_key}_break_start"
    break_end_key = f"{period_key}_break_end"
    _render_break_switch(f"🕐 هل يوجد استراحة في {period_title}؟", has_break_key)
    if st.session_state[has_break_key]:
        b1, b2 = st.columns(2)
        with b1:
            st.session_state[break_start_key] = st.time_input("🕐 بداية الاستراحة", value=st.session_state[break_start_key], key=f"{break_start_key}_input")
        with b2:
            duration_key = f"{period_key}_break_duration"
            st.session_state[duration_key] = st.slider("⏳ مدة الاستراحة (دقيقة)", 5, 120, int(st.session_state[duration_key]), 5, key=f"{duration_key}_slider")
            total = st.session_state[break_start_key].hour * 60 + st.session_state[break_start_key].minute + st.session_state[duration_key]
            st.session_state[break_end_key] = dt_time(total // 60, total % 60)
        st.info(f"☕ الاستراحة من {st.session_state[break_start_key]:%H:%M} إلى {st.session_state[break_end_key]:%H:%M}")


def render_aggregate_tab(period_key, period_title):
    company = st.session_state["selected_company"]
    _render_aggregate_break_settings(period_key, period_title)
    uploaded_file = st.file_uploader(
        f"📂 ارفع ملف {period_title} (CSV أو Excel)",
        type=["csv", "xlsx", "xls"],
        key=f"upload_{period_key}",
        on_change=sync_file_cache,
        args=(f"upload_{period_key}", f"period_upload:{period_key}", (f"{period_key}_result",)),
    )
    if uploaded_file is not None:
        st.caption(f"الملف المختار: {uploaded_file.name}")
        _classify_aggregate_file(uploaded_file, company, period_key, period_title)
    else:
        _show_aggregate_results_from_cache(period_key, period_title)


def _classify_aggregate_file(uploaded_file, company, period_key, period_title):
    result_key = f"{period_key}_result"
    stored = st.session_state.get(result_key)
    current_file_hash = uploaded_file_hash(uploaded_file)
    if stored and stored.get("uploaded_hash") == current_file_hash:
        _render_aggregate_results(stored, period_title, period_key)
        return
    try:
        df = read_uploaded_dataframe(uploaded_file)
    except Exception as e:
        st.error(f"تعذر قراءة الملف: {e}")
        return
    if len(df) > 0:
        df = df.iloc[1:].reset_index(drop=True)
    if ORIGINAL_TEXT_COL in df.columns:
        df = df.rename(columns={ORIGINAL_TEXT_COL: MODEL_TEXT_COL})
    if MODEL_TEXT_COL not in df.columns:
        st.error(f"عمود النص غير موجود. الأعمدة الموجودة: {', '.join(df.columns.astype(str))}")
        return
    if st.button(f"🚀 بدء تصنيف {period_title}", key=f"btn_{period_key}", type="primary", use_container_width=True):
        with st.spinner(f"جارٍ تصنيف {period_title}..."):
            tokenizer, model, device = load_model()
            texts = df[MODEL_TEXT_COL].tolist()
            preds, confidences = predict_batch(texts, tokenizer, model, device)
        result_df = df.copy()
        result_df[CLASSIFICATION_COL] = preds
        result_df["نسبة_الثقة"] = [round(c * 100, 1) for c in confidences]
        sales_col = find_column(result_df, SALES_PERSON_CANDIDATES)
        time_col = find_column(result_df, CREATED_ON_CANDIDATES)
        claim_col = find_column(result_df, CLAIM_CANDIDATES)
        result_df, duplicate_stats = remove_claim_duplicates(result_df, claim_col, time_col)
        has_break = st.session_state[f"{period_key}_has_break"]
        break_start = st.session_state[f"{period_key}_break_start"] if has_break else None
        break_end = st.session_state[f"{period_key}_break_end"] if has_break else None
        if sales_col and time_col:
            result_df = calculate_wasted_time(result_df, sales_col, time_col, break_start, break_end)
        result_df = result_df.rename(columns={MODEL_TEXT_COL: ORIGINAL_TEXT_COL})
        result_df["الشركة"] = company
        result_df["الفترة"] = period_title
        if has_break:
            result_df["بداية الاستراحة"] = break_start.strftime("%H:%M")
            result_df["نهاية الاستراحة"] = break_end.strftime("%H:%M")
            result_df["مدة الاستراحة_دقيقة"] = st.session_state[f"{period_key}_break_duration"]
        st.session_state[result_key] = {
            "df": result_df, "sales_col": sales_col, "time_col": time_col,
            "claim_col": claim_col, "duplicate_stats": duplicate_stats,
            "company": company, "uploaded_filename": uploaded_file.name,
            "uploaded_hash": current_file_hash,
        }
        st.rerun()

def _render_aggregate_results(stored, period_title, period_key):
    result_df = stored["df"]
    sales_col = stored["sales_col"]
    time_col = stored["time_col"]
    render_duplicate_summary(stored.get("duplicate_stats"))
    st.dataframe(result_df, use_container_width=True, hide_index=True)
    render_period_charts(result_df, sales_col, time_col, period_title)
    render_pivot_section(result_df, f"agg_{period_key}")
    excel_bytes, pivot_added = build_excel_with_native_pivot(result_df, period_title, f"agg_{period_key}")
    if not pivot_added:
        st.caption("⚠️ اتنزل الملف من غير Pivot Table لأن عمود المحصل (Collected by) أو رقم حساب العميل (Customer Account number) مش موجود في الملف.")
    st.download_button(
        f"⬇️ تحميل نتائج {period_title}",
        data=excel_bytes,
        file_name=f"نتائج_{period_title}_{stored['company']}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        key=f"dl_{period_title}",
        type="primary",
    )

def _show_aggregate_results_from_cache(period_key, period_title):
    stored = st.session_state.get(f"{period_key}_result")
    if stored:
        st.success(f"تم تصنيف {period_title} بنجاح ✅ — {len(stored['df']):,} مكالمة")
        _render_aggregate_results(stored, period_title, period_key)
def page_classification():
    init_activity_state()
    page_header("CALL QUALITY CLASSIFIER", "🎯 تصنيف المكالمات", "اختر الشركة → اختر الفترة → حدّد الميعاد والاستراحة → ارفع الملف → ابدأ التصنيف")
    render_company_selector()
    render_period_selector()


# ==========================================================
# تويبات 2-5
# ==========================================================

def page_placeholder(eyebrow, title, subtitle, icon):
    page_header(eyebrow, f"{icon} {title}", subtitle)
    st.info("هذا القسم ما زال قيد التجهيز، وسيتم تحديد مصدر بياناته ومنطقه في مرحلة لاحقة.")


# ==========================================================
# تويب 6: الداشبورد
# ==========================================================

DASHBOARD_SOURCE_KEY = "dashboard_uploaded_source"


def _run_neglect_followup_pipeline(new_file, old_file):
    try:
        df_new = read_uploaded_dataframe(new_file)
        df_old = read_uploaded_dataframe(old_file)
        
        # حذف أول صف
        if len(df_new) > 0: df_new = df_new.iloc[1:].reset_index(drop=True)
        if len(df_old) > 0: df_old = df_old.iloc[1:].reset_index(drop=True)
        
        # البحث عن الأعمدة
        id_col_new = find_column(df_new, ID_CANDIDATES)
        id_col_old = find_column(df_old, ID_CANDIDATES)
        last_date_new = find_column(df_new, NEGLECT_LAST_DATE_CANDIDATES)
        
        if not id_col_new or not id_col_old or not last_date_new:
            st.error("تعذر العثور على عمود الرقم التعريفي (ID) أو تاريخ آخر متابعة في الملفات.")
            return
            
        # تحويل المعرفات لنصوص لضمان المطابقة
        df_new[id_col_new] = df_new[id_col_new].astype(str).str.strip()
        df_old[id_col_old] = df_old[id_col_old].astype(str).str.strip()
        
        # جلب التواريخ الحديثة
        mapping = df_new.set_index(id_col_new)[last_date_new].to_dict()
        
        # تحديث ملف الإهمال القديم
        result_df = df_old.copy()
        result_df['تاريخ_متابعة_حديث'] = result_df[id_col_old].map(mapping)
        
        # حساب الملاحظات
        target_date = st.session_state.get(TODAY_KEY, datetime.now().date())
        
        def check_coverage(val):
            d = parse_date_cell(val)
            if not d: return "لم يتم التغطية"
            diff = (target_date - d).days
            return "تم التغطية" if diff < 8 else "لم يتم التغطية"
            
        result_df['الملاحظات'] = result_df['تاريخ_متابعة_حديث'].apply(check_coverage)
        
        st.session_state["neglect_followup_result"] = {"df": result_df}
        st.rerun()
    except Exception as e:
        st.error(f"خطأ في معالجة الملفات: {e}")

def _show_neglect_followup_results(df):
    st.subheader("📊 نتائج متابعة الإهمال")
    covered = int((df["الملاحظات"] == "تم التغطية").sum())
    not_covered = int((df["الملاحظات"] == "لم يتم التغطية").sum())
    pct = covered / len(df) * 100 if len(df) else 0
    c1, c2, c3 = st.columns(3)
    c1.metric("✅ تم التغطية", covered)
    c2.metric("❌ لم يتم التغطية", not_covered)
    c3.metric("📈 نسبة التغطية", f"{pct:.1f}%")
    st.dataframe(df, use_container_width=True, hide_index=True)
    out_excel = io.BytesIO()
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="متابعة الإهمال")
    st.download_button("⬇️ تحميل تقرير متابعة الإهمال المحدث (Excel)", data=out_excel.getvalue(), file_name=f"متابعة_الإهمال_{datetime.now():%Y-%m-%d}.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, type="primary")


def page_neglect():
    """تويب الإهمال: فلترة الحالات المتأخرة في المتابعة + حساب فرق الأيام + داشبورد."""
    init_neglect_state()
    _init_promises_today()
    
    page_header(
        "NEGLECT TRACKING",
        "⚠️ الإهمال ومتابعة الإهمال",
        "ارفع المحفظة وسيتم استخراج الحالات التي مرّ عليها أكثر من 7 أيام دون متابعة وتحتاج إلى تدخل سريع",
    )
    
    # اختيار الوضع
    m1, m2 = st.columns(2)
    with m1:
        if st.button("🚨 الإهمال", use_container_width=True, type="primary" if st.session_state["neglect_mode"] == "neglect" else "secondary"):
            st.session_state["neglect_mode"] = "neglect"
            st.rerun()
    with m2:
        if st.button("🔍 متابعة الإهمال", use_container_width=True, type="primary" if st.session_state["neglect_mode"] == "followup" else "secondary"):
            st.session_state["neglect_mode"] = "followup"
            st.rerun()

    st.divider()
    
    if st.session_state["neglect_mode"] == "neglect":
        # إدارة حالات Sub State
        with st.expander("⚙️ إدارة حالات Sub State المستهدفة (الإهمال)"):
            available = st.session_state.get("neglect_available_states", [])
            if available:
                st.subheader("🔍 اختر حالات إضافية من الملف المرفوع")
                to_add_options = [s for s in available if s not in st.session_state["neglect_sub_states"]]
                selected_to_add = st.multiselect("اختر الحالات لإضافتها لقائمة الإهمال:", to_add_options)
                if st.button("➕ إضافة الحالات المختارة"):
                    if selected_to_add:
                        st.session_state["neglect_sub_states"].extend(selected_to_add)
                        st.success(f"تمت إضافة {len(selected_to_add)} حالة بنجاح!")
                        st.rerun()
            else:
                st.info("💡 ارفع ملف أولاً أستطيع استخراج كل الحالات المتاحة تختار منها.")
            
            st.divider()
            st.write("📋 الحالات المشمولة حالياً في الإهمال:")
            cols = st.columns(3)
            for i, state in enumerate(st.session_state["neglect_sub_states"]):
                with cols[i % 3]:
                    if st.button(f"❌ {state}", key=f"del_{i}", use_container_width=True):
                        st.session_state["neglect_sub_states"].remove(state)
                        st.rerun()

        uploaded = st.file_uploader(
            "📂 ارفع ملف المحفظة (Excel أو CSV) لفلترة الإهمال",
            type=["xlsx", "xls", "csv"],
            key="neglect_upload",
            on_change=sync_file_cache,
            args=("neglect_upload", "neglect", (NEGLECT_RESULT_KEY,)),
        )
        if uploaded is not None:
            _run_neglect_pipeline(uploaded)
        else:
            cached = st.session_state.get(NEGLECT_RESULT_KEY)
            if cached:
                _show_neglect_results(cached["df"], cached)
    else:
        # وضع متابعة الإهمال - هذا الجزء كان مفقوداً في النسخة السابقة
        st.info("💡 يطابق هذا الوضع تقرير إهمال قديمًا مع محفظة اليوم الحديثة لتحديد الحالات التي تمت متابعتها.")
        c1, c2 = st.columns(2)
        with c1:
            new_portfolio = st.file_uploader(
                "📂 ارفع محفظة اليوم الحديثة",
                type=["xlsx", "xls", "csv"],
                key="new_portfolio_up",
                on_change=sync_file_cache,
                args=("new_portfolio_up", "neglect_followup_new", ("neglect_followup_result",)),
            )
        with c2:
            old_neglect = st.file_uploader(
                "📂 ارفع تقرير الإهمال القديم",
                type=["xlsx", "xls", "csv"],
                key="old_neglect_up",
                on_change=sync_file_cache,
                args=("old_neglect_up", "neglect_followup_old", ("neglect_followup_result",)),
            )
        
        if new_portfolio and old_neglect:
            if st.button("🚀 بدء المطابقة ومتابعة الإهمال", use_container_width=True, type="primary"):
                _run_neglect_followup_pipeline(new_portfolio, old_neglect)
        
        cached_followup = st.session_state.get("neglect_followup_result")
        if cached_followup:
            _show_neglect_followup_results(cached_followup["df"])

def _run_neglect_pipeline(uploaded):

    _file_hash = hashlib.sha256(uploaded.getvalue()).hexdigest()
    cached = st.session_state.get(NEGLECT_RESULT_KEY)
    if cached and cached.get("file_hash") == _file_hash:
        _show_neglect_results(cached["df"], cached)
        return

    try:
        df = read_uploaded_dataframe(uploaded)
        if len(df) > 0:
            df = df.iloc[1:].reset_index(drop=True)
            
        sales_col = find_column(df, SALES_PERSON_CANDIDATES)
        substate_col = find_column(df, PROMISE_SUB_STATE_CANDIDATES)
        duedate_col = find_column(df, PROMISE_DUE_DATE_CANDIDATES)
        lastdate_col = find_column(df, NEGLECT_LAST_DATE_CANDIDATES)
        net_col = find_column(df, PROMISE_NET_AMOUNT_CANDIDATES)
        
        if not all([sales_col, substate_col, duedate_col]):
            st.error("الملف يفتقد أعمدة أساسية (المحصّل، الحالة، أو تاريخ المتابعة).")
            return
            
        # استخراج كافة الحالات المتاحة في الملف للكاش
        all_states = sorted(df[substate_col].dropna().unique().tolist())
        st.session_state["neglect_available_states"] = all_states

        # 1. فلترة المحصّلين
        df = df[~df[sales_col].astype(str).str.strip().isin(PROMISE_EXCLUDED_SALES)].copy()
        
        # 2. فلترة Sub State
        df = df[df[substate_col].astype(str).str.strip().isin(st.session_state["neglect_sub_states"])].copy()
        
        # 3. فلترة التاريخ (ما عدا اليوم)
        target_date = st.session_state[TODAY_KEY]
        df['temp_due'] = [parse_date_cell(v) for v in df[duedate_col]]
        df = df[df['temp_due'] != target_date].copy()
        
        # 4. حساب فرق الأيام والفلترة (> 7 أيام)
        if lastdate_col:
            df['temp_last'] = [parse_date_cell(v) for v in df[lastdate_col]]
            df['فرق_الأيام'] = [(target_date - d).days if d else 0 for d in df['temp_last']]
            # فلترة الحالات اللي بقالها أكتر من 7 أيام
            df = df[df['فرق_الأيام'] > 7].copy()
        
        st.session_state[NEGLECT_RESULT_KEY] = {
            "df": df, "file_hash": _file_hash, "sales_col": sales_col,
            "substate_col": substate_col, "duedate_col": duedate_col,
            "lastdate_col": lastdate_col, "net_col": net_col
        }
        st.rerun()
    except Exception as e:
        st.error(f"خطأ في معالجة الملف: {e}")

def _show_neglect_results(df, meta):
    sales_col = meta["sales_col"]
    net_col = meta["net_col"]
    total = len(df)
    total_amount = pd.to_numeric(df[net_col], errors="coerce").fillna(0).sum() if net_col and net_col in df.columns else 0
    agent_count = int(df[sales_col].nunique()) if sales_col and sales_col in df.columns else 0
    avg_days = pd.to_numeric(df["فرق_الأيام"], errors="coerce").mean() if "فرق_الأيام" in df.columns else None

    st.subheader("📊 ملخص الإهمال")
    c1, c2, c3, c4 = st.columns(4)
    c1.metric("⚠️ إجمالي الحالات", f"{total:,}")
    c2.metric("👥 عدد المحصّلين", f"{agent_count:,}")
    c3.metric("💰 إجمالي المديونية", f"{total_amount:,.0f}" if net_col else "—")
    c4.metric("📅 متوسط فرق الأيام", f"{avg_days:.1f}" if avg_days is not None and pd.notna(avg_days) else "—")

    if total and sales_col and sales_col in df.columns:
        agent_counts = df[sales_col].value_counts().head(15).sort_values().reset_index()
        agent_counts.columns = ["المحصّل", "عدد الحالات"]
        fig = px.bar(
            agent_counts, x="عدد الحالات", y="المحصّل", orientation="h",
            title="أعلى المحصّلين في حالات الإهمال",
            color="عدد الحالات", color_continuous_scale=[THEME["surface_2"], COLOR_WARN],
            template=PLOTLY_TEMPLATE,
        )
        fig.update_layout(**PLOTLY_LAYOUT, height=420, yaxis_title="", coloraxis_showscale=False)
        st.plotly_chart(fig, use_container_width=True, config=PLOTLY_CONFIG)

    st.subheader("📋 جدول حالات الإهمال التفصيلي")
    display_cols = [c for c in [sales_col, meta["substate_col"], meta["duedate_col"], meta["lastdate_col"], "فرق_الأيام", net_col] if c and c in df.columns]
    st.dataframe(df[display_cols], use_container_width=True, hide_index=True)
    out_excel = io.BytesIO()
    with pd.ExcelWriter(out_excel, engine="openpyxl") as writer:
        df[display_cols].to_excel(writer, index=False, sheet_name="حالات الإهمال")
    st.download_button(
        "⬇️ تحميل تقرير الإهمال (Excel)",
        data=out_excel.getvalue(),
        file_name=f"تقرير_الإهمال_{datetime.now():%Y-%m-%d}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
        type="primary",
    )


def _render_dashboard_work_settings():
    """إعداد استراحة Dashboard النشاط؛ تُستخدم لخصم البريك من ساعات العمل اليومية."""
    st.subheader("⏱️ إعدادات حساب ساعات العمل")
    st.caption("سيتم حساب زمن النشاط من أول مكالمة لآخر مكالمة لكل محصل في كل يوم، مع خصم وقت الاستراحة المحدد.")
    st.session_state.setdefault("dashboard_has_break", False)
    st.session_state.setdefault("dashboard_break_start", dt_time(13, 0))
    st.session_state.setdefault("dashboard_break_end", dt_time(13, 15))
    has_break = st.checkbox("☕ يوجد وقت استراحة يتم خصمه", key="dashboard_has_break")
    if not has_break:
        return None, None
    c1, c2 = st.columns(2)
    with c1:
        break_start = st.time_input("بداية الاستراحة", key="dashboard_break_start")
    with c2:
        break_end = st.time_input("نهاية الاستراحة", key="dashboard_break_end")
    if break_start >= break_end:
        st.warning("يجب أن تسبق بداية الاستراحة نهايتها؛ لذلك لن يتم الخصم حتى يتم تصحيح الوقت.")
        return None, None
    st.info(f"سيتم خصم الاستراحة من {break_start:%H:%M} إلى {break_end:%H:%M} من ساعات العمل اليومية.")
    return break_start, break_end


def page_dashboard():
    """تويب داشبورد مستقلة تمامًا عن التصنيف — ترفع فيها ملف النشاط بعد التصنيف
    (فيه عمود التصنيف جاهز) وتعرض لك داشبورد كاملة بالكروت والشارتات،
    وممكن تتنزّل كصفحة ويب HTML مستقلة تفتحها في أي متصفح."""
    page_header(
        "ACTIVITY DASHBOARD",
        "📊 تحليل نشاط المحصّلين",
        "ارفع ملف النشاط المصنّف بعد التصنيف لبناء لوحة تحكم متكاملة",
        centered=True,
    )

    init_activity_state()
    break_start, break_end = _render_dashboard_work_settings()

    # 💾 الكاش الشفاف: لو فيه داشبورد محفوظة لآخر ملف مرفوع — نعرضها من غير إعادة معالجة
    cached = st.session_state.get("dashboard_result")
    current_source_hash = st.session_state.get(DASHBOARD_SOURCE_HASH_KEY)

    dash_file = st.file_uploader(
        "📂 ارفع ملف النشاط المصنّف (بعد التصنيف) — CSV أو Excel",
        type=["csv", "xlsx", "xls"],
        key="dash_upload_v3",
        on_change=sync_file_cache,
        args=("dash_upload_v3", "dashboard", ("dashboard_result", "dashboard_source")),
    )

    if dash_file is None and cached is None:
        st.info("ارفع ملف النشاط بعد تصنيفه، وسيتم بناء لوحة التحكم فورًا.")
        return

    # لو اتشال الملف والنتيجة لسه في الذاكرة — نعرضها من الكاش
    if dash_file is None:
        _show_dashboard_from_cache(break_start, break_end)
        return

    # 💾 لو الملف ده اتعرج قبل كده — نعرض الكاش من غير إعادة معالجة
    current_file_hash = uploaded_file_hash(dash_file)
    if current_source_hash == current_file_hash and cached is not None:
        df_show, hint = _render_slicers(cached["df"], cached["sales_col"], cached["time_col"])
        _render_dashboard(df_show, cached["class_col"], cached["sales_col"],
                          cached["time_col"], dash_file.name, filter_hint=hint,
                          break_start=break_start, break_end=break_end)
        return

    try:
        df = read_uploaded_dataframe(dash_file)
    except Exception as e:
        st.error(f"تعذر قراءة الملف: {e}")
        return

    class_col = CLASSIFICATION_COL if CLASSIFICATION_COL in df.columns else None
    sales_col = find_column(df, SALES_PERSON_CANDIDATES)
    time_col = find_column(df, CREATED_ON_CANDIDATES)

    if class_col is None:
        st.warning(
            f"⚠️ لا يوجد عمود '{CLASSIFICATION_COL}' في الملف؛ ستُعرض المكالمات دون تفاصيل النجاح أو الفشل. "
            "تأكد إنك رفعت الملف بعد التصنيف."
        )

    # 💾 نحفظ النتيجة في الكاش (شفاف — من غير أي كروت أو أزرار إضافية)
    st.session_state[DASHBOARD_SOURCE_KEY] = dash_file.name
    st.session_state[DASHBOARD_SOURCE_HASH_KEY] = current_file_hash
    st.session_state["dashboard_result"] = {
        "df": df, "class_col": class_col, "sales_col": sales_col,
        "time_col": time_col, "source_name": dash_file.name,
        "source_hash": current_file_hash,
    }

    # 🎚️ السلايسرز: فلتر المحصّلين + فلتر التواريخ (للعرض فقط — الكاش محفوظ)
    df_show, hint = _render_slicers(df, sales_col, time_col)
    _render_dashboard(df_show, class_col, sales_col, time_col, dash_file.name, filter_hint=hint,
                      break_start=break_start, break_end=break_end)


def _render_dashboard(df, class_col, sales_col, time_col, source_name, filter_hint="", break_start=None, break_end=None):
    if filter_hint:
        st.info(f"الفلاتر المطبقة: {filter_hint}")
    render_full_dashboard(df, class_col=class_col, sales_col=sales_col, time_col=time_col,
                          break_start=break_start, break_end=break_end)
    dashboard_html = build_dashboard_html(
        df, class_col=class_col, sales_col=sales_col, time_col=time_col,
        source_name=source_name, filter_hint=filter_hint,
        filter_summary=st.session_state.get("dashboard_filter_summary", {}),
    )
    st.download_button("🌐 تحميل لوحة التحكم كصفحة ويب HTML", data=dashboard_html.encode("utf-8"), file_name="داشبورد_النشاط.html", mime="text/html", use_container_width=True, key="dash_html_download_v3", type="primary")
    st.download_button("⬇️ تحميل البيانات كـ CSV", data=df.to_csv(index=False).encode("utf-8-sig"), file_name="بيانات_النشاط.csv", mime="text/csv", use_container_width=True, key="dash_csv_download_v3")


def _clear_dashboard_chart_filter():
    st.session_state.pop(DASHBOARD_AGENT_FILTER_KEY, None)


def _render_native_multi_slicer(label, options, state_key, empty_label):
    """Multi-select نظيف بلا Tags: الاختيارات تظهر داخل Popover فقط، والزر يعرض ملخصًا مختصرًا."""
    options = [str(option) for option in options]
    widget_keys = [f"{state_key}__{index}" for index in range(len(options))]
    selected = [option for option, widget_key in zip(options, widget_keys) if st.session_state.get(widget_key, False)]
    trigger_text = empty_label if not selected else f"تم اختيار {len(selected)}"
    st.caption(label)
    with st.popover(trigger_text, use_container_width=True):
        action_all, action_clear = st.columns(2)
        with action_all:
            if st.button("تحديد الكل", key=f"{state_key}__select_all", use_container_width=True):
                for widget_key in widget_keys:
                    st.session_state[widget_key] = True
                st.rerun()
        with action_clear:
            if st.button("إلغاء الكل", key=f"{state_key}__clear_all", use_container_width=True):
                for widget_key in widget_keys:
                    st.session_state[widget_key] = False
                st.rerun()
        st.caption("اختار أكثر من قيمة؛ لن تظهر الاختيارات كوسوم خارج القائمة.")
        for option, widget_key in zip(options, widget_keys):
            st.session_state.setdefault(widget_key, False)
            st.checkbox(option, key=widget_key, label_visibility="visible")
    return [option for option, widget_key in zip(options, widget_keys) if st.session_state.get(widget_key, False)]


def _render_slicers(df, sales_col, time_col):
    agents = sorted([str(a) for a in df[sales_col].dropna().unique()]) if sales_col and sales_col in df.columns else []
    date_min = date_max = None
    if time_col and time_col in df.columns:
        ts = pd.to_datetime(df[time_col], errors="coerce")
        if ts.notna().any():
            date_min, date_max = ts.min().date(), ts.max().date()
    sub_col = find_column(df, PROMISE_SUB_STATE_CANDIDATES)
    substates = sorted([str(s) for s in df[sub_col].dropna().unique()]) if sub_col and sub_col in df.columns else []
    class_col = CLASSIFICATION_COL if CLASSIFICATION_COL in df.columns else None
    all_agent_label = "كل المحصلين"
    all_state_label = "كل الحالات"
    class_labels = ["الكل", "ناجحة", "غير ناجحة"]
    class_values = {"الكل": None, "ناجحة": 1, "غير ناجحة": 0}

    # عنوان مستقل للفلاتر، وكل Slicer في خانة مستقلة بذاتها مثل لوحات Power BI.
    header_col, action_col = st.columns([5, 1])
    with header_col:
        st.subheader("🎚️ فلاتر التحليل")
        st.caption("كل فلتر مستقل؛ اترك الاختيار على «الكل» لعرض كل البيانات.")
    with action_col:
        if st.button("↺ إعادة ضبط", key="clear_dashboard_slicers_v5", use_container_width=True):
            for key in ("dash_agent_slicer_v5", "dash_state_slicer_v5", "dash_date_slicer_v5", "dash_class_slicer_v5"):
                st.session_state.pop(key, None)
            for prefix in ("dash_agent_slicer_v6", "dash_state_slicer_v6"):
                for session_key in list(st.session_state.keys()):
                    if session_key.startswith(prefix + "__"):
                        st.session_state.pop(session_key, None)
            _clear_dashboard_chart_filter()
            st.rerun()

    slicer_cols = st.columns(4, gap="small")
    with slicer_cols[0]:
        selected_agents = _render_native_multi_slicer(
            "👤 المحصلون", agents, "dash_agent_slicer_v6", "كل المحصلين"
        )
    with slicer_cols[1]:
        selected_substates = _render_native_multi_slicer(
            "📊 الحالات الفرعية", substates, "dash_state_slicer_v6", "كل الحالات"
        ) if substates else []
    with slicer_cols[2]:
        date_range = st.date_input(
            "📅 التاريخ",
            value=(date_min, date_max) if date_min is not None else None,
            min_value=date_min,
            max_value=date_max,
            key="dash_date_slicer_v5",
        ) if date_min is not None else None
    with slicer_cols[3]:
        selected_class = st.selectbox(
            "🏷️ التصنيف",
            class_labels,
            index=0,
            key="dash_class_slicer_v5",
        )

    selected_agents = [str(agent) for agent in selected_agents]
    selected_substates = [str(state) for state in selected_substates]
    date_summary = "كل التواريخ"
    if isinstance(date_range, tuple) and len(date_range) == 2:
        date_summary = f"{date_range[0]} إلى {date_range[1]}"
    st.session_state["dashboard_filter_summary"] = {
        "المحصل": "كل المحصلين" if not selected_agents else f"تم اختيار {len(selected_agents)} محصل",
        "الحالة الفرعية": "كل الحالات" if not selected_substates else f"تم اختيار {len(selected_substates)} حالة",
        "التاريخ": date_summary,
        "التصنيف": selected_class,
    }
    filtered = df.copy()
    hint_parts = []
    if sales_col and selected_agents:
        filtered = filtered[filtered[sales_col].astype(str).isin(selected_agents)]
        hint_parts.append(f"المحصلون: {', '.join(selected_agents)}")
    if time_col and isinstance(date_range, tuple) and len(date_range) == 2:
        d0, d1 = date_range
        if d0 != date_min or d1 != date_max:
            ts = pd.to_datetime(filtered[time_col], errors="coerce")
            filtered = filtered[ts.notna() & (ts.dt.date >= d0) & (ts.dt.date <= d1)]
            hint_parts.append(f"التاريخ: {d0} إلى {d1}")
    if class_col and class_values[selected_class] is not None:
        filtered = filtered[filtered[class_col] == class_values[selected_class]]
        hint_parts.append(f"التصنيف: {selected_class}")
    if sub_col and selected_substates:
        filtered = filtered[filtered[sub_col].astype(str).isin(selected_substates)]
        hint_parts.append(f"الحالات: {', '.join(selected_substates)}")
    return filtered, " · ".join(hint_parts)


def _show_dashboard_from_cache(break_start=None, break_end=None):
    """عرض الداشبورد المحفوظة بعد شيل الملف — من الكاش بدون إعادة معالجة."""
    cached = st.session_state.get("dashboard_result")
    if cached is None:
        return
    st.info(f"📌 لوحة التحكم محفوظة في الذاكرة — آخر ملف مرفوع: {cached['source_name']}")
    df, hint = _render_slicers(cached["df"], cached["sales_col"], cached["time_col"])
    _render_dashboard(df, cached["class_col"], cached["sales_col"],
                      cached["time_col"], cached["source_name"], filter_hint=hint,
                      break_start=break_start, break_end=break_end)


# ==========================================================
# التنقل (Sidebar Navigation)
# ==========================================================

PAGES = {
    "🎯 تصنيف المكالمات": page_classification,
    "📚 الوعود": page_promises,
    "⚠️ الإهمال والمتابعة": page_neglect,
    "🧾 أخطاء الحالات": lambda: page_placeholder(
        "CASE ERRORS", "أخطاء الحالات", "الحالات اللي فيها أخطاء في التسجيل أو المتابعة", "🧾"
    ),
    "📊 تحليل نشاط المحصّلين": page_dashboard,
}

DEFAULT_PAGE = next(iter(PAGES))
with st.sidebar:
    import base64
    from io import BytesIO
    _AHLY_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAaEAAAIfCAYAAADZgGNaAAEAAElEQVR42uz9a5BcaXoeBj7P+52Tl7oBBaBw6+7pHkxPz7AwMySFESkxpAV6g5ZoilKsN6I6Ylfr0MoR29yQwnY4vLGxv5hV+ucfjl3Z4Y3geMOy1muvjVJ4ZVKmKXJIABQ5M+Q0OLdG9XTPNPo2DXQDjXtdMvOc7332xzmZeTIrq1AFFLobMzgRWXkq85wvv9v5nu993hvxeBwcee8dGnnfzb0PUg4r12mH32GHZe7k3nHXb1fvnZS/F3XYyzJH29T7TgCkswvh/KVrPHPyHwtzl4jz57F88rAWLi0LSxLBnYwLH2Iu3a+eD1PmVvWkWi3hDOzSHGzy+nl7Ds8Bc+u+vIL40kvLcUy59gjr+XGViR2U8bDP+G7b9biUqR323Sd+JHhyPO6HforbtvnhujSv556bSG9M/Zvaemc90V99SvPrWXbp5EL3ZGsxx9JPX3+0Tp/mpZMrYfbebNLMkYZ4OLyHdTRvd7J51DsA4haL8k/D+GObxffJ8VNwJI/RRHxU1z/ovdzFd9rF/dwCaDgiFfSPc63TyXPPPZfc6N6wg7WD/vbbb+cvLl3I96hv+AmNEQFwYQFcXq60d3FR+Jf/x0bG7v7E0GwHetpsrnpM7l48flUQVJauB6wjP+H+GVpwWwB++SvNMDeH9M5G1gzIm3JLkQat3cs3Du6rO4DutsD9yY/jrststVoEztvxq6ucPXaCwCXMnZnz69cvaGEBTo6VmD7t8/mTLvNTedhjsgveToTVNoPCHYi/2MU12/3OdhODO6QP7gtMCwuDdrdaoAQe/cV9zZv1Wwdy8sjN+q0DR5/Z15TEbeiNrX5T9+nPndb5Ycrsj8XCwgLn50/3v5NALC4ybU5MexqfyU1fSJPk+UY9OTYZ8pmkuV47u7xgo3Oj1Wphh3Xe7RzgXpa5sAAuYGGo704uLPDU35xO23fCTA3dOe90n5b4bHQ8bZ4c8k5sChr6jYWFBS4sbPnsfCrbPgxAIAA789xzyYt/84vNX/n5bPoXTn5m33O390/9taf/Wm15eIx3tanZg3Y9jmU+AaFPQBIiHnzXv9VrN7/DHdSDD1AHXrs22CkvLcHPnz8d7m209yvqM1b3zynHM+0kzmBxkVUAOH16x+160LbseZnXrl0jcL2v2+jtfpMQ90ePzwP6CsgvyfQcasls0qjXAYTRMldWVnZTv93MAex1mfOnrw19vjA/r7yRNJjaQXn8TDB+XuQX6DghS4/QOInW4tCCNDt72R51PR9xmTh+/CqfOpLXwr7ufq+nx2uN5jMI4Wj79v6publrQ+0tAZcP+Iw/aLsexzKfgNDP+PGw4Dj2qF/vpHnOOVKfB/AlmH+OhlmcXBm65wtfeKAH9VFJtjs+DhzoDt9zcoUR2T7Sn5Xr5wB9gcIzCj4bamo8jafDqNxTgNmn/5i/Bh7/wheG67q0KNW8aeRhB0/A+QU5XqD4GUqHgDBxHjBV+vbYsY3qZuWxO5ZKvd4UuvUk+GFHfAHgSbo+F9SYvX79sFXn0vz8T8di/LN6/DSB0FYWIruxVvu4pLldPTAVCm7TETYmJ+g8GhG/QNO8y58Xceib76FWve7YrXmePr3rPuQjGJddlHlh0yfv4emaSzMCDhv8KUhPG3QUwoF0wicPTeXp4gj9dubM9n346TlOY/bYrWH6UUDWtclMOmLAswROAHjWXUcl7ndD48wCDBpM8ZOPyQO7ALAyVEPP6DPP1K2bYQo5nvKIkxB/gYYXXDr0NBBarRalntR0io/H+D45ftYkoU8LAO2FJNRv0+HDw+1xT6ZBHoHwnIDPQXjGqQPNA9366M2j9z7Ceu5JmYcPQ8du1obrPHO3SY/TdOyHcFDCYQCHCRxMDNMTqaej5Vy9uvrY7JKv3rzXrysJYXGRMDYIHQBxHMQzRhwnNCdgRvD629ffNpKPpcXYyspmU+xWC/zM6gdphE87dBzA5x34oovPJQr7Z4CwtLTUN064cmXjiRT0BIQe28X90wZU963r/HxR17NnF8K7ZxeawTgL8YgcT8nxFBxHJR3oJJ2ps2cXQn9xu9zs39va3P696ivtRZk9pfr8PDSHOW+VUsGrZxdqzs4+kfsp7IOwH445dx2T4QjMZp15oyx4QE/deozpqTMwGBsw7SdwGNBTJI6BOAJpP4jGjXs3HtvneH5kXFoAFk7OJweAJmOcdeGYg58R8BkXjiLazGQ68cS15AkIfeoBaCcLoe7z/6fiaLWAhQHXoKUlqNVq4en3ULued2aj4lHCj0I4BOGAoAPReUgxPXCisVHv9c2tExe9x7WvLCxQGmrvXvTVbvtz7PUFABXtXVqCzuOCL7ZauPi1l5OZGUzVQ32O4iGJ+wROSzgA8iiAp4yYM61P4kyfniIA3Dz6GQKnP7WTeKGs5/kLFwD8eOi7S9dXGgQmBewTdZDkXAh2iMBhUAeia+pgDLVhaarLCxf6hfPT3PDRcTm5sMBa5zON2Ez2IySHITsq4YiDBwXbr8DpkK1PVteAq1dXVDHjf6IfegJCj9XxuElCAKBFAM0D3bo8O2hmRyDMQZiB0KQ4SXB/EvwA7mGi1TodAGB5Gd5f6D/lktD8/LUhC0AAOHYMaVqfmInywyJnQTYJSy2EBsmDEI+5Yy4DJi9dX+nRUwKA5x+TyXgY0M0PhunHxr76BKkpCtMEp5NgE2liEyD2Adhn8mljc6J6z80DNY0Z60/lcfwLQ1SpFubnNT3RbMDjQUjHBB0BMCOoKaIp14zXsE8DKV+vvz6YQ090Q4/f8TiItTsJyTN6nXZZzqdZGtq8qzu5wthpNxVw0COPGnkI4BSAGsEmiX0uHkwjp08dn15vAe2lYW9z7kHbd3T9Zhed4lhc3Ny2xUXg5MkF3Lp1ma0W8E96lT65wkk7VO8g3R9Ncy6fJdSAmDTrCbLoM3n0IxAOEZi8W5hp93//8zdH6rKEIU5y22OpuBw7Dx2zZZu3Ok6uAPPzwMoScHW22dOL2N//5V9LwbgPkbMC9wOYbk4kSAKx3s4mAE1H2KxRs+//zsu3j7dvdfjSsq4e75cBFMr/nY/toL0fx/NQOqJe7P+OLf0Tf/9f/DtNJ+ZIPCXXURJTLiUAJyQckHj4vRn74JVXXr731a9+Lb9wYQiEuLz86WU2nhyPFwjtxCN6Nzvu+0Ux0A7E+Af1wt9pmRyzWOv8+Wt9iVUCLy0jZBanLA+HCRyTMEdi0siURMOAWTMddsehI43k7mKr1V5aGiwtc/PXuMu2a7f9KRWgsogWlkfMxUt5B1gcaWuxJGN5dpkvX/kNLLZgWATwtat2by6bTMFDch012gHRmxBDLTWAmshzzBl0KEeYeWoibQDYKMps4dLfXNE//mBaC2VdFlrz286b5cKvCAvz81psAVpcEnnfOamyzdyyzcslCzU/+P3lkyucO3ONi+fPCK0lLZ884ceO/QYWFlaSqXcx0/UwJ/phOA9AmqknATSARA3gNKQ5BD/mnbU77wE3Wq1W5+Tsih9r/YYWAVw8/q94efYEFy5t3eZeewHgUmteLSxhcRGjbd7zOb+Ahb4pnwAS0G+1fsuy9NWpaPkxOT4D+BEza8boJDgJ4rAzPM1259rs5VsRwB2gkJaLYhbQai1raWlP14+dtP9hQO9RlPkEhD4mSWivy9EnVLftyhuaoCR07p9ds5olUyTmABw1Yj8MKQkCqAs4AOEoXXMTNX0A4NYO6rynbS8tl7RU3VfvpFCJc+f/C3Jp2Xs7c51rGfP3J6JpTtRRCAdBNEiQRiRmCRVnAM0Gch8nJyZ7bebSkuvsQvzSchHkUwL50u7aurS0s7b3Fu3dtLkwM26RLy55Fale/ru/keTW3KcYD8MxB3EfwDrJAu1AEzQJaA6wo0jitSzi3tLS0kZFqIHOLvDUwrLvQZv3fM4vzM/r4omrAgYhHxZOriQOn4bzMKSjRuwPhiSPcAANAIcIHReS920CN0sQ6tepdGTVx/yMfhrLfAJCnzIaS4+gXO1ylvWjXS0ubr07W1wsrlpcBMtzzl0/bK0WrKcf6X5UC+nBOCXgMImjlnCGRpMAEXU45iQ9LfAddbO3yl05W62izEvLh215GWq1WuVvLWp8XRZZfMdNC+0jGzRSLbV8yEfo+klh39sTznDY5MednAPRBAAXRCMFTUjaT/CA8nxW3/iPbuCv/9/bJGSVKNOVmGOfjklKCPwnm+rUvRdqZmE/o+ZAO+jQFMEgyCUBggmYJHhIwpHcdaAJXN280p/1PTDf3nbO9/x1AKGM1DFuMql8J7CERbSAkyu8fBmURJKSWvb2v/zuBGD7JR2icJDkRGKGLhwQ6iBnAR4pKOg4IbUMWBIWW1zEEja+3wxnMEy39p6pyv+bGYelkdhHDycpaa/WjUdY5qdyof5pA6BHEXmXDzJJBRAtcPnkAucuXSPOAG+8scpTAC5f2eCl8rrjx5t64YUp4XzhYHkewPQbqzxYmwpvP4f8zPkzzqUl/8Z/+bcOSPgVM/sNGl6spfZcSEMtiw4XOpBuALgk4U8Ckj+8dqj73dXV6bwX7uRz1zvpXdyN1+fmfFx9z4z8fx7AmeuHhYV5AUt6pAs5AfnZQL7UB463zv2DRnD9Epy/DvK0hBfMOAWwNtFIZRDvrHY6EP6S4NeTJPwb7/K1vJl99NkX/3lbAnt1rp5/XBuQ+5V59uyCVdMxEMD1ry88tdGxr0TaX4X4Vbm+YmbP7ptOJQC377Uh6X2Cr0P4Do3fUtcvnvh3/oe3h8QAtYxc8j18ljROksPyClHOr4tvrBI4BQA49cLrQ9dffGOVl2c3iEvAieN/Q9fa79q//ZWNiDNn/L1vrtR1Mx5Trr/uzn8b0N9M0/CZpGZYb+eAcBPkewZ834RvGPGNp9O11883N+IZAIvnr9uvTCXpldU78W1MOlaA479aPlO9iQzgjeOFMcSVr2/w+GxTV45NaWXlggCgYmW3W8bgofz/HuH6/CSVw6cEkPYKgB5IuiKgIsXA8oAn2brMYrFcBF/sLZxnF/yzLy7HnnRgsgmH7wNwgML+YFZjD+2IVMJ+AEcAHIr0qcN5En79peVO70de+e1T/N0rK7EnWX2axkkO4vyloYcvdJIJWbbfoEMEDgmcMTJxAXl0GgEJAdCUkXMuP8YQrgNYBdAeAp3lBZOWfRdA9MDjvpMyFxYWeOLWrKGSjsHPnU6udJJppw4DOAbgEAxNUciis5QYAaBJ8YCowy4cShJN+bnTCV+sRFA/D2u1WlhaeiAguq+5c9GPS2P65OKWG7L/ZAFWLPYr3n8+eQEfffPvp6u+vt+FAxD2AYUURBGEAVQNxIyEg07NGjRxJU/Ciy9e6M1t+6f//vP8H7//TrxwoZzby1vmu7rfIs8HXNw/LWD02EhHjxMIPexCcD8pZqvyt5uwfMiB39KcubpIsrJLfuW3T6V5wIxF7pe0D+REEohYGgsEowlsRPl+EvshTc4kcWicv/qbF7NHNUgCePG3X06OHbuS1g4cDrzbobp1dZ9OYvdex7ONGXVXP9DJW7O6CBQb5nK9unfldZ1fBM6cXLFzrdPJ9PEv8PCXJxNt3JhlxEE4D8A4k6YhIYludEQvic5ilZoUMReIozJe6UZdP3fu9OqLxaLMc+dOhx9tHA13L77s584NdujXrx/WpUvzWlxc0vLygv3C1NFkX/ODNDRr9uMfAdfW7+RXL17NrnztYhynKlGrZfi7V8N7l28ld+cQahs1A4Bus+t3ryPOAPHk3LxffOMqT50CLv3uLa4AuHX8smZnwe57t8K5c6dV1hMfdGt1d9vv8GMCngZw2AxNCMpyr0RUYAPQAQhHDJxzS/ZdwuEGpDUAOHd+MbwNJCdPrvi5cy3voVJVqsUiePH4qXDwhS+FO90YkjsdzjY2vHvjoN/o1uPlr3/NXxqzkPcMCR5E0MXyUP4j9f6m9ywVbQbw/SJnIDWDWUG5giSRwjTt0Q+5dCCSE0m9U30G/T/8z3/c2cUCvVX97WNmj36m/ZqSxwiAdrPg7xREtIPdLndQ9tjvH5L6GXt0Ym021P2IpMOgzYBljLiSlw9mMIN1IycgzZDav7GBGQB3HwnojLZRwrHf/c2UydxkdyNrhrSWdJCzdi9Gv+d+gN3o07N+bSLRwUlg4h6w/tQaAOCp555x70T/oB7si195Bu16x2pr2UTG5DMwHKNrVo6JxAwuFTlUJYCgGQlwAtJcBJ8S/MNJJTee6T6TvfYv/72N2mRi6q7Xs2Qtmb29Ho/XTvgqAGAVx+YS/fKvXou3vv6y/7WZGLx5s77RrjfoZnP7lM1Mzq59+a8f9OeO/UZcWtq86z8P2BdvvFWz2cnGTJxptNFpNBOz+kbTm7PKQu7Zh/F1P/hUne+8k3PqKzXMx64a4ee8HTs2/blg2c1n8o9+7++3NVNXdrt9IDcck/Mzop4xaI5kQwLz3EUQ5aJcd9csySMuHDfo6FS3fuXNr/9mAAD5en2yMRFO1WdcN98s630AP5qCGv96Jd7O/472/VXg+MxEcu/Oem1aCJ6autbMbC7rHPMbnXvzp7vAhTg85iIWF4mlpT2b261Wy9b9zRnPdZDBDsAxA6FhAYgRBAGaUgDTTs1JPifDvsQmGwDW9n4vtWtJSTtcNz7uMp+A0Cd4PCorll0NfKvV21Xd33lksXdycoUXb83aqSvHIksaRa2WffPpb+5HyJ+W9CyNR0FMATCX4BpUyoxgVI3kNME5b+P49//bv7P25X85cRelebAt/RP31m/ZeZy3MyuHhfn5nfXX4qJQmOFpi+959xfzMIVuXc4ZWT4dEpvIcqTWCLahrhyMIeSRXVPX5KFmSgLlbq5a4g4RM8YghByYpvAZCM+y8JVJJPTjdUqCRBgJCQ0RByQ9DfjtCLRTSyeTuq9TucHSRkBMDI2YdXIVwfUaZTnrWrcogaYukiQY3ZUh4d3JRpInSbKBt8d3yfTxq0wmm9Zdb9TFuK8WwqyLk2JMTS6a5Z3MnJYLDUoun2DqwaKnNFK0FDGuOTd0ryMaD8nxPKBnAM4JmCZJQYhe7DZIgGQCcgLAAaOOQzihLF8T734EACnZaK/dC4m5IzX1H/kc8EBNxDpyBsZVBCONRDRoI6XdyZTdTaab2Zkz97CVsZ9ard07u59cIW7NGkbm9ptfefNQp909DgvHIJ8DMEWz4ABcxWNnRgNVV9Q+AAfhNpdt2JzOtu7xpaWuWi07D9iZxcWIxUXi5AoxZJq+tPlZG5q6/XQh99to3m+N2C2YfBJlfioprsdVEnoYbn8rScjHiOZbccvcqk4S+LWvnUoOX8knMZ1ON2vpJGB177p5oOpJEukemaYxMDpKIsGC00JKc4+dgLyWmthOUzZ1UPLnGcJJAj8P8iSBp+tpYg7QBdTSgCRQna63SVyS4xsuvSLxRxMebsYaotxUS4NyIOQxs8mQejsxodOB0gk1LDJakjhjYh5NTtVZ7yJBu9PE2tSt9vozC8vt0lp4oPQHiFaLH/zKm80s2P6AeCT38DTFZ2SYEzQFt8CAjEJXzhxEBsnNAgCX0dwRaQyUPACcVMRRI56V+DlAx9OaJS4hFkIBzIAYBZc2INyUcFXAO4F6h7TrkLedNANrLg9GxNLOGYDBoQT01CjzCIHqkHbbndfkej+dCD/hndqN43/3tzd6lmY9izASOtdqJS/83Yu1ePvArFzH5HoW8qcIHIRpAg665IJlDnQpZUZ4Yi4A9AgD4e7smAV51CwMJyR9RcDnammYBIWYC+5Fo83IEFjQka7bLl024ZLL3wZ0yx1MDDUQBpcMdIFyOEg3IUmilIC00m6zbaZbkl0113ts2AfJgdXbxy83O3xpOVbpt17CxB/9L/9BrbH6wWRMbSY3TiYeU+ROryHKgidRLkvcc/dizuXqWmTDUkYiOhHRbiNYLckijprhBcG/DOArIn/OiGP1WkCeC9EjaqnBArC6kbUpfUeyP5L052yHt9pca9cBdLOUHtyZRa56bmkWvQ1AMRPabeSKhtQSNwvtjhuTGIHQ6bZ9Y1/N16/fztZfWlrZKlPtTkNRbbfW7HbdfdgynxgmPAKpZlxaBD1AOTv1JdhNDDSVptTqLVBn/6M0icdr+6at+zQUjzm0PxjrwQTFPDKE3GOeSYpKqNQoORkpeIpoYOZOtzSvR7dDISTPCf55ks8AmgHMolTa4BKF9S4oqAbZQYGfEzwH/ECehhsQukZ5DiCHkqSesi3JkHusJyJzz6QAdmsm1AiaJcpyy+4i+g3m6Yd3gevLywtdYDn2AIgk4A4sLuLonc9lV+be6HS7dBROtJ+R9AUYj8p8AkCU1BWZGdghLQpuIOmEgIAIGEGDqwZh2sl9pO8XmESJEPrI5wPorwnYL1dKYkbkM5DWrPgdEggECS9DGBEqsEgpgKakBEQH5IeSfmxCW0nygda68vaabMjUWQAWCSzpPOBnTt3rXv6jqTy1mnmOfVaMz+fgOgagCZKgOhQ2CHVJFkYYJoIywUBj7oWOqynHARBHjGgSGmRq6G/VexIgEKFJSk9LqEH+HIENgygxoQCQcpdIxUC5i4mkBsFaFEnDhhk/dPjbRnXqtXC13c391uXL+fGFwnhgeWHBdLYw9+4B8atnF1LWcSAJ4ZnE4zEEzLhZAsRoVIwJomIek4BcdI9wBScyOaPLaZYpTYSIugUdEfk5CZ+n8DSAKQDIo8O9GuiQMLAm8CCk5+XoRMvm3K29LlewXNEzudwoMTO55IqAmNac9BCj1/M81owgo23kMb8dhOt3V3W9NhGjhKxCMavVAksT7p3S9A/K0jyKMp/QcY8QiHZrLqsH/P5hAAqNmYz0pIaEMxKOEjoOcpZCg0CQHCQzBmQonrdoRgEOwiKBDKLLUDdiFtAxgscBzIFsAmBBV7DYqkpwBwAGSfsEPEsyBeyIU7dJdBxwKaeZBRdpLJZiKxRLIhkgpIU2GF0BdyF8EIJFRN0NUwhzzeGoC+7Ovo/IJeTJmbwta6xnsk65eDYBHSys21AjYCyAJKcpgkaSA4MMgSh4PyMRAAURgSzbqyG9VO8IkJoA6gD2AXiKRAShciEvjcqqSxoLKIASgDmhmwA2JBjJrsR2Eti9AcThwV7k8nIRaWBpacmX/gn8rT/+B528EzsGK3fSbAA4CGAOwESpt8tA5ECRlJuiyqGDSqpNoBmRiEyNsOoizEqv+wCZUkEHyhhzOQkfekYKvVnRA4JLCAJSFdlq2wA/krROIBgRoysLCbonL63kpZOrXbp2jZcG2XoBALWpmiGLdcj3C34cwFFAMyBTuRsAGJlLymDMAUZAcok0OaiMDjnUIHRQwlMEjqtwSK1LRIwqhFaWmw0HCDNJ+yB8FlIC4Bkj2xJddCkWuyJCFIo/RkBGyRFEpCYCUMfBm0Z7300dWrzrtUYo/Yj0CSz6P5NOq592EHpYv4xH4bC6mzI5jSm/V887jJ11hNqGHDBiRuBRAbMkGqQEsmvGLoFMYKQBIkUplqtIWuwOtQ/gPgCTkFIUuhCAPRtnIXq/phMQjhBoyngE1Bpo3f4SDxjkQLGzDTQmVsRdCwIigXsgr0PoCogKasvU7uRpPmKUwMXFRZ5cWeGl+XktLS35G7/8a9l+1u7ltfCBOd/MgYmCv4ITOGKB+6YaNQQr1kqBiD5YWPtYIUFerJxRxbu2iNdd2CbQGGDBmIRgdeu7Uapv0mVW6M2MRMyFThbRydQVdMOId116k+BbjLxK4+3U1b5+6ZpvK4kLqN2rd7tp5w7EK4BPA6qRDA6BwvF6I5msp5ZaOVYx+mCRrQSEL3Rehf7HXQU89VRBleWxFwmdJENiSSASwkpEY39O9PSEwcrfdaGTCe3M1wXcJvEeiTclvhWCriSJ3eGt2MXSlg6SAsDuatenp6wdc6wJ2FDhNnyQ4JygGRB1wqOAroEdCRnBWOx3TAQiSEmsC5oCMKti8zABIClRk73ecQcklZZymgRwnMSEoA0AXVAuh4oyRSPgFEAEkAmkIIIQc8DuQPEaDfcgZG7eTi20vduOlbn9Sa0bn2SZT0Boiw7mGB3QTqxMOKLY2wvP8W3LXFyClio79K99bTX+fDvcXkuynwg1y+CpwH0Aj4rYV/q91A3MSXQEZDTkLKDBaOytn4FACqgmsAYhIQf7YlZ+s9hREwJqoPaBNhHIDFAOIIoU0aeWRNAhpQktdVedgIO8a8C7ErpO3UgSf98CriDmN6YnJza+fOawlgc+N1haWuICgPlSN3Txveu+74UT95i136biuiXhA1BvmeznovRFj/78+kZ2dGqyhiQQLiKP3q872HtHQcx5tde1aSR715LF6lGATG/16lF2ghHwysBl7ujmvmbkWwBfJfgqgDdg4W3L7FrWye98kHXbZxYvRC0WpBoALS4uYWWlb31iLQDHj1/NfnQtvzkR0hjpd5DhCozvkfgixJPdTv7FgGR/sx5AErGQC0pakX1ptlC+FH0AAmVAAfZkoYqWmv2EFSRgPZQqGqkCG4syjbBggIQsi+h04j0Ge8OA1wS9Jtrr8PhuNybXJpLs9izQxcKC4V/8i0J/duGC48yFisFNi7Wpm9k+v3FzNc2SrAtDUA1uBxz+lIEHJc0WgKPIgLYcmYRIMGCwZwKJUMxp1VEE4Q09HC3FxWL3ot7eCSiiaitImjFD7k4n6BGSARIBETJSglIQtWhIzZmJuOVQm+S6pOuBuFJL06s3pnELV0IHAJaXYa0WtLRUhDBa2pkeZ6s1inugE3rQMp/ohD5hSWgvuNbtdFGbyrRK2YVYfzG+vIh7v/u1Y3lzo5FRASAagCYITpa7vgbJSSMnS/2Gl1QUi4Vp8BqsOdx6xvUkI9II1EjUWIohQ+X0pATJS+uo3nerhNYguwbqXQmXo/R2V7hyvXnt7osvXsills3NnbdqCJTlSt+cuTKlw1fm1rm0tKqzCx++NzPzLpPwNg3vm/s1B29luX/x9t3OMTObStNQ0ohyiEU3jtgEqQI4w1qCYWquoCSFbHBjz57OQCHK4XKn0AXsppFvC/wBZN+W9APP+W5m07c+/+0DGStOngNLR7CMrSYAXFhZ4Zn5eeJ3j8UXlpbuArj76tmFD2cbyTtdS94N0k8AfQRhtd3OX2i388M0a5RGXypMA0qyjPf3R9s08VRwhbGiHIOgEjALxVcGrivCo3cAXAf5Jg3fAfA9gK+prrc3burWl17677qlIsQwD5M7SaoMPN7fhJ1cWeHnMZ9h8T+7vfEHf6sbw3TmTCn6FMVpwCcINAFNs6CNpzCQhG20NUPNZuWNg4naG9vy+gCiSaIpEkbBS/643GQVezTKSVrpDewCO4TuErwK8O0AXU6IdxA2Pnzp/7yyChXjfOkS2N9k7Aw8tMv1Yy/oub0OzvoEhPZYetrJ7mEvJsq2IvHSEsqIvlfXv/GfPnv9nnuwoES0KGCd5D0Iz0k4YoHTzVqAGYxkQT8Uilj0Qs/5SDa6sbOvpJ3AgqIKRpgVqvnBeY+iIQSau6ObKwK4AfAnEN8sret+GIE3U+KDH/xt3HmJhc9ImYph7LEC8OTKYXG5WMBLZ9ubb51rrU+Gj9rduL4G4bYDN0B8DsDTkg66MJ0EGlHoAop1o6BkWFgRGLdCX1QoKkoApeKGoj9hZgTyGAFgjeAt0a9RehvkG4h4lbBXN3zqrRd+4z9/IJ+qKmB96aXlLoCPvvv//o83Ds3dWkfM1z3JVylcc/GEhOMA94uaDMYACDEWApK8z7QZyqX0/m0upkchPBT9RCsW9zwXAKyBuC3iKsHLgH4I4w8S4I3JNL5z4FeX74y2RWfPjs7t/nN06do1vrS8HMtIp2tX//W/9eGG708hq9E9grxH+G0HnoZwJEk4UU8DjLRC6i1o1n5CKReie49yG3q6hh7aYm8iFr5hxfwVIC/En/JV4pwjy906nYh2N++QuEbiHQE/gviaUa9D9vZ0hutf/Y9XVvHk+FQs2p+2um0Vs6rqn+A7+Hzb+FfYeTQE4f7Or6P1rOY6Cf/o9Olmp702a/XaEQLPBuILQvg5Gn7OyM/V0rC/Vivomjx63weGpd5EGCiACrDhplepJyj8hfrUVO/c+p8nRiRJseCvtTM3s/cAvgZwxWg/FPFml90rNdQ/suTWvRd+/fer3uhW5m7RNlLrplAxH577R1OOtUPexVNu9jkRX4TwJRp+TuAz9VqoBxpi7ojRCxFGAolSnVYl6jVsmaD+cqXSDEEErJ4YgxnkwkYniw6/CtObkv8QsFcpvREd77CdfPjcd5+7y61D3Gw59luHxhGv/8v/61TOW4eyxvoziHxews/B8CUIzzt1PE04EQIL3ZcTeS4RcElG27QMb5q96uuHJBAMATQraT0BWaYuhQ8gXJb0fTkuGfjj3PheSn70mW9//s64NvcCi45pOxcKydcHHd6yy19fmc7vxINJSJ8Klj8r6gUBX5Twc6A+N9lMm7U0wCPQzRyxGND+ELpXdX7aYms/MItkH8AcHh2SF3ah7iVp6eh0I+6tdbuALsdcr0bXq4Le8Ji9k+XJNeDuzb/9V+buVkIdFe3bfV6icSbTD7u+PkyZexkz8wkIfYwgxIecEFtdw9OnwTKelQPAuX92uhHbnYMhhs9KnBf58zR+heQJAIdCsFqP9jfrqSIolnvjKthsBULV81JVAzPSbKg7nMBdUlcE/tDAi2LyfUvsxxtx/YPb+2/3wt6gR9MsLu08WcECwPnTp7l04UK/7Tp3OrmN56Yik2PtqM879PMufRXkFwkcozjVE3l6Fn8Dqmo7EBoaCIESQUusUI97rjaga6J+DNOrFvTdGPMfJCnffgNXbr/44oVcrZbh5Arx0rJz5/NkyOes1RrQdS20uNgCcAZ2G29P3dnIj8vtiyHor0j6eYd/XvRjAqYSMzMYPJa0ZCHIcCdESykgCIDRBJcjujvEDdKuGfgmhe9L/La7VjoxfX/9evPuqSvHIpaWeubHvY2F3+cZGdpwqGhw33Dj1bMLk1N1zIF+woGTkn5B0FcEfRbgbBoCC+oVTtDGSFrYzli1R8v1aFd3VyxAiIVEGRHdEWOMcr8d3d+T/Acx4i88+vfB5K32+p2btdXZzpmlC7HX9tOnYdUEeR8jYPxMg9CnnY57GLv83eYV4R7Xr7oQ49rI9S/+wwttAO+f++3THXazdkRYM+k2yQ8JfpbAMRf2g6qXfj9uNBdgvXRjFRXRfWgawQp/SbpIFO6OGYl1ADcBvCPox5BWXOHVbo430/yFK7/40lJ3tMzllRWubG772L5dWID96uwpe+Zk0377C6f4m1+7WNBzL17IgQu39crL3Wv3YtYBu57nbSNvuvD5AD3twqwRTYiJ+uqB+8ZAKqWmHtAC0eXRsQ7ptoAPIb1N4IdyvVar67XUNi7Pvvg/3a7wp7rUWkjaL5+ys796wi9dmtdisUjvaKFotVo8c+a8HT++yitXpoQlOJeWHEWw2Nvvnl3oZJON3KAc1JoLtwB+DsAxSPsFTAKoWWkG5xhYhI0b6R5XaT3TQxcgdCWsA7gj6BqhtwD+kNCrmWOl04nvfOml/36Ifjp7diFcunQpFON2EsvLy3HM4rdpzBcWwOVh5Ym+9NLyqqS1d37/f9v2btom8rUI3KVwndBzkA5CmAGRlkNb0K6CVVSV3O6JLCW/grAmiwkeQbnkrq6kVQAfQXiXwBtOvgrp+4wbb/6v/sPvXh/3/H9hFbyws6goDxIu55Mo8wkd94ikoq2kmdH2bKX/0Q7KH3eddtB/9yt7E+VxrnU6yQ+t78tDeijJ7ZmkFp6H25eD4csiPgfiEEMRH85INxoLM1wKfaFoS0moZPDIYJBLVlhLMafxrpFXCb5hhu8a9Jqoy3JeWf9o7s5Xf/Nr6+MGotVq2crSEpc377Q0nqYCT64scOHs5sjVJPDBH/+jqU5c3dd1HQsMLxj4FYF/hcTnSRwm2CTKVlih7SmQBhWxSH0EIukAQhIMgtTpupO6auTrcL1G8DUKP+qi897sJD/c9zf+u9uj9VpYQJifhxYXMTDH2n6sqyoLLrZQ+g+Nn6tX//X/YXK9w4MhxGci/QskvwTh5wz4LMljEGaSYALEPAoqjDU4CkM9/RdBC4Xdt/LcCeqeXB8K/rYLbwTDDxDS19nJ3p48OP3R3N/4r+6NGVo7fRp2+DA0vwwtDeYq77dwlgO/ae7rlZfTD95Zn11PNo662XNU/AKELwViXsJzIGaDhRAdYNnGsqiq18Fw95UUY0HbyT2qCGckZ5ZHyL2tGG9J+onL3/AYX5XiD7vC5Yk7G1cwjTtbBPCt0stV6n03tPxWrIx2sG48ijKf0HGfAAh9XNc/nHg3HPSTrRZ4/PipcKJrMwknngkMX6Lhl2T4ead/VtQhEvVgZokFQEVwgZ2CUKnDLRS/UtfM7prZT8z4BsHvmvEimb8xE+K143/3dzd6Tqf92F4j9X2YvhoK83N2IeDSvLi05K+eXahNz049DeGrIv/XJH6ZxGchzvQUQUb2Fp+xIFRY1sFBJvVaAACsrmU5iNeN/AYN30qAH9Dzd27c6dwujQcqksCyKhlFtwvrtNOHvH9fC8DJswus5g164/f+/kzqOoHEfpHyr1L8eUkvkJir1wIkoZs5BESoVA1ulnSdQEiTwtik042QdIPSjwX8APKLIeFFmN78zK8t36zSqlUKDVvn39pN23tUpFd/572TK/X2VO1gyLvPQfHLBL4K4EsgngU4C7DW03tW/aIG7dXmH+6BkMsEoRsdeYxtRb/pHt+B6zVI33OP3/OoNzdSfHTmH17ooNwgcGlLkN3JuO8m3NcnWeYTENrj+u2GXnsQENqJxPTQ4LO4CJ45c9pwHjiDIkld9Zpv/dNfm+km3eeSxL4UpZ/PLf8yoBcAHA3BJmshgUZAqNQbDXxEAJixHwqhp1+J7jmImyTeJfhqsPC9EHCpniavv3vx2pUXl4Z1Pzi5wvNlEr4zZy7EbezSdpbcTyLwkp0/f43AGUy/cZVf/c2vDSypz7WSd7o/+WUz/DqAM6K+CGE2FNxUD4QKjTML04Pilx2lpFBop8m0WU8AAnfvdTNB3zXyD5nwfCPoB0de/OcfDNXr7EL40dTR5O573/D/5OsXfWQ3vNV8um/bBfClBdj8/GmexGH7hV++x8//27/frUpeb/2rf3A0SfWLHv2vS/olQV824nizkcBdJaggL5vfN01nH3vpJJJaajASG50c7v4hwVchvAL4t5Ike+WZX////aRat1fPLtROzs378vUVXbo0L2CpJ7n5yLO2o2dJApeXF+zEict2796Uzp8/41VDDbVkl//K/26Oyp6T9CUYviLoKwBfAHCkniQBQGnJ51GF4Tp75uZVKagXPzdGSRKjuzJX1z3eiLm/JY/fB/QdRP8hsuzHX/3Nf3N1mH1AApzG9ZOHheVlLGNTMjttowPciq7jLgDj4yrziU5ojwFyO5H1ftTZTkJwPHLv415wg0Vc8CJi74VNv/PLN3959evNP3unMRM7GbVm7lkEmgRnYJzs+1UU3JP1yy0i7pQ0Xd/FYqgPjewKvAXXmwReicFe6Xay95oHZm5WAagnBUngmYXC3+nFF7fkoHee3K/oAD9zRlhcXOQZHBuOwnxmMYb/5f8ULVh05D1fxSGmp2yol30wMJQqbbjRj2VXVLO0HHMCmUlZDBY3VeylZX//XCten72oUim/E6DRTiawluGLrQu8hNM48d7qptQezUaijZgBcpnRqw5QZQgf0eA9c2ZWgj8M/D2pqpQkUEY6qEiZq5Zsquv1uWuOM/O+gGUtLAyHA9qlLnUwt7XsQEvAkl58cXhuc4muV16++dq71/KaamtmfteVO6UZkYeSgCAQWe5FoJ0yEofUN4frhZQqp1KhJVPJz9LVJvCRUT8mcVGRF0MDV23q6K3Rup5ZQoQKx9vFS+Dy0rZ0I7dYd0bXo506o36cZT4WvkKPk7PqqO7hfo6sD5Jg7uNokFD4Dm2ip86eXQi4BP1b/7ev33nrn53uXOmCueOQgV9EEYSzCDUzEjBAAq3nUbJ9V7qBGyCvAbicyH70C//737nZl3yOXw3VEPvVII571Ffqg4OEi1/7zaEKXzv/jyeZ+CSICYp1UIGbgYySQj+2XBVtWTUZLiMFFLckIJsEp6yDab3y8hp+91i7305AOtO/V3s4PwqDhiWo1TqDUy+cHwJdtVr2bvetKWOcFjGlIu5d0qeb0Nv1w3px84bHvdCBEaWurKyiEUHQBIAZEDMxw5RGM64WS7DIHeUE2jEQVbOsDs3thYWA3z0W55e+duP933l5o60bJuC4oK9A8BA4CERb+T1y015qbF2soCzXSX5I42U2Oj/+0ksXVgdS9ukEhTQv9rNv7Xhd2Gun0E+izE/tYXhyfGKHBGJ5wc6eXQjnWqcTFHHbAAAn/uGFdtfRFgrwIZkkoaBcytXNS+5cFR1QFYm8CBEtL+k5BDODkFhhMdedATb6lTm5wrdrnYDjV4NaLStzxvHR90HLdHYhvHXuHzTaWXtGxH6H73f4tBlSM/RCXUdBTgLBzOq1wGY9QbORoFFPkBZ0lBWqesldchWkjpmaNOx3Yc7FA/fWs0mcXEmkhaKtH9fDO71KCTx7diHo7EL40S/fTK2whpu1IkPqPkh1AcqjyyUvVT8hSchazdBoBDQaAfW6IU1JQgGFj43HKBcgM9YEHARxhMIhj5jCRjOcPbsQpJZJ4vT0F/io5/b586eDzi6Ec+dOJydmLxuWFgUAx9u/2kGStEF0UURUKmT6vsxLlXFqx85t9f1ci7kdjLCAQCIh5WbJ+hAAtVr2NpCcP386qHBlfpB8PsLe6ooflzJ/pkFoHC9veLAI2p/e49K8FhaW/czihXjp0nyOpUVB4jf/6a/N1C0cMAsHLdgMac00KRxNVQRgdoCRRCTgLvXSEjgBR+Eq4hCrD2oKYjpKB5jYgdWQ7+/vOF9ajs9Nrme4cixicUnVcDx7uTAJ/RhoOHXlWASWhEvzmroT0tS03xEPOvygoH2WMA3GnsVbjsKpfkDPlKG4C4mn5G+Ka2OBQS4zkMYpGo8Y9VSEH17NOhM4MStgXljs+cic5OJi65EtzIuLS31OaWFh2bmw7JPvdUKeZ5OIOCTgiKSDEpqEVNiQKKqMBaAyQQdNYCjeSzsyCYjuioXuSArGBoAjFJ4W7Ih5MvXe6rQVRhFLAonLl2/xUWPv9euHhYVlP3PmQjz1tYu5waSzC+FK4+uziH5YxEEFTdBkJdoW4hzhINyKuV1EkCgC4TlBJxD7RijF3KYZGyBmCNsXkc+89c9ON6rU8nPTq/H8+Qv+EJmOHzcgemzWx0+7JLTT6AaP9VGEu4KWlpYcrUWe+y9emmzX8iM58BnQnhLtoBmbIVgvwCVVhKkLABIQobKiWBE1GEk/4k8JQsEsBbCP4Jyk4xlw7Fu/9/dnqkDEpSXf65TkQ0drsZ8Mrv9bJ1eY1dSQcADkIUkHRM0kiaVJYgQQIASJCQHmufvaRhbvrnbyO6vduLaReSeL8MKPMZBMAASp75s7BfhxQc+SOApxEvdeF1lp6/lLj35e3Zsq4mqWwS+6hzs1ETOC5iAcMWK/GesAzV3BHWmR7kPI8ujr7czvrnX97lrXVzcy7xYKFBHFtdEVykzn9UDOQjgG+BG49sV0ot7vd0Anbs3q45zbBOSt37J3J9IZ2erTpE4Y7WnS9oEMvUgJZcS3oCJFQwIwkD1mFSYogEhKn6JibgdjMNYh7HfosDuP3rD08FBFLp/wqtXezwLR8kQS+gTX9G1eD1Pebr/b7nsChaXc8smVIWuY8zhvNbs9kVCHitxBOAxoGkSigaoAEJIk0GopkYYyYnQRxR5JStRqRJKQAqx0aPRQQFWThn2AHXTXoebG+ozOtmq9bJm9Y2UFXNy+L3f6GjrK9laU6CDmJlKGMCXxAMBZElME62kwhGBIE0OSWNKohZ793x1Il0F8n8D3BVwWcANAZsbQqCeopWFgLUg2XDgC6mkBRyBMbhqRN67yUcyrIarv+uGBvqR1OgmJJiHf79TBCBwIiTWSpKBck2CopwlqSQAcDuEWgbcJvAZghcCbBK5DyGsJ0agFBGMwI0IgGjWDpBm4ZkHM1oLPvPrqQq1HRd27cky72NY9UNvn5uY5RD0vAt0sTmXKj8P5LMHjBKcJMhZOpqXOEGmamNXTgDQxlC5QMBOShExTIiQkKJYSkkIR12kCjllAh8105Nw/O73/7Nn5moBqmm9icVfrwbbP8AOuLR9nmU9AaI8lIsfOTa33Aoy2u4/3kc62urf/2coKOHfpWvVBLbU5eTOSByAdlnRAQFOCPLpLiGU/9AnyLI/I87wj+T13b8cY0ZeVIJDFPSVVlQicALE/IOyPMZn5UefHjX5Cut6xsICVhT0BIQ4oqVavvYPfarV4J5loeshnQRwKZvuMoR4Kvqmft81K00KP2gD4DsFXCJ4DcQ7AtyH9WI6beVTsiVrRC18SAwOE/RAPgzpEhqn36ofTzUO2tJOHe8fzqppAj4SwsFz6XYk4ebgB+CyJAzDug2EyJCSt5yrDMn6gIPEOwcsEv0PgTwPtT0h+G8KPIN2M0TGIFVe0uYzqmQpxSooH83qcm7k7M3X+fCsAwHmgLwXKtZWe5IHbjsUWpwfA3v8+0Ceiaw7gcTgPEzZBUHLEImWUvMdbuooYc3nuHXdfy6N38nJuk33DjQgxqrArTOSaNnHOGI5M1pIDc0UCRWJxYDCxfHKB0o6f/50CxG5A7ZMs81N3/LRF0f646bqH1k29cXyVVeriXAtA6k1Ah0AcLbKRskGALoR+1DRJnSxSUpfkrUDehNQR0cxdh/KOZpMQyqUJSd+ksBCZJggckHQ49ThrE/wIwFAYl/lL17jyCET6ant7R97xCdAOCThKYY7gJIkEXgS1zHKHXBmENsD3CLwazF6R4R0AcOkzkO4B6ALI1tv5kSSxWrGa00jWjawDOADhgIMzuDvTRNUw44VjWnwZVWfVh55X7iiyzVbSvRfU32L4sB4mQT8AswOC9gGYtGBpoWqPyHJHJ3NJukXwTQivGvA6zK6KiJQdkvwWCOXRLVvvHiq4LPZy2IFFHp0pEoc86gg6+Y3n8PY6gHw42Orio4guP7SJJIFz586Hp+OBSVg4JOoooYMUmyjyrFpvF+buvtFxQGgTuk3oDqEMVDM6DuZd3x/MTKWOWD1jNyEBMCnokMuPQvgw35i8A2C96mw9d+kaFy/t2IR5t7Ha8Ckp8wkd94jotUfBke7VzmFPkkpNH18lESYAzYk6TuKQkY1Qhq1PE0MIhKTMpWsAfgjo25L+VMSfCvoWgO8TeAfCXUKop0QttTKNA8Eih9FhQk/JdDizbAKLi0P1PH78C3sdfBEAMDu7sancDccEiDkVKR2eAjVbTwMIII8uQHcIvAPguyC+CelbIeA7TNJL9dB4lcBfgvhzEX8O6DsOveHya5K6aUJMNRPMTKQorM+wT479qfKZV88u1Mr1i7i+speL8FZ0KwDgvet3UwSbVMAswQMG7gtmExONBGWkBAm4ReLHBP5SwJ/D8ReU/WUj5fcC0++nCN+R9G0S3wL5CsHXAd2AENPEUK8FWLAGwH1F32rOo2bad0K6ua4r3PO2AwAuDo99t1YP4hRds4QfJLQvkLUkGGrBUE+sUG5KGxDeB/CqgG+J+hMAfwro2xBeI3AFrnUjWK8R9RoREqIwY2ehW4w4LOeBSYVGaxPzuvpTp1d+nI9PMwjtVPTfLoiftnjdr3xucf9WD+ZWEYbH3dv/bH4ewqnhgrqzJ5KOx5kIHXbgGAwHzIpAj1Jh9tbNHU59BOKHIL8B6g8o/C8G/L6Rf0jwTwj7jqT3IK32ggxE916uliakI4I+Q/ixbtaexvJLVjVfPQVgAQvb9eFuXpXjZI95ZEkLGWucis6jAp7xIs/OlCBsdHLE6HfLIJTfCcY/cvc/ZNA3g8U3agpX8+n8SlrH63D7C4HnSZ4j8E25fgjpeh5jAd5F8Lw6hGkXD2UJ5yzJpn/7a6cKNmDh7ICeku5nxnu/eTVgNTfdKGqi3egAswLnQBwEMJUGQwDh0UHgrlGXIf05wT8y2B/XIv98Im+s1BTfmYh6N3Qbb4QkvGLEeTn+CNI3Cf5I0u08L3IiJmkAjVMgDhl1OO1ms/tq9xqjjbn4tcu2C8vAbdveKsPiLC4t6d6VqaqXMa3dmFKI+4xxP9z3QZggYb3434pCzGMG4QNCr1rgn4j8fVH/s6jfN9gfmdk3KF4CdNUVNyAv4nBToEEi6g7NApwz4QC9NnGmddp66dEB4MqxjScg9ISO+1iotr0ye9Qu6rYbp04tLgLnz08NX3PnThOG/QIOEzgC4ECRClpod6OLyARcB/gjGr8XwO8A4RIseR9mOVPuZ44blNqQMlJ5N8ufzV37g4Ue4zFF6CiB26AdMfj0JSF8qdAdDRbQ+fk99zc4cWK21HiULNipq7Us+qwBxwE8DegQgLC23s1J3DLjTwi+RuN3zPTdWtCP0Y3XD9/stHHpOZUOp+tvnfsH9ywNN7nBm4LuuuK6gbkk3lntHG+kSRF7QJghcFRRT9XS5u0XZk90gIux1JfthJ7aSZ8QKIw7zp+5NjxnW4vMfXU6pDxK6BmJRwk088xx514X3SyuAnoTwPdCwEU4Xs3c3g737MahN5/Z6Mfza7VW3/vbd+/66tVbQLgJ1x3AOwS43smtk/msF+ZmTQKH5Xpa8Hcz2k/Uat2oxo27fGWDe/VMraz0o3Xo7MogasK584uhicY0PZ8F/ADI/SQmSSCLjjxXRFQbwlVCPzTyO6K+S7MfusVrlqUWkB5m0F1HzNxjRiFrd/Ons+hT5Y8aiEkAc4COGnCwwfrk9Mn9RiL/BOj6J8dPsU7oQRfGTyqUxdjfNYP+h/9hYDF1rnU62ehk00h1AMAhAYemJhJ4BPI8dkjeEPihgLdJvoZgr5rzh6nX3/rFf1ikI3j17MLdjN0cjI4cbQbdBnjTXc8CflCwfY1akmbdeFhF6ohDeQzT6NbTUqcCALj3wuvCy78tLC3tRdtL0F0UsOzA1/pfriXZZJ02B+FpCsdCCHWXr4H4iYS33fVmYniNCCsJ8MZBXP6Af3vU8x/47Iv/vC3h6tV/8w+z2LaMyDYEuwfirlw/565DgCYEzBJ4GsAzSRquPRUatwGMRAxf2W0akK3p1TdGqJ+TK6wlYX+EPQfh8wCeLh1Ub7rrhtyv0OwHpL5nwKV0Mn3rGTz3Ef/e0qawSljChtTqXDl/tZt3OxnybiZpA+Ld6HoWwAEQCYEjAj8r8H2DX/7RL9/8CYB+gsKrN7vEgb151q5dG8z1S8v96BX85vLdNEkxBeCAwDmJh2YmUmQudKOvAbhO4SqAN0muQLpkxBu3OfvuV//e8joAfHj2H63H+r3YyVcjgTUHbkq65e5PQzgQDJMQ9qtIH3VLxGFamLpbcQJ/wLWGj2D9ehzK/Jmn43ZC72zlQ7SbBHQ7oJAeCCDHUnwF5dUzLigcF8u7mD6TNYPFAxAPEthHoEECWe4gcE2O1wD8qQl/GMzOCfjLRhLe/oW3f6GfjvpLLy13N56tfZghvCbEP5P0RxL/WMC3HXjbXatkYdJL2TQd+ww2A6wNmS2fOXPGrZdVkw8EPEUWPgzSNRc684W+xHH27EKo1+MMYQcBHAQ44a4ugHdp9gqC/TFlX5fsz8yylZsbP7m+OfRM5UcJHYufuRWYvZVldhHGPybwh3J8I8/9Msg7AOogDgE4lpgdSumTao1oDc5f43B2hgddlBdKYrMqCs6a3PbT/RkAz1J+AGCHxJsS/lzkH4Dx6yn0TTbwo8PX12/yxaVt2rzkx3HsNrLsskf8hcy+7vA/EPRnAN6AuApgCkWInKeQhAPAzXq1jGMHalpcKqzHuH1bt217qwXizOn+tUtluiOcb4UDjQ8mFbAPDAcAmw0hTIRgaHcdAK4K+D6M54HwhyQvsOvfn/L973/17w1Sihx56f+5mubpT7o5vwfxT2T8I4IXPOJ7HuP7eRa79YQw8gCMRwDNOfKZI0CtWs/jx5taWhqEd9rBs/wga4I+oTIfhDV6AkJ7AEa71TFtV/6jpxAr5tA9PcS5//p0PcvSgwKfInQYQANA+85qhm4WPxT0OomLRv8GjN+sJel399dn3vryG1+90wuH0jt+5a8vt9eO3r5aq2+8BtO3CX4DwLcc/D6Iy+sb2UanePgDxGmKh5JGcuj3fu/XKovTUjVBwgP7LSwCXBy0V6wA25ePT0/AkllIk2VazBsALkP8gZn9RZqEb7IWvp0l8bUjL/7zD1749d/vSC3TuVZS9Wsa0t+8uBSf+tMTt979ixOX88DvwO0b0fWNLPfvAfoxgFuEjMS0iP311CZw5r7PwZ7Nj2v3btVpmhIwA6kBoA3qPQjfc9c3zezfIPDbR+dm3njqxf/vR7g0n+vsQiifVUML1jMGI8p0GOfhz/7Gf3fr7ekXfhwTu0jan9L4ZwC+A+BtgPcIBICT7piywCG90K2rTe1V28+M3tta5I82fjxRN5+j65iIOZBNd+S37mVw6T0BKyC/LcM3xPCtSe77wXP/Y+29Q995Zu3swkJYWOi3H3P/m//q3pWDN97FdPoDWfIXrvBNj/iLGPEqobc32nlemsZPAzgo4OCtjVv7qpW6cmVKe7DwfxJMzsdd5hM67hENxl5Gy36w+0cspgTwDza6MzWvPQ3isyYcKYLR4ENBGxAvA/qOBX7X6K8rhPe++vf/1d3B/Ys8e3YhzF26xjOLRZDGF3EhB7D6ym+f6kzPnshWvZO70E7INYDrAJ4uI91Mk35cFp55aqN5992zCzeefWl5YyRqwkPRmIuLGGL1BPC9d/6jRv6jOwcQOAuJIK5C2IBwk+APLNh3p2cbb0x9+Z9e6wMXgPPnYU9t3Azvn1+E1PLyQzt3rpDeyCVHsat3LOH29T/9995cv5MlmcuTOm8y8mmxnPdEM4veRHg7ATCQNs4c3hNd4vz8PO5d6afANpxcST6MYVJCDURX4nVAH0F4U9BfetSlRj28eexv/ffXqrSbzp1OfvvlUwVQHgeunJ/S4tmijhdvzdqppUJSerGQmK69+/sLuYsZwExgBPB5AAHEhkuBgQ298nK6+LvH4tLSkmZvnfBRS7YHPc6cAXpSBgCcx3n7TH54PwzPgjph4GEnXeL7hO644w0Luuhm303hb3TSw+/P/fp/3qcKtbAQMI/wqy+fUi8zb5lu/s4rv/Ny1sw+iqK6iPkaiHuC2u44LoEO7DfFI5F++Fv/zS/dTtZ/YeN3f/NrcWVlML6Li2Ozxu6W8npY5mQvy3zssq3aTxn4bBf1WTuQih7kt7aaOFtKbucvjSqrW6x3032EjoN4CsI0hFUIb1D8MwO+DrNzYPaXCmvv/bV/9/fvDqMq9dJLy/HMUgFA1e+++psXs9raxI2Qhh9K/FNSf0jTOcG/D+qaKIPhoElP0e1oaCSTjpFs1ss7du7b3PbWuG+F8ObaJHObhXPCiHuSXiN4AcQf0cOfJ5j40VS3fbMKQL2F9rurH+Tnz8OLiM1LwvkzXuQ7WtoUlmXub/xX9+6ttS8jtW8I9sc0/hnob4C6FXM51Q3XMDmyGVvW6O/upu29qBZLS0s638sHcwb2UW26rphMWEhyInwA6lWnviHqTwz88w7wox/WPndz88p+IV45drF4XbkYcf6CY2HZsbDsp17+Wj6a1fSZvz1/mzF5C5ZcVIzfAPAtgj9A9CvI0U4yppcu36qfKZ//hbPL3ivD9eCL1+IidGZUJ3b8C8wU9lN4BsBnREzRcNMM36PxHGh/4MY/ceEHPBCuvlABIADg8nK8hPn85a9d3ERJfvXv/ZfrSsPVrNv4gSE978AfCLggaEXQDUIGYL9HO6zV5oHG7Ot1tFo4e3a57/R+cmWBCwsPvbZspxr4uMu83xr1RBL6BKWgR7FD2JMw+Odx3hjQEDkJIAW4Vvj48Aqk182SS1R486//wz++WeHfbWk4O6Q43Eb1d5JvP9fl0tK1s2cXbnzB8tsEOpA6AJ4nWAcRJUyY28xaZg20Bg6WD9s3i4tj+nxxkdlfQxoSppRnAj408F0Yb9HxkSy5Mef77vCr/4+8Srn1ALaXnXQgXS2pF+ygzOAwVPcvvbR8U2rdvnL+6kfMuzcdOOHCNMjbEeze2bhjj4Da6ANRsRpfZYIYNjoIcNwDcVnkuxBuuPtbmGi8+8UX/3l7ML4tWxpOqTFUp3H2Ir17SjC+9da5f7Ah5bmZ3YnuB43MIN2C52oCYfr41WLe8NE9h43ZW0zJuqBJkSmAVVDvA7hswV4j8bpb490v/Pp/O5Dsh7O+YmlpyZfGzAO1fstwCetc+ierr579revMXrkd2ckRPQPwnEsZIVGacNP03TuNu4vDjquftNTyKCWhJ3TcIz4ehlL7xK1Irq8M0z1zOGzX4zUhYAPAdQGrlO4QfEfQ5enYePsrv/k/DyXnOn78VDjXmrLrJw/r0qVlLS4Nsm0uv7Rg/d3dpWvkUrFol4v3T777P/3dNLhJwg0JswZ2BawGmtNyjtKFWMZmh5fd7ZHVC4nTiyFWUz3vsrtmciYhuHfze8HCrXZt4+5nX/x/tQfg0zKch+H8eUhnHFwStxjzouwWz55csYW5azw/oG5QLswf3fnDf1f3EsuQ+T45upbgbvdOd0iCeiQBXE8dU/zzH7t18k5k/SO4rSGJOXLe+nDdr/3K3xkA0NmzCwFYCTq7EBcvzasEsq0zuJYpQZbLe1iC9Gdf/OftV37nN37ylB3sRmX7E88NCOt59PZMreMnXji29+0cQ2Wax5yJ3ZP0gYDrJK646Ue1Gt+87en7v/C3/z9rQ2352tVw7txpO3PmjANLvQCopZS2YOda13jm5GHhEnom+vjSS0vdc+da7zY++LO6RUHShwTrgm6DzJwIvm5jreTm53ctNegRAMKjKPOxW8wfpzrvJnX3XgLY3mwTK7s5CDz3X5+u5xvxeAjJM+5+wGBu8pu1tHG1s9r56MztM3dH04BXlfG9FEPjvitTgg218xtnF5ozjY05j8kcos0mDHWJnTzHzTDVvHJ98u2PXnyxmuobxoeIQNxrr3oet4uL/OBXVpuKNycSpomn0dduhvYGbm186aXl7hh+qwh/s1gsSPf7re0ARedOJ+/Vn5/BjazpScPV7mxk785sfP4/+M+6/bzRj2LMzy6EbwK1Y3PderZRS6dDnZnq+bV3Qvurv/m19U3Xt1qGxSWNG79xc7pI2dvi6DwBgPd/5+WJWIuNxO8GTzNvZJPZvQ+T7rcn17OeVLmHc9vIwVzRK6fSd68985lc6Qm5DlKeKeAD1sJ7E16/cawCQJvavnn8hmjhcXPh9/6bX5uZ8uxog34o7/gMTXRpLTD9aCOP185cOXOz2kclo7Bbc/ydrCH8FJT5BIQ+pSD0aSMQee6/Pl3vbnRnamltJuaxRks7Br8zN3v47rhFeS+Os2cXwonGRj3J0qmmmlPRYiIPndjA3RuT798bAqExFNcD8gR9ENpaaFrUIwWDrerwqH9X4PLygs3NFfrAM9cPC5fmhaVFjSRMfQRtFtFaLCTc8vdR5vrZa6lvaINVgv6Ve9P7N1ibBVTvgJ2g7q2r0/duD82xkfse9Gi1Wnby5Erymc69hq92p/PAyaDcEGyje8/vjG7o9up3nxw/eyBUPX/sQej3/rPna83wVLo2G1PcAQ4cytu/8tK3NkZBY27uGq9fL+i3HedGIdD6LdjiyQVevHXZTr38G5sU+O//zssTua3W79wBPFr31jPvdV7cxh/noQEAwFaKf6llWF7hMoCFS2XSuYdYJCQRyy8Z5uaJ6yviFjt/SXyUIHS/hfv8+dOhD0wP0WYBXGy1uFhSqlu19+Py2pZa9t7yv65neKZ2N+3yvRC7f+/v/av1aj387EK4eOKynTr1G7FHv+2037DY4vkz5+0MAJy/4FWJ/ezCQpj91VtTE5H1ZNLj7WvW/lv/lz9Yr869IvL4EyB4cjy4NPS4h+CgBJ47dzo525qvvXp2vvbbv30qHdcqFXG5hmJg7Xaha7VattUC/MY//ffrb/zev19/9exCTWcXgqRHlt5bO8jB84BpmMcvtgI/1jTeDwhEe9Xmvey/PWnX2YWgc6eT0t9p/DVlSvndzu1eO1ut8Za+r/z2y+m7/+lC892zC81XXj6VPonZ80QSenJsBiGMptJeXOxbLD3yPdonRU89OZ4cH8vcblUCsy5tbdTy5HgCQh9Xm/SIf2vXURzIflSbod3d8ksL1rNIW3g47p7DQg8ItIjl0gJuGeC/+BexykkIYi9j2BYP7Z4oSIuoB4tcLuty6dK8gKWes+NOfCR245hcYm2LJ0+ucKFoOkpqc6eK6Z3UYdtr+nTZIjBo933r8MBtb7XAxcWqs9Zi1QfqE1Gel467HOl/PeTvotWCnTy50E8UeQYX3Jbg2vkaoQd8xne75nxSZT459njx3yprqW3x2ksKr1qWjbzvWfk9qoF7108cpX4eIJ3jXqVN7wNuqwVrtWCt4bEa97tbjee4+oy9fvB7LSspnLCLcu83r6rXhErZQ9dUqdVWC7YwXIcHafu21/d+a4SeG1fmw477aB/wfuM+pk07afuWv9tqwRYWEMqyucN2bJc9+UGv+TSUycdlMX9cAaiK9hrz3XY7CO1Rv3EPy38UVn8/TWXu9neflPmzU+Z2z+P9Ahnv9vrHqczHQiJ63DKr7hbdd0IZjPudnUpeu11ct23XwsLgtcud0G5+e1NZPUlkG8Xw/epiH8OLu3jt9voHHcvdlrWb+j9M2x+0X0c/f5B6jHsIWTUg0IM9a7v2B2xhyGhhLyWDRyFp/LQYW/1USULcQV3vF8JcO9wlbTfhd7Ir4X12JzuV2B4kj81Od4UYU4ft+pNblMlt6tu/9hQAnDqFqakpra6ucmpqOHLx6mqRZ2djo0io1u12++XWfvxjYX5+qNBms6mNjQ02m4OIzxd7ZV64gNOnT28qc2Vlpbhwfh7zKytobtGfF++fHkTbjJl2MY4EgF6yg1WAGyNlrszPAysrAsBTp04N2nxxEGB06vTpTb+5bX/WapuuH+3P0fHBhQu4sPtnr9dPvs2DQmwftUTYXgekXTATu2aHdyiR7CZlwidV5mMjCT0OIPSw9NrHAUK4D42wW9pwN1Ke9vCB/DSX+TBj/aTMJ2XuJQg9DmU+NiD0aY8dxx3slD7t9dYet/9BwHu3D7xvMZH3ckI/KfOnu0zeZ15ym+d5J5Z/D5tqZCvJY7uN46e9zMeSyvu0S0LcAe10P3PercrQLoBjOyOy3dBwO3lI7zc29yu7R/lw9dQp4uJFbADE/HyfosmybGzZk5OTDgBzc3N+4cKF0VDzP+smn8QT09idPK9Duo3Tp0+rRxUCA7oQKCjDWq2m3rys1WoqKVSdBnRh/IItjKc7xz2D2mLcdjqO2sVmTvd5VvkJlPlYPLePCwjdb6HfjbXMXiwUD1O3n0U6ZWwup1OnTtnGxgar4Bhj5IEDByzPc8YYGWMkALg7ASDGSEncByC6092JmRl477y8VhIny3OXqGZzU31711eztPb+r9Vq1itn9PttB5FUt9t1MxPJsXmJep+Z2ebvNjZkpMxMa+W1vevMrDi/exdmpmCmO5XyQggyM4UQdPv2be+d975L01T79u3zqakpHT58WMvLy9Xx0QPOT40Bip3O/4/Tqu1B6KzHrczd9v8TENpl3bTF5N0pCO1W4f8wdXtQJ8DdKIA/icNw6lR4/s4dW1tbC7N5bt19+8zdOTk5aTeyzKbcmee5xRit0WhYBRhMEpM8DzFNg7tbCCFJ3C26ByWJSbLgbhZCcLMgyXqvIJmbDf4H2DtHCJRkZlbxB5IhRkgyAxjSdDgMUfldmd2M3m8gADOmxYrfK4s7TnLnrixG0b2Yl+NACEW2NZKOEIY+j1kmB5ykYpkOPZT/k3TEKMboJN1Jj4BoFllcW3xu5jHGaGaRpFuMHsvzJElyy3PPzTzP82hmbt2uJ0kSO6RCCB5C8I2NDRXDYErW133VTHNJ4nfMVKvVvNls+ocffqiZmZm4srCQY0z07k/wsPswC0Xm3e2f2Z0wHNwDEMJ96vkgILTd2vgEhB6wftrldx8HVbKVFd79dngEgIWFBVy7NsiwOmrd1LdaunhRF7bOFPvIJ9gpIMWxY+mdycmkXasl6cZGGrrd1EJIQ5IkeQiJmSXRLIntdhJCCGaWxBiTEEJw92Bm5u6JpED3hGRKs8SB1KQAKRWZyD2hWSCZiEwcSCAFAoFSAJmwuN7K701SAGAkA0kjGSQZ3M0LKYokGUIwkTSAiBFehCuiF86j8OHViyEEkty1EtkleYzSNuNT2j4LpAwQQoADoqQYoyS5AIeZk3QCTskNiJQckouMAHKQEWQO0lV+LzJSykRGkbmZ5ZIyStHMupJyuUczyyRFd8+DWeZm0YvPo7vn5h6NjO6eR7OYJEkWYoyeJHlM0zyNMXf3DEA3z/NsbW0tf+GFF7oXLlyIH9PiN/rMcQHAtdOnubq6yir116P/AGDfvn1+4sQJLyXBrSgsPcAapR0A207W4L0q87Gghp/Ejvt46MS9Eo13a3HHUbCrgl6PCsuyjN1u1/I8t4mJiaRerydmlmpjo06zGshaRtZYvpDnNRTnqZM1ACnJ1N1TkgXIuKcgE0lJCQyJyARSSiAFmQpIKSUlCBXnQNIDIQEJinsSFlEFincpiEy8AKAgqQdCobzOJA2BkJlZiSqEO+TeAyBuMoUsrkcJQrs6JMndIW0dm5kDaUg0E8yg4la5ex+EaOYAChACokkRkqMAoCgyZwFEOaRYgk8BRmQGMi+vySBlknKadSHlkvI+IBXnWe+cRXl5eX1efp6F4vpMUoYQeuddAB0AHUnder3ecffM3bNms5nleZ53u918bW0tr9VqnqapeibjW5mHj1CFe0WnPyoK/XEo8wkI/ZT1zf0scUZB4VFYLfWdC0+dOsU7d+7YvrW1cGdyMul2u6FWqyXdbjdJ8zyZnJpKulmW9iSWUjJJLMYklv9LCm4WzD2QTM0soXsCoA6yRkfqgTUAxcu9ACAplXtNZGIlqAQy9MEHSFRKMRqEsUnUA5oSVCglKgAokCyAhQyggsRQglMBNJIJCCCDS31HShJGEuxlgZMg76WNAIyG4nsA8uL7rbhcElbITLs/BHhZ/v0Gj73K0XoAVt47XtamJEheCm4uMrcCjCJQgpMgFfRdXoJRhJQDyCUV4AREAZFAVkpPkWQBNmRk+b2knMV1sUiVja5LmYAMZFZKW1kJQl2DdYOFjkw9QMvKV27umZnFSLqZ5ZH0IOUhhNzds8Q9jyHEGGNeq9Wye/fu5Wma5o1uN++ur+f30jQ2Go04OTnpJ0+ejMvz8+LSkuuTXSN3qpveq8SbT5La/YyBz5Co3xPzt7rpwoULwOnTQGFVNnq/A3g4zlzi/EsvpTdefz2dXl8P3WYz1DY2kqTbTUKaJp2elNLt1kOaNjKpEaSmhdBUjA2YNXKzJmKsG9BA8aoBqIlMJSWUAkPoUV6FNCIkJBIUUksKs4BSYnEpcfcEJVVWShuBpMm9WPUBY0l9qed5TxoliiWAAIRg6p0TVkRPFaXys4Il652PnbmbBJcKEPS+G6wK2vJp7gEZd/B4aPM4FZ/dJ0nNaFrciiS1dZgPjda8X4ceo6jyHxFwkKXKSw7BS/2UqwAysaDwvLTM8JLSK8oafFZIZlIUkMs9h5RjQAcWLyAnmZPMKEaxkKRK8MuDlLMAuRxm3R54kWwD2KB7G2TbgQ2S7RjjRkq2CbTzLGsrTbu1GDOv17Nms5l1u9041e3m7YmJeOfOnfhLv/RL2fLyQ2eJ5TZUOk6PPP89Cr0nyV0YPPu6jy7oyfEEhD4xMNv2WFhY4LVr13j9+nW7detWMjk5mSRJkkqqA6iRrHW73XopraTI8xrJlIUU0gBQF9AwckLSBKUJkJMCJkROCJgE0CTQpDQBoA6gpgKIklLvEghYqXephmwpdDNkYQxQUl7R3VDQYX3Ga2xo8CfHY32UVGFURRJDCVwki/dCAiuktOL7qIJCjKU+qwAhsiOyA6BNcp3SOqR1kOuS1lm8r5FcJ7Ahab1H9zFJOpIyd89SoOvumdwz1GodSV137wLobGxsZFevXs0B7EY/9dg4eD4BoU9PHe/nz8NdlPsolf32/PPPp7VaLQ2rq+lakqRpmibW6aQds1qIMYkhhLpZ6iGEGGNS1b0AaJTSSz3G2EQINQA1A+oqQCgRUCdQg1Qj2QDQENCg1ATQcLIBsgmpQaAOqVGWnQhKgQJ8hvQgHJEORjbuBX2kfo5kIwsBZisQ0vDJyL9jpIrN+34NCzIakQYG33LkHu1kXCty0g5tGrmT6cXteJRBzblJPFK/oH7vbx6dzYorllITx9Ro3Gcjdb0vzThEEBGbqsZ+n5fXl1KVFEtDiq6AjEAHhSTUhtQm0HapTbINaYNAG2YdSBsuZSjow65LGWLsWglMCKHDENolUHVIbgDoKsu6CqGgAWOMIcY8hpCnMWZeq2UxxixrNrPJGLN2u52988473YdmLHYuYWkX69pWdD8esIwnIPQQkoTuMyjjYpsBAE4BxKlTY2NrAYWjXLPZ1MWLF7Un9BmAFmDnn322dhtotBuNCcuyyTyECXOfdLMJc58AORWBJoC6mTXgngKoOdAw956k0oR7QaEV4FEYDAA1mSWl5JKo0JkEmCWUklIPkxIIAz2MEoIFpUYY+6vHGJaL958sPf6nR1/tTJWvPlumMaOqMaOtEZZNWxS7zVcYw8RtXbtdGNZzD56ssXg2+JDcBQhyu+u2xRveTxoa04FjQGioD7WpUwVCQCxBKYcUewYWKqi+PnXHwTUuKbqUu3tExSCCZBclCFnx2YaX3wHoGtBBjJmHsBHcNxzYkNmama25+3qSJKt5kqynMa7vX11df+YnP+kuFxLUXqxhhlOnbL7iD9dbb3rnKysrOgXg4vbS2INKaU9MtB8xnfUo4jLteHI9++yzYd++fWF1dTVpNBqh2+2GJEmSPM9DmucBZjUB9ShNkJxG8ZoCMANgCsA0gH2SJgU0STbZ09MUUkwNUg2F5FKHVO9LPGYpyBQFZTbUbvLxYFkfhrF7qCdLj9/Tw0dQ0OPCxfcka0nw8kCha+qS7JLswqw4B7oA2ije+0AF947INRa03yqAuzC7J+kegLsA7klaDe6rBrQ9xqxeq2XdELwdgtfzPO90Ovnk5GTe7Xazp556Kp45c8aXlpY+tY/XExDa27ptF3dK2FlYjl399vz8fNpsNpNbt26ltXY72UjTgkLrdtMshLRvrpymdWRZAyHUEWMdITQQY6206KrTvRYLHcwkyQmaTcBsiuQEpUkvwKgJoKDJgBSFaXINpfkygUTFeyBLaYNWqOkr23Gpt8/kZspriPkZoptGdOG9/S5HJZG++KKRaaOeynwYVShtOShl8vIqxaehTTs5TCFxWBzgsIgwkq1tnITCStlbRtnSNhzXcGHaiVAlbOqral+wB8TcQorT5vGr1lIaljE0eo1682IwEiJHB2Ks8MhqM6ty8iZhV+OkZ24SojkklBGb/xSGgmNEPJJ9atajw+Wba13ookqTdWQSCtN0IAOUSegC2IDUBlkAEbkmaU0xrvV1UOS6mXUQYxdm3VzKZJaRbCPGdpIk7RjjRg3oKE0zX1/PYprmaZrmeZ7nzW43j1NTWbfbzX784x9ne8Co3C9Vxl5LTk9A6D71NACcB6z7/PPsUWhTU1M6fOGClocH4MEGf2EhfOH7358IIUxmWTYRQ5hIOp2mkqTp7hMmNc1sIi+kmwkVYDKlwhhgktJUSZvVBNThHiClEFIYU5qlVvrY9KScilFAQsAEGWiBPcf6/qNZrCrkGPpj0w63ZyY1rByRtlOmjAIJh9Qy6n9fWZ00wBoNMS+qPAWssjIcEln7q1LxDTlIclWYVhefWwWIei6k5aJFq9xLaPC9Kgsb1f89U2WV5HjwIu8HLoBXzLzHPUwqQUjamnLsQa80XLbKQVW1X/uDQZSCwaB8AF6eS4Myq2DZL3MzDTo0TQYitYYsBFkFEo7uEjVKT45izCYIH7JALDdX42nHgZLJXZU5ys3zf7BVEgQn2TOmiALy0lw9A1lISoWRQ0dSBjI3smuF5V4WgU5J622wAKlVM1uTtMrSgMLc12MIbZJtM1sPMW7EEDYkrbbb7fXJycn1lZWV7kPReeWUfP75560X97Hnb7VSpP3wJyD0yVFzu6bjehZoPWfNtbU129duBzYaaQY0umk6CWCG5HRJl03TfYrkpLtPg5wkOSlpCuQkyAm4T6uwQpuCNA2gCakmKYVEFLHH2FPaD/wmR2g0jNnO75DPqq4i3OLznU3H3bolcAflbPF0cQAqPfDogVD/nCrfB+ejL6vcy6HzzSBULX8zCA2y+m3X/VJh/hVdcO0MhLaSCjcBxhYg1H8fKbN67uM+r94PlNMRY+v08S0R4+fYzqb8ZrP7ndO/m+TKqoNw4UdW+DG5pNyBrqQ2gHWSqyRXCawCuEdpTeSaSasi10Wum7QqYBXSKoE7kO7B/R5ivBekdtzYyNaazXyLOH56RJ38BIT2uJ477dDw/PPPJ2tra2FycjJJ1tfTjTRNJ0Ooea3WkFRXt9tAmtaR5zUA9UDWc6khacLJ6ZI+m4LZlJVUmgMTlJoimyjOGyAbgiYANgRMsDClHnRsbxuswbbQbODHqlGxZLM8M1iXVFmrBJXOmCq2QqzcoP6+tAShavwZcpje4jDlpSFqZGD8pNKKAX1QYLlTHgWMXjn963rSSQk8IAYOQv1rfAiErPQVpQ1Aw+jDAFQBlVFJqCo1sUIbjbZrtyAkAO4FAGWxeO+NmvHBQQgVwMAoCGEzCKECPBgBIWiwNR4FJK+AkFcnV4U69L7kxTGgNwbMiv9VBbwqFVkF11K0HlCIMvQIwwJUNDZGIiuCOXtzeGDKWQ5ZvxAOS2AcdPaQtMW+dFVEUWL/pcJbF0WUJEQA6zTboLQBYL2k9TYIrEtqC2gLWJO0TmnNpLsl3bcqadVDKCz2pE6SJN0gda3wg+pkZLvdbrcbWZblzWa2traWNxqNeOjQoXjx4sV8h+uePQGhR68Tur/59OnTyfzKSuP2wYP1ydXVBicmapIazPOpaDaZklMO7EdhGDAjaZ/ICSt0Nk0na3Kvu9SAewNmdSObJOtmlnoh2aQs8jAlqISVQRGuZmhB3ySxVBQexDgrLI1l2jW8O1aJRiqp8XIBqprFbtbBWGVBNg4W+T6YsHeNhsDBSod+oyOU9yVWfB5MCFaAT7Diu+IaIVDF/1ZIKma93y3KL4CmCgyV3wU2gw2KOgztmjkafWDEFnYMwGy3495JkAQJcBeyKHRzIC+loSrADe8puK213yZwUsUwW2MUnwIgG86xoc106LAkNAAz94F05X0wYuV88L87CzNR5+CzEnyiswDj8r7ohEBEL37Ph0CsD6B9CrH6232aURwwkxqm8kY3GCVFWyKOhuCGY5SCHChNN60qAwtP9p/PXp16JunFZk45wb4TLoFchd4pj1IO9y6kLqWOARsg2yoovbbMOgQ6AtYDuR6kVZndBXnHgTsxxjtBWo2FXqqNRqMzGUJnvdFo11ZX2zug9J6A0COq2/2yndqzzz6bpGlaizE2QwhTFuOkD8ygm1Y4Zc4AmCGwz8mDlGYFzAKYBTlFYAKFH04Rxsa9CDdTeP8nVgTVHArp/3F2nLYYMmknZFmFzrIBsIQeGFSBoXpu6ANJ7zyUoJJY71xIbABGW75Xy6n8FjkMQuNptCoXq7Ezd8vkLLu0BNspAxpd6OZCJwOyOAAh431iOu1gvHb0eUUBqG3K1xi6T9oChHwAPtEHIBQ1AKE4ck30wbVxzDU9gCl+h5tArw90fXqQI/Xefoy4RW9tSdM9jDnmNpOj9JeTyrh/lKIVMfuiyOhALjIj2YW0Hgpa7y7IWwJuSboJ6SbIOyrovTWSGyQ3JK0mMa7FLFvrhrAGoP3OO+/kO9B5PwGhPaqbjXRmOHHixGQD2OfATJQmVDhj1i2ECQEzdJ9SkkzCfbIn5UCaFDlp0rST+yjNCJgGMUOwCbJhPTar6pDX1+H0Qk4O1LwcVexLFX11yUwMK+1ZPR83BuQWLh5V/Ue5A9wsrVSApLLIB1ZApS/BDINBMB+AlKkv8QyDUOXenpTDEUmoKv2MfE4blN+rd9VYYPR9vA/MJrOu/i5XI2vFOC/mnSbMuZ/FW+5CNytBKBfiNiDELQoe54K7Xda8Ib2TuGVlOerNq7J/qhQbBpQdemARB3ReIc1UAYIVoEJFOiolIAGxVH0OAVUFhPrSlg+DUKyAUCF5Db4bBcpRKaow2a7QgH0pkhrT11WTUI30rypk9pCdRPlMsgjAPt4qoy819Wi9ihRX+kTBe+NY0HDrlNYcuEfyLqQ7Au4SWJW0HoENuG+Q3DDpbql3WoPZPYWwhjxvG7mRkxtN93s32u17V69e3dhi2jxJ772XCrZjx47VSM7l0ucJPEtyTsC0FZECCl8cs0m6N0FOchCWJi2t1VIWIXAKXxuw3guVPJDgDeKYmGMjon1/zrLk0ElQoKBQmgaot2Ptm69o04askqaGI4vWQEooF32GkgoLoZBGkoDiZUCaFFJJYkASCvApJJaBBDMEQqXOJhBgUB9kjF6RhEqgsxGKTBV9z1Z6mIqkU0aCGyiUOLI4j3hkahvRRmO2TuKw8kD3lSQexJRimGJDEBBV2C/2pAxuBgxtQZhoTNSBLU1CxoAXMaTm2EqL0l8QObIvVunJ3C+fg3V6IC2pYlFJjOancA1LVsKw1OPjQGgUnPqg1KP3iIiBpOUC8jiQunqv3IkYhehg7uyDoo90ScXqkBro3rRpk6CBXrFCnbPC0I0VilRRQNFKB4mKC0DP+yD0ry2izguYgLRPUgelP5OK99KsvPB3kvsqyDWQGwLuwX2N5KqTtxL3Dx14p9lsvg1gCIQWFha4RSTyJyD0MGxUrVarAZiD2Rck/byAZ1hQbSkKwJlEEbKmF5YmRRGKZlt+padsHr9gaYyW5r5SPlHZMI1TdA8r+Id1MdX/jQP6KwlAGtR/T5PyPQC1pHhPe9dY+QoDEKrScVWrM45QZEMGA0PK/e1ZDWm8bCFuwRFovNjyME+NdmCsx91yYRheUOgAc4GZwDgQccn70Avc+efa6UOxCwlvS6Tj1nXnVqQwt5bpNCq9jFjuuTByPgJCQgEyXkhleWTx8uH3LKL/f4wldShwuPxqHfqfU1vtVri5LZvmD7Vt32/S9UqjHUeQtdIxfWr0XhZUShEVorC8awPoklwHsC7yHoEPZXY5uiuE8BGAa9VyyvQtj0XonscBhPpHmqappH0OPE3g8wA+W5pLB5IJyEbPumUgxWgTfzt0Xuj4XRXxx4fmTmkEULJlPU/OqjQj9RdrDlFmFUV9VS9TUGQslfzep7BC/3rAzPsGAGkJJD0pp/8eep/1gKcnGfUkICGEQZm0zVJKEed6sz4GHFiygcOyfTVsjqoSio8o5TWqUN/86ifzKcWZnn0UR9yQqzv5ofPKQ87etlSbgWNI/KyITtROxKWBlEov6LgQAcsB84EkPKDjdF808FHv2pFFXSMSEgeC9UDaG/O/Rq0vWBGHMOa+cVHNOGo9qAENpYr14+bbhsbaewyYho1m+p9VqGwX+1RdHNJLATH2QKkApt555ijBZwiECgmqV5YPaMS8SvuVoLfJGAOABhShfMSwome30G8vyy4fbDiHnKv63Tfsj2t9E78et1Ix4DOJMiYQEhShvXrLUgdkpwz2OgsJMvsgxjgxOtV6Eb6fgNCD64LGcpm1Wi202+0JFBZuhyDNlVk6B6DDMb4tVeV2ZbTLmUSJobdM9WR5je5phsxTe8aim60ogoF9MAgaSCYVoEhDAQ5JANKSVgtBSK1Ks/WU/iWocLNif/hVlWYq5tJVKsyqOpfBU6ERHUzvfx9HhfXwoqKTUIVG63MhPlj/qqDCWEoSAhDL9bGM1Nc7770KEZX9zHP9a1WRSKpl9wd0AC4cBb7eCjAKgttuGzlYyx3IHejEIkBZt2DlKqbmW8jMHJEIOaKQGKEieytbFSRUGaveeW/8igx5g3OS/fPq9f1selZOeKtcU9Kq1XJ7q2p/de17G7BqgF3FqrECrlWl4k2mfaW0YgNKsCrNFCF7qrqins6plHSqOqqK0YR7IS31DCZyr9J65Xsk8vK7PKJ/XVZQgIylcUap9mHv+eegLeo/Oyo9pMW+xR4roM5yco6LCrFZwixPBunfDUBTRWSVphe/dgPAvlqtVh8tZzSr7BMQ2qMjxmhl1s0apDqL2GmoUNTWNxCoWLJt6/SxhSA+pL7gZjG8uuGsKv579FiaFBRZrfpeoc56+pweQCU9EAqbQagHPJswdou2aCuOh2MiV4+GltEOaJwRRThHpB+OAkQFYCxWgCYOg07vVb1mCJw0DFD0Mb9VAawqCFXBhhgGsp3G1+61JXeg7QWx3ymwtIwivo0kNCK9aAtpZAiEOAz4/XMbc159txGwGfqeg8+DBuc2AMMqcA3XgePNKbbSw+3EOpHcvFEcVp+BWxCQw4wXK07EA5ApgGUMCEUM0XpZLIAnr5xnvWuKe1iVnDAijA9XSGOp6k0+gdzCukSjfMJwd6hQMzRENkXWEWPAY3w8ViDk7lQlm+ZgK98f3p5lmveVqhXppTRRHVD45VM0xrqsT6v1/V6qlFnPV6YqqfQMBEqgSYb0NIX+pmYDiSgE9PU8wQpJKBB9HY713ivSTO/pHN7cDyuGq9QHKlLDgGMcpcEqC3UlU0x1cYe2AAQVdNQAeAo+w6oSzQhY2BjQgQ9fb1VpqXIvMAJCGlPPEZqt//m4dWw0t/cOQEgqQMjKMgOAHKOS0P31PptACONBqPp9j7XsAQ8wABeMSEIYIx31v+udh2Hg6oNNqNw3Wg4r0tsIwG0CroDNkc84pm1bXEMOs5W98E6jDCJ71Lr3DB9KOq6k8lzq65u8lJD6IOVAXtEt5X3gYemMPJCUYiwlrVi17quWO86qr2RYKt9VH9Ue49rDfw4H1+ubwYz0kKGagVjiExD62A8rMxIP9b1VwjOyqsgYVoKKKolcUwEkYMXyzAZAUAAG+0BST4RaWkg59QSlVDNKqakPUEnFQTOxXla44jdoAzPlvkNoiYijYWlGFfoqZ61Y/X9kcR23aPfOYwke/cVfA8kkAix1HcwrkkkU2NODlPczDkstFjdLN1YBEmAYtODjwXDomvKpHZJeKnTfKMCMRk0dr/MZCWR3XxAayJZeWjq5ymxpHAT3MuwchIYlpC3EzxEJQyNKGI2TovpRFvt0c0GrVkCrX26Vqhs53wRCBngYgJiXUlTvM4WRawLgCYe+w4i0Vn0VtCCLvqhcW9S/QoeSfenbvOJoWun5foQNKxa4HtVXDTkiL3RWfVCSBtZ6qPhCqUfRWSkRAVk+oOx6UlSeE1lU35Kvd20U6E5EFs9dFCWV5rMsjGNZTlpuNmTqDX/YioTgT0ECvscKhMxM7u6iIsG4aaWobJKqppWmUSXTwDy5DzZV67IE/fOCSitejVSop0A9BRopSqqtR6mpEkFgM3XGLSzltIU3jINj3dG22vOwOjMrdJeV4NEHjNHzCnhYD4Dyyvno55XvbZsyx+trtgCFkXfuwPSLn8Cjx8pSFyrieNgJCD2iQzuwuNvS5H30mjEU3DAIFQ4ICoAHlu8Y+64AeFI5D5vBTAHwCkgV7xy+npvrWTUR38ZVuRfddiRs95jtRTXYblUNPETxqS8RZfmAsuvmQDcnsrw4r1J6eQlGhaTVf7Fneu4aVjqOCXy5lZzukHoRGxyPORB92kGIo891sXcYSjHcE1xVXdNZca5MCFiV9rIeddaTdkZAqPrqmUEnhSSUlsBTq3zfs0LrWaBVFUqbJBVstharerUPPhtML8aipaPSi42jsoZAReNBopRoWJVoRkCruJZjQap3/Xb6ndHB22mGwu3VduOXm48ry9cQY8SBcr4XKqZn9qRRrdCQo8/9A7vu1oLnfmVpF+Vv2hdUpKQqkFRBBKNgYgPgQXkeexLUCPj03nvXaRTMyOIaAm6Cl2V4xciioBuEXjwo9kWhkV2LVcCJw2NaNb2hhq0qVa44A5PxHrhoADhDJuMjIJQPaL8slrTeqI6q58jr6Fv5VQLej27LKtHBXYPoiTvOuvoEhHbxjG3xhcpoZ4OpVIim6ulylJRAUU+ARiLUa0AtBeo1oV4CSY9SGzV57n1WDU3TiyRQ9bfpmz6TpYJpYGZW5c5V3Wn25lHV+gsV3UdPfxMr0kxWSh8ZYJlgGRB6/+cqpJSsIp3kGgKW3jlGPmMc6HJGFf8sPQ/HGRdg1ACg+v/YAawuyrzvAji8uHIscGy1AGuPgGqrGnKTccHANn84M/pQ/osSf+4ft2frfPXcNZjev3+Gdepjx663CUL5bhVppKe7sTFGDARUKs426ZZ61J0NG1IglPooA9STtEyIJYjlQcgSIU8ET0qgSgklhBKACco8woXj6MBuWn3dKargVW1DJZDsIDQ6KwqAAtSMxabWpYJuVE/npL61nlcBRgPQigLy3PpA1o2FVNXNqxIVComqD1CiVGxxSh1SXy2mbbIGlFlc+QSEHoEkNGZTOpDUS5Cop8BkA5isAxO14tUoX816QaXVUxSSTcCwM2dPctpEq3GsActA5196Yzur8tn4VUD3txDrU2E5ELol8HQB6wqhW3wWugNQsm6FJsu1uTwNS03wrRT69983aywPOBK5YAvKiLte9h/8890mmeAOpmM/nHl/7AWvWMBsK508wLLwifAs2v67qvXhTsd0yI9pjBFFH7T6wDSQevIwACAmAlMhpkBMSxBKCdVKAEpLlCgVsD19E/sSEyoh3CsGHEMqQ2JMOr/BHtP6eDnGWrU0Ma9QblX6LUYfgFAJPp2MaGfFe++8+srjwBWup2pwjTXpeCIJfeKIxULCmZkADs0AB6aAqRJ0ejRaowbUK3qfJPSUmOXuqZxgqFikDVmhYSSsvvq5FMoXCyX+OKV//13j9S4VAKlKNiFzMAdCBjAr3i1T8X+uvpRUlX6qklV/wfDtPOK3ljC0E30EMTakDjA+YsJOF3/tZsHcoZZWD8PHDQWQLVNOx1hCkYNQGSes4ogIK7Kwh/KdlVSqO61M39GeYxY97S34cBv92xi9HXc4ln0fozh+jjkFmRBZnA9swVSESQoCg2CJoARAilIKIjwtV7OEUGDFCghgYPkysP9/AUR9+q8PftxkPNEzmiiUfoWjeW9LMhpwl5WdaT+UkVeMo8qAtzH26Lkye16OvhTUzoC1DnBnHdD6IHLEUGDfn6LjcbWO6ytTSPbJHmMh4eyfBJ46ABydBaYbBdD0jG6SwKFgnANHzoqZU5VawIhup+dAWInJpai+tQ57UkkpvVgGhJ60kvUAQ30FP7vFecgH1Bqr4OSAuQ8kmJzD4Fa+9506fZjY3y4n8E4AZ7td/aa0BcTY2L3cgfJ8u5BIHAUBjmlJz4qOO9gObiEuVRdWjQuzpCqlw5FgBCrhpycr9SI3DSha9qK4Fh7SJWjtFIErITl6ikc5JO8lV9850Ixz6N6Fgm40KKuwu4Rf2uJTlxCjIwcQ4chZJPLJKUQ6IoVYAkOJ6X3AsIo+qh/0sAQhSwwhIUISEBJDSAowQglmnrCg9xLA0/K9BLb+d2mflxuKNFzVnVHDEUf6qqiKCWbPSbjQcZW/B5V6oaIPOjlwd91gVgDSeoeDSO07UQY+AaFPVhJKAzDVAA5OA0f3A9PNYtSyvNiRVE2eNRoM3rfX1FY3Z1X/mb7lWCaEDmAdIXQE6/Ros8F5H5RGQMhGQajiK1OtpbYKLbOFN/6WuoHdKkweglJ7mAeGm2JwaTdV3RuOb9OC3bMeKUCHRhApDOlQRO9NNpk9jTS4Cw6r8o9U/mR8OH5T2l5pdp9Ng8bcq/sp4oaiTo/2UCFH5hByePGu4j1CRQbb8hptwcmPSuUsE2QxECEISWJIEiJJgJAAlhS0nUrwiWkBCLEGeK3y3v+8uLZvbj5Io1eJ17ONtSI45IVLFMJdsjloKrqx8D3c6BI37hXfuzYnDHoCQh8fO72rJcys8O2pl/qfZm2wm8418Ax3jjh19qSK2NOZqA8wvYCVA9pMFfqsKtmUIJSpkIS6PSmoOK9KQH1LtbxqjaYh5f64+ezYIbBsBUJ8QBDabqHXw9y8jZg2KhJp0PpqGoe+N29VRN2qnIeajaUuSA4pQogADVavIZmcRNJowNK0IGrcoZhDWRex3YFvbCBmneIeEIYEtAS0sAOOkCUARTgyCFkpRRkMNZDpoJxR0WQzeTeIsfegjB3vz1ayIpKqjHvXB5xyx1/VqcX+y5GrkIQGJGdPxhyNWFixThQH/loiAgiLhGXFeWLsh8qyIFjPCKK0xBuShGqApyrApycRpeybnHvC0qKPfVP1gXm5BpaCZF9qK+gXDEeRl0oVlfoSU89nsJ4SaTJIPvlpVR/+tIOQxq8D6jupbkaooQzXg/DxqHqblzxxj05ywFTqYboDSzRmADMVINIzCOiU71kFbLKSequAEqIGzp49/dCoscBIFAJo2FppuAO0ZYaqoSx/W3Bs3Ib31wNgBR/iMeC49KYa3U6P0Ho9tzxa/94hTayGt+Rb5fC5n4JIo0liUDFKGZLCeqE2EoSJBuzAftQPH0Z9dj+SyQmIBmUZvNNBXF9DdusOuh/dgG7eQsw3yuXW+rt1qFr8mJDsrOojIyIKMGPprRSUjO+3LRR9qkbarfaBtLNJcT/DhV46g0phkvqgEqVSyvE++BSfD4w+HF4+uz1qszq5BxR8TzsTYKWxAJEUEI9QeZkTJhbPaZ/R1OYQR71QRknPfJwD36YShGLPIi8lvMYCoFIg1grgynuSVVJY0DFhX4ugUI552d8sQ3tLVUdaVky07xuTfexGPcuyJ7HjPilJqJdqOOtZn5TMRV5aqPSkm9AzTe5JMh0UFFrP6qxbnnfUp9dCu7imdz7Q+2jIAm2IStoqj3NlLg2bb4/rhB3Mpwc0O+NOAONR7LOkXQEEjGAIYJqCtRqYFlNXWQbfaEOdbikpVXQme1xpuZeRBgirN1DbN4V07iDSp46h+cxTaByeQzI9BdDg3S58YwP5vXvoXv8I7SsfoHPlA3Q/uom4tl5Qc1L5XiWXxnVD8bs0Q2ATxkZ/MWZpZiZpF5uD0phhiNXb2QTaiZOwhjZNA0mnR6/1qLYe9RbHUG0ajspWCZg6/ImVrx7wJCBSDoOQVSSmfn/E0W3ryHPUN7TRsBVd6FF0JfjUiVgjYg3I60BeZC8rkjT0rPVqlVeK0vy2eO57Riw9t0Bi4PDaSy64w8f7iST0qUKtagbHyoTqSTZJH3R8ADgVKadqWBC6DvZotVL6Cd0BvcY4fvkY7Ns4fpM6LnbYDnQ5O5lu3ILZGs7O+UkOUiV2Sh+Ih9y+SunAB2JNMNhEAzY9jTC7HzZdpGHxO3eRffAhYqcDoVvuhWtDFmgPd7Afn7C3VDJtoHboANITn0F4/rOoP/csGk8dRf3gAYSJJmClJNTuIK6uonvjJjofXEP7J1eK1/tX0f3wOuK9NbhnBT3HtDRaGM6TUYSliRCIMNFE7fBB1I4egjUb8PUNdD/4CN1rNxA77dIwI1TjKY6lJSUvjBr6zgUGMBSUnrafZByzDa/uFlUBlM1UW4VyG/lMGPUTU8WhdOAITLKksApwCbAShIjAAnRSWgk+BTXHviHIgFKt6sY4xhuLlU5jX/Yq/o8mxASIiRBrxSqqVGANYCowAZgWfkuWEqqzYj5uJSAZUDegZoWJuQ2ca1XJvfSAE1hPQOhjOjY7ERWpPApHUhYOp6XmT7GQeur3HLXbjvSuI73nSDYKwGE20M1YNtDVFDSahhw8rW+RtjXrM5RnZ1tOXb3o71vSaGMdDEept20Sbd2vDtwiddmWMc12NDZbpUPTkOqi2phi92uD4HDwcoefgLUUtm8a6fFjSJ99GsncIUiO7L0r8PYG4o0bELogUggpaCx8tsZam9xH8uK4zUSvPwxhehLJc8+g9ld/EbVf+DLqn30Wtf0zSCaaYJKUmmSH8gjvtBHX1pHdvovuB9ew/tY7WF15Has/eA0b77yHfPXuQI9lBWcseKXaJSGVJEgP7MfU/Bcw/fPz/3/2/rO7kSPL+kd/EZEOHvRkeSPX6pl5Zv7Puvd+/1d33fWYmbZqqaSyrKInPNJGxH0RCSABglWsUqlHrSa0IFIUmQAyM+Kcs88+e+NtdMmOzxj+3z9SjMboZDxXTpvBd6tB2MF/CiGkq8Dmdr/LzoVVp9V1VZJArMCY11U8q5Bb6TFUCVTVmumKh0uVpV1uVjO4Tc0gNzELRmLeF5LV6sdW7mRrsWvB7upasx8GGAyQl866mcEo6wJT+dUx9KyjgntulskxEBylnFBCQ2HbHnQ8aEioyflYiLKVOcV5CFx51+XJktfQFXzft9ya2v33Bic3jyaQJR9bZqAKSzC21M4N4Zkm6Bm8qSMOzMtzuyAjrJqhLe1XK4rWnwyL3SCPuW4Ow36Mveav6gKVU+2zHk8lS63yrO2cMeIhaxHezhb+4/uE33yJd7CPzTKHf79+U24f2WLLFOLnIXKzasJWzORQyFqIt71F9MVj6v/xb9T+P/+T4ME9VOC7Qd8ZBXOWxhuD1RqT5eSXPaL7d/FaLSevbDTJW9Dj6UrgWUBmQgqkH+B1WkR396k/fUTjqy/x2m2k9Jg+e46QEmsd4WHhVb/QarbWIoRE+B4yDBGBVzK8NTYvyqfGzvAfUdYea+QxF4Hj6sCurrLbsOQrPZ/FWrFrksl1YNvCAU6V1Y6HxEcu+j52EXA+mFRdQ9xZ/pjiOgRz0a8q+1uFMRTWkBdmBVpcxHUU8wA0y7EIJbbjIXIcYSEQEJWQs1woEP2MrcQGQfAPsyP8JoPQ/CYuG7+2VKaWWuAl4I8twcC6IBRbpKkukJWm9/zfyzfnkhz9TYPQr0iM878zQ3CS5X7Z1xFYXUCWY/OimoYz1y+SAtms4+3v4j95SPDlE9TeDmY4ojg+QYRBeXr1ovL5ucnAPAiZkl0pEGGA2ugS3D0gevKI+pdPqD19vEy2UCuCx0ohlEIGAV6zgarXXWVeFG7T8T2S14fkwxFWZ267Varc5AuEDPA7LcL7d6k9uo+/2cXmBdnpOenRCXmvj0mzEixa7Kquz2DLflLJpItC/I0uXreNjEIQFj2NKXoD8os+xXgMJeFBiKBkuNmKQsQCZrPzs10hFlTIBrNAtMxuW9/lFMyYbYuvqszx5UoQUmUAWvR7FgGrOqG1rnKwFbxPrN6TletuKzXR7L0ba8qvC3KFti746PnT/f4SPmmAfOUdJa5apaYhVU7PZ3bdZh/6n2hT+IcOQlerkDLwzCbLSxpMtdUgtUAWwg2V6oXY5ioEZT+k1WXX4+3L2Nl17lc3qH7EeuW16//mwyyCv+sxV6lrs/MkJQQ+sl5H1iMQCpsk6PEEm+dYa5AlZWkG3EgpkPU6amcb/+4dvHt3UBsd5+MThQsIbN37tGveZ9Xqe62ag5g3iy1FyWjxUY06/sEewaMHBA/u4e9sIa5rBl7z8Lc2aHz5xNFzgwDpe1hdoNOUIp7Mt1VrDVAgVESws0Xjy8fUnzxC1Wokh0ck746ZPn9J/Oat6wdVLLbcR5BlH2SWm3uIKMTf2iS6f4dgZwsZBBSjEcnrQyb6OXo8xpAvJNeEdCHEzoKOdQOjJblAX+nrVPs7C1q1XYHbRGW8V1QqoWqgWfR1FsFGlf2rGdy2nj96tdixV3HmOSZ9xd3cLnTmlmnjZh5UXQCaBVezBC2+XyOxfGFdeo2XkP8so7VzMzPLOpNosUZp2/CPPz/026yExLKopF3Sf1tEpJmIojBlM7CcwBZXwNcPvNzPkYP+rT/s1WsjfA/ZrKE6bVAeZjzBlNAa1lzFGYVwPaFGHdluIVtNRK2G8H3m8uVyJtUsPsM5rxo5uW1ICG9Rjd2/i7e7Db6PyXOslOjRmOKih5lM559RBD4yDFGNuiMsKIWQkmB3BxEEyDDE6oK83yfvDTBpNu/TzLY/6fsEO9vUnz6h/vQxVmvyy5ek745Ij47RkylCKqQIlqjW876QXVA2pefhNRuEe7vUnz7C3+hSjMeoWkQ+GJKdnmGyCcIItHTVk5n/Y5eGSIsVcsGsOrLX9m7FGrhMXIHbZlCbX6l21FKwEFdiieWGyuT2GshuRa7aXBmeNeR25fMKu+a1K5/Rrsh0m9X9aWkrmhNzhFzV/7mthH4N7Z2PR/ZnE9NzmZMKTL/O5XHW25GL71frePGhzfWfNcjcNAjNzdacrIWoR8h2C3ynhyL6w0VuZ+2y9sli8GsuOzOna/s+wvMRykPk8pqk4WNZFdU0VOO8Uy2yHuFtb+Lt7SDbLUxeULw7Juv1Sd4ekR6+Qw8GboA1CpBRhGo08Dc3CLa38bc2CHa28JoNgq0NhPySYjAkefuO9PgMkyTocex6NOSAQQYB/sYG0b171L94MlfCzAcDdJKAgML3KYZjTJKUQ60GIeXVDbcMTiqKCHd3qD28jy6Df/z2iOmrN5jeCGsM2hYlbVjPg8xitocShrpumHT5rK/CbbKsY2a9nNmMjzcnHbivskKxrt5I9gqd+/1XXKwJPNXjVKu2cgS5AimaJTafKf+/FSs1T2U4bRE/VkuXdVrl4krSLJZ4IvZj9spbAdNfIACtg3/sdWd+njXPtbpK6Q7pJOGFEFdldywIs6ZDufoqHxoEFT/3k72/v/XBn4uPe4Ff9JhXipEShJHOQVAEASKKEI0aQvnYOHHU1Vkeag1CqFICRzo4KM1dxTQau9/vtN1xajXkDJIrgZr5hruKt62BEFdjz9LvCFFm+C4IiShEdloueHo++XBEcnHB5McXjL/7G/Gz5+S9SxCOTq1qdfxOh3Bvn+jhfeqPH2K/eIx68gjhK/yNNrUH92h89eU8CKVvj8l7fazIwUqE7yOjGqpRJ9jewmu38VpNZKNOsLnB+IefiJ+/AmPJihyb5QjMfGSqCtbYPEcnCTbPkYFPsLOFCAOMLohevsJ79QqRTcgnU7Q1aDIKq13ck2JF6cBU+kPLcNvqllutdhQCJWZVTrXaWVw9iUAKee3oQ3W7t6zjoF4dTRBzQSUxh7FmbD1tF6QKXSFW6EpVtFzpiQrMu6h87JWqyy7vH9VBartcFM3uOSEEUghEOcUuxLU5uF1J1Jcet1YOv8pKqGQOyfWXTby32XRb7XzapbPLjDclFwEoihBBUPqar8JosxM9c0KzmDhB9wfoi0vMcITd2kB4ygWgMHQV0dzjdE10+Ril6SuSC+U8je8howgR+OgsI+v3mfz0nOEf/szwD39i+sOPlSBUd0Go3SHc26P29oj8/BI9dYOq0f27yMDH67SpP3lEdnZOMRqVdO4+1uj559FxQn7RI+/18dotgv0dmtaJn5osRw9GFL0+cqRcT8eYitDmIgiZPKeYTimmU4zWyFqEv71JLctofP0F4zeHxJMJ+viYZDohy3PX4xGeU4Gwdq12h7gmqanO93iVp5qz21YJBtd1Ua6ve8QNE6Rqb2q537OA3BbPRSfNru0YLj6v+FkOieuqoFklVKVb3Gj7ua2EfnWRa20QquBygt+eJvqvMQjNN3GB8DxkrYZsNBBR5JQkZ6Kexl4DpDjZYZummOEI3e9jxmMoCoSnEGHgnr6bIrmRHtEnYopCKYTnuU1sMiE5Omb603Mmz34ifvGKtHdMQermNyYT1CRCD8bo4YRiNKQYDMiHA4rRmEZ/QHT3ABmGRHfv0Pj6S7LzC9J3xyRv3pbnQ2CynPT4hOGf/oJFkw/6BDvb2KxwVVIYIgMfqTzkHIKr6sfJ+SfQRUGepeRpQlHkWCnwWk3C3R2av/uKyckpk8GAUTwhjUekNi0Poyob+fJcz5J+2wpbbcFwk/MZn2oF5FVIBmLNnJD9wOZ7VWNRrGG1LmaYDAu22yz4mBJaLOZ0clPpcYmlCl+8B+qzH5Mti1UkTqyQquwSGv1bz4H/4YOQXbpxK/MNsByEqg6Y1xVZ9soBP15p2H4AmL7p7/+cY35Mr+YXOaaoIJszDwoXMGSriWq3kbW6+5MkxSSJoyyX12+5cV3Oz+cZZjrBjMfYOHYsOs93lVAUQuAvZf0/6/3bq7dFtUqzaUahNdnJCcm7I7LzC8w0qXQ7hBMWxcNq64ZJdUo+GpJdXpJfuqqm+S+/o/7kEV6nRe3RA5J3x4z/9iMyiiB38KJJEuI3h+STIcnJO6avX1N79Ai/0aQYjij6A3f+tJ6zQq0VS7HIcR0sxhqELsh0Rp4l6CJ3m0C7SfOLp8T9PqPeBV7/Aju4QE+mMA/tYnmCf0W6eq5gIK5SqWUFbnPyjbPAI64sL/uBFJ81QhB2BdZiyXBywXDLbXWY1syZfMUStGix710W4r2B4cMdSTtnwy0qnkr7SFRXwUKTfVk5j8p5vw1C/1iV0FLedoPN+fbxMy7CLApprC1cDhwGqG4XtbWJiEJsHGNGY8xoNB88navLrixhWxSYOMFMp5g0dZDTrL8U+JVKyCwgvZ9l9maXqzEEaINNM8xk6ja1fp9iNMJkGcLz8GQTa6Ky/+W5eSgBJs8wvRF5r+dgtcGAYjzG5DkyCGh8+QR/s0t4sEews4VqNpGjsKyEMvKLC/LBJcVo4GDJ0ZRgYxOTZqTHJ+T9ATpNMEYvHI3ErNE+Y7cJrNUYW5AWGWk8JRuPsdYifZ/a3QNa469onRxTf/eW0dkJ2TTG2AxhLcIa5Ky7IpbTuGq/x5vruC0YbnIlPbgq1Gs/uOxuIrxuF2lPpfpZwG1ZhdVnVuSF1kOJnys7+9BSWbGIf69x920l9A+1Cb4/CN1icb94EIIyD80QNkBGIWprYy65UwyHrs8zGGDTtPw7tVDLri5+rR0klyTYLMMa6yjPM4ac72wRrLHLQUh86nW+GoRsXmCnMWY4xgiDmcaYPHe/O3svCEcKMKWdgykwNkEzdZDYeIr+KcGaAuH5BNtbhHs7qHYTf7OLv72Jt9FBXtSwWQrGCZxKGSCl51xJJzHa9ikmU7LTc7JejzyeUOi8dCcFLVzlU8zDkACrMaYgzVPSyYRsOKSYTPCbToevfu8u7S+e0nn1munJCcUkJhleYjMNZQ9K4aGkctYDdnmYVF4zULqa0S/U4W4Gt637fiGLNbPXsEtQW3WIdkY4yK9o1dnKgOsq5CauJUV89pR5pSd0xRnvNgj9g2FzswA0u75yhZggrlrkzluWAtZPH1YylpW/uW7znS2RD93ANznm8kDlz8EHf+FjLsrQMg7YhfQOIGsRarOL2upixhNslroez3CAzbWD7KQTM52Jm861s7TGprmD77Ky+S6ko3gHAXi+mxUymlUdZ3cgW0GS1p9zURl4tTPxUGbaK2CzHDMcI/pDjC/BaITnITy/NKzSGFE4Vrcth6RLBQhJMH+tIhsTv3uDCpvUHj2g8cVjolqEqtUcjXtrg+SoQVFokBKv0STY2SK6c0C4t4uqNxzJYDAgu7wk7fXJ4jF5oTFIjBIObtKaQhRoa7BWImyBsQVRlpFOJ6SDIelgiN9oggAvimgcHLDx1VekF5foOMW+MKTnl1hyFJIAiS8857gt7MK7p4TYpFiQEeQauG3d/XNdL+XDdlDLJIM5scAuKOXVgdKFdt26ThIrHaC/ByiyGB9hZq0hZy0DuyTBdF3qfDus+o9WCYkbCNb/xqxz/+446JXNvcS/fQ9Ri5CNOqIWwXTqAspojIknZQAqB0+XdoBZJWRcBZSk2DTDFqXYn5QlYUAhlMQW4jNXQnLOUrNJhhkMEf0+tOpIz8drNlHNBuKy7yRydA4oVBCimjVkLUR4jlxhMo0eTcgnfYrRkOTwkPjFK5K37/A2OoDFa9RdRdRpY9PM2Te0WwS7ewQHB6hOG1PkZJMxce+Sab9HPB6RFTEFYKSPQTrPHltQ2FnHwyIxSKPJ8oxsMiHt90nOzvFrdbxGHZ2mBLU6nUePKIYj9HhCMRhiekOMTvERhEIRCA/lyPNzAVGxBsBanMXPA7ctJYwwD0CzIdp8hd1WHaJd/3riV1BrrGfH3VZCv67a5tNCwnVB6BaF+2WDkK0EgLI3ggRRjxyBQEkn6jkLKEmKJV/SPVuRfq7AcRl2mmBjB8m5ashp0QmlSt22lSD0qQvZrgtCKbo/gF4TfInyg3Jup4UMgvktKz0Pf2uD2sO7hHf28TpNrIGiPyR++YbJyx/JBifk/R7pyTHpyQnRvTsO1fN9/FaLYKPrKj9rEGGIkdIFj+GAfDQkOT8jPjslHg3IioScAo3AGokRM3kZU4JxtgTELNoY8swFoaTXZ3p8ghAKFfhOo8/zqO/tYr7+iqLXI3v9Fvv2mHySECCJUPgopOsyoWZUYluV4+Fnw21U4LZFj2cxn6RXWG7VSmgx31OVWb3aDf64Ov8XWi+C91C0Pylrug1Cv0AAWiEnLZxVr8WZZkOp5eCXlQIxk6ad72v2+s3nGnZcVXPso12mrkG5rjvm9apt19C2bjCw8Ises8qMnqtJg1AhImqguh1HRsgL9GCE7g8x03hR0cyPYZfFJIUbYrUlKcAmCTZJIc8XitWVSmgZ3qygg3aVvPueuMOighJlOx6Ec0otqyDRqCE2Onj1OrJWm4uOCiRes0nt4X3a//Nfaf7+d4T7u1htSN68Y9j6I0U8phgNsVajpxPHnpvEc2twGYV47RY6ySjS1Enl9C+xyYTCFKSjIelwQDYek8dxGWhkedsaZwchZurfYmEZb21pL5FSTGPy8Zjk/AKTpKA1Mgio7e9R29uh9eA++vyC9G8/YV+8IZkmKGvwhXIePXaxWTqtqzWzPOuJdO+F265juM2Ci16SzzFX+j/Lm4ZdAdqWsyX7wZXxiU6Ra49pV6D9CgtuyTdJzP8flUC0kBZaHgmXt86qv0446Co7bqXxd/v4xfBtJxvj+jJCesggQnbaqJ1tRK3mxEqnMfr8sqQ1i/I2/ACyXVoi2EK72SInbe0UMZR0FZaUFc23n5sCVSuhMsNPM/RwBP0+cquL2Ogig9AFj7LiljIg2N6m8eUTOv/Pv9P5f/0/RPfuYI0lfv4SrCU9PSY7PUUncRlg9aJqkxKrFMb3KDxJFhfkyRQ9Hjh7hJJUkKcJes6Gm4mXlvuR1fOgI6qbm7GI3CmWU1o46PGEuDeg6PfdDFSeE21tEG1t0npwn/jBfYrdPbzeGDuJ3bkwC2klK6p0cPsz4Ta7wpK3S/I5xYpVhJ6Lh9prRljFrxvQEpVgtLRH2St969tK6L83rNy8SyOqxdBi+IuV79/PjrttCv2sFTVX1yrACEQUoba28fZ2kY2Gg7QGQ4rTC8wkdluF8ErhvjXp8aqnjXAbNUqB5zm5Hk+VQWi9B84nXdN5JSwXzK4sQ4/G2H4NNZpAXsytG6wxIEHVG0T379D45kta335D8+svnX8P4De+pRgMmTz7kcmPP2EvL9wAbxCgahEEITbwyYUlzlPGkxFJ/5IsjUsrAYkucvIioaAo1Z5VqRRfaadbiyzVDURZnQkhkMag0hyVFXgGPKWQxpJf9khevMBog+/76KeP8XZ3qO3t0bh/n/juXcxZn0yfY7IMa3WZA3hz+HQdG/46iE1U6hRbrV7soppZwG2GovyqqxXRnHCwuFjLE2bXB7//rtV9nZzBdUHoZ0TQW2fVz42YVk/qqnbcanE/U8ueZ4BzDarF0OqqTPqyie8qXvbpl/9a5psVP+ukfPIxxS94zCVmoQNQBAGiUcPb2ULt7CCEoLi4RJ9doM8unOK0dbehWPI5XwQ1h/YIUD6iFiLrNWS9tpD9CWYipp6DxCpmdPPPYPlA0rHuGlb6CGIxq2SnMWYwwvZHiNEEHQQUcYrOMqwSqI1OyXh7Qnhnfx6AAEQUUH/ykPqTR4QHBxhj8Tpd/I0NvG4XLSWFJ4mLnPF0zGDYIxlcUti0DCQBVoiyGyOXUFMhcXNTUi4EQUolCmEF0lik0XhJjp/mrr8TRgRBgE4ziqNTiuGQtNkk+7ffY+7fR7VaRPfuUHv0gOz0giKJ0Rcpxs5s1L0SrjRrF4pdA6+tLujVgdLVwdK5YGilJ2RWILebVD32E3Zv8ZkD0BLXptIyWF5Di2+qg/fz62zLKqkyiLwOQ7h1Vv3vjl7i+mHV20roF0zzLEvCjcJTyGYdtbWB6nad0kGcoC976EEfm2blGfeWmwCzTd9oR7mWClELUd0OancbtbOFbLcWXkKecs85HGc/3zUVC5gRY8tZoQTbH8L5JXngU/T7mCRGSIm/0Sa8s4e/vYk1hvToFKEkfqfthnWbDaL7d50ZXhRRf/yY2r27+Fsb6OmUrMiJ4wnTeEpaZOTClp5nBmlNySD0kGZmfGdLmybpKqooRIUByvMQhcEmGTZOsdMEYTVeluGnmlAowlqdoF4nlxLGE4qjE9Lua5JXh2SPH+O1Wnjb24QP7xMcHROfn2EuLtA2K00XQreZ2ptv7lUDBLPiR1TYhX5bvtLnWZckLWm4XbOe/xFW8tVK6GeT426dVX8lUWjhzyG57Qn9ogWrXao+BMJtlFGEbDVQ7RayXncKztMYMxhi4pGD62RQzveYSgXkTJ5m3QBhAmTYwtvbwX94H//eHdRG11VCJTQn/DIYSTG3av9FPqs1TtG7N0QfnZAqSX52hkkmKKHwm038ThukJD4+IT49pRiNCLa3qT9+gF9vEBzs0fzX3xMe7NN88oj644d4rSZ6OCAZDkhGI7QxeJ0Oqtkgz2LyyRgzdcCUkDMNglLJwAo8pfCjGuFGh3Bzg7DVQhQGfTkgOz4jnyQYtFM0SAt8K/H9AC8MUZ4HhUb3h6Rvj4mfvyL54im1Rw9RbefqGr45RP30ogwexZywIcSHq4a5dNDS3A5z3bZllpupSOusc+uxsAZ2+0ddO1cDUIXwcUvR/lXk1otJx9kPr2XHiWW3RCdSVXmapZ4RwlaG0mbDlSw8Htai2Uvv44OF+01UY252zOuMkX8OUvw5jllpSlvr5OexCC9Ahh5qo4Nqtx01m1KEdDp1g6omLWG4wA2oVthrQsiy4e26AyCQ7Sbe3QP8Rw/wDvadCOqMwTabQYoijOdh8xlBt3pv2BulxfYaYGZuC2ElNi8oBkMya0gwZKNLTDpBBR1UGCH9AJOmxCcn9P70Z6bvjgj3dtma/k/a33yFt71F+9//FRPH1A/2HXFBCrLLS+KjY/L+AC+MqO/u4jXqZJMRo5evmLw7psgThJBI6znw2FokAt8LqNUbNHf2aD1+SH1/H2kgffOWSW6Znl2QYVGmQKYFQhsH75RKDygPU2iy80umz18xff5yTj+P9veI7hzgdzpIz8Nmdv7arIHbWIHblh1YzRVX1oWw6GKux6zc/x+qdMR7rt2n1kX2RkCd/djVUmVmXzE5E9WWwupQfQlNu6/ufr7VjvtVZ+fXseNua5fPe5pnG7zGagtSIcMaaqODt7/n3FOtxYxG6P7ABaC8gFUrcCkc3GUN6MIxxhAIQlS3jbe3g3fnwBEcWi0XuPIctHZq0o0GolFH+B42Tpkpd38sGieu3Y1mLr0Way1FHJMWGUmRk2RDDAnSb2OVwgrhrB4GA6ZHx4xevSaLY2r371J/cJ/a3i5eq4kQgnCji9dsEJ+eMT05Yfz2LWm/j99o0n36lMbdOySXl5DmZBc9TDZ1KgVCuP4AOGBM+dRrDVpb22w8ekL7i6dIBNOojjy5QKvn6Nxt7SIroNBY7SjuIopQzSYiCCmmCcnhO6Y/vSQ82MdrNvA3OgS72wTdMshmaskor5o82RWatLnCcHNDpXotrdquqRNudIX+4QD0JQXylfbBbSX0m9kfF6Uu4rp57tvHZwlCMPejFEhks4l35wD//h1Ut1PK8wwojk8xo0lJ71VzWUtRSpY4rbXcBTMswvdRnTbenbsED+/j7ewgazU36zKeYPMcm+eIMER2Osh2GxFeAJPSs33FYtneLPgsjMeWfzpLYIw1FDon0xmJSYmJ5wBTLsF4ChGF+J024e4OeRzjb3RRtQjp+3iNOqpWQ4Whm2MD0uGI8dEx46MjkuGAcHOT5oP7bHz9FcnpGcnhWyY/PseOBigkvlBuVgfwUUTKpxbWqLe7NA8OaD9+jFQKmWYk3z3DjyJkXm72RTE/d9ZaZBi6gdt6nbw/IDs5JX71mvoXj6ndv4Nq1PE3N/A3uvjNBsU0BF1m4cbOfYZmZnerlgnVgdJVl1K7pqL5NTLbfqnls9JB+DkU7Vs/oV/oGkmqrmLcgB3HMgNOliTV5YssVjzvK/3wVevDG0BzN+HULP/GTY7JB3/+c5g8n+uYizPlcl6hJGqjjX//Dv6Du85++uSM/PAdxdtjzGjsrpANHHNRVKbtBY7+S47AR3W6BE8eETx9jP/gLrJRx0ymFMenkKWuH5QVyEYDb9sRFvTxCfqy5wgNRi27hK+av1wXgKgONTO3RjDWCYPOsvlMWBIMcUkjthTEaAolUO0WjW4HopDGk4fIKKL18CHhxgZeFCGDYB6AksGQ8dt3jN++ZXp2Rp4kiFpIbW+X9qOHhPUa4709Rq028rKPMgIfD1G+qicUkQqIwhphs0W4sUm4u4MMfPLLHsHuDl63ixwdoSmwpkBnCTp2FhrC9/HaTbxWk6I/IO/3S8vxY/Rkgtdu4bWa+BsdvG4Hr1fHGGcDYYxGG40WYEQ5TGrtCtw26+9UrcDX00fER92t1/3Gx64k1kLTn/+Y4goCIMSqfUMVyVlWoRDXI5G3zqq/6kpoaXDtthL65R5l9eJ5yE4Ltb+D2tlyMNxoRHF0QnF2hk1zx3ZDVei7ZgnWA4MMIvw7d4j+9fcEXz11ASjNKN4eYUYjzN62Iyf4PqrjoD/v3TF5vfQp0sWSIrIt4asP7Rt2CVRiHl7nk/vWkGJIJWRKkhlJYiEH0BnTZEoSTzFSUD/YI9zfRScJQimCVoug3UaFQRncLGmvz/jVG4bPfmTy6g1Ff4gSgrDVItrZora/i6ckzf19WptbyNMeMslRlZaXJySB5+NHNfxGA6/VRLWbTnlhexN/bxdvdwf15hWaIcZqdJqgE6dYIYMAv9sh2OyWJIuY/Pyc7PSMYjgk2N1BhgGq1UR1W4iojsknWGvQxp0TLUQ5HWYqMjrLfZ6roV7cIAD9k4AKayqhj8wubyuhXyNMdKsd9wue3xU6tsvcSuO6Rh3ZbiLqNRiPMJMJ+rKPiYeAQqgaQiqENa4HZLU7iikdWIMQtb1J8MUTov/4N/wnDzGjMcX3P1AcHiFCHzN5QPDFU/yDPdTWJl6c4O1sI+u1Mog4Bpe7D+RSpVxdrst623bBhrFmzuiqNs8LDCmWREgyJcmVILeCXFtknhGPhkwvzkn6fWr7e4Rbm6gwRCh1pYms45j47TsGf/mOwZ/+TPLiJWoSU+t2aW5uUdvaItzcRAlB8+CA1v4BHF9iegOYJvOxTSlABT5evY7XauI1G26eqhbhbXYJ7uwT3j3A/2mDYpBgMWUQcmZ4KgwJNrqEW5tk9RrpaIju9SjOztH9gaPSS+muZ7uBrkdkcUyRZWhbOOHUUk2uynDTS2oGV/XbrpPssf9sa0lcZcj9MznO/LbYcVT6PwvHrRXevVgeEKMq2f8xoNTHMWQ+vrT/8HsQn3RKf+Yxq5P5xpZMHRDKR/gK1Wo6sVIpsXnuPHcmsdOJI5+PBAtZ/mFhHQQnyk0qjFDdFsHjh4Tffk3wu69QO1tkP/5EcXZK+v33IBUWi9rZwb93F6/bBmPID3ZRGx1EGCJy6WCQKh4nlttDtrQeXbV/LlZ6F4W1S+ytTEAGpFa4bF9Kp6BtLfk0JrnsEZ+e09jfx2+18Ov1q+fVQnZ+yfjH5wz+8w+M/vAnijdvqWlobWzR2dohajRRYYBotagfHNB68BBzckGaF+Rx7KZprEEJ3AxSq+F6O5E7/0JKvGaD6M4+9ScPiV+8IE9GkGl0nDhjvSRF1euoRh2v3ULWIqzWFIMR2dkF2XmPfDTGGI2NAkyzTtGsOeHULKcgL72KPKyUZfBerX2u+vV86O78WBj573b//4xjWhamDPZKr/oqS25pW1upGCWUQq23Vg7/UJUQt5XQ58MOrAseWOuy5DBCdVpukLRWc/I8ZxdOnmc0AW2WUzuxWJ7WOOBGyND1gR49IPyXbwi+fora2QIs+rJHfviO7MUrhFSobgfz1dBRlJsNPCnwD/bw9nbc/xsP3EaoC5fFlHNIZk6kWASVat+iag1QrDhxzqb2CwuZtmQY8iIvK5LSu8oYJ3Qax+gkdbblFUsJqzV6GpNf9Bn/8IzRn/7C+M9/Jv3pOXI4Jmp16ezs0ehsoCyY2OnrBdtbNB4+ID86QfcHZGfnaJvNyQAEHrIeIWuRq7ImU2f5LQTBzhb1xw+ZPn5AcnxM3uthJlPyXp+812fWpRRhAGGAFqDThHQwIDm/ILjsYT2FlgLTqFE0I7JAEltNQY5FlsbmYqn2ESvQ+O1j/VoSayshwT/LZvWbtfe+NgjdLobPE4RwjXgoECZ05ICDffx7B8hGAzMco0cj8jdHmP7QycdIr7RFYMlmYbb9y6CJd7BP9K+/J/qf/wP/0X2whuLdMdmPLykOj9AXPYRU6JNz9NkFdjJxZIhuB//uPsGj+2R372CGA4rREGNSMLkbgFU+VkmsKeEiazDCznXIdKWhXgh7ZZaFkkhQWEthCnJjyEkpymxU+j5eLcKv1fDrzqPHq9UQQmDixDHPLi7JTs9Jj88YP3vG6P/+F8lPP2LOLvA9n/rWFo3dXfyohhlPyM4uULUaXrtF7cE90rdHJG/egpROR3RW0fkKEQUIT2GShOzswpEfAg+v06b26D71Jw+ZPnuOHo3dsY/PSA6PMFlBkScUVpN7kkw5irk3GjI+PUWdnKJaDTQW6jVo1jGBNydoiHm1QwV4W2cb93FVzz8HGrc+CN1StH9F9QwfoR0nZguhwrWXYqGYMIdkhFhCp0UFlLOwRN0Ry9Oki2+v02Gza7xwlv/0RscUSxJudu1yvu6YS9Im152tTz3mrJq0szHEHAiQnSb+/Tt49w5ACPTpGfmrQ/K3py4IAeAvwANjF/PFs3fRrOM/ukf4H/9C+K+/Q7WaFEcnJH/4M+mf/kZxfO6IDUKjL3oU744pTs7wHz9A7WzhHewTfPUF4ctDzHiIfqXR4xGGHGSAlQFaOQitMLqc1l9M8euKB43Fsb1mdnBuTyh7TLokGxs933A9GRBtb9M4OKBx9w71/T2irU28WoTJc9KzC6Y/vWDyw0/EL1+RvjshfnPI5KfnFBdHqFQTdrao7e8RbG+CgPTkDFMUhHs7yKhGdPeA6MFd/L9tOv+iYrFPyVIIFSEohiOK0RjheQTbW3ibXcI7B9QePSTc3yU9OqIYT0kOj5D1BtlkgvYE8XhMYgqm0hniMRkRnJ4g370j3N3GGON6TY0GIggqTj2GxRwVJcS56mD7c0A0e6PN/FOPea3T7mc6JvP60Fb2qmU9S1Z7QSzPCs2hYztXkHLX/TYI/d0Ckf34P7qthH6Rx5LygOuiCCWdPM/OFmprEzN0fkH54RHFybmzXpDKERdmhzG6VN0EUAgRoLodvIM9N1+00cEMR6TffU/8//1fpH/+C/r8cg6Cm9GY4viE/OiYYDBA7W7hbW8SfvGE4t0JxXBIliRkaUqej9E6xRhBYRW5LiisXnLetGt0ylZRfYEp1ag1EoNC4IsQPwyp72yz8eUXbP7+W7pff0XzwT2Cdgub56QnZ0x++InRH/7C6I9/YfrTC7LTM/LLHtnFBZYJnmgSdrsEO9vIVhOdJBSvD8nOL7B5Tu3JI7ytLsHBHv7WBqpeR05naKhaBCHrVLHz3sBtXU8e4XXb+FsbjpywvwNRRD68wJyeUQQ+ajLG1HzSeMIknpKUEKRIY4KLc7zjI4znjANlGOLVaijfX+pmSLhdWR/dQ3p/JfTPcj5/wxRtrgQh+0+Es/6iQchUvHak5xhx9cipW/sBVhvMeILuD9B6hMR3dg3Sc3Rsa8ogVHp+BiGi3kBtbSA7LYRSmOGI/MUrkj/8meQ//0j++iUW60QzATOZUJydkx+fUFz28PMc2WjgP7hH+O3XZKMhcjpBJxOS84QsnVLYlCIXZfARpUH1VcvoipBTJQS5v5gZxHkooqhBtNEl2Nml/egBm//yLTv/8e9s/P531Pb3ENaSvjtm9N0PjP74F0Z/+CuTv/1AcvjO2WbnCZoYMEjl4zWaeKXQaTGekJ2eOZsGpQgO9vG3uvg7m/jbG3idFqoXYXSB9HyUHzgWHlD0ekx/fIHJUqSS1J88xD/Yxd/dQW5vomsRiSmwwwHqRCDiMSbyKUxOPBqSGY0RgizPiPs9wtNTVKNO2Okgfd+pJng+3FKsP0uKLdY6rLJESPjI6HZr5fAZU4aPcFYVK32g1VmECg9lpey1S6+4UnjZay1RP3gnCPse99aPOOay8emHf+djb9kbHbNi3W1FKf8ifYQflDMpkeuZZBk2TjBJ6r5HY1GVGR3h4pAtVUaFh6jVUdubqJ1tZC3CTKbYLCP78QXZ85fk747QDJFEQOjmXJIp9vICdXSMf3yC1x8QNhqozS7+l48IkgleMoU8ppA5ydEReRGXtgDOMVXMw42t1M+28rPlaaFZxi+ReGGD+v4d/MePaH75lO7XX7Lx7Td0vv6K+sEeUiqyoxOGf/6Owf/6T0Z/+AvTH5+THrkqzeq8fKVSNWLmKxRGCCTFcET86hCb5ahWg8bvviLY2cRrN/G3unibHbzjFnoyRfoBMozwwgghBPlgyOTlS4rhENVq0Br8G8HdfWSrCe02RaNG4gnyIoaBwcYjrO85ZmCWUhQFWNB5QTYakfQuibY38RuNecXlbDMqjhklHP7+G2gdbHWzvfpDx7kJjPbxS+TzH/NKFVSVGFuK5suBaNmoE4SwwlowzonrytsoVbRvrRw+YyD65EqI20roM6ZtM1DagBQIFSCj0NkqhCG2KDCTqQsieV7CnwugpjoB7ozQDEJ4yKiGt7WJt7XpmHVphhmN0b0+Jk6caCM+Bh8jBYUp0GmG7F1i371FvHqFenjPwXmNBt7dA4IiJ8xSQp3hS4vyFMXpGWIydZUHnqORG8u6efTqnLqUEmFBWYvn+6hWE7m7i//kCbVvv6H9L9/S/eYrWo8fEu3uIID09JzJs+cM//OPDP/PfzH52zPX44kTrDEI/PL1lZM8CnynAi6cJl4xGJIdn6KnU4L9HfKLS8y9A4RSqGbdqRfU25A42R3h+aAUxhjS8YjJ6THp2Tlqf5vOyQnB/bsYY7FRhGk1KOoR6ShDZzEmm86TOKcD586PNRadpuSTCUUcO7adVEjlfIt+1gK9fVydBaqoudxEnfy2EvrVX2CxBo67hQx+Tta2LE2pEH6AbLSQnTYiihw9OY4x0zIIgZtXsWI5mytzZgevCZQfuCHLbqe0fahhlEJ2O6jdHeSdfXTsU2RQ5IZ0kpAzRfQvyN68xvzwDHmwh7e5iXr8AFWvEz16QFE4/1EbhqjNLsnhW/LzC4rpBJvkLmjmuaNvG+sEQa27Z5SUzg9JeXhhhAoCx3zrdPB3tvAODggeP6L29Zc0v/mK1pNH+A2n1lBc9pn88BPDP/yZ0R/+zOSHH0neHaGTqYMwUa5HpiQo634WlkEIi8ky9HhC0e9TjIbkZ2cU5xfo4RgKjQyDUmanhYlzCAOskmhrIM+I4ynj0YD48hx5fET78BDv3l2KJMV6HqLdRLSbmCymSCZonB24sLKsEMtZFmsxRYHOMkyeu+Ap1RLScPv4PHvVfM9CrIz0/vYfv60gNEd8KsXzXKaWZUXtawfL7HvBgA/DCJ+uynYjdaolxp39KPiC62CQDx6zPH+WSvteuSqo20ZtbCCbDVexJAkmjrFp5lSa7fXvbn5ppHTuqPU6stNB7e2gpITQR2c5WkrM2TvXJxkOic9y0vEEignpyRH6+x8Qm5t4G5t4rSbh3i5+rU79wX2k5xG22zQe3Cd+c+g00c5OyXt9irHL8JmpcWvnzaOkwvN9/DDEiyL8Zgu/0yHY3sTf3ye4s4c62Mff23VqBAf7iwDUHzIqIbjB//7PsgI6dRI5mJKiLh2MZcpz6UmErxB+6QxbaEyWopMpxXhM3uu54HnZdxWokMh6HdFuOn+mwENLKIxBFzlJnjLJU8bxBHF5QfvtW4K3bxHKxwJes4nXbqFGA/JkUrkesrIB2rlFx1Vt7OV7aTaEaefJxc3vf/sxC/sTSoLP9fuf/ZizdkHFYVVQ3atK1MD+9svNf5JKqFIR3eZdnwQbCCHmQqVuVru0297sorY2Xb8Bi0kSbBxDXpSLSM6pqKww0Zh/NWgB1vcQraYLQu0WansTwgjbamIPDzG9C7KzE0QjxB5JitHYVTVv3iC7HbyNLl7HVVJ+q0XY6eDX60SbmzQePiR5d0T85pD47VuS01Oyfp9iMsFmmXu/hUZZgac8/DCcz/sE7S7B9hbBnX38+/fw7+6jtjeRzQayFjl3VyA7u2D83Q/0/3//h8H//k/G331P+u4YM03KDV4hpO9YgSWsKYTz9JGBj/S8OSQ2U6Y2OqcYj0nPL0hPzhC1CK0NNgqhVceMI6yvKKTTbRPWuZKmwhKbAm8yZnx+RuPkFL/RxBqDiiK8eh0VhEgcHChRSOGXscdUlpF01Y9Uy2oZt4/Pu1dRgeKu1eyxv8kK6TdJ0aYiBlh1WV3N+m8fH3MVZl0S54cJFlGLUFsbpWZbhBmNIElgGiNmQQg1ry9n4cdJ4jiFMYHGs06TwFOCKAyQnTbezpbzIvJ8aNaR9+7gDXuo0yPkTzvInzpM3xySXfaIe5fw04+ukmrUEWFI49Ejwu1NpO8T7u0S7O4Qbm8T7uwQ3b1Den5OPhhQTKdXgpDyFF4Y4kU1/FqE12rjb27i7+/i3dlH7m4v80mynPT4lPEPPzL8zz/S/9//yfiv35O8fYcZjrDGIpTnNOxEOaxrSpBTSMduCwKUH7iAppxbrPUUxlry6ZTk/AL/5BTVaZMXOTrw0fWQInJBrVACoyRSKaxSrndmDVmWkk0m5OMxUnkOUlMK6fvO0E5I90QiRdkjqyAKUjpIUirXQ7tdP5+2fOwHl9e6YdWPqq9uBUx/wQB042HVOX+2CiGtNv7ginbccuS6qZ3B5zJP+BnHFB8rL38D2XlxdcmIqqMjBsgRwiLrkQsW+7sIJbHxFJuk2Die94QWEI3BWmc+UFhNhqagAJvj6Qydp6gsJSwyajNqdBjg725B6OPf2aeWJtR7F9Tu3yfc3sarNxh+/wPp+RnTd28RCKTvI30PtEEGPn67Nd9Qg91tROgTbG9STCbo6RSdZlA4kze0cew3JVG+71hgvo+KImSj7lxi260rpzE7PXcQ3P/9A8P/+hOjsgIqRiMwOQKP+SRNqTHnblVbarwplBegohBVq4HvQT3CBj4FkCYJ8WCA17tEKVyvS0ly3yNV7p4uSg8jr1bDC0OE8hZD27K0uQf3+lq75wwOtAuKOtYlClK4npgqGXvKD1w1NKvS7PqF+suQtcU1oJ74GcexN4CvP8cxlzXzlt7/0kzQsv8ZLBOr5mME5bm35Q/krZXDrxk+4iocx00aqh9feP3WsrbVhSOWorkpDewsouYa/t7uNrYoKE5P0HlGkSQuW18yODOl7polpyBDk2OwtsDLY8x4gOpdEFycE/X7+NtbLhNv1AnqtXlmWIzHRFvb+GEdoS16MkGPR84D5/Urxn5AGIT4KDwE4skjvK0N5z0EeO02frfrVrEu3UW1C0CUTqWO+eeqEaR0rq/lYK0x2gWO8v1k55dMfnrJ6A9/Yfh//sD4bz+QvjtCT6agDQIPKX23gWOxxsx17KxYCN5YpZzDaasJYQDNBkXgk2LQeUqSxARJjJ+lFNrJhuYY8vL9WF+h6nX8VtNJBnk+SkiU5+PXG/hNJ4ZqjUEnzkvI5MXcNXZ2388V9axFeh5evU7QbOLVakjPcySFMoDdPj5PJTRL9MSVfYtbK4dfweMTW3IrbkL/ZHa5nyMILQ9rOqq1QSAqtmQiClDdNmprwwl2+opUFyR5Sq4zdPnPQhpHoq2lQJNj0UJgrHEzPxfnqMM3eC92CPZ3Uc0mQaeDVwlA4JrqjcePMUlO0euTHR+Rv30H55eIUQ/9+hWZXyO2En+awOk5wf27qM0NZLft6OSzqtjzPmoFmEJTnF869l/hhFGzix6TZ8+dHM/zV2TvThyLDeNEPaVCSlXaSZT2ENaWtPOSpm40SoCNQsc0rEfYdoPMV0xtgadz6tb1zTzhbCZ0UZDnOUWWOaUET6EadYK2U+1Wno+UCi+ICNsdwm4XISRWa/LJhHw8wWQlgxFZvj9XsRrhZtxkGBB02kQbG/jNphsiLtlyzn79dkX93M1tOWlefv6zPH4zw6oL5bhq4MFpx5U/s9U+0T9J0+8maPJ1vi7Ylf9nmWusgYXAQ7QayG6bwpcUoU8qLVNbkJmcwuZlGLJgJdbKkltnMEJghWOJ6SQmOztjGgaoeg2/2UIoj/r9+4T7u24DrLwNWYsI9nao379H8+AOWXsDwSGWDH/QhxdvKApLfNHHvjqkePoY//Ejggd38e7sOcLDx54/XQqpvnhDfnGOzgu0lOTjCfGr16SnZxSDESbJmFkXCNx8Eda4qcKKrprBuoBsDBROmL8IPUSn5ewwOi2ymk+sIPQlphaimnVkGGLThCJNKaZT9DR2mbMUjnDQbOK3WwSNJkGjSdTpEm1vE25touMEnedkoxHZeIzO0vJ9OkkljMUIgymZ416jTrS9TW1/l7DbwQqcnNB0is6yeeW4fI/8sijCqiLbp0Jz13NYP98xr4KIq6YxS3IJsKJouaxxuUxXuLVy+AfI55dqofc2+/6ZKpybBiy7vCCtC/F6BgEBHoZCSWwYQC0CUzj4SAliYUjRFLbAyYSCIyjM+g8zecuyOshz7GBAIgUy8PGiGmiN7vUpzg8Itrbwu11UveaqFyFQUUi4uUF9a5us1UV4NXQxxdMF3uUQU7whuxhgj07Rx2cEZ5foi0v88wu8vR1kq4GIwpLZNiOtlAh7CdU5mWqLTTP0ZZ/s5WvSZ8+djYIQmHqNwlqys4uFdYKUSOkvqqfZcO/cMG9m+ObYbJnJsQUInVN4ToPP39pAbXSgUcPUQmg18LY23Hmo10mHA/R0Sj4YUkwnjlVnLNL38JsNos1N6nt7pL0+zXv3aBwcEG1uEp+fY/KMfDwmn4wxeV72IVylZtFOjUKDJxRevUZtd4fGwQFBp006HFAkMcVkgsmyNcDt7eNq4rfOq3fFPaha/cwVEW4roX/8TXctRfufb8hOXAnO64POouy0FYfRirI4goKCpAxCPppQWBeIfA8CH+P7aN9zTXNJaYdgyuRYOoqCNSvvxb2aME4lIH37jonnw2RC8eYt6YzRtrdHsLuNt72Ft7mBkAIvqhE2mtTCOkYFFEYhjUCZAgZj9DjBjqeY8RQ9GFEcnzi/oe1NZMu5j4ooAs9zPSBZMte0xmYFNnesOZskLgi9eUP25g15f4JtNWF3G1OvOxWEEp5Clsw2LMY4ewhTkjIW/kWmdCHVpORYbfCKDC1BNmoEGx3CzU2CjS7BxgbRzg61g31qe7vzs1ZMpuQ9FxRUGGELDVLgNxrUD/bpfvUlUnm0nz6mef8+QbdD0u9hity5quaJo2aLACk9kM7iopSCAWvx6nVquzvU7xyggoB0PKKIY/LJBJNmS+neB8vtf9oAVGWGLq9LiUNqrt2rPh/YcRuEfsb+eYUd5wkxUx1b/mW7EJes0rHntNIqF/86czu7BgT8WMNVey128HE/vwlWaW8AWYjVPs9sZKe6QKo21hWDt5mjqFXONdQWxBgyILCa0Gpyo52eqee5mZl6w9lABz6k6fx11gGo1pYXzLoAJQqNHQzJXr5GnPco6s9I602inW3yBw+offmU2ldfIC3IdtNN3kgPT0h8K8rrL1CzocuiwAxGFIXGjMbkh29RzQay2UQ2Szp3o46IIkTgKMvWWGxRuGHbOJkP35rhkOLslGJ0iU4F4s4+MgywnudERn0ffA/rSbSGQpvS+tqgLRjpqqKZh5EREm00GQUYTZhnTrNNgBdF1DY3aB3cIXt4SWNvn/aD+9R2digmY9fXGU/IBgPyPMVLU4o0waQZQinqe3ts/evviba3aN67S/P+HbxahDUanWfoPHfafTO7E6VASKwyCA3S4irNTof63i71vV1HRrDW9ZNGDsqDVcPu95kufOKNfuME62MX5+c75mx92WvCkK38lgSUA0BRQsy/XyVQLcmOVdd8yVI1wgosyGv65nmei9sg9HkD0UdH+VWHwvW07N8OvPZh22R7ZZHMsrNlZ1GzZOamK6SCnEUQ0mgSnZNlGVoXCM8v52k28Npt5PlFCWdp57s515Fbxd0lUqhyMYLKDLI3xFwMyIocg8J0u/DoGDEcoyyoqIbHvpvtMW7TnD0dGy1wfjeFxhqNGU8wkwkcn5TqDAEiCl0l1Ki7r2EpymksNndByExjbBxj0gQznaAnQzQJVjYdZTtJHFwXuFkmWg3MdEyWJ2RZRmGdbJAREo0ooTgzJ6SZcl4Ki2MUjsfkoxEmzwkbTboPHiK1pb63R/vBA4J2i3w8Ih9PSHs90uGQjAxvOiEdDIjPzsiGQ4JWi/bTJ9RKT6NwYwOdJmSDAdlwSBFPMUVWIq0+xhSuAtJFqRAOUXeD+u4u9f09ws1NsuEQk+VkwxHZeOR6QuX1myvt2X/OUsi+JwBXutQIJArwkPhIfCFRpefZFUfV20rotwLJcZV58qu/mT8cVMQacG11WsGug9rswrTNVCA3vcbm2qxURrbsYeS4UVWFJS9y8iRG5zl+q0WwtUV0sE/85g3x4SHW5qU+nEEJhVQ+AmcJLu2sKyQclVgIlC1ZWbkBmwFT11fpxxTPIQ9Dsk7bDcc26ti8cBVwmYc6ZW9nfTCrajC6tA/Psbpw7L5UIsYekgDCEBmFiMBHSumwqEK7QJSk2CLFkKNJyicYoxDTGDmNEVmGCQNMs4bttiimQ9LpiMRm5MKURoUSIwXGwKyKF1JgjZjPZek0I+31mR6fUN/bRwUB3adPqW1uEW1v0bp3F4QkH4+JLy6Iz85IyBw0Gk+Iz84YvXhJ8+5dOk+eEG1tUtvfdcrXUhGfnDA9PmZ6ekY6GMz98CQ5pnAm5qY8h2GrTfPuXVr371Pf38NvNsiGA4rJhKw/IBsN0WUQE0LNg4/9J5MztWsg7asEAlGan7vvPZwNiIfCExJvbowurizwWz+hX8c1XsuOW5cl2DmvftkeV1TZJ2J1/uW628pek198WLdtGSK8Sc4iKvCavUHlI9YEowXvpgqvGdxg23KlU/63NUtBZhGcWOkJzYbk5HyTmpenaUYxGaPjmHCjS7izTf3BfZLDQ6bPX5CUlY+U4CkPz/PdgjTWsa9KKGK+SGdNWQtoWb527vLGTGNHE3Svh77sYUZjx4gsCqzRc/Vnx0Qzbv6nMpMjSiCEeciyGApEal2DXiqsdJAexrpjWF0RGFIYFAXaBaIsxQ4HiFEL4XWwjRp2u4tNRmTDHslIUyjr4C5bzh1hS0sL6WBiaZGz+aM8J7m4ZPT6LfW9fRoHB7SfPAYpCNttVBQxOTkhHQxILi9Jh0MKZraCmmw4YPT6Dc0Xr4g2Nwg3HhFubiB9n3w4IrnsEZ+ek/b7FHGMYU4VwU1uuTMT1Jq07z1g45tv6Dx9SrS9jVCKYjIlveyRXvbIx2NMkSFxag3CujtFWFFCxPYjtm9ugFP/ehAH3tNTXe73zKA26YKNEHhWunteKKRwd78ScsV4k8W+JcTSil4V9Cnv+Ft23N8pEH1yJVSla/+aJ4U+Hl676qey0GWjElwq0Jq1ztq6AruZCixnrzm+qHjFzIR4JECSogdD9GgEHBBsbdJ4+ID03TviZz+RvXqDjDM8A4E1eMKg9CwASfx5EBIVN9xyg7bSVQpohAhR7bZT2K7VQAhHBsgLzHiMyTKsLUNkKbopSkaakLIMQP4121sZZE1Jo6700qxQWOFhBGgToI3nZpyER6ZzitEAOWgSdhoEmxsEjZDUZHB5Tn5qyYpsDr/I2aDqTC9OUJIZHOFdT2OS8wvGbw5p3rtLtL1N/WCPcKOLV6uhk4TJ8bEbNE3T0g6ivBbCxxrrKpXBAB0nCKXw6nXnL2TBZDm20G4AtVajmGQlkKbKEKvxgwatew/Y/N23bP3rv9B+8pig1UKnKWm/T3p5STYYUKRxCdD6jlUnfruVkL3R2l0mUs8CkOfAYXwUPgLPOWshy3/beQ0kKp2lX/9edQvHfUwAuqFy9q8l+KxTK1heCOsrFrsCm800DWYqBVXlgmIFbquy4RavsiwyIrBIWwUcykxs6lhjxWUPm2Z4mxvU7t6h9eQJ2eMXmFeHpEcFMp3i2wyVadf3QeHh42FLMGJRDwoMaFm+68ItzyDA297Cf/QA/9ED5MaGU2g4v6Q4O8eMp2UlpJbzwpKYstr0ncvTlPRpawzGWvcU5TkUOAhNuq/aSrT1yK0is5Y4j8kGl3jNCHV/n+bONn4UYCQM3r7Flpu/05aIUcViqUkrsUZiyOazHhQTkstL4tNTkvMLijhG+gFerQaAKo3khFKoMCJoNglNDtYShXVnDd5u4zcaeLXIuayWlb8MArx6nWhzi+bdu6SXPeR5iIkThJEIT6AaEfXtPTZ/9y07//E/2PrXf6F57x4qCEguLknOL0guL8nHY3Rp8FfiiiBMhXr/24Xd7NKaECuQm6hUP4unhyzhNwdhy8pfWq6iNUvkqU8PRNb3fXsbhP7em7m1S5lE9aLaSnNodoFtJfuYb/praBB2Caq116ZIttqHWocZr9TrYh152r5/AdiVQKKtqQSbZWhNr1RFqzRsswIjCFbF+Ss/tQu9a8c+sygrYDJFn16gj08xgyHiYJ9ge4vm0yfob4+Q785ICos5OkHqFEFRLlYPWWlpz3XNzOwTFGUQytwybtTw7t0h/P03hL//Btlqoc975G+PyA+P0MPRDExyS72c+HciW8w/AxXXVFv+P2PdaxbWVYezr9pax2ozAl0y22yp8pAazSSNSUxCUAto5BlBu0VzbwdhDIOfXjBot0gH6RwygaK8vVR5Hs08SZhfW61dfybPsbnr1VSviBdGhJ0OjYN92g8fIs/rWGupd7q0nzym++UXtB49JNrZRoXh/G9VFFDb2ab95DHZZIyQkuD1a9KLHhQGVYuo7W7TefqU7X/7Vza+/Yb2o0eE3S5FPCU5O2NyfEzSuySPp0sQ0Gy2xVp75f4V1+iq/by6g1/0OO8HBG3Fg3fGdJPziscTciUIgRRyDjeLijr5TFV+vhdUWwlUbB0qNdY68u41UJy9dVb91VRC4mol9AvNCYmPwJDFtbPUdm1DrFrtaBaQWjFnspmlAGTWVErvf9/iGkiwkp1Zt+wcoGCdNts0RZxdYt6doM8u4EmGt7FD7cF97Le/Q533SQrIZYQ5u8DGztVUCOcC6gQ0y0A/W3Cm1GkTHoIQ2WriP3lI+K+/I/qPfyN48ggzjSlev6U4fEfx9gg7HJcwSFAB9yowZYXPvuqOY0SpCy6gsJCvnFtjbSkrV/bcpCQ3mswkpHmCHQzIR2MoCvyoRmNvj40nT0iOT1A/emSjEabQ2Bk8WPaDhBAoLYDMvfNWh/r+HrXdHcJuxwURa50Ct3QVXdBp07x3j81vvkEgaBwdY60h2tqi/fgRG998TferL6nt7iJ938GSQiCDgNruDt2iQAY+YbtNfWeX6fEpJs/xGw0ad++w8buv2fz9t7QePSBotzFpVvap3jB+c0ja62OK4jcNFdn3ru0FzboKufkl280rA9IqxFxN8VbBb1HJnFfnhW6+o9zCcb/iQHTVT+iT61s+pE0t3tvbsZVbcabAdtUuzCxVKqZS/VSfxQp9Ws9deZbhumpoEx9cWMu4tpj/zFUs0soFei3Aw8ePM9RZD/vuGHN8ih1PkPfuIPf3qH35FDmOCYRH2u5SHB6iLy4wyRSbGowuXJ9Cl59SSpCeg5yiENloIFtNvN0dwm+/Ivp//wfh775CdTpkL1+jLy4pDt+hj09hOnWgngzK5HJB9jArg7emAlMaMasYcedUmPm5LeyiX6ZXNozqsYrSZmF6dEyjVDTY+v23qCCgdfce47dvmZbUaZ1ns7IIqSTC85CRj99oUtvdpfPkMZvf/s5VM9tbeLXaokoEoq1N2o8fI6Qk2tokOTnDGoO/0aFxcEDrwX0ad+/gNRtQWnMLKVFBQNBuI5XCb9SJNjdoHBwQn56jsxQvqlHb3aX9xWM6Tx4TdNoAxMMR48ND+s+eMXz5irTXxxrrmur/YHRs8YFgc12idhVyYxluE4tKSM0DECXIzFIFs/o6YuVFxKcPqt4Gob/rQ83xjWtLkapEj1j9eYWlIMQ1t5xYf1vaa+9qsfbt2NXKrJqdz/o2tkKHtqZSzbAyp2MrvZ/VjN6uLdXfvwDt3GZbzhXEmGdxCje/MBuqc7M+tjQ/cwtOZBr/YgBvT9CH79Bn55hHD5yp3J1951Rar6MO9shfvSZ/e4S+vMD0R07FYOosH2YeNyLwEfWaE0Xd3cG/e4D34B7Bl48Jv/0atbOFncbo0zPyN4cUb9+he5eQpS4nld5c3dnaBRmjCl3OZ6HE4v+tBidbVkCrnbKqFYh0g4KYLGdydELvu7/hRyEbX39F+8kjGvfv0f36a/o/PKP3ww9M3h2RTyaO+m010vPxmw0nPbS3R/P+fdqPH9J+9IjGnQOi7W28em35OipFbW8XFYXUD/bJR25wVdUiglbLiZe2mhijKSZTF4RwEkcqClFhSG1vl6Ddor7v/l7nGVI6Be7azvY8AAFMT07pP3vG5XffM3r5mmwwLO8TR2YQQswrTLHWrOBjnX9vUp9c1y1djzRc//r2yr9Xgegqy20WbFSl2lmsDbEML1ffp5hB+nYOBi/1LKv6cBXfrllT24qrAfEG7Dhxa+XwK6qElh1WP+6yiPfAaB8yMb4OBrNl4FkdCK3CarrS11kNQjclcX9I06uqrSeXGqlXm6szqqkEt+ikQhiDLTLEZArnl+jDd+SvDvHv3yN48hDRbOA9fYTotJEHu6j7B6g3bymOT9Dnl5jewBEKSkVmoRSiFiFbTdT2Jv79uwRPHuI/foB3Zx+10cHmBcXRCdnzl+SvDykuLtHTqdsQpYeQAlP2r4yFwrrpl2Xosvy+cg3sDTZNUamvXDbs4WEpjCU5P6f3ww+owMdvNakf7NN8+JDmvXvUdneItrccnDUcopMEU+TObK/TcQHo3j1aDx/QfHCP+p6by0EITJ6T9gelJI9EegrheYQbXaLtrXmwnaselP+dDwbEJ6dz+MwLQ4JO2znNtlsEHScLZI0pGXvuHEhvsR1M3r5j8OOP9L//geGLF8RnZxRxUhLqxZKu4D8aH+F9idoq4eBDkJtcsyesrn17k41m9fkzt75bxYRfT1PoKtZ64+BzdRsXa0U9qgjvSvPfVqAgcbW3Uw02Zm1vh5UAtDwQd11PqhqE5MrikpWvck0QWqqGxDLrRwBKSKTwyq3cgs6x/SHF4TuyH37C29tF1ut4+zuIwMe7s4fotBDdNnJnC+/sHN3rY/tDp0iQ5VAGIaII2ayjNrr49+4QPLqPf/8O+D4YQ/7uhOyHn0i//5Hs8B35aIxz1pEsJmb0fAbKUdLNXCy0qt2ml+DLauv3qsxR9fzZcjhWCg8loDCQT0aM3x3iBQHhZpf6/j7hhlMq2IgiB52dX5CORs5ML00RniJotahtbVPf26Nx94Da7s7SvZUcnzB5+46sP8Bo7aC0nW3q+3tEGxtr712TZcSnp/R/eMbo9Ruy0Qjl+9R2dpyEz4P7NO/eccrksqTDVx75eExyfkH/+2dc/OnPDJ79yPT4mHwyKm3Aw9L2obzTf0WwnLhB8KGCHojKGq+uhVXITZVBaBlyq/Z87LXjDR+d9Yqf9fH/IemJ/5hBSHClc3+FfbaqgLEkWfsefLaSKs1njNbRpss/XaVMm7WqAyx02CrBya5Aa1dZbB+WlrfXBqLFgpJVZYL5ApOVxbSAElRlcV3pE82YUJXFYjDYyZT87THZd8+QrTYyihChj9roOgS1UYc7+8h6DXOw65xXpwk2zZw6gTGu9xE4uR1Rr+FtbuBvb7oABJjzS9K/PSP+rz8Rf/cDyfEJWRI7XyIMxmqMtRSlcbiZ694tdPGuwpeWVfWz66AiW/mXk92RJXJi0EVGOhwwOjzE/3MTv94ECxtffUVtf5fuV19inz4p7RcmFGnqILCSOu03m4TdztIrp+cXDJ494+LP3zF+8wadJASdDu3HD9n43TdsfP01YXl+5++xMEyPTxn8+JyzP/yBy79+x/T0DCkVjf09Ok+fsvn736GTlNaDewSdzlIQKqZTxofv6P/wjIs//omz//ojw1cOhjM2X6wDMWMYfmhou1qdfy7xRHttZf++o183yVQlGbjB0uX+zjLLbbZOKuBjhfCynOTaa+qslQpphTglKpvX+4ZKxFXk8LYn9GuuhBArJnc3LJPtNU3L1dmcasWiVwRA1+mw2TWZ04ebox/CuasVDksVzNIim0NuciWj40aVny1neWwpyGMRUGinMv3TS0QUIcPA2Wh//cU8EMlahAz8RQSfuXOaSgYxczIVwgmCSrdki/MLkr98x/j//heTP/6Z+MVL0n6f3GgKZEkmKNBGUFjtREKvdHSunlv53vW7/ueLTXeZ9WSKgrTXZ/DTTwgh0UlKMZmy9W//QuvhAweHlaZw1hikUgtx3QqsBpBcXNL72/ec/9cfOfvPPzB4/pxiMiHodJi8e0c2HFFMpjTv35vPEVljyUdjRq9ecfGnP3Pxxz9x8Ze/Mjk6Bgv1nR3is3PnJTQcEZ+dEm1tubmiKMIYQ3J+weD5Cy7/9Gcu/vwX+s+ekZxflJYPspL1/7oT7vd1pcSVJK06zyPLodLrWG4f2hvEJ1Von6kgWqqEbueEPm/N89Fl5iJju74PUjXBs2uyp7l0jV21OWCJNbWsTlBhXq2phqrQ2rpGqF0bCKqV2DJeLdYGH7lS/awfpFtAbFVLhfX39ZXNfHZOhHIyvgbMZEp+dOwCSekaqkdjggf38bY3kd0OVHoO+P4HF1oxmZCenBL/9JzJ//kD4//8L+Jnz0hPjyniiSMaKEFhrAPiLAvrhPf29q7bHD5l6bvtzFpnwx2fnYF1/RydJBTJlGwwoH73DmG3gxfVUFGECoJSNsFSJCnFcEQ+HJIOh0wO33L5179x/oc/cPnXvzE+fEMRx/iNZukFNCE5P3dMuEbDzb2VQWhyeEjv2TP6PzxjcviWeNJ3igxxgslziiQhubxk+PIl0dYmQbOJiiKsMSSXPUav3zD44RnD5y+Ynp6QjydgLEKWNbKBXyPqY9f0ZViCWsW1CdqCcLBMOpAVOHtRy9krs3s/4+5Zszet2Hv/Ezx+7VYOa9a8sMK+v7FIBWRZnpbnmqK9MtNoV+nQZqlvU4XUqlCcWQlQq4GrCq2ti6oLbQK7NhsSKww2VYHX5EqjVJbdIFkJTvNJ7BKCEKxOZLvQPEv0nRZYJRTNSDsVbTDXGxBOdSDL0b2hUx/IM3QSo88v0F9/RfjVU/wH91DbG+Df7JbLJxPil68Zffcdoz/+hfEf/8z0bz+QHR85vTpTYEvNBSPc8OmsR24/EHw+dcuwa24YAQg7SwkMOktJLs6xuiCfTkguzhg+f0HzwUOad+9Q2911m3+rifR9F4TimOTiksnRMaM3bxi9eEn/xx8ZPn/B5N0x6bDvfICSFFMU5MMh4zeHRFubLgh5Els4e4f0/IL47Jzp+TnFZDq/xqZIiU/P0KkLlIMffyTodPDrdTdTpDXZaExyceFUGy56FHFp1Fd1iRWsHUz9ELhmr115Nz/v4kYdn+WfVAdLvUrF44mrSVn1uTz0/h6W3zXQyjpB0tXjLKu6LNoFy3vZwqvJitsg9I9XCXH9fMCiKjGVQCEqsjemnBcxa2wNrqNKf3gyXFzb3Vmue1ZNf69qUl3FruWqDtssq3rPsl1n8PDhXaHcUkQJJ0mcUGieY/pDiix3PZ/eEHM5xIzG6P4Ab38H2Wo6ryHfc8+yJ2F1ydQyZl4BjZ/9yOC//sDwz39l+vwF6ckpejp11gPziHjVAEr+Xfw+FwFbCieM6oQ89Xyjz0ZD4pMTBs9f0rhzl9aD+zTv3XP2CN32XFqniBPis3PGh4cMnr9g9OoVk3dHJJeXFNMYY5xygskd5JcPh4zfHeHVa6haDeEpbKEp4hg9jTFphtEagcQnmr9jnaZlgLlgHEXu78MQISVWa3SazbXpTF4wH66dJRtLDdi//45ob7xhiLU9nxnDzV/DcFvHKLVwDXD+S2x0lX9uK6HfUFuoku/PKiEXPEzJmnJ2xov5EFGBzcx8MNSsaK1VK5+rrLj3Z+BLAWLt1xlksAq1lU9xlUa9XAGJeQY3q2AWTVTWwGs3W+jimmA6O7dWSLAaWxQwHKMzTZ5k2GlMMRyQvnmD3N1CtFvYWgRRCKEPXgnfFa6CMmnmYKmTEyYvXzP5/gcmL1+Rnp9TxHF5JRaLVKwhl9xk1vxzbinuvUinKmEtxhZok5FPE/LpmLQ3IDm/ZHpyyujNG2pbW/ithtODk9LZOPT7TE9Ombx9x/T0lGwwRBdp2YOT85kcrVMKrbEZiHGpEqEkVhsMWdmpA4GPEh5Cea560RpjnDm70cAkRk76SGb25gsgeSnlEbMgtK7W+HX0elYh6oVMDnhi9kmq/Z4FgrActK5Cectj5vD5Dc2v+s18osblrZ/QfzdgZ1dwmJkek6xMIFvhLKdzYcgoyG2BJp8LfFornA2zNQ7egbWMqtXc2763x3BVn20WLNRKc3Q2HLqKXa/2euaBSlxVOVh3V1aVoa/qe638vnj/Xb0KYs7mwd0mOaNKW2ySkl9ckqcJXJzCD21spwXtJrbZgEYdGwXzasgW2mXh47HzrLm4IDu/JLu4oBiO0ElShv9KqF7ShVut5cQ1enjv20Kun+14X0nuDGIXWnWLK1Tq0iUx8dk5eTwlPjvFq9WQoY9Q7mparV1faDJ1tPNJjDZZRalscfFWP5MhR+hFlbJUaVtKPT5cX2d+p9jK3+srHY7lrbnigHuD7fd6mod9T1q2btXwUT9fYrOJRaXjza1Cqv3R60dyrjOAfB/sN/8/67xb5i+0kixVHJ7FEnFqhS0nVsHFOWwvrmlNOTj7dk7o1xGgxEpPyArQwpJhSHDWyoUtFkFoTli4mfbfOnO5q7Dg1SHSVcaauoY0IFcC0pJ1+QdTIVupfm62WYiP2FxsdYOa/YKUJWOu1GTLUvTFlOLiBO1JTBRgGvX1QUhrdBy7Bv1gSDEeY9LMwXxlpSGtXMlE7dUYtBSM/l7r0M7nh0RZEYmVpWWyjLSXkfZ6159sOwseoqxQlvsHc4jpBjmvECtup8IdU1j1MbkdVRHbv2fFY28Itc8qIK/CavMqLFBPyIphwvoekr2BysMvnkn//EHV20rov+F+/EAEqmrGuYxiFoQKXCDK5rCcWeOcYz+4WV+F05Z7PtWbfwEVMGelqbU06uWhOYFAiuVjr9JD7XuC0DpF8FlOLT6wyFcDqOWqCOiMiWZF2U8TYtFT0zkFCQUJOreYHPTIx3oRRBEE3pyWbY3GpBl6OkUXCRo9h1aEDZYHK225KdvKJ7cfSRwW4mp+XY3YQlwD84llTWO7nPDMacyi0lA2BmtzjC0W58su57di3r8ot1CplhIOIaqyLmL+fmcyRZQ9OlHS2oXFiabaCkg8RwgqgWXNZxazczmTLrLwfvXEzxbKF0F7XTExWw9LSAIVpQ+5NH7g7BPkWjXvK03+dcmdfc8PPuOWv3am8fMhl7dB6HOfSGttddq0skMse3hfTSzEfMJrTp22VVqCuCFQs7owlgOGqi4QcX1lc72CwdUAt6pzZbm6Zy5T2Va+LsEAs58vGutXbCWWBnWrg5qsmYtyVWQBbj6ndHItrC5DvCk9dWZAXYEtEhjn5UyQXGj1aYO1zuNTLinaOTjJClNJLKTbtWeeQGvKPntthbDsLwROrdoKloIKUlROrZ33v0T5WgJTCrrOemNyEaytqbwJW4ZTVYJhZi2sU6XLY4xzemWxMy2JW9oZI7AMNGKl/ykca2/mmTSDCu1S4F4DJq9le31qc95ek8BZ1im4rXu16nqQc6iN+WyPWoLdZjC1nK+pdUHOLmm0VawU3vv2Z9RAFvDkz+BoV7etVSbSHPEQYu05uYGVw2xO6NbK4ddQCVWx1+oBDVfpvIsZovffOous9Sq1s0qZXjW3kleMBt4/HLpYOB/B0bEV0tin5EhVmvOK/cEVp9bKs7BXmYROqsiDCgS0VK8YlxIsn2GP9/L5ZlRsqz9p8Vtrlz7X2p1hVkWY9advfu+UBnhzb6KyerNrNnJ3Z3kf10ipVCJLn3fte7aORaffU/TdJGD8wsibvUFvRaxU5PKKSdxCz+2qjM7yCa2uHrtafn5oUXyuIaAbBKSlpPkzVEKln9BtJfTf+qjy7oVYyrzsCtWaK4GIKw1/scy1mwcbuaZ3U9WfWgezvT//u05FofSzec/KmPU/hansTWJe+Cz/tpxtlmI5w7JuczVLTMBZC3s1CLFsLyGqg7pckSRxL+Ro2Ng1WktXdMJZOqvzmaS5LOlsTsfxnpYqkTW5tOvdVEeJ9Vy4R+AhrecqLGuYmTpQ1jCz4xl0abknMVZhESVr0rnASkBYb17B2ZVUZ3X813k1yYp6wkISh7LSWXzeRVycQ7s2QAhVqobn5VVahvmUlWWPSS4gPiGuz0D+DgHIXpPciQoLVFaV3CtBSAHKVqsfsXyVxfX129xbyrBCNVr3vlgCYBfyVWJZFv8zbVarQejW3vsf7mGv7M2iEoBWHU/XUY7X6aipsrGpWGaoqff2fa72iuQaIzuuXZR2DRoibrY/VJI9K8GokqZtQWqQZqUPMgsyc0sJStsDu3Y+SlfVIcTMBcm9RtUjiZKuPId8RDXKLS/+qmZD6W+6MunuL4GrdmUgWM5gsSqUxoqb0vyiVyVkq7YWLh2RliVzh8VxZjTmhe3DTLZoVQdDVgT2F3oZ119viUQYt6UuNOmWjT9M5R3Z+f3p3DqlVfPPtfo7snw/c8FOIRbjPp854oiPWKn2Sg+ykrAJiS9K+4TSRqTa/xGljYawVy2w7ZKAolj0D2evYi3WiLmP1cIg5X2yWWKlzpKV/pz9MFq5BplYUi20V9PgxQiC+Izi2rdB6L+3KlprWzALK3Y+S7AINFdx51VKtVgDHVzHlbNLW8qnYWXv61qJCq5oFBgPdAhF4IKR1KBS9xSFnWfb1bmnqtXBqv3B8oCurfS07docoBpyxRVoY82SsrMxU7+ErZbpwuWofrkFLP8OV0DO6yPzok71WKZzL28ykuBqv6RCW9GVDFyWbrOiQpFeHEut6+atzbTnQXaJei5XPu8qgLzQSVcEpXn4ckotrLg5HPt50sAP3MeLq7WsXl0y2qzzyPWsvFLtyGtEfO3qfWjfX9m5PX6WKMiPC7HVhOoXAeT+HsPWt0HoY6+O/ZQ/WnfzKyHwhcQgkUKVGeIsCKlSLXfR81kMu60w1uwKBAAVX3iWNpOrMu+fllVe2WjLCkdYl8EbAdqzZBFkTUvWhrzhUmaZgzcCbwhyYhGZxWqLFdZprokq1HbV/mAhSbQCKNllFuESbX0VVqiWo0JU2L8aa51jp9eoo9otVBRgrcHECcXMAK8oEJ6H12jgtZvIMHB6aKMJejzBZvl8uHMZ+jLzn0vfR9ZqqFro9NtKUVE9mVKMJ9jcef34rSaqNJUzeYHVBhl4yDBES4FMUsxogslzvMAnaLrhU1vk5JMpRZK6TTMIUIGr5GZ9HaE8lO85gdZSqaBIUkyaY0thV1sOjUrfx6vX8JoNgnbTyf0EASbLSS57JGcXZKMxQknCTotoc4Og1UL4HrYoyMZjsssB2WBIkaaOUCHkZ1iYYm1d96GBUmeDsVBv9+ZIQ7nObAm1mZmZvF0h6VSHO9fgVyVhA2NZ/aeaikqpUKFCBQoRSKQvkZ5wau6iohRuHLvRFBadGUyiManG5Ivqtup++9+8V36Cf8RtEPro6l4IsVY9fh0bbg7LlDfVTG0gRCKER4DvBlptGYSEd4USLVYgOpZ0paqvtWa+WtiVTP9Tb5FKwKugS7MXkpW1mIcw6Vgm25Z4G/I2WE/gZeD1ITizeBcGOdCgDdY6a2tjLUau8UFaUrO7+n5WhfwsVTr3Qkh2jsVX2F0g3AIXudscvYBwZ4f6k8cE25tYXZCenRG/OSQ9PqWYFM4MbneH+uOHBBsbFJMJ05eviF+/Ic+ysufjNhNRqqvaUk5HEKDqNYKdHcLdbYKNDWQYoCdT4sO3TF+/psgzZC0kunuHaG8XEBSjMSbP8JpNwu1NtOfhnZ1TvD7EDEf4nTat+/cINzoU0wnjt++Iz86xxhK0WwTNFtLznZqBUs6+oV4DIcjjCcnFJfH5hbNMSCxWzzxgJSpybqitB/dpPbxP/WAfv94g6/e5/Nv3XPzpLxRpiqpFtB8+ZOObr2ncu4sXheSTKZN3bxk8e87w5Sv0ZYHVxUJ2aV6ULmjdn7pI19tlL9/oVcVqJeQcbnMaiKWKQYmECmOuwLZLd58og+kK020GAzvqup2DlLbS/VNSIVsBXjvAb4X4TR9V91ChQgZyrkaFdYmaKTQ6MeTjjKyXkl0k2FGO0WaFz2rXqSav7R8J1lGyr3afr9vnnD6JO1VyzS/eOqv+SishWeLMCHfLW6HKKsiWVY6bzbjJYrzeJuDDSLv4ORNpYhmCqEJduQdJwzLZtAz3LOMdyFoCq8DLIAwhwhIUBpUYSPTcBtuI98uyXJnEXwbbrnmzdv3vVAYoXebqdMqk7xPu7dL+/bfUH97H5BmT588xWUbeH8B0gvA9gu0tWl99Re3uHbJeD5OmpKen2OGgrCDUYnWXTqszkoGq14n292g8eUzt7l28Rp3s8hJrLcnZKXY8QkYh0cE+zadPwUJy6jTrwu0tGo8eYqMQ++o1k8mUwhhq21t0v3hK484B6aCPtZYiSdBphlevE25uEDRbqCjEbzQINzYImg2M1sTnZwzUS/LplHw8rkwfGNf3CQPqe7tsfPM1G7/7hsadA4SQjN+8YXJ8gqpFqHpEtLlJ+8ljtv/9f9B5+gQVhqT9PmG3g9Xu/Tjx05GrtqpQZsWm+1MguOucc3iPS6ln3XPJs2elal7MWomra0BU4DZrr/gauSREORdgKRBKIEOFqvl4zQC/G+F3Q7xOiNcKUHUPGbkgNEPnXAVkXBCKC9QgRZ7GEEhyL0GPcuy8IrLra5Kb7Fefx1X1thL69baClhlaaoltNUObbQkPyPcaYq3P+D69rBM3CE2rbpCzfr8p6bszyE8ryCKYtmGyaZlswrgliAOnBuYH0GiBzQR2Ct7QwsiAthVCwfIY63UabGL1XEgBnkB4ZcanLTazJQV7NtD54Z6NrIWEezs0v/yC1tdforMUsCRHx0xfvXHXUSm8dovavbs0njzGOz1j+uqVU6SubolLjEgzNxDz6jXC3R0ajx/RePoEv9UiOT4mOTtHRZFTFggCgq0tanfvYouCIonBWvxOh9qdO4hWkyTL8F+8RI3GBO02jTsHtB8/Irm8JD4/Y3J8jE4zbKERUhJ02tT3dmkc7FPb28Ov18nGY6TvMT09c7Cfzleo2Bbpe0Rbm7QfPaT16CFerUZ8csrk+Jjp6Sk6y/BbTRoH+7QePaT96CHNe3fn56qIY7LRmCKJsUYz1ppsOCwrQyq6dDcjKqxbC5b1w6QLvUPmhnHzwVIzg7Rn/EJmqoAOLvMUquYhIw/peQgBpjDoTGPSAqNNWQo4BmHVQUqgUL5C1X2CdkDQDgk7IcFGSNCO8NsBXjtENnxsTWFDhfUk+AKUU1YxxlAU2lnnaoNMC/xxiugGiIZCNTzy45iil6JjjbWmYkj3cZj7YnTt4xygb3tCv7LHUpMclpQSqj+TS/PpC+bY7KawM6zLrke3F1DbNXDENQzqpRRlzbhCdTh09sZsOfRjrvgW2blzqMZSlH2gtC1IujDtCqYNGAvBOJboQhIqgVAWv2XxOwJxKRChgGmlaW8XWadY+aDrfI8sgCchlFCXEEkXkFKDGGnsVC+UDOwKVFGNF+V25TUaBFubRHf3qT28j0lTsstLgq0tvHodqTyElMgwxO90CHe2sbrAq9edPfhKXl61VHe9GOHguM0NojsH1B8+wGu3wVMEW5tOkVp5yCDAazbxNzqYNHNOsdMYVavhdzvIbgf/rIOKIqTv4UUhQadNtL0FUhB2Oig/wOQZ2XBI0GkhlKC+t0Pnqy9o3r+HVB6To2OmJ8dYY9Fxgk5jrDbzkWcLCKUImk1qO9uE7TbpYEDvh+85+V//m/7fvidPU2o72zTv3qFx54Cg2wEhSh26EdZYos0NmvfvUUynpP0B2XA4n2mS+NWLfSPF9XUCn4tqp+LeW2q4eUKiZrBbiTQIwOqF5oYpGYQShZIS2Q4JNmsEmxF+I0RIgUkKskFK1ospJtm8P2MrEsQgkZ5EdUOC7QaNO00ad1s07jSp7TeINmt4TR8ReRRYUmtItCUrNIV160pbS1EYiqxwvUoFnjWoPCDcDFEtD6/ukUiwmauSqiPvAuFmyEx5bt4ztycqg8hL5pKzvUt8QsZ7G4R+RZWQYEk7rmoNt7qZWlZJ0B93xT81d7nevdXO+zNzwoAooTO7eBbCkvmWpAnJhiDtCpKGZCIkw6lkMJYUmaRRE0Rti4kstCQ0BTYSiESUtO2yaLHivTvRnPYsBIQS0fAQbQ/R8RCNcjZmrLFe7r6fcg3UYyukBoX0Q1Sz7p71GqoWIT0Pv93G73RQjQbCcxumkBIR+MgoRIYhwvNuVGkJpZBhgNds4Hfa+BtdvFaLfDBA1evIMEL6PkJ5CE8hgwDhefidNgIINjfwux1Es1G+rnNHFUohAx8VhXi1GiqMnDp2kpIXE/xWA51neI36vBoyhSbp9bDGOPO7yZQiS8tqXc3DvRACGThyggx8iiRhenzM6PVrpqenLjDW60Sbm4TdLtLzyEq/ofHhW1fFlZ/dBU1/fn+JD1Q6193ZqzJGVx1Kl7XclHXPVQ6jlQLwUGWDRAYKr+Hjd0KC7TrhTp1gp47XihBSoKc52fmU5N2I5HhMdhFj0hJmRaF8D68V4G/WCPcaRHda1O+3qd1rEd1pEezWka0AHUgKa4mTnMk0YzzNmRaGJC3Ic0OhDXmuMYXB9wWNlke76dP0fYKujx8KpDbofkp2HJdJsCnxlE+A42DJwuGfycbhHyEI/SzpCbFU6i4vHgGfTbbwKmx1HbR2NZOszrvMIbZq1TMLQqIMOqVY5qztWghLHkHSgXhTkrQFU08wyCQXI8VlT6ELSW6g3TCYuoCGwDYlti6RUwsZCGM/0HG2S2rcIpSIrofaDVG7IWLLR9QVtrCYfoEJFEYIEBk21qBn0FxVlqDMfj0fVaujGnWE5zlX0jh2Qp6+q0q8pmOFzanaQoByAWBmCV7tA63bSIXnIXzfBbDAd/+tnN6a8FyAkmGAUMoRHITAaznoz+7uUr93l+hgDw2l/bissLRccJTeDD6SmDwnT2OykVMBRwhU6Fh5pogxWUYRTymmU+cFVLLBrmz+AoSUqCDAbzaJNjeItrfI+gOQjvGnggAVhiAE6WBA/8ef6P31O/JkghfWAEHa62OybAZ4VcQWrtY5lquqz9WKRwpH/XajDItZOm+u2yZQUqIsCM3cqcvMldAlMvLwmiFBJ8TvRESbNcLNiGAjwt+ooTohohNiI790hy0ILqZ47RAEFJMcO0oRnsTv1Kjvt2jcbVE/aBIdNPH3GqjtOrYdkDUDxr4iTgqmI800yRmPMyYT94wnOdM4J0k0Wa4pco0Cuh2fg7t1wk6dTscnLBQ6LSjqHsKXc2mnn2N7flWSVfy2B4P+gYLQh7XjVqJJVQttVXtp3XW+rpK5qYPhNfS9ZSYbVymstjJvo21FdUAs9NiqunYuGJUU7MqGYbEYT1A0oNgUZBsQ1wUjK7kcK057ivO+chYVAWwWlkwIdCgRTYltS0gscqwR6ZpTXm1Yl7ClBQgUtu0hd0PEwxriXoTcCiBU2NQgWgXSk/MNzmrrApFdaAqZqi2DH+K1W3jNJkiJHk/ILnqu2lASr9XCb7eRYQRGO8UFYxY6b7PqSMqF66e183pXlOq1wvcQZXXjlBEstihc32YWIGYmcdq9X6/VRDUayMCjfnBAtL9HMhyW8F+F6VeysRASlKuQsBZtCoo4dkSFJHXV0cQFprTfJx0MyacTdJatcYUV8wa5NRYZhjT29tj45hvS/gApFcn5hfuU5WcAZ4s+PnzL5V//SjK4QPl1/FoNnWTOrhuBKpe+g4fNUuC+4oslKtRqUYHZVmfoLEi7sBlRSiLMTMB1Bpo5uFGFHt5mjdpBm8aDDo0HHZr320Q7dfx2iKh5GE9SBJJcSoy1+IVBDDKCVoDNNVkvoYhzRCipP96g++0u3a83adxtE2zXsK2ARAoGacHFJOPsOOWsn3DRT+mPUsaTjDQtSFNNlmripCDNCvLCgLE0A8Wd/RpRU3GwF+EJ8AGbG3RcUCSuWqpeMVsN4iv7zeqGJu3KRNBSwryq8nLjcfXbIPRrqoR+bmXzaVHTrqWrLijPyyKgy/prpjKnY69ahZcBayEmCjYA0xCYNuimIPEcDHc5VFz0Fb2xRChBlFomiSCuCxpKIJoS0VWQWMgMNn2PWrJdkAtEIBEdD3YDuBOi74SYvRC6gesPJcZluYVFJgY71Zhx4YLQ0uFnqgSem4VpNly14/nzSsjBZ6GDztptVBSiY9c3sVrPn0YX5fdmueE2/34W7HxEECDD0FU9nlcGL4H0PWQtRNUihO/Nab5CKVQjwmvU8Te6qHoNMZlUzkspJGpM6Q47e08aUxQYnc/htmw0Ju0PUGFE2u8Tn56R9noU0ylGF9ffV9o4p1Qh8JsN6nt7NO/eJT4+pZhMnW7dbL7IGHSakg0GxKdnTC9PUKpG0GoBAp1mSxnZOuHXK3JVtqx+bIXhJhYOv2qNMK/rMc5EVsshY+mhfInXDgm369TutKg/3KDxeIP6oy7RvTayE6I9SWYMWWHQAlASz5MESuB3NZ4E3U/IL2NkIFF1n9bvduj++z7NLzfxtmoUoSK2lotRynEv5vDdiDdvR7w7HnN6HtMbpkwTF3C0NmhtyXKDNgYpBfVIwUaEzkOktqjcICYFZpyRH8dkxzH5ZYpJdAX7+OciFNz2hG4cRcQHI9z7LLaut+FengtaUhJYCTp2HbFgJQgtS+LMlKftfN7I2mUGk5WgPYGtgW2ArkMeCGIjGCSS3lgxGEumqcTzIUlhPBWMaoJ6JJENCDZBJAbGAsYzc7brz5WQAtlUyP0Aezei2A/JWj65lOjcadYpIwg8SVT3UG0P2fQglA6au0L2MEv9Cq/RwGs0ULUaql5zVUgYEGx28dstN9OTJA5/Lzd9k+foNEUnMYbZBuuvraWFV75Oq4Xf7eJvuB6K12qhmg1Uveb6TIFfQnLWWWUXBbYo8Ot1VOgGRWeV0kqV7gZf05QiSSjS2MkgJQnFdOqqn14foRTpZY+knA3SaVox7LueeWOKwpEdBgPSfp98NMbkBV4UzSsvq12lOIfv6g28sInfajp2WZFjytdbGIpUsvMZWDYb0p4RCcxsgNuiMHhWoOajDbPngmk3Y6uZcoV49ZBgMyTablC726Jxr039bpvgoIXaacBGjWkzYGwsvV7MYJSSZZowVHTaIZvdiDDw8JREbtVpPOpCqmncb+M3A+pPNwmfbqC3avSt5bIXczFMOT6dcHQ04t3bIUdHY87PpvSHKeNJQZLreRByxbXF9wTtdkC7GXH3oMGDe032NyOaCOxFSnI8If5pRPJ6TH6eLIJQdQDYfsJ+9XntG26tHP5b6qRVmG1ponRFFoX1VNOqiaGwCwxq1R1hVQ7EVCZRtDBztTBjV7TWyiW5ahNuoLJgF23j2VtYZSGBE6a2NdAt0A3IAkGCYJxLBrF7TjNBYQQSS5bDcAKXoSTywKsJvA0QsUH0NLZnSjYPa2aBykcokVs+6l5IcS8kbQUMkIz6hmxQoJQkUtAGlBJ4NQ/VUJhQggJbiCUr5dksjFAeMozw6g38VstVPp0OwUYXm2f43S6qWXf9GlllL5aZti6rkOrlt2KJyuhovz6qFuG3WwSbGwQbGwgl8TttV4XVImQYugZ+GLi+w2hIPhq7Q8UxaI0uCtC6osi9kBVaVEV6fkWNLgPTZEpWQnlJr0c6GLgqKC9YDDxfHf8UwtlK6DQhvexx+bfvOf/jnxn++CNGG4J2E+Gp8nNbvDCkvr1N5/Fj6skefsMF8GQwxOicbDxGo+cqIaKEJmd22GrGaJMKTwgXgKxFzO0rqrp0K6tHCicOau3cil1FPsF+k+aTLu0vNml+sUXzfptwp4Fp+EwtDDLN+emYt+dT3hyNOO85WvzOdp2nD7s0fAV++RlDRXinjaoF2LxA1X3kZo207nM2yXl1POLl4ZC3J2OOzib0LmP6/YTRMCNJCopCo7GIcjDJCAHSEniKbtPn3kGdr7/s8vuvN/jiXoO9hk89zineTkmfDZh+PyB9PaG4mFVCbjjamoVdxqpUkrUrY99LBkLLw6vzNqN9rwPse51VSxXtWyuHX2sltKr5trYvtNTBuSpFtUQoWAuprYp/sqRMbddo+HIlji650C33CTywdYFuCvKaIJOCpBBMEsE4cQEoN8LZ7ggoChhPBL0AGnVJo2ZpNC2iLaEhsaHA5qvDqHYhnqoEoqmQ2z7yIMDuBExRnPcs5z1NkkMQSLoNiWpIB/mFEhnJcgBQLEQ0q1IqAmQY4LdddeI1Ggipyh4LyDAsoboGKnKMNGRJAvB9R+ve3CDY3iIfDLF5XvaGTMVzzjHdVK2GV6+jarWS+VZaTEiJ8DxUWFZjzWb5PiTFeELy7giTpKANfrOJjcJS/ma57yiUwqtFhN0u0fYWUXsTOzxHeYGrDvKcIklQ0ynFNEYnqZMEsqvSNGtYfcr1qZLLS4bPX9D77m+M37zBb7WoH+y5c1rCh+HmJt2vvkT6jk0nfA+dZXB4SHx6Wr6CmddBSiiEsEhr8exCv02ZhWyVLOOLFALpSYTnYF4nbG6xRdn7swu1BCl9vHZAdNCk+cUWrW+2aX29Tfioi9quk0Uek1xzdj7l8N2I12+HvHwz4M3bIcNxRqPuYx51OehE2J0GKjcoJcBXeNt1/M0a2lpyAUNtOBskvDgc8rdnFzx73uPwaMR5LyFOHexmjUVKgVIC5ZfvXwp8C4Ev6TQD7u3V+epxm99/s8HvHrc5aPmEk5zkIqX/05DR9wOSFyPysxgdOwh1poAuxG0ldBuEbsxiW8wL2TWmv1W1KVMyuEwl4LAEra1UQui5xYFe0l2rHlcsNzDXcOrEOjKFrWiy+WDroBuCPHRVUJxJpokkTiVZITBWoErxgKKAcSzoB9BOYLMDOhTYhpwz5WyGY7FpW6nHBPjlHNCmD1s+tuuRR4rhBE6GhqNTTZJamjWJNB6tUKADAZFC1BQiEK4SWjrPjpqgBHjNBuHuDuHONtL3yS4di8ukKcHWRjm303A0at+f06Kd+sE+ra+/Jj0+RSqP9PSMYjhCZ/FsBNb9fUlu8Op1BAI9nZJd9rC6ILvsYZIU6QeuAut08ZpNhFLoSUwxnmCm5dcsQ4Th/D0IIZwBnTZIzw2Wdr94StrrYdKMwevn2FQ7arQQK/0juzxbe425jpDSzUgJiSkKismEtNcnNQl2KijiKTpNMUWBDEPq3S4ijAi2tkh7PbLxkOnxqQucxpbiu45cEKDwhYeQFmksXqlD6IoEU2bjEoXCb/oErZCwG+G1AmToYSUUqZvfSc9j8mGK1RZZ8wi3GjQed+h8uUXrqy1qjzaQB02SRsB5rhmMUs57Me8Oh7x82efNmwEnJ2OGwxTPk7RqHnVfUvcVNV8SBIrAVxC4CjQ2lnGcczFKeHc64fXbIc9f9Hn+osfh2xFnF1OG4xxtXfDxPYnnL0RurHGrvxYoNjdC7t9t8fXTDr972uGLe03utnxq05zs7YT0hwHT7/vEL10AMnFRSarWYde3vZ5/0iB0M/2LhdWQqEBqJaZvV/sz1QFRlno5q+KeM9+dZQYcrLrKXFd9VUOQXWc9MWurSCAog1Ad0kAQW8kkdUEoyR0M55g2bsFmBVgtGAYwTGCSQ+oLwkjgtRSmY12kmpq5+ON8mqQmEVs+7PqYTR8dKiY59EaWk77hqGfIMvcajaYlswIdOAhO1Fw1JLzFePhSBSgFqlEn2NrEa7XQcUp8dIzJc2r37tD88qlTXpCumpGe595fnrvZnY0ura++pBiOEAhG4nsmSYpJR06rDUdE8Fot/GYLIT2K0Zj49SF5f4jJMtKTE7LTc4QxrofSbLieUF64/kk2Iz4UmDxH5E4JwZEXbAm1OYZbvdul/eQJOsvRWY7FkJxfIIPAfWatMXmBKWZ9JVOxvFils5TFqNboLMNkGRQGISSytDo3RUo+GpENBuTjCdZa/Labq5rNMmXjEflgQNEbIOKUQCiE9fCQhELhC4USAikswizYfmYOvQlkw8ffa1C/26Zxr01tt4HXCEAJsmnO5O2I8bNLpq8H6GmOv1mj8fUWm/+xz+bvd6k96GBbIUNreduPeXU04vB4zNHphKPjEe/ejbi4jClyTbPus7vX4MsnG3z5ZJM7+03a7Qi/UZJfCkOWFPQmCW9OJjx/M+DZyx4vXw84Oh5zeZkwmWRkxiJUWdl5Er8UKC20W+dYCD1JtxFwf6/Bt190+ZdvNnl6r8l2KPH7KeMXfcZ/6zH6rsf0xYDsPMaWFZBELk0crPYHP5rpZNe0GCrcoDnblA87q94GoV95G+lqEJn5sAgMzNlpumIBblZM3FaDkL1BJfwhF9W1Egtr+l9WgPXBRqAjyJRgql0QSlJJUYgZyjQfY9AutjBNBeMYJrEl9gWRJxEtBZsWmZdVUMrc1VNIEHUJWx5s+xRNRWoFg7Gh1zP0Bob+xJLnFqUs09SSWTC+RPgga06l2E0kVqyRq59augpBxzHZ5JLJi5fkgwHJ8TEmTfGaTfL+AJsXpbpATHp+TnJyQri7U8rsbBJsb6FmAaQq46LUvBIpJhPiw7cU4zHC9zFpSt7rk52dU4zGrkrJC4rhCFsYiuGIYjzGFjnFaER6fg6TCenlJXrqZn3y0Zjp6SnTo2O8cng27HacykGn64KDMaViQR+T5aT9Aflk4qR9tK5c/eWpM1MUZKMR05NThJQkvR46z92MU1mpZ6Mx8dk50+Nj4pMDvFodFfgIo7FJQnFxSX50ijm7RMWZG+wsKdYBCs+Ug6Rz0QE3g4UAGXn47ZBwt0HtQZv6wy71+x3CMggJJQnjHG+74WatPIme5kT7Tdr/vkfnf94herpJ0fAZxAVvTyf88PyS73+44MWrPkdnU857Mb1BQl4Yuu2Qewctvnyywb/9fpevn25ysNuk0w6RoQIDaVpwejHl1dsBP7zo8d2Pl3z//JJ3JxOG4wytLUoJ/EDhlaMCsoTLtLEl/AlhqNjshDw8aPLNF12+/XKDL++32K17eP2E6fMBl388Z/jdJemrEfllgslKuaOSVUllJOCz9LZvK6HfUKRZ++OFX06OoUCjbVH6Zc6sza5Sp+2avo5ege0WocVe2VBuxrhbGa9c8X2ojkTNg1AIJoRcCeLUMeHSVKK1C3GqDEJus3Kyu2kOkwRGE8E4EES+RDXB27SI1CCmy/MOeALRcIOpuoThRjn0Rob+0DCZWpLMkmeWiWcYx4aksGgpEL5ERgoRSpfBLmqrBSvLgp5MSY6PnU7bcMz0xSvy/oDs8gKTJvidNsm7Y/J+Hx0nZJeXjJ89QwROUdvkBcnRMfmwZJqZFaaZMS5YDIck795RjMfI0NkrmCxDT2LMdIopXD9JxwlFHCN9p7Cd9QdgNFbnroIJAkbv3pGen5OPx1hrGDx/gQx80l7PiaIORySXlxRx4voxxjI9PgYLXhiRTybEZ+dOOTutDpDKpSRFpxnT4xMu/vpXJifHJBeXTM/O0LoAqbBGk09jpscnDH54hpSS7OKCoFajGI1IXh+Sv3yDeXeK7I8JMo1E4Qk1V7IWxsF0c4aeUHjNgGAzonbQpH6nRe1ui/BOC2+ngdioYZoBuS9RSqG0oVH3Eb4i3IiwSUFtr0n9d1sEj7pM6z4ng4TXb0e8eNnjp58uePGix9u3I877CeM4J001QahoRB7723Ue3+/wxeMN7j/o0m6HjphRaEbDnLOTES9eDXj24pIfnl/y08sBr98NHaNOW0eQCRW+L1Gemx3T2qKNnfeGaqFieyPi0YM23365ye++6PL0XpPdUOH3UpLnA0Z/vWD43SXjFwP0ReIyuZIJJ+QyGeHTe9a/DCRUqmjfBqG/G/q25E1tF6wlayuMI1tWOZbUahJbkNmcgnw+x40VWCvmkjhVO4PlatleUzXba4OLuAq0rP/Nld7AqhqDndlKei4I6cCSS0i06wmluUA7wo6DVypwgQEyDdNEMBxbeqEgbEv8hkBq8GIDfQ2egLx8375A1iW0FbapiD1BLxacjwyDsSXLXTNaG0ucwjg2TFIHyVlfOGJC5C0CUTG7Lk5KFivILnuMvv+B+M0hJo7JznvoaUw+HFAMB8hazTHLTs/REzdTY/9kiU9P8FotEAI9npKdX5Ken2GzDIk3x+xNVlCMxqANxXCICPz5BmK1dhWWdu6sSOko2kEAUmJTN+eDtaSnJ0wODzFKMR2NiM8uyMaT/z97/7kmx5Et6cKvu4dMnaWgAQJkk2y551zBXPBcxfz/5pnzbdHNbrIFNUTJVCHd1/nhEZmRWVkQbIIkeiP5JAsoVKWKCDdftmyZUWYrrLUUF+dc/HmKSWMPjGfnrJ4+p7y6AqAucrLnp+jAeJHCMqNaLLB54Rv5BB1LKd+1tHnO4uuvqYqCIE2oVhnZixfUWY6WBtiLivzZC64s1C9OWUynxHECVYU9v6J89gL1/JxoWYAFrXxultHNc9Xdml5h+iHR3QGjjw4Y/+aIweMpyZ0BapRQGkXmhFWzsEdGGPQier2QcBgzeDD0djfDGH3UYxkavnk6509/PedPfznjq68uefZiwcV5ztWipKgdxmiGQ8N0lPDgzpBH90Y8uDPk1lGf/jiB2MCq4uoi5+tv5/z17+d89sUFn//9gq+/n/HiLGORVYhShIFC6044YFOs1NZ5ABJIQsPRKObJvSG//fSA3//6iMf3BhwGCvNixeLzC2b/dcr8z+fkX86wFzli7SbUpTV83QGgmzKWttes7kSq2hFDCjf5ku+LvWt7zfsouaqq3kc5/BLLIddUQSWOHA9EldRr6XQ7brdtYiKvELSot/OyO8nBov3ZtqVc0+IdfwOoFZROUdaKqlY4p67FmLTnvIifGbpaKPox9BJFbyAkQ0GNvZOCa5RyAIQKlWhUqpFYk6O4yoSLubDIBGs3FVdthaxoKLna+zRLqBuBQoCKvLTZY+2mEqouZ9SrlVcZOYdUfvDUZiuqqwsPBtatwYI8p17OWX39tXc/0NpnI5VND6eqOvMvCikramuxq5WnUPQmD0M6/Zi171dX9ebabBpQRkNgsEpRWOun663FKeUrm+fPmhmjpl9T1X7OqPIZPlUTPreOsGil5bWl6z/draVtUbJ69oLs4sKr/prZKMqaULxXgbGCu5yTL3Ps90/JwpDIBBgBXfryVxc12vp+0lqH1yykrh0mjQKCcUx6b8TgVweMf3vC8LfHRPdHuEFE7uBqXnC5KFjmNRoYDyKSJCSapITHfShGFKWldMKstHz7bMEfvzjj//7HMz77/IxnL5ZkRe2rEuu8sWsaMBnGPLgz5NMPD/jkwwMe3BsxnSSYUENpmV1kfP31FZ99cc4fPz/jz3+94MtvZ1xc5RSlB4g0MRij1gny4oS6Acs1AMUBx9OEJ/eG/ObjA377qwM+uj/kKDHo04z555ec/d/nzP50RvHtnPqygKoJGNTGH781BfcDaLT2xbWuH24Td/+29MHvQejHq3V+8AN0ma0uHVfimvRQt97JbI+hvp6I4GU1z+uyhVvu2soDTwtAYrz/lm5TrNq2SvNvFqhFUVmFtWozdNp5KboT1lKUMFtAEikGAxgPFaO0MTYdaq+YK7XvDwUKFXq5tUSKooZ5LlwthVUDQlqD1r6CLGshL/y9qBWpUUhq0IMA3TN+wS23Z55cUeKKFXQGKEFD5ZC8bjYCm/xNwVv6SxMBQBPKvhsisN5bOgeuTcN0W/Ma7e+rrSTW7STO7nbVNdHnFRqLxqH8519ZVNYNxdgNgpe10/NGVdWNS9y/cRVrsbZEcttEjmgMASEBWoVeOi2gKouuKlj5fmaFTw8O1pH1Bq0aOyEan8JOjIfuh8THfQYfjBl/fMTgk0PSD6eo2wNmkeFsXvDiPOf0dMX8KgcnjAcRUagRrQj7IQxjH30wLzh/seTL7+f8+fNz/uuz5/zpL2d8892cxbL0u3ejsM15GoWGw2nK44djfvPxIZ/86pB790ZE/QhbOS7PV3z95SWffX7Gf31+xp//dsGX3845Pc8oK4tSEAWGwGjCUPvKp/ZGpNZ5R2xESOOAk8OUDz+Y8JtPDvjNR1M+vDvg0Gj0ixXZ5xfM/tP3gJZfzrBXuacqGwGNMi0Ft9sD6mYy7NQvrbbdo6L/tFsQsm7TrP1hPaV9xdH7ntDbRvObklV3f7M76tJaqbfqt26vZ59rwqtfiHrNn1Sv93jSBUzxbgghuGBzV7UQ5AqTbzbyqskhb9KM13cRteP95Q0n2zdbVTATCBcwXMLhBIpUEacGxgEydUgpqEK8PDtQECpc4KXfy8qxyHzFU9nWil6aix/ywrFcWVY9RYrG9AP0NMTMK58FU7r1cq8bJaCSXbn8NuGpdgTsagtGNj+7L/F5My+ktiaVOlrErefZfHb7NiKyswHZfg1qz9zXdsew42m3Y1K6fRY077KZIdJrjzZDSECovLpNK/+oq5BbGgABAABJREFUWnwuTzt8sHEx8MOoWunGrFX7eR7n1j0g04uI7gwY/OqAg9/dYvq7Y+KHY+wo5ryyfPX1jM+/vuLLb+ecn3vbpINxwpMHPjZCB3qzy6kcZ5c5X/zjkn//43P+87NT/v7lJS9OV+Sl9V6ATbXpam+Rk8SGw4OURw/GfPTkgAcPxgwnCTjh8jzjb38/50+fnfJffz7l839c8u2zpRcyVBat8R51uhkOd4JpKIC2AgLoRQG3D1I+ejTmd7854re/OeTx7QFTBfrZgvlnZ1z95ymzz84pvpnhZoV3nkBt3BDWFNyesYqN5HZ9dBtOfC1gkHXXWe8kDMv102AdxSXrPrDqOjJfPze3vhmGobwHoV9IJbRPnNCtdNR6H7xZ6uRnRFxRYEOh6kGdCjbxIGQqkFnzA8He9felzJ7yeV04B7WFvIR5BrOV/5qFXqQQjQwcBh6A5tYHfWl/BTi/xlBUkJfehcE1YKi1WoNhUQjLpWPRd/QDSFODHoeYUQBXGjdnbRvTqteUM3t2gwbVWvBc+6zMG3y4qgNb5rWPxc1nEFv1i15XY3sEJjcORL88wXZdnymDJlh7tgU0sdhdz7ZmiFTtLoq7kC4berGdfwuSkOT2gP7Hh0z+cIvxH24RP5lQDSLOs4ovv5vzX398wb//8QV/+/qKVV4zGcVETw4ItKKXhiSR8TTpVc6L0xVffHHGf/zHM/7vfz7jL3+/4OwiwzrxMulmB+WcoJUiiQyHk5T7d4Y8fjjm4f0RBwcpALPzjK/+ccGfPjvl3//0gr/87YJvny2YLyuq2vnHC/UWBVfXDqcUtfWVEECaBJwc9jwAfXrI7z455MP7Iw4CDU8XzD475/T/PGP22RnFdwvsvETZxgNR+6FW6fadb7zYpJOfte4B+MgK66txkE0TR6nGiV2/Iorkxn32v4Sm7t3rCVlQOFGy1pCpV7VY1Gsi3et28t7mkXcBVKlQjIVy7Kj7QAhhoTYKs8YJQRqb/JaS0Wp7o9ReDKrDFqjmOqoF8goWmbfzmUeKONCogSE4FHTePECgmtwXDzC1g7L2QFTVoIyHA6W9v5gTKEthufJKuVFfESWGcBTAMECirkpu88LWdjy7n7Dq5pl3rHnUnt1kN0BvLxB1G8PXuXrpoPZ2v29bKeItVVobm83r0WqjLNxY+nSdfbrWUU3lJNv2LLoBM0UTkaDAdJRsRhqyUtjKBm7nhhAvuHDWbdF/21dDA6S9kOi4z/jjQyb/z21GvzsheDRmnhieX2R8+d2cv35xzh//+II//umUb5/NMaFXsPWTgMNJwsE0ZZCG1JXl/DzjL3+94P//X8/4jz8+569/v+DF6YqirAlDgw69TNo5L5hJm8d4dG/ERx9MefxwzMlRDwLN1UXG37+85I+fnfKfn53yl79f8M3TBZezAtcMnxrt539aEGopOOfANsPAaRJw66jHRx8e8PvfHvG7jw95cnfAQaDRL5ZrEcLsz+csv57hZuU6aRnjnTnQjYO8Y/cE2RJEIa31VddLxTXD367ZeGlwkT/3gwDC0H9dH7/rUeU3KJgaA0VxaC3vQehtKwraQ59lIsaI0l1vT1lHN28vSl1+n2ti6pc9obzhi3xT4Gqn45XsZAsZqBOhGAqrqSU/FOqhoCJFnGkfaieKqAQdNbSRbcVygtGCNoK2ghO1NeCmVIeEaijqysIqg8s5nIeKaKgxqaI3VZjCI5U4hQSNd5n4Kqqq/O/WDowS78qArBeYogGh+UpYJYp+bIiGAWYQ4JIm/6dp9m8pGt2+NtvGu6/9n7SWsS2N+TrR1N0dqpKNBH63Umgfb5+BnvLn2/rcumlAcSufp/utXa/CTT5P0HSHjNLrCidoaThp8npkXydJdXC6iclwtul02h2dSyu/0Zg0JDzu0/9oyvgPJ0z/7YTw0YSrSPPlswWf/fWCv/z1gq++uuKrr2d8f7ZiVVqOBhEnRz0e3Btx/96Qg4OUINScnWf89asr/t8/Pef//Ndzvvj7BWdXOVaEMPTzOiKeHnNOiELDcBDx4O6Qj59M+fjxlDsnA4LQkM0Kvvzyiv9sKqDP/nbBt0+XXC3KpgLyoNvO/ai2ElQK18iwVVMB3T5M+eiDMb//3TH/9vsTPrw3ZKpAvp0z/+Mpl//+jNmfzymfLZFltaZJ28RTf16q9ecsdELnWqVPc3duc1FvVh3vo7LZABh0lKD7PfSwjx70UEkMRjduGrKm7nazbkW1LI3/SaWU1eCU7drHvweht3rTSont9O3lNQqdX6KDxr7X5PtA4quggZCNhNVIqAagAkWlFUmhcDVQCGEKznizSAOESogCIQgEVW8G8TcrkGyBYLvxyku4msNZAEmoSEeKdAiqDKAUpABJNNLghm0Ebm5n0L+rvisrYZk55pljVWuqREMvQPcDdGy8e0K5s0K/xPlK9lFi8k98+q8Y7VhXkHt/Va65YOz6Cu6DRLWnP6Q6WTxBJ5m0+7220lE3BSWuKzi7GQQ2oHUrRPA7eqkdtvAiD20M8SSl98GE0W+PGf7miPDBiCwxfH+24k9/OeP//r9P+ctfL3h+ljFbVmSVJUlDTo77fPBwzJPHE+7dGdLrhxTLim+fzvnTF2f8x5991fL8dIV1jjjyggFvHeU8XeaEXhpydOBFAp9+dMij+yP6achyUfLNd3M++/Mp//HZCz776znfPFsyX5ZrSi8KNcb4jCEvvxa0Yh3JgECSBBwfpjx5OOY3Hx3wm8cTPjjuMQHk+wVX//6c0//fd1z96bRDwTX9t2YQdQ0KbEfTC7u7i26B0iY5N2arKvLOIdqgdIROe+jpEH13hLo9QE8THwZpHIj1YwPtBu1mik4QnBKxCqxSSt6D0E9UGWmtxbU+NO/wh74XhIxQJ0LZE4qekCeQGU0uClcbYqepQoWMHMYKOgIXeQrLAJEW4lAIA0FraeadOlRcJ9VUd1bEooSrBUQGeilMhopx4rOGpBIkA9c3ECivKO0YVe6wXWsRUFUJq9zTcasKKqORxHgQah0USnVzWfm6R1b9hDwp+9N4N4EUco3mU6K2aF4fj0ADPh5aWnptEw6nm7TSRlfXGIheM9PoUI8imwgQLYZgEPp00lGESb1kvF5V5Gcr6mWJ0YbenQGTTw4Z/+6Y8NGYRWT4/jzjL3+74E9/PuMvX5zz969mzFd+/iYIFONRzO3jPg/uDrl3Z8hoFFPVjm+fzvns8zP++JdT/v7VJafnGVleE4YNZWb0BpbFCwf6acidkwEfPZny0ZMDjo962Nrx7bMFf/rLGf/12Sl/+dsFX3+/4GpRbig4BaYxIdXiLXisddTOrd0QkshwNE158mjMb351yK+fTHk4ihnMCsqzjOVnp5z9n+85+/fnrL6+wtmqqT4bJWMj41duI+GX9a5u45vSLd3bXFkVhT63Ko5RcYRKY1TfVz4mHWGGQ9Skj5pGyDEwqaGXIyYDqZvd3atBSCnlRMQpcD+QzHkPQq9zrf/P//k/5X//7/+9/ofcGKdFO63bVbCZxpQ9uPRKmNodH90VZ8srf2fzp9fxQNjeEouwJQt2oVD3HGXfUSaQa82qMixqQ2UDwlpTiqDTiii0mEDA+EtBOyEyjjR0JJEjyE1zvastNZesF8NNZ7OsvTjBGBgO4KiAaaqIegY1VUgPpGeQuNl5Vh1uXG0LH9reUVX7Smi2EhYl5ChcbKBnUIMA1TOoXKPq3R6Oet39SOevL9Oj/TPsr7qxklE7kLTVaZHtUDi9FpzvVjl6IypY/yxexcb2Y1xTmaxBqWkM2U3tpWJDcOSrnPT+kGiaokNNdVWw+OqK/PsFWilvLPrpIb0Pp+SjkG/OMz774oI//vmUL/5+ydPTjNmqoqgsSRIy6EccH/a4favPyXGf4TDCinD6YsWfPj/j//3jc/7cVE51M4DqWVcvkW4ryDAwJLHh5CDl0b0hHzyYcPvWgCAwfP90zp+/OOc//vSCz/56xtffz7mcF9S1Iwx12xJt2y/rjU8tQt3ExydxwPHEV0C//fTY94Bu9ZlYwf7jitkfX3D5H8+Z/fmU7OsZtRT40eBg57g3QIR05slaAKrXJl7tkQWD6qfoQd/79o2G6NEQczBBH00JDg8IDo4w4wmqF+Kiitpc4uQU585wtmiouE6kilK7Ke+t56XX6SllOxnDWyf1+yiHt1gJ+Wtde02q/GsYLonCg1DqqHpCGSlK0eRZwKKMWJUhRhR1bImGjjR1hGHT17AQKCEJhTRyxIHvDa0vVnV9F99SZ63QIHewyOFq6UUKi9THMoQDhe75GSGihkKzXgJ3TYzeyMClAaFV3nFPcOAi7R21+8aD0MwbUV5fYH+0FuI/fdrtEc7u3dt0VXKbnk1XVr0BnqDzVe/J41Q3uLpfe7vrBrYHIqMNJg2ITvr0nkwYfHJI74Mx0VEPHRqqixwzSQgnKVqE4a+m9D4Yw0HCRV7zxZdX/Md/PucvX5zz7YsVy7xu5nm8Bc7BJOHe7QH37ww5nKYYo7maFXz1zYw/fn7GZ1+c8833c1ZZ1ajWVGOeC2Vl15VMmgQcH6Q8vDfi0f0xt477RJHhal7w96+u+K8/n/LHz8/46ru594FzjQ9coAmM7/l4Gx6HUgrr/DAq+EHUo2nCh/f9vNHvPz3i48dTjmOD+n7B1dMFi79dsPz6inpW+Gqqjhp2wKzj4T0lZq+fjlr5aqcdhFYKFUSoOPY9nskIMxlhxiPMdIo5mGKODzG3jwlOjgmOTjDjEQRg6yvU8hvqeYks56hcbTm9vMZJL/8qCrl3qyektaxv/KtAkO8HuVCwiafkqtCDUFYELFYhszzy8zRS00stA2WJjLfdR0HghEQJqXUkgRBoaZ1F9i6kaw2HbGaMigoWK7iYwziBaKQYNLLtIFI4rVAl+4v/FoS0YB3UtbASWKwcy9xRWt+/UolG9QJ/jw1Um9iIn7XkvhHa5Fpm4q4qXq8F4xtaTXXk0wa9roRMUwkFnRHWPRF2W8+7W2uthRNuM/xqooBomtC7P6T/eEr60YTkgwnR7QFmHKMCjZkmHqQOPAj17w3QhylzEb4/XfHFPy75y1/P+fKbGfOspm4Wea0hDDWTUczdWwPvaD2IqErL6bzgH19d8fevrvj26YKLqxwUpElIGGjAV0FVo1pL4oDhMObBvRG/ejLl4b0Rg17IalXx9XczPvvinD99ccbfv77i/DKjsh64gkYJZ4xaO3OsVXBOcM4RR4E3I70/4jcfH/KHT4/45MmUe8d94tpSBYrAKKI0ID1IiRo3brEOW1lcIdjcUS0rn5zb2Blvhk4DT7MlIaqXotMElSaY3gDV72FGA/RkjBkP0eMhZjzBTMb+e4cTD0qTA3S/h1BDBlrOUUWEynTTtu3ExYvsN1ffnCV+j6mUqP37ovcg9FOt3/JSukzWaqZ9A403xd2pG3fa+ym41yXjuoKrrdgGLbjA+8DVkVAGkFtNVhuWRcC88GabGEU/r+nnltBArL1lTmCExDjSWkgCRxRIx7y0Ixve6i3IhnYWP8awLOBsJvQiRWjAjBVxqDCxRjUScN2O4anN5yDrv/vdb2UFWwuLzPeFstJRisFGBvoBahSgLgMoHZLbtYZZdV+Q7Ps8b3KjUD/46lOv7+q3vSGC1v0OmiTSsCFmNBuFm96qeDY9oX0AJDuUn3QVEl0E7OZLGUMwjuk9njD5wzGDT48I7g2RSUwVB9RxQBBqgsgQ90LSWwOv6u2FLI3m6emKf3x5xT++uuLb50vOZ8Xa0UBrz0AFRjNsVHEnBz2i0HB5mfPltzP+9o9Lvnu2YLYoKCuLaVRe65USDxStqODooMeHH2z6QHXtePZiyWefnzUANOPF+YqycoSBFyB4A17f51Ta2w25ylHVfrGOQs3BOObh3SGffjjlt58e8fGHB9w+7hHHBqoaCTXRcY/hrw6ID1NcXkPlxRrlvKS8KCjOczjLqa4EV1b+KCuNCgwqSdDDAXo6whwdERxO0QcTgunUU2/DAXo0QPdTdC9F9/qoNPFVUhKh4gQVNjNvbfKuczuDr6wrMe/os2MptQEn/yelnDjn1iqjzmndGJi+9457G5WQtVZEqbYa2l+nvqxGfZOBoJ+IaxQNEvi+kA2FWkNZafLakNeGwmpEFEFhuFyG9EJHqGCYONLIq+KUCGnoSEMhDoTQCGUzQCp7+lbth6cbNZVqRAoXc4gDSELoJzDq+36RlsYjTu/MI+2IE6Bxbqj9QOsqdywzR1Zp+lp594RxhBnVuFWN5Hbdkfs5DsrLO0Jq62sXSFpQbG0tQzTRjqLtpaq2a9XOywnB9fdatZZWmCggmMT0Ho0Y/faI8f+4RfRkQt4LuSot2aLEZDWjYcRkEJGMYjRQV45lXnN6lvGPr674218v+O77OVcLn8ETBpow0F5yjAeDYT/i+CBlOvapss+eL/nibxf87ctLzi9y6sa92nQjOzrnmdGKfi/k9kmfx48mPLw3Ik0Czs5XfNb0gf765SWnFxlFZdFKNYOoek3B1VbWfSYnvtpPYj9r9PjhmN/86oDff3rEJx9OuXurTxoHUFmKvCZX4A5T4tgQFjVSeo/CelURXOaYFyv0syWqH2LOQ+qVRZyGsKHbhgPMdII5PiS4c5vg1rH/8+FBAz49DzqtMCEM/cXS9pIEpCyRqkJsgctWSF4gdb09hkDLTsg6O2j/oiFrSk7ecUrunQIhY4yUZSnoG2LA2iz3tl20z5BWbxwAfglA1AUhCXxFVCtF6TRFbaisxok3Jc0qzeUqJDaCd0qpvUhBO4wSkkBIQ98figKhqATr1BY4qz00dwseZeVFCqGBQQIHQy9cENlEzBizASP1EjAXGkPT3DGfW2ZDTS9W9Ach5jBGzWqqixIn1Z6JiLcNOnKtAtnAiuq4aFx3o9MNSRN67KZJlyHAV0KG6w4c+57vZkW6rL/sS+5sQxO1CYgOEnpPJox+c8j498ckTybko4in84qvni6YXRYMkpAP7g+ZjmOCSYIKFLKoWCwrvn2+4q9/v+TLL6846/iwmQ79Jfie0GgQcThNGA0ilquK758t+PzvF/zjmysuZwUCRJFZn09+ZqaZdzY+UG46Trhz0m/6SglFYfn6uxn/+dkL/vj5Gd89W5AXNUo1EdzN6/Dy7o3AwYMS6/7Sk0cT/vDrI37/6RG/+mDC3Vt94ijA1pZsWbFcVuSVRUKDGscg0TpBWEqLPuwRHaWYk5TwdkZ9UVFmBut6uGiISge+z3MwxhxMMEdHmOlkQ7/1U1QcQxg0BrV63V+iqpHaUw1tbpTUtQcfa/cKfTetoZvcGRpphvODqo1Ee+vWCBPeg9A/uzldLBZbV+ByuXTGGNdRJKxH1tvNwVrh6Csl1Wjo1lP57d1vZ1WHBpE19SHyOg3v14mxu07i7q58guC04EzTFwoUNVA67QHIqbUgqrKaWR5gtA+QC0NIIw88ofHAk0aOQeToR46yVpR15+m69I5sG0Y78TEPtfMD3OdLuFjCQQ7DFGKfCed59cBb5htz/d23OC+6CdHLHRdXlrO+pndgiHoB8WGMuaqQpwarG6UT3Ulx+QGnjHrF9/dXHrtUWHd41LQCAqVbAa5P0wRKFAFC1Zw5XRXcyyeGbqq6rr+4LnOq1PZrNYOQ9OGI0R+OGf3hmPjxhGIY8vSq4E9/u+SPf71gMau4d9xnOoz8MGfiS96isry4yPjHt3P++tWMb54tWSx9dkdg/ObMibdiSoxm0AsZDyNG/ZDAKJariu+eL/jHNzO+fbqgKC2mqZ7aDY21rhnshDQNGA8j7p70uXPSZzqKUQJnFxlf/OOSP31xxj++vmKx8q8hDL1M3UnriedvrSFpa3p6ME744P6I331yxL/99oRPPzzg5DAlNIrVsmQ2L5jPSxbzgjyvcM5tznmBQCvCNCDpR4SHMdG9Hr1FjV1qyqJPKcfU0RHSnxKMBphRH91PUIkHHRWFXoYdBn53plSj2mvc0RGw1su7Ud67D39xKN0MbXdfkOqm0qyVo50dWmdDKcqJOCfOOePcNXVcmqbv1XFvk44zRu2tVGWnqXe9KaN+UZXQup42TSVkmqFVp6jFR3QLPnbZc+uKvPZAFAVCEnkKLgkEHUOghV7kGKWWeaEpaoUT7eMdWlpkD4XWFSnUAlnpK6LzOUz7wiAWdF9htAejOIIogMCsc762mnO6TdJT3kfuamY572vGfcN4bOhNQvQ0xPYDCDRS2h9hqPgmYf3L3dB1529dRVvYOFB3wail2tzagW7t3bDuE23TwvKS/tPrbFw2q9JGtqvRaUB80mPwoa+Cko+mVMOIF/OSz/92yb//+3P+9LdLbO3FK9lqgmqzb0rL4qrg2bMlX3835+unC84uc8rGi23dEBchMNrHLIxiBr0QrRRZVnN+mfPsdMXz8xVXixKjFVHkQ+ToZPdYB3Hkq6gHd4Y8fjDm5CAFEV6crfj7V5f89ctLvv5+zvlVjlZePecpOB+9YK1P991U7oo4arOHBjx+MObR/REnRz2i2LDMKpaLktPTFReXOYtVxTKryAsvtmjXdO97FzAexxxOEibjmDQMiZxC6oTSHpHzkDK8i+sdYgY9TBqgTJMl1IijBbwzfG23z7iOe4bg/PCqYTMprq/3Z99gt+KXDqVuckp4Pyf0I9+ui5iciNqYy27THO+AbK7bkHYaXCBIwBqExClPwXXcsLVqxuREUbRAtPDVT2TAaEscCL3YMelZVqUmqzwQlZXCOi8aMGq/KqzrQlJb7yl3eiUMYhgkQhJ5EEoif49CD0ItL+o6DgN+g+eX3LIU5gvLxVxzmAvlVEEvwDQ2Piox/gmd7OGqtl/pq0IyXpYApXb6OttO000SbUfVFiq9NUBqOpWORWGRNnBi/bz7PtedeKjXhyHV+APajQ+dUhrTCwlPevQej+l/NCV9OMKOIs7yms+/mfMfn53x2V/O+fb5ikEvxFnnKUUBakeZ1VzNCl6crXh2uuLFecZsWfo4hND4YePaoUSRRAEH44SjaUqaBORFzYuzFd8/X3J6kbFYeSNRHRo/SNo4eNTWP0ZZOaJGWffk4ZiPPpgwHcfMFwXPmvmir771iahV7YhDnwcUGI1tHBZsMxIAEASaOAyYjmPungz48OGI+7cH9NOQ+aJkuapYZRXn5xlnZytms5KstORlTVZYitriHW68cerBJObu7T46UvSGCbqviMIQVIpijNPHSHAbGx+gkwgT+v6OjyOpPdXWUmrraR25vitT0liU7NTg4jp3ef2FQ3kHHxFkbV//Psrh7azVWZZdu2YDpUSJYFvzyI0flxK2009pjpEveuRGPYLaWaYUbza68nrjjnto/waAXOgte1zQVEaq403WibBXjTrNiWJVas6XgZfQGn+f9h1x4Bj3HFllWZWGZaFZFQ29sp4ulb2SLF8gejHDqhBOr3zlM+wJo57QT72zQhor0qYass011LoztM4A7XxSVQvzleNq4ZhlwsqCjTVhP0CNQn8vLaqwWwDUNTe9rijcT4dedzRoM0rpgA6bQdFG1Waa722LCjpqNqXX/aCbRDBbqrbOOdQNhlCvUxV1zE7XCkTl6X8TmHUfqP/pIfEHY+ww4iKr+HsDQP/x+Tlffr+grIXjw5TJOKbfD/0OvrKsliVnVzmnFzlns5z5siIvrM8FCvzMTV07dKjppyG3DnvcOuoRh4bLWcFyVfPVtzMuLgvqynkXCKWuMQ+uUcQZozmcpHxwf8TDu0PCQPP10wX/9ZczPvvijGenK2orBIFBG71uyrc36zytZbQmjgOm44S7J30+uDfi3u0Bg37IfFHw3fM5Zxc55xeZf53LkqKw1E6oaqEoLXlVY62gjWbYj7lzkhIGiqNJRDUJEBv4eTgToHSMDiKUibyFj3WI1F5k0GamKLzBaXtxbjlodzsGsnZyV63/W1MhiXQcEmTjMXhdpt0GVap1sKunSESqm9fO9+q4t1YJdTLf3pz/+nm1JGqHitsCoGa13ExGbyuMunHdtVMsCoPWEAeNIi6oGaaOXuSY9h2r0nKVaRa5bnJ/9n8EGwXT5ntlDVcrIbqC6UA4HPqhQaUgjWCQKHqx3/UWjaS0FS9otZHt1LWQ4YHIy7WFKlZEsUGNAvQ4RC9qVC3r6O+t0uyG3tvNwL+bO7QRFJi9/mz7B0i3L1+1o5KWjnGLrCOWhX3Bh+qHnSCdU7/lnnVkSE56DD+aMvj4AH3SY24d33y/4s9/PudPfz7nH1/PuZyXDAcRk3HM4UHCYOgXUpfXzC4Lnp9lnDVUVWXdliqrtq6RRwu91EcgHB+khIHm7CLn++dLvv5+znxZorVae8P5PtDmfNVaEYW+n3QwSTg+7DEcRH7A9dsZf/rczwMtsorAaEyi17/bbXFI5/HSJGA0iJiMYoaDCGM0i2XFxazgq+9mfPWdD7pbZTXWeo+69vetlQbM/OOkgUbVEcZadG1RpeAKgzUREiTUJsIpjbgaqXIPrFQN+XrD+q66547arni6BstdxGmroA7wdmdWZX+RLGglyudx7NuDtRLt95XQP3vb/SCNMUJd/9Bg3V+eAqMJr7ORYAMfZrdu7jcbHaV2m/7+wnLOJ6kuC8P5UogCR2ggMDWDGEapIystlyvDPPcihdptekNwPaxJd0zIrYXMeiA6mwmnMyEM/OxIEinGfcU883EQXoXnH6xNr1aKxnVbvCNDMzO0KoSiD0mkYRigJyH6qoKioTm6S7nqXps7rtRsp7N25dMbG5yND9uWe4HSneFSveVq0LVSkhtkD7Ln/mPtUK7lAalNOJrph6S3+gwfj0kejsgGIU/PMr74xxV/+fyCr7+Zc3lVrLN20iRg0AtJkgCUIs9qzs4zvn++5MV5xnJV4cQvzLo5+Na1UdhCkgQcTBIOJgmguLjyAPTdsyWLVeX7M2Fjkoo/1u3uPY4MUePhNh0nxJGhrCxnFxnfPJ3z9fdzXpxlKO3NRo3e9IGkw+1q7auHKDQkkfEVm8BiWVFVlqL0Iouvv5/z/bOlp/Yq5wPurGdMjPK5Q8N+yHQQcXwYc/9Wj4e3ezy6lXA0ikmjFNSQijFODqhlgpW4qUwqn8NEvVaJKKWvA4v6IZZRL1PH3cTSNcESSjU+QvJeov2Wi4b1Lb68dOWw50QZwTlEKRHVlZ3t3NdHctPgXYtRXlF2qX/mhe75AVGdWO7G70x0MxcUNUmq2p+BWoRAeecDox3aKayozbxPQ0U68eq3WWbQKsQoRRhAFPj+0Dh1HAws89yLE1al9vTZ2shU9taZ7e7RiXfZvlgKTy+8Meow9UA0TBWjHlytYI63T1FNedqVbttmqn1VCPOVZb5yLAea1GjMIERPI4KrCreqccXO9dTJcBB23QO2Z3hMS7Wp1pUAAsy6j6MbZZvelVGr7QFSePX4ueLNM6heW4igOq9CtbFJBpNoosOU9E6f9E4fNY2ZlZYvX6z47O9X/O2rOeeXBdZ5f7aglTgH2gsOgEVW8fws47tnS16c+oqBBoRMJ2zONdxqHPu4hX4vZJXVXM1Lnr5Y8eJsRdbErpsm8tpXQm5NwfX7EYeThLu3Bgz7EXlh+f75iq++n/P0xYqreUFe1D6O2wlo2QxrdqtzowiN9hVXoLHWcTUvWCxLqtqyymrmy5L5sqQoLVorwsBgncM2w6Am0IwGsRdH3Bvw+H6Px/d63DnucTRNGQ6HhPEYHU2poyNqfUCtJ4ge+jNLVLNBMJvz73XBZ8tmfXfYvXNvJtnXZrTrOaGOL+ZmKqhpTmHFOae1drtP+16i/ZZuodZSNeu32xfkJDu7ibaXIm979fiBlVAry24qIadaJ34h1EJo3MaCR7azK7pVUl4pLlemEQ4IaQiTviMOhYO+r4isU8hCsSwUtd0IBdXL1sMGRBYZPL/yIFRbSGLvpNCL8YOyao+eoHMcnEDZmJpezS1XA00vVfT7IcFhjJpV1Jcl7qrygWd7Lmy107vrzuF0XafDjpIt3HEuWP93w/t2LyX69lFtP/YJ0cCsk0biC0ordGQwo4T4dp/4dh8zicm14mxe8uV3C/7+tV/Y88IShj7qIIq8WEAr1aTeCous5sVFzrPTjPOrgqJohkKDTSVE09MLjFo/jjGaqnbMlyUXVzmzRYlzsk41pcnxae10jPHDrXdPBtw56RPHhvNL/3v/+OaKswtvcupdGdSea7epc9V2cJ1zQpbXrLKavLAsVyWrvGMx1CjnFAonQtL4zo0GXh7+8eMpnz4e89GjIQ/vDJiOe8RJiugBhRtRyJTKHFGrMVal0Kb6rq81/ZNst7eqoP0KHEGkSclTbt+c0PtK6Ee+NHdWR2nUptueSXK9jL02F3JtWPXnQSG1D4RC8ZVQA0KRdkTGERqHbtymXRsNsBZb+AW+sl6KbVZ+bigJQWlLPxbGPYcVixVFWSuyyuDqV6eDa73RL2QFnM03P3kw8gAQGK+QawdX1Q1gRhM+tsocF1c1533NIAxI+iHJkcPMSuSZodSeu9dNrajEW7R0B0a7NFtXRLD2ZevM9ATrlNImDmHH3md3cFVevT683VvjDODEYqmxOIwLiMOA5CglvT8kOOlRx4ZZVvP8NOP7ZyuenWZczksc0A9DL5cO9Jpm083w6DKrOG9ECVfzkqpxJlC6aTSIrJ0K4kZyrZqo7CyvmS8r5suSVVYTGP9zurFkd3ghQl07VAKDXsjJYY+jgxRjFC/OMr5+Oueb72dcXOWISOOMrV96GapOT6coLVVlqWpHlltWWUVe1v6xgia3qKnqwsAw7nlD03u3Uh7fG/DRBxOe3B9x786Yw+mYKBk2ADTAVQOsHXkAIsU1FZBqzRLVT3M2CK+i4rYoOVH/Ah6av2QQkn0lZWOgLvsotLap2W1Irq0fWx5uPbC62W1vbGNep/W978TZcoK7GUZlh/bR0qmEPB2n8K7YsXEkQU1sAowOUHaTXaMaj7ZuxVdbTVbC+SLwlvcKbo8dw0Q4HDgq61jkjnmuKevtiGvZ6Q3Rht41tF1RgV36Haoxfi4vDmjac6wzXrpJ3Eo2IxG+hwWrzHF2UTNKNcOBYTQI6R2Aviqxw9C7dVebekfL9uCoUbvO1Nv0mqYbj32dZuOlR9eD3o+OPq9MmVB0Y17bQUXb/KfRBMOQ3v0B6cMhHKYsnPD8LOPpsyWn5z50rigtqsnt8Z+7WqtCxQllaVmtKq5mJZezgsWqWofOKb1J4fVu1b6fFBhDXTuWq4qrRcliWZJlNWXVzL20tJSwFWfeNv9Hw4heGpIXNc/PVnz5zYynLxassgrwzti62aWsfdKU6uhSfA+zDcIrK9tU57J2145D01j4+MdwNajQ95nungz4+PGYTx6PeHK/z93bAw7GfXqDMYQHrGRKZceUMqRUPSod44ibcG+1Z4V588MraodqVdtdzWu9ReF6S6HLBHjl6Vo86wBdlu9B6G3dXjH1uz0n1HGEdju7iK1KSP28VVD3jHXGS7Rt2IBQ85I8CFnS0BIHlmAdzaDaTPO9bYTaKea5BmVQuukNhZa4oecOh45F4bCiKaomyqHjIaf2tCekMTctKphnQhR6OXgaQVZ6Cka1goRtVmlTUTWy0qJwXM5qTnuag8OQW0cK+gHBJCIYR5h+iFk4DNKEvfnQt0D5CifYERFsBAddsFEvWShkz8LxM+8jZeMUobRChxrlFFL6+iIaJvTuDxh8MCK5N8D2Ai5WFd8/W/Ls6ZLZrKS2Xpqo1C516SuIqnIURU2W1SyzilVeN/5s2xRYC0JhaOinIUYrsrzm4gouZ4VXnTVx2qojYmkPuO/H+CqqVc05JyxWFacXGc/PVlzOCt+zCA1aa2TfWE03xVSgds7vPnfOT59ZpBo1nVp/bziIuHvS4+PHY/7w6RGfPplw/3af0bCHDnrUasTSHpJVU0oZ4/QAUXEj5qepftxPQcC+WSW0vUdqB1bfRzn8LOu3EvGqR3VdtSS/fEe/1i/OGfFzQoHgw+L9PwZKSAJLL6xJGhC6aS/W9nbaCzavGmpOe6BIImHaF+IQjke+IgI4m2sWhaKumwFu0xnJ2Q0JVb6SKWsPREb5+aHaCnm58ZbbWpw6INTuJKtKmC8d5zPL5cKyqgSbaNQgJJjGRNMEs4CgsAT1dtqoUXoNTLrTB9pUPNtV6T7HgpfRbj/JMtMV/rXThiJrWxcdGcw48pELgSChIp2kjD49YPjRBHPSYxYozs8Knj5f8eI0Y7n0fTRtfE9EdQ6ac1BVjjyv0Yqml1JTVtaDSTO57KneDRXXT0P6aYBS3qLHA1FOXngPqHbxp/m9VnIZBv77vSQgCDZ9pMurgstZsRYPhEE73Kq3mAt1Q5vMdeZu2ive7TijRJGh3/NxDvdu93jyYMivPpjwyZMp9+9MmIyHBNGIwg0o7ZCsHpO5IbX0EZWiVdDMEtqds0T9xGtDS8nuiXJQnc5DO2Tk01Xfx3v/VLdnxkivRrRyIqI6erfdJWZfg+h6t091Ej2lM856c1Lnq7//qpTV9v+tX5wNwJrGOaF51EALcWBJI3/31ZCjdnpL3edoE1p9deSaqsVVmlkGz2etq4Fj1BPGqSc0G19FslJTindDa8GsdXfpzMpuRXcXJVyKN1B1IlTN7N6us/bmd73E1jk/g7KywtXScjG3zBaWVeItX/Q0JrnVg0wTnZUEC4fGy2uV0pvEUbnuZr353KWTAiE3OkP8gE7kWwGi7hMprTCDkOhOn/h2DzOOUP2QdJoweTii/3BMNYzIlxXn85Ln5zlnlzmrrGrmaHQTKtgu0s0mobAsVn5Ic9EMpta1NAOhzfneDPwr7ZV1vcSQxAGIMF+W1LXj8qpYg1Bo1HpmzPs1bkAoTQJ6aYgC5ouSVVZxep4xX5brOaJu850bzGu77EVb4eom38g5T9H5NFVFEBgGvYg7J30+fDjk04/GfPRoxIM7Qw4PRqTpEGumVO6QXMbk0qdSPcREKAmbzewmqH3/ta1uODVu8pBUO4dcdqru7SFs1SxnrWVS6xS+FlipzSNoEXEegGolYmtj7PtK6Ce6GWPa8vOHI/8vYM+wds7W0nz1VVBX+RYaXw2lYUvLOWrnvJWPgOuc0l2q2alNRXSx1M1CoUA5Dvq+KqqtY1VAVqpm17Wh3fZ1xVqrK/A/sypu6qpwLQpoDWANwDkHWeGtfC6vamapIjGKdBIT3xKCpSLKFGZZocT5Dk87kyFyU6sNkLeMHm/nRGhpMJ0YwsOE3gcjeh9NCE96qEFEPIzoHaZEk5jKCvllzuW85HxWMF9WTX+mMR5tzwHZzPtkhWW5qrCNuKCqvAxedWivjWOzB5fWB845WVdC80VJVTkfMhdsIhvWzgjaz/L0eyFpEuCccHGVU9WOy5kHMK0VYWgwWq+rr9c5XLv0XFsB6eY5R4OYe7f7fPRozG9/NeW3Hx/w8N6Y8aiPDvoUdsCynFJxSMmYmhQhaMevmk6z/YWcO/LKYVWPaMor48B11sT3IPRj3zrWE7tuLIK84zkaTR606I3MrzuLbTREgSMNLf2ophfV1E5R1MbTE83gUdeLde1gjTc7XeTgxFcQUQBJ4BikwqQn3J4ItXMYrZhlNGan/vcDs6l8hG2TU9uCVWfmaqOylc731dbrUp0FsqyExdJycVlz3tP0xoZwGJKeQDwXwjOLPrOI9S3i3Qr1h8KN+rmPd5cz7vRhVOSD6eK7ffofjhn8+oDguIeLAkyoMf0IlRjcqiKvHPNlxWxRscxq7/Om/DxNi9FOZO18kBW+D2SdIy+sH+R0L3mZDU2mtcJaR144Fo0nWyuFNh1pdbtrb6XUSRwQhYaqdt5MdFWzWJUehBSN0/YG/NQrPrL2PHLilXfOypo67PejRv3W56MPRnz8wYSPH0/54MEBk+kEEw4o3ZDCDsjcmFLGWHo4FTWSbrupgNYO8+rHPVvkLaLVrkL4p3jW/2Yg1E0HXH+oVinRTaogCpEb+DHBe5zKbhncpYyUuvbtl1utqB++vMlOgqf2d7Rs/OKaKkeUrGm5JLIMk5qsrrBOrw1Mt+TarWC9AxjioKoVC+clunEgREYDjiSC46H31dPKYUVRLPz8UGA2YOY6laNcWxw6qtWOC/cWgbGP6G/imZcrr5Sb9DS9RJMMApIDhbqsUaMAiTWy8r5hojrgo/ZyWj/KeiBvAGDyA1BI6WYGyG2G11RoCMYR0e0eycMRyaMR4b0hjCJq5+eygvZzEyhrxyq3LJoKpaqbfKEtCzMPQkVZk+W1r4ScsMq9ss26TS9F7e85YJ1Q1o2gIa/JS7tWQyrV0Rx2/f5aHzmgKC3zRcnVoiTL/fP719kN+pOX6otVM6xrlD8ZXVPdaa1IeyF3jgd8+GjEJ49HfPLhmEf3x9w6HjMYTHF6TGHHFG5KwYBa9XAqYZ2Hq3als9KhyNTeXs2rCPq9Z0h32OdVEqv23lmx1sdJulS0Fq0Eq7XXa1z/EN9HObztm/Pj1aLU9qf/yiGvLXXcz0jFKdY0nHhdcaN2lR2Fmu8NDeOK0mpqpymtzxnqRih03+o6VAxwjcP8MleczjVG+0vpeCSkkf/qRChqP0w6z9VGZehumB9SXsTQ/czbz3bfKF/39XnZsP9bljvOr2oGPc1gFDAeKUZDAwchMglwA4NkdiP93euk8I5VvmsBiSDWu2IHvYDkpEf6aET0YAhHKWViKJ2QZTWUzpefmmYY1EuVi9J6QLEbSfa6EmrmdYrS+uHOvF4DQ9XInbf2ZaqbirupotrZnLKy1NZtjEVvSNRt00/L5neWma+gfFheG1CnXgvJ2/kgsX71tc1rjkLtE1qP+3z82Kep/vqjCU8eTTg6HBEnQ2oZsyhHZHZCraZY+jgV+uUbB9T7m08/WTnc3d1d3+W9ZqX/L5Gq+osHoWZO6CUfsr4GQFsSbdlPgf2cEm1Rm/gGF3hxgmhpdnxg3OYCBAiNox/XWPHgk1UBeWWonWzzxmq7B9Py3bYZZp1lmx2W1sLxyCvmjobi01MRgpl3R6itajj3jiHpnv7Tlr2K3sixXRMu6Nx2xIPpBAnmhQehJNGMDy23HNQ97QHoMMROAlhaVG43WTiy5wW8a0DU0pY4jDZE45j+gyHJkzHmdp8sMuSLilXtyFY1gROwYCJNbWXd62lByDXzMm0F0oJQ+zOrBoRoQKibqbNLe+3+rlaKsgEu6/ZXLaoRCyjVCiG8eWhZW5/h0/jQrelZtbN52fcRNarXdvjVNfxhGBrvon2rx5MHI3790ZSPn0z54MGU4+MD4nRMzZC8HLGyAzI7RFQfdNw8pqzZ/C63/9NBz66Baeug7bZf03oduzES0feERJxSysn7KIe3WjS8+gB3jqsT2dAdsuOovrbHb4ZVt3Douhn/j3b2dWyDfBXUSLIDqIMmS6gJplkzW645LbWXa/fCGlBUVrOsQvLK94Vqt5HYdimyFph0Q2cIUFSKC+eDDbR2aO3ju5MIbk/AaCHUiu/Fe8VVVmFkY0japQZaZtFbc0nTIFYkoU9bra0iL8XPIjnWdkSeivFmp0UpXFlLeFUznVmuckfW1/T7ARxFuFsxKnNwDqwsSq6HgKsfeYl489/alx+//xxoPzhZ/+fFCPFxSv/xiOiDEdkk5jSvOT3LmC0qqloYxgGRMfSG4VpwUNZ2XW348DmDVmqt7fL9EyEvW2+1Cms9HbcBhe32R3vutUOhbdVSVZuKS5omzpZPudpQbHXTB2r7SVXl1hWw3pVO3ngtq8bUxItmrPNGpFFkmI5TPvxgyq+fTPjkwxEfPhxx5/aYyXhCEB+Suwm5HZO7EbVOgRiImsdtYl12XoNCvZ4A9nXOnht+V6QzxdaJcmj4ikal27S6G/Dx6jjahOjtShoRJTgRscoD0bXt9r4YnPcg9ANur+I1d5pAzQX4kkpoq/nzM1ZCgY9vaIdUrfYEgWpd353/OY0Hh0A7UJbS1kyqirL26qJVaRpnbLUu8roY2qYHt9Y+eQWXq0bOq/1SeDiAXgwnY396V7aJici3wiGvZ2m08xsNUPVixaTvwSgvhYuF27hrI+uKqq2UrBMyC/OV43LWKOV6ml6g6B2E6DsxemVRhUNlXgeuUNff5Ltyu3ZOKkwvIDlJ6T8You70uVSKZ0+X/OPrOReXBcpobh2mTIYRB40yTRp36Kp23iKnM3ekOsoq5/yQ6irvgFAjZIB9BtCtCamn03yP3oNJ61iwf05tM5vk47dd5xyRRkW3WUFf5TEjIlhhbQOkFUSJD7J7/GDI7z4+4PefHPLRB2Nunwzp9YdgxmR2yrKakNsRVg0QFYFu3TPsno3mT14A38BV76uENnlCL/lNp5SyHXXcm8agvQehH3ILw1CUn1T1/epOa7W1ivmlD6u2IOTpOD8jVKMorMZZha4gRAgCIdTOW/BoCHH0Isu0V3kzUjz4LEvjU1MFlJZuBtyGmnMbsUJRKS6X7RCpvx0MPDV3OPTKN6PhfAGLHCq7ifDW+vqVoPBihmGqOJkYeolivhKKSpgtGy6fja3PlsO2+BiI2dxydlbxItUkowA9DkjvxuhVjbqq4aJCnLxbV9Y+EGojurWGUBOMIuKjlPikRz2JWc1Knl7k/P3rOefnOUkaEgWaVSOtDgLvDmBtc3eyVql1q+D22LSeb7NGXr1YVVvVyXUQ8o/b0nyqMSZ1thkolRtWuGb5c02vq63C2xiGfQa8N4GZc27dt9JKkcYBx0c9Prg/4NMPp/zukym/ejzh9q0Jg+EEp0fkte//ZPWAwnn6TWvTVD5uTcP9VPTbmxM+P2i1cspTcvLewPQtbiD2SrTrWozGiSjXRNtuB9B0WPc18bGnuS1ve7+gdi+4lpvDK+LaOG+lqESxqgPyIkBXkGjHQFfEuK13HweOSVo10RA+T6iyCufM9W5m88m5jsy6Da0rKsXFoqH9mvt04Km5WxMf3Z1EwrNLxfnSA9F6UHSHfxDxw4ujnubW1DDqay4WjvnKcTZzqIIt9ULLLSjd8P7Wy7Wfn5YMYk0SauJxQHwSoRY16nmJfWHgynYom1/KLqN7vnUVVJsGXatWa2dblPLhdGoYER6lhEcpehhRB4pZUfPsMueb5ysuLnKGw5iDabK22RH0OvLaNsCgul49bbqtbKqarInzzsKa+aqiaEUCpvVta46HEkTUWtRg7fZ4fvv+WjHZfrWgHxtQXQd7XidLUq39Cl3zGqwT4sRwctTj17865HcfH/Dph2MePRhzeDAi6Y2p9SGFnZLZEYXr43SMViGivPuuehnwqBu4M/WGfNzrPqaSnX9/WXXE9kDU9acRp0Sc9SeWrqr3IPS2brsS7SiKpM7zToL3qyYNfmin6S2DU6vEbGaErGiyOmCWR7hC0zPWVz9BO5PmL6fQOAJTNxerpqwN1nmzxbJuJh46u03Xeb9G07gceMVcVvhgvJa+tOIroSRSTH0SNPMcrjJeKaMNDPQTxcFIczDUBAZeXGjSSLHKN8Ow3UA9P2vifz/PLKdnkEaawdAwmQSMRwHqOIKjCHlaIpnFlM2AUpsnud7+vwOFUJMNo5TGxAHBQUJ0q4eZxtShZllYLuclZ5cFp1cFF1clNYrZqiJr5nsCrdZBnO4lTstrpZvzQoGreUkQaLLC93c2lanauwZ6Z4PXi7ToUmxKbfdYNv2Plx8iEcHZNibeU3hJEnBykPKrD8b826+P+MOnRzy6P2IyGaDDARUTinJKVk8pXB8hQZRuBBqu41Dw81BwP8FC8rKmw3sQeuv1q1r3duUHYMArTse3e0xV5ynWCCuK0hoWZUiRBayUYBuH4pGriEIhMh6EtBa0BiuG2laI+KbvLNPktfZiBbwCTneec22r01Bz1npXhavV9ucx6bE2R94o3rYFaWpH3mt0ky+UKEY9TxNOBppRT7PKhbyJorBuM1tkTEumSqOUE6JIMzkMuXNLqMYGNw7hOMIdR8jKoa8qVO1ulur9IteKbm3ug8CDXkhynBLf7qMmMZmCy0XF+WXB5axktqy4WlSgFZfzksWqoiwdcai3SRx5SfyEakHIAiXaKKpKqGr7UhCCroPCdr/oVW3U3Ryom2yTdmeTXDOThAMTqLUE+6NHY37/8QG/+3jK4wdjDg4mBPGIzI7I6jG5m1DUAyp8BaSUoDsKuM148zsAQGrnuL4aYprWxDuPQe8gCAUB4up10eoHT2QtUFXd4ntHgq86I6lqx+RfOoHOcsOqpm44meWG+AZR26Ovaj2Eo1DOq8/aZcU5yCvDPI9QTlFajRVNbTWjtMYkdp2oGmhhEFlkWDUyaoVIiM2U7xftn83d2rHqBgSKCq5WviKyVsgKn866Knw1pFsar9Nr6Jpkqk48k5+Ib6qioeZwrMlKgZXv/bgGdEzbzG6MUavS4WoIZzVnVzUX85p5z5AkhvAoxt2tvVLOOrhqm7mvkYErN/3lhsiG12Fi5DX4fLWvCnI45TAa7w5xq098p48bRSxqx4vLgrOzjPnc+7vlhYVFyflVwcWsZL6qiAJNXbcMdKtU6/CwO2ygdV4h1zqdWyeN6wHrHKDd8+J6daVed/186Ue1rx5RrVCl6VeKQC8KuHsy5LefHPJvvz7i0ydjb8EzHqLDMQVHXgXnhlTSR3SEFrO1H9k2s30Lg+dv/EnsS9HcmvLdiitaq+NaMYiwu16JEhFxTlTXkK9za0Rd70Ho7d30jSGH27kce6jnnyHJYQvUGmWmsl4Fp6XdvdG4IRiqyqveWhcFR4lSFWAJW8+swDExNYH2ctZu76el5toFZXc212h8UHFjwZOXTQy3VWRl45DdfF8a9dvLQrbaCftWhJBGnpo7mRiWuVDUXqhQW2myhzZWP47GCqgWlrkPvTs9qzhKDHFqGE5DgnsJOnOolUWWFkq2XCLejfLdg4fSinAQkp6kxCcpZRpwkdU8fZFxepqTZfW6wikbi57zq4LzywKjFKu8XgsStu1z9lc0frbIdn5G9lQ1co3Ke9vVgzSzZL4XBWGgiKOA28c9Pn484X/85pg//PqYB3dHDPp9nBqwqsfkMiWzE0rXQ1TYbH5kEz7HT/P633Yl9DK6dd0L96mqTr/jlNy7Pax6YyzuDXYga4M13kq8903WLlv11UaV6WXHDQj5uG7/uiurEQJ07n+nFk3lDEVd0Y8daSykoRCHtkk2VWu/ucDALNNklaayIE55aq5TragGwKWJafHO2/4l1s77zKE2yrh2l9ZKQNZeoh3j08oKVe1ffxL5SuhkYrhauubuKy2lt6mdtiKSxlPualbz9HnJKNLEt2LCQcDwbkKQOfR5iT0vkcqi2z7vL4oVV9vyf6W2XDpEFCrUhIOI9CglPEhYRprz89yD0FlOntX+WDZecFlec3ZZ8P3zFVXlOJ8VlJUlMD7KeisSHHWNB2vnTjqb7k7Z0z1Tr1eHW3Taj/UZN1Jy5zzIIoIJmjjwWwN+9XjMH359xG9+NW0qoAmi2wHUCbkbU7oelgiFwqht9du/Qv9nPQQuL21J+Fau1jZQyvHewPQtFhCvWOu7VNOmjN3R2ne1pZpNxPceek691CVsn3/U9djo7T/LtZNLNZWQdjTpoYJRjlD7OO9Se+oqqw0ui8itYVUFrKqAqa2ZUGO0JTBCZBzj1GIaJ2UPSgGy9OKFepeladM2O5VTu5u2DrIKSrtZgJzbBp9tusPPdNTWVzpZKVQ1DBLFqK85GcPFwnF65TifO4pS1mmfXfJEa4UohXVerv39s9Kr5HoB/UFA/yBEL2N4FuNOS6R0qEI2C+raF0i91Hp4L00kN0zP3yB5kZcsrtt9qg3qK4NPRnIanYSEo4hoGqNHEUXtOF9WPD3LODvPyXL/4YehXsulzy4LvvxuwWzp/eKy3BIYTRIZbKOfdx0DTvXqxs3mCO5DGXU9Z/j1Wc/9Sontj0d1hK2QJgH3bg/4/a+P+bdfH/LJR1Me3J3Q7w+xekwph2QyIZchNT3QAVquJwPzCoL2Jv+BXVr+VT/zwx7zOlDKzkZ1Az4breXeCtcnIVotUotSttTava+E3tLtTUz4umWsk5eoFtQNFO3WD8gbIKTwOlkjawwUhXEQWEVt/Z8NfiYoNj47qDKWSnx/Z+UCCmsoarP2j/NxDgrrHEnoVXTj1GJMhTGb6AWz8plBtqOoagdhpZOo2vWCs24zF3QTj981LPWVm5AVwiITVoUw7ntKbjrUHI8Nz4cehMpmeNV1o1vYjiHIcsfpRUUUaYaTkOPDkOkwID0IUbdjOCuhcHDZ+Kp1o2F/1i1Tt2nWWTyc8vNBzvkMpFhjUoNJA1yoyUvL1bLi4qpktqgoG7ucKNTUym+qZouSb58vuVqU2CapFLwjte8nymajsAcb1U6U+8/5EdXNvJcIBE3+0J2TPh8/mfA/fnPE7399zL07Y3q9AbUMyYsxRUPBVS7xqcFKUMruQMG/jvrtNRY9AeXFrUo5o/X7KIef/uYEUdeGVfdmtr/18kygM5Gw2wfa1FnNV1FoC7oGV2tC6wgEQuWIGhAqjMU6P4RqRTWuCC3QKmqrySvDqrSMUx9Yl0aOUWoxxgsL4kBIQuFyZVgWmqxUVJbNLElX/dYFcbdj+8P+xa0L/LX14DNbCrOVNDNHimFPczQ2nEzc2kFhmcu6f9QONAZNwqe1nqK5spYwrjk4rzib1RylhqRniG7HMKu9i0IpHoy6ARg/tlx7X39EdTcq+wIU3VoJJ7iNrQSCIkCHDh0BgaIG8sqxaKx1ssZWR7WRB82bWeU1Ly5y5ksPPlUT6LbxfJOXnqRbnqE/w1LVqjOt86/dWSEMFIN+xP07Az5+POHffnPEpx9NuXdnzHA0wakRRTFmWU+p3IiSFCdBE8Aoe664dxhyduYXXz1Xtf3je4ZV30c5vNVbXUtjJOabKVsaRdnTiVH/JCDdbNy+72/dZ1fSJoB6GXWAwjQJdnXtoBKK2hE5R6SEWHsgigJHVWmcUms5tHOKrNI4CcgrzaIwLApHVtXUYplS04thmFhCA3Hgh0+TEE7nIKKxTlE3vRR9g0BDdY3utsLEpFMnqi3z4drBqhAuF35AddJ3pLEmDBTTgebOYcA8F8pacM6yyP2wJXozK9R+zLXzvaPZ0nJ6UfHsRck0NoTDgOFhTPBICAqBpUWWNZTNi133+XYaIq9otL9MQLdlDNuA3HZOUiMFFrdJ7JN2Qsu7NctaLqwwsSYYBKhegA0UhRVWuWOZW7LCUpSOygpB4B2nJfBvp6occ1exyuo1CK7VU1vHSZoQk80LV7vAI/Jm6NRdIG9S8+x7TNk4yajGC86pxttRII4D7jYU3P/49TGffDTl/t2xT0FlTOG8Cq6UIZYUaCsgNspT6brVqB98Lb/+RNQPeMzO3F477N1Ywq3vatdFu8PorE1M1eaclNfYSwwGg/fquLe2q1LrIS3XBO/8KB9219hDXjNjXm2xv9d/TjcAZJqclVA0IRrtdANCjqyyRLX4ysX4FNW8suTWULvW9Nvv/pxTZKWhqAxZZcgr11B0NZVTTJ1jkPgK6KDvRQuh8WIFf1dkpc8N2qp8dtnKmxSlXE9NRXwFs8yF87ljeGkZ9TS9xDEZGPqp5vYBZKVQlEJeCnnlRQw4wWyv9KhmILOshaurmqdPC4ahJjQK3Q8Y3fFKOX1R4i4r5KxpTLvdMKMfqebdXU3o+HHIxpVj/XRaoVQAEoCLPNCGBp0GBIcp4d0xejrABgFF6ciziqLJBbJtpHOzaWkdLpwItro+t7P1dtXuDuhn3hI3val2+FUaQ05jFL20VcH5QdSWgkvThoLLRxRMyO2EWhJAo3d773ITYfyOVUA7A0G7c0IvOX6+Atpv2fO+EnqbPaJlUYhttprqR2PdNrYkwrYJy+5p3pUyrKmGBmx2qTfT/pvyzsCBaCIMBu3zWSpLXBjiypFooR9Y8qgiqwOWlUMwXlHV2QHWTlE3/SLX9Idqq8hqTVZZDvqOYSLEAYxSR2AsSQiDRDhbaC5XmnmmWJWKqt6AUBsl0U1qFfVqtwSkzSwSzmaOOLL0E80gVSSRZtBQcpWFVS5cLR2zlVvLv53z0RYt+BnjpelepFDz/TNFvBYphAwmIfpOgjotkPMKVzrvMWe7evTXHGLdtZHeknx3A5tagYtDsLj13TWkm0ITgApRUYiKIwhjJEogTTCDlGCcEJ2kxPcTzP0E1wspq4p8VVLklQehDpaqtZO0fxNrsQ0bTU1LxSn1y1uEdVMxtmar4gRt/CDq3ZMhHz8Z84dPD/n0yYR7t0eMRmNqxhT5iGU1pWJITeIpuFYB1zQ05Z0Fnn0g5LZBaEfhu1+h3UxuiPeP074ntHU7OTl5Xwm9ldv/+T/Ujx6JDkNpevMi+zjx7tip7GQ6rHNNbhoW2nYB6wKRpkOvrYGmqXSaP6+rn52fQwmB0oSiGxCCqgqIC0tcWHqh7wUVccXSBoRFiKpYR3lrkbXE2qu7FaWFeWEorWJRGha5ZZ47DvuWad9XRcPUkUYw6gmjnuZ0Zng+05wtNHPnveekqbi07rxfvdv1UtdaI1r5uAnrhLISZktHGCh6sWWYKp+YGivSRHM00swODGdzw+XS94eqyq0HVlvZuGkMW8XBMrM8FdChpjcOOTiImExD4mmIuZfirmqkEniew7zeOsZ0PNSuEaddukp1EFC6gTf+vJKdkrHxfsBRNX0f5akiHaB7KWbQRw8HqNEYRmP0dEJwMCI86hMdRySHEByWSLSgvLqkWBUUmTcWrbsmoeqa1oFOEMH21zdo9NwIWHuo1zd9TJFNRdg6aYh4YYb3ggu4fdznt58c8v/81suwH9wd0+v1sYwo3CE5UyoGDQXXVEBtpCjbAKTecPOvXm8C+Ud+zM4r3o2z6FZBXcRpAeimURO/W7aI1NZaV+3zjvtf/+udqYjeKRAagMyUEveSfPWt6lZ2YKQb46CuU2ttBSM7J9mm2uEayJhrd73+c1sdeS5X0ChPx6FQoglrTVBqwkoTO0saWfqRolfXxGFNUITUSq+7Cl7K7Yfz2ltZK8rasCo1q0KzLByrQpPXliPnGPeEOBTGgSMMvNmoMa2XnPZO2bW6LkTfI+3u/ht4LzrVNJzLJp31aul4dmHpxYok1sSh5nCsiCPF0dhw7yhg2fjJXc5r8tIPsSrxi5ZpgvEsQlUJs5UluKwYvyg5mZZMI00UatKTGFYWXQmqdJA5pLbrbJbtDYbasz7I9Z6HvGIB0hqlQpQYtET+9QYBOokxoyHhaIgZjzBTDz76YEJwNCY8GhId9QgmhrBXEoQzqrykPneURU1VeFcDJzcv8i1W/sKp8jV4+n6V71tprYiigOODlA8fjvjDp4f8/tNjHt4bM+gPqWVIUYzIxecB1SRND6itgN6kR/OOVUO7ZM6rZh395+BLKKWc9qF274dV39Ztr4v2a1KsslsBvQSE2m8HKHwCiTQVT7fy2QEhtf19s/N90yjh2hfntoDMS7W1U5haE9SK0EGsHYm29KKKXlSRhSHOairZZAYpNjHgDk/L+WFRT8tV1qvpKqcoaseqbNVzHjQmPSEwjl4E4x5crTSLDLLK5w2VdSPl7swGdYdV27tWEAQQBR7ATMW653M2c4SBJQiUN90ERj3NqK95dCv0FKP2O+SispS1D9lToQcirf2QrQClFeYry4uzkm+f5gwjTXAUczyJSB9CUAoyq7FXFW7ufMXoZCPZ3ql414162SRbSjM9LNituY9NSpNBBQEqjtFpgqQxURzh0hjVS5F+n2AyIhgNMaMhetzHjFKCcUQ8CYjGAeFIQSIo41+fzX1kQVX5r9bKVgrttd5jZ9amXXN+aUtPm7LaUnDOefPdJA44Oezz4Qcjfv/xAZ9+OOHB3RHj8RhRE5blmFU9pWRAJQmOoLn2Ni2Pfw0JtnojeHrFQ4naZuy28oSev0OIHbyjR0x2W5Otg3SrGqILRJ0ogXZQVTViAdNUGQZNiMEo7auNnYqnrYZUB4zW31M7Muydd3Dj1K1T3rrHKQJRGITIeBAaJSVVGYLTLCtD7TRWNrO2m2QIWUc7O4Gi0swyD0qLXDhf+mpo0nOMUyGNYdJzDFPvmn25gvO55nwJF0sFORR14+rQgl6HmvPptaCMr6r6iRc8lLVikTuyArJCeH5hm56RonZw7yjgcGQ4nhiU8t9bZo6rlSMrd4sT/+Z8iBtUtXBxVfP19wVxoAkiQ3QSEx3HhIWD8xK5rHyWzaru9OvUZmhqXRQ3ikNx/gpuhqg8ANUNGCk0bdxtgE4idK+HmYzRR1OCgyl6MkZPR4SjATJqAKifotMYnWpMXBGEK6JwThjOCULr57sqi7MFtiqorWs2D6yjGXSD+GrtaCDrIQDZK71WbwVMfshi2Q4++wwiR2WdzwM6SPn1hwf84TdH/P7jKY/ujRkMBmC8Cq5gSsHIU3BKexl22/zYJb/+iRTU11pe3spjvtqeRe3pUHuvSN8I3NEKCXi3E7+puu5f9b+vA9N7EPoht90ohz3s2bVKyMl1F+Cti0T5aiVQXqnW1jFBA0IBBtNUNt0ez7booE2yfBVKbuvt2oqIZqnzMm5fEakmmM4oIQkso6TCVQU4hZWIVeUrnHaxbv3mtN78ua1S8kpR1Ip5DhcrzXApTPuO46HjcCiMe16+HQVCFAqhEYzZ2ProAsrKe7rtNsZamx6tNxEO/cTPMfVyxeVCmK/8PND3Z5a69sOs/ncU06Fh3NfcPjBczAPmK9e8ZufzhTrhdd24h1VmefaiIDSKpB8wGgUMhgHRUYS+n6JmXpygToHMQu38p2Jko7DYYuC6BpIaJSGKcF3i6SRGJTEqSdDDPmY8Ijg6xNw+QU6OMYdT9MGYYDREhiNMv4+JmzkgvUK7c3Q1R5VXiL3A5cW6i6gFqMsmlcKLS1wrflD7Y+Z/ySbhG3eNxgOxMd4MA8N4GPPw7pDffnzAHz494snDCZPJENFDltWIzDWDqNIA0NoH7l81guHVMPSac0K780G7yarvK6Gf6NxX+/pBXaeebbG1B5ZQaeIGZixBuzQQEDQ9HTaUWvPnbhNSbVr116aS5MbW5P5B2haE/GS970hFgWMY16i6wFlN6TSV09TO4KRRJuiOF1zDnTnXmIE6hReLKYrKz+eUte/95LWwKoVh6o1KUZBGfsBUKV/dRCuYZ97Cp7bbwoE10APG+Fjv6UATGMWo8lJzxHKx8HNDPuJ7Y5pZVEIv1vQTzf3jEOsgjhTPL2sWK0dZu8YHz/eHmqKFonRcWCEISgbjgsNJyDBShD1D716KLgXduKFKmSOu9ro167eMsja4dOtWscKgTIROEw82vQTVS1G9HmY0QA/66OEQPR1hJmNfAR0d4iZjzLAP/RSVJthkgE4SggACvcTYDPKMuroiyy+Q/JKQkjgwREGE0gEG11QcaivDjBsF/7/UPpDf2LUO3VL7NxOGHoAe3R3yyZMJn3w45oMHY44OJ6hozKoes6q9FU/lPAXX5AW/PIzuXx+D9q5j12sr5RXCSr1PVv3Zj9za5+p6wbSvfDIoIjRKGUIV4FSIEt2YIba1T0Nz3bggqD0gI5uW4RZdcv08k0570TjvIaesWiuyQiOoqEZbhbWGogEiJ4qy1luUVTvNut45dQJc2yZxVcMsUw1FB2dzYZB62fYggThQxAFM+0JgNqqm2vnfdeJdtzciiw1tEzRANOx5ujAJPRBX1nI2E+aZQ134haq2cLl0HI8Dxn3N4dg0Dg+edivLiqL0MKE7WhKHt3uxVrhY1Dx9UTAdGHqBIjjywXep8k4UZJZ6UeMu62Yxc526tOsXFKCCCDMeY6YTzOEUc3yIOZig279Pxv5+MEGPhuhBH9VLsaG/bGqx1CicjhEdABVGMsL6HJc/YzF7zsXpBUWWERvHqBcxGvjqc2u3u3POqLdkAPG2+kD+LljrZ52iwDAZxXxwf8JvfnXAbz4+4NGDMdPpgCDxcQyFTChkRC2pt+LpwO9rS7DVWzAt/dEfs5uqur/p17UabAUJrp0ZE7nmT6cQP3Yl4pxzdp9E+z0Ivf3bazvz7I6dmjWxpgkwCKZTJ+m9J951M54f73zXO5WQNBRbrBxBXCFNFWQbSFyqgMpu/OPakLr21rpkmzU0+1PYq+gU8wyCha9+RqlwMBAmfejFTR6Q8S7aYeDnhmh6TVp2nK+b79vm+XuxIo01g0QwulXMCZdLYZE5vj+vyUrH+dxwdeh4cBJyNGmtfQIuF47Z0lHVQmU3c1EbY2pPWeWl4/yi5JtIkRhFGGrC45jwJCYpHcxK1NyH36mV9mo5pVC6iZY1ChUGDc02JDg6Ijg5Jrh9QnD3FubkCHN4sAVCejJC91JUFCIKqrLCrFao5QKKCqkqpLKIztFcEpTPKedPyc5f8PzpFZfzmiDQHIzgpDKMBoratnSqN3BV7+imv51d6vZik9hw6yjl4ycTfvfJIb96MuXoaISJBhRuyMpNOnEMAUoJBvvjA8q7uri9zrCqd4B1Sil5D0I//dZLlFIO56zS2qn96XM3lkMbek116JANuyev4OBf92i/ql+0BiGn1vdWTdf6uoWBQyUVrhFUGC0YLSyKkLw21Fb55mQj225zemRHge5cG+u9CbxrlXC19V/7CYRGqG0TaFdvD05u0S8NNVfV3q4nK/33x33NdAhx6KuorHSUte8PXS4cWeFY5t73zCvfoBdrnIN+qjkcBSCwyK23+LEb4qx9TuuE2aLmOyAwmrgXMhiE9EYB0XGEeZSiMos2AfpSUJVGmRDdS9GjFDVM0YMBejjAjMaYwwOCwwNfBZ0c+apnPMKMR56O6/dRZltdp+MYZS3kGVAhziK2Bp2hWWCqOZLPyOZzXpwu+PbUUjvFeBByOa+5fdyjlxic81HWxlwHoV+6I+WWFBsPQEop4tAwHSc8uNPn48dDfvV4zL27U/qDCbWaUNRjMjf0cQwSeYEQttN4/O8NQD/kqOv3BqY/4e1//k/011878RbmNSJ1M0a9dhXYihuQPdYYrQdTZyZx3ahuzJ1elpX2ykH8Pfbya5WP2hk4FIV2YJxqIh063lJNVZMElklSoQ0ERtaNeslpTE67Dg7bgWVr67TGscHoDbAgHnyuMkVphWgFppnQLy1NCJ3aigZv35NRbXieMFt5ii0rDMZ4IAoDL/de5n6IVcSxKhx5BSpzhFcWEbhaOgapIY78tmAyNL5603C1rCks6wOltX8PToSscLywNSooSAYh42FAL1CYSJHeicEqdNrDnAXoPEFHA8KjKebOIebWAcHREXoy8XLq4QDd66H6PcxggOolqCRCRzEqCl9eYavWstyBq4ESrUqUKrGuJisqzmclXz8ruFxY4kjz4iLn8bLm9mGyDgNUbNSaim4vQG6O737FefimY5lv8pitKk/ENVESHkTTJGA6Snl4d8iHj8Y8eTji3p0R48kEpw8piylZPaGkhxCgtL4W6/GDGylvoznzoz3eDdb9ss8Al22XjpvLpebSVu+8h+s76R0n4JSIVWotMntpSbs1C6Y2d6e2Y71fVVS9lfcjnYHw5hVUTUKBWAgVJKFFBzT9moZV0mC0kFfemLR9z44dP8nmkX2VtC2bEPHpqUWl9i9EOy446/yh5gfLyoNQOhfOF47buY9xGPU0dw89kAFEoeX0yrLMvRJutnJkhfDswjJINQejgMnAEBhP6cW5w+iNY8Fataf9TFRphVXpOJvV9J8VTHoBqQJ9YDgcBqSPAszAEJ6n6GKM6R0T371H+Oguwb3bBCfH6HFDscVRAyQaZczaOnwrD6Zj3urE4aoaKWvEejsaxIKUKHKgwLmayjqy0jFb1jy/KHh6VgLCfFlR1kJe1CSxYb70lj2qAdqu+8AvvRIC76AuIkShD6a7d3vARx9M+PDRhLu3hgyHA1QwpKwnrOoJWT1AdOg3Rjj++96adUf2dxbkhzzgexD68W83JasqpUTW9rP7AWhtaux2mkJtymVT8chLOj0/BTHQBUXBe8PltfJDqlaRGCEJHWloG8cDSAIf3dBfhczygFVpyGvt6TmntrOCOjQdSrZCAG3T02k/q/Zdt0mspk1h7XTI28pIGs+4TISrpXA6c7y4cgxSL8M+HHl3gSTS9NOaXqx4dmG5WjpWhXgDV2CQ6rUCL40VVb2tppPd520sl2xbEZ2X/CMyBAhKIoLbCfFBhBnGBMcjlD0hHD0kuvcr4gcPMbdOMJMhKjKvd4Cc8w6tsonolroLQIJqQUhyhJy6KinKmqywLDPLbFlzMauwzjYOApqitPRSw2xZs1jV/mI0vhpaA98vdFnpXEY45+X3aayZjmM+eDDk0w+nPHk45fBgjAmHlG5AXg8p3IDKJWgVNEnCbi3G+e8JQdIMSr/EwNTdADdKtTrZfT2hd4qeexeTVV9+aGVjF9JVmKyrArUZVlVb5bF6nSSp/X+7KYJSveIdbQd54kRhnSa3hrIymFpTGofF0jeWyDgmaU0cQC9yDGPL+SriYhVwmQUsC0NdeyBqq/QtK391fTfbZgatB1KbKkCrHe6x6wvWuVic9cOpp1eOb09r4tDv6Cd9w/FYk0SaJFJEzQLrpKaaWR8F1ER6z1eenotD/wxF5RqaUXUm5mnEGI2/XCNUmC0tXz8vEHFoI6S9gF4vIp4YgmmEDnuo6QHmzm3M8R3McIgyb3BKad0Mjjafp7PgDGi7Rmg/11KBKxCXU5U5RVE2Kag1eWEpK09dLVY1T08zVllNFCpPaS5KnPgIB7U+jxu3OHklM3ATXfODyaabH5M1pd2q4rTRaKMYDCJun/T46PGYjz884P69Q3r9KbWMycoRhevjSKCpghR2axh39wJ5nbTT1yHU5LXEbq8mI3/4Y25skZWSLTcx6Th2qM6u1HU2h7tT+U2FLErEKRGnwTKbvdPV0TuZrNpMCKutBsu+Sqgz07KXnv0p4VS9hIpr1WeAFUVhNYsixBUBkRIyW1NJzSB2TVaQw+jaV0aBT1ZtYxtWpVDWeqv3A51std01tvkszGuuUrspqyK+N3SxcHz9YvNccqSYDAyjnm5OM79oGaOIAsVVo4RTyjsiXM5ty4ptUX/dHolrrv52hsg5KGvH+awCccQhDHsBvTjEhDWTiSUZOdSBQ00cJG4t2mhdsbfsIG5akLesf6TR1G94Sj/bUqOkxNmSqqooypqibMFno4RzTTLqMqs3KsPGuNQY1Rmo/WVNyXStm7zFkKdWw0CTJiG3jno8ujfg8YOh7wONxzg9YVWMWNV9ahUjyuwoAYX/rkKEXd/+GyuhG+aEUMopcIiIeV8J/XLLKJFfyNGQbeRplxflFKZWhJUitHhKSUEtimUVkK9ilFMsqpqsrhkmljR0xIEjMN7xYJx6mq4XCeOea1JUDXmlKWpFUXlPudrR2I4353BrN9QRdOyuta9SCLYgYR3MM4dz3oy0skJdKe4eebFBHCluHQSEgWbQMxyODM8vLZcLyyJzrHJHXjo/j9Q4MQTBhpraXbJae5i2eVtWjquF8N0LRS/JCQKD0xqdrgj1HMwZzn5PlUdoZwmjPkrFIAaw6x2pdwltxAGtI0Yry1Mvy5XyAXZIhXMVta2pag9AZeXB1nVApaoddfM9rVXjl6ca+yX5ZXnCdWXyWmGtp1JtYzjb70XcOe7z4cMxHz0ac+/2kPGoTxANWNVjcjsmr/uIiRoGQv77wc6NWPtP5UC/04q4f2EQ2qygis2AWPc/deNBV9fR65/sDO1LtFw3spRndMJSExeatNLEAkGjUiidZlEGVGXAvA6ZVzX9zIfe9WLHIHKkkRAFMEosg1g4qB15rVnkAbNcM8sMs9y7a9tKrUt8OlJwrRRKy3pCSu+SjrKPsFDrSsUob7xZ1zB3gnV+0S1LWOTCnUNhOtKkseH2gWHU1xyODNOzmu9Oa74/rz141fjUTcTb2DhQWm4EQNcM9elGXFJb4WxWo74rqMVnmkaxIY4CIu0921S8JBxkMHpAmMYeRMX4tPjWV45tYYJ0PojuWPTGlt8PwCqpUa5CXIWzPqSurJ03aK0sVS0E2oOm3kp/fdlZ/KYn2Q2Umryh/Y/siQtvKxjlxQi1E4I4YDKKefJwxG8+PuSjDw45PhgSRD0q6VPIiIoRVvVQEjQCnHUuRWc04sfgkV5DDyjqDZ/hR3jMLZxRN7L1++oi2V2nmh2YaL9vsv5AS/beMeH97Y2ASW3C8rRTPsah0CSlIrYQIgRK0MovxnltWNqARRkQGUccWgaxZZRYRqllmDj6sSMOhCiwpM4H2oWB8Wq6hqaLKkVVN0DkNv2Va72p1+iLrYUPmo15apPNVVTC6ZVQlDWzlXC5dNw5Crg1xSetJr7H0v5Oo08gMI6iamaItmKQb1gj17+3iQ/ISsez8xJrHQbHIIDYWcgyBqMl4WABdU0YBBgToqOxX1i18TNaa2ruTWki14gTasRZrHPenLT2gFw35qSifQUXBAqldGcOa6Nu/KXubbvq4dafMY4MxwcJHz4c8cmTKQ/uTuj1R1SuT1n2yF2vMSWNGmC3jRHrv0ow3VvjcF7nRJBGoPW+Enp/+wGnmBI/IyQQWEVUaaJSr2m5WDlSY0kCS64deR2Q115RFFWasjaUtSWrLKvSMYgtvaYqatmjOBDGqfOVUiqUVlHXisr6odXSegeF2nt9UjcisBZMNk4F+xejVtTQ5hNtkl99pMPZzDHPhKuVcLVyzFbCycQxTM2a9hv2NMeTgMAo+qljmVk/T9RkDIm7uSW9BYRaYRsn6lVuORNHbIReoFDWUuUFd04qJk6IwwAbRZQimN4dgniEDmJUo1a4Zm76GoulolXIWQ9CtgGfFoA6SXWqod42a42sBWI/WjL5j0gsKDr9uObc0FoThorJKObOSY9H9wc8vDfk8HCMicas7IjMDSjFJ6OiWhpu4+DxHoBetsmTV+QJseWY8B6EfoFkHHTnfzaNIXUt3I6tvLtXld5vPCR4ozquY6HplM8TKjVRZkhSRz8UJkGNTQqs1VS5Iq8CrNNUVsgrcE6xqjRXmRAHjl4oxJGQBD7ELgi8A0ISeUNIUeCcprJQVopVrckKRVb6wdKiUuQ1VJXv8XR1DXrPHm0Tx+1VbWmsCI23o5mvHFdL7xuXV8Kq8DLup2eaUd9XQ3HoK5hBqghMQJo4ZkvF1cKhlSMrLJW8ujSTa+IToazg9Krmc11QOaF0ggpCouSSKPkWUQpdFgRlTjJ+RJgeo03nXTq3E/3dhinsMLZbf7CoBoTq2lI3VVAL6rID4k64EWTfaFX5Z2i313jM1s+sFU8oBb00YDSIuX9nyKN7Q+7d7nN40CftjSiZUtZTCjvEEjfdx03QoLCb87X/qlJbSa+vu3zTJUxf+lHtula/6i//1GO+bCO1prs3E+btyET7ma8t59bnmxZxzgk4a60zxryn436Zu4n9w6rbXfif/7VqwFhFUGiilSFNHUPlqMMK11fUKAp8qF1VGxBFZbWvFEq/cGolRMaDTy9yDBLXmJMKfYMHJeMb6LVTlDXEtZCEkJSQN0Dkwcg7KVS2qYzkZmV5+/kGxs/7jHveXmeRaZJLx+nMzwSdz7wA4fvQR36PeprpwPeH0lijlSIOFUmkySMhL1UzQCuvt/TINjWHwLIQvjuvqUWhjCFNS+JkBfqUXmUJy5JYHGEYYkyAiico3WgEG4HC629+ZG2W6l0EfAVk3SYtVe2jtn7pmzrVgoF/P94dWzPohdy71efDR2M+eDDi6HBAnPRwuk9pR5RuROV63phUCUrZnYiK97eX0p4i656n3DAJqZRyTSXk3nvH/Qy3Jm1EXrU7bmeF5NqwKr8Ym6p1b6gwpIuAQeKoAqHuWapeSYZiUQeUpUFqTS2t0q3b1xEKBXntHRQK6yhqR2mFovapqmHjtiANoIE3KjVaSMOWkvMAlZWwyGHZxDnYpv+id9wTfKKr4Jx3wp4MNMOep/jGfaGXWF5cWWYrxzJzXC7AGEs/0Vz0LdOBYdj3EeAC5JU01YPwRrOasq2ac06oalhkfvELw5I4ygAoyopbR5aJCFEU4OKIGofY25h4gjIpqsnS3oDtbhjHzVx+WzW0AHQTxdalOn/JWUFaKaxqZNnWg9B4EPLo3oBPnkx4/GDCdDJCBX0K16OwfSrpYSVGK72OZhD0e4R5zXP5FRLt9iR0SimnlJI9ldA7hfTvHAhV/sLwxc5ub11tOFUPQJ0eR3vRay813UxpXlsatqMZbirDX2fpEHnJWbF5Li2KqNT0lgGjSLCRo0xr8qimn2gGeUWVh7ja4JxuVDFeyba2/W9AqbQguaasNcvS5/uEgbf6CY2XBAfazxdFge/pxDEY4x/LOlgWinDuK6G8btkpBXrzsXlDVMFVirzyMQthAJO+Jo40474wbCK9n557257ZylNUi2zztTfTxJGfHwL/mEXpH2+f+exeZlNtzgKRbSl6ltc8PXOICHlRscoTrANjNFGg0VjqYoEpFoSDBwS9E3Q42PIgVDvUbjd2YPPa1DZlKd3+UqdiEunEvv+w7KBtxb/cxA398Mds32c7B9XuPLSP6j46THjycMTHTw54cO+QwXCCY0ReDyhtiiNGVNCdfKE7cnrTMOyOqdSbLamy3S7ZR+XJ69Ds6sd/zPak2RLH7ZbD+/4sQtdPob32NOKsIEpEbgChbtPhfbLq29grOG/75rx56b5h1o7Dr+t4w/1cw6o3dI7aE14JhJWmtwqwoVCnlqxvWYWOfmgZpRW2Kv3iWgWU1iurWvds1ZG8OvGpqkUNi8L/e4u5gRE/7Br6uaJ+7CMc0nijoguMByJfJV0/i9cspmz6GkXVumn73X8vVgxSzbCnGQ80476/n819RVTWvuLJS6GorPfBM34AVanmAL9hJbRZulRDzW36GYuVpbaFHxwVCMOIIFiCCMOyJO5nxHW19qdTOFTo6ST/httkqfaUex0ai1c7ZvxCq59uUqrgGqshRRwZpuOYuyd+MPX+nSHT6QiCMYtqyMr2qIkQPMX6L9Az/ykaBzs9stf4JbXJCr5BmPC+EvqxblmWbaH5ycmJfPfdd05ErFLKskks239415Y9v8yLobvrMlaRFBq3DKgSy6pnycKaMnbUvSYrwYBbCSoPKasmFrqzW18vHrKRYW9RK7qVa3sBQln7PlBSeVl3FPh2SFF5Sq6VT19rn3UwXZrqZZE5Tq8UvdgRBZrpUHE40gya/s/JxHGx9JlBs5VjtrQsMw9EVe2l2dZ22FL1w6+l1oZIK1/NVbWwcHbtShAGK2rrWGUld45KDqYWrRVRqBFd4dwKkiMwQ3SQonRb5zSovI4euDlCvh0+fdc6IOukVPEhda4pu6PIcDBOuX9nyMN7Q+7e6jMdp0RJn9yOKOyIou4jOmpc22U74fH9bT9Tsrdx/WY3a+07U/W8ayB07fr9X//rf/H48WMXBIFFxKom3nZf/estSrvOcPsCuNX+beBuWvsP3Mjuc2Jef7/ZpbdUmmnyhCQX6nlIljjKAERbVFJhAlABiPYRELiQovYx3ohfOMza4Vq8FFjaPsWG9nFNWurSNdVS7jNtWmdupf2gaFF7h+3Wq6011HRuUwkZ7YUVIjDPhKfn1s/+1GAl4HhiGKaaYaI4GvtqaZ45LheO87nlYu64WljmK8cy92DkrvHgaosCeZX+a+0T2DqKr4cjhbywzRwRLFYls0VCXjicU40rucXVC0zvCtN/iIpvo5MDgqiP0Tu1uNj14iGdk8f7qam1tdCWK7Zsj0p33blvRIRX0Lz73vs/9ZiNjFyJ72vVTVLqaBDx4O6QXz2e8vjhhKODAVHcw9GnYkjFiJoeirDZRLj1C5EfcMm/XhTFq6lIpW6iHNVP9JiydhdTW7/cRs033nHXvMV+0Hr5HoR+7FtZlruQIEopUc45lGruL/ng1zHJv1BbfLVtIqwbWi5ZGQaXIbXxjJAOakxUo1OFshpjNUYUi9JQNM7Z6/ZIe2+cF7pz/xu3AV8FFTWdwcFtKqlb/Wi1LTGm8/3WTywvhdPaU3NtIJ4InEwVo1Qx7mtGfZhUhsnAMRkYLgaOi7nlYu7dtZe5o2jECbZxZ/Y+ZT+czdJKETasmhNYZpayKsgKS1HhkV0ZnDiquqSfLUiGK2JbY/o1Cosoi4Q9lG4uF6Ubqk7TpED5OlRptNIYrQgaINLq3dSDtVlH1gomVhyMYx7fH/LxkwkP700YDAdYUqoqJXc9aukhxCi0NyftUKPvFXEvq4Qc/6xWsqmE3tnbO9kTasjmfxnvpJaaa5VycWEYLASMImh21MFAMKHDJBWhKKJAiLKAZRGQlZrSKqw0EeHdDB7YdtSmawjauvWqDSvQ9JCCpj/UKuLoTMxvDbK2ppaOxg/NrcGsTWk9Gmr6PUMc+vo0MIp+4hfvwCiSSNFPPaW3zP09Lx1FKZQCYmVrPun1N9Vq4xTe9JqsFarC4VwNqsSYjKp2LLOC+6uC44OSg9r5ykhVKDJEFtTRASocoUzigUuz5aQtXvGC0tpXls1da96J6O71XminCtVKkcYBx9OYD+71efJwxJ1bY5LemEqGFFWfQlIckQfmaxP/7wHo5VC/x8D0NRY2eXnJ9E6ti+8iCKmXUXY/HT+obijJf2BF1O4cBUKrGWQBAZrQW42iXI0aOnRkicKSNHGkcchs5bhahcwLQ1ZB5bw4QfmZtmbYVG2rvVTr2rCpctYAI5sIb99cbheoTcaPQl07Clr7n7EOFpnw/bllmQsvrhwHQ82oZ+ilijjylYJSnvYLA8WwZwgDIY0Vycrb2uglOOd7RdgO1dQBxX38iLrpUu9kLCmgto6LmY+AWKxKLmYRVwvLB5kHT60cuBVSXqLzcyS+jU7vYJIjnO41oLPpxonSgEFrQ6ANQaAIgiYjqBNWt03rqDdoR99AtSn1T5x7sj7eaz83kSZbSnnrp0BzOE25e6vPw7sD7t0eMp1OEDNlWU3J7ZCKeC3Bbj/r7Zbe6/XGblxRb1AA3vioN12QN3xfvc6r+MGPqTqvVXVes+zADuv+9bZjwh6HPemK7EQ16rh31kn73RxWFVE3h568Y4gqm0pIlKAcGAdx6asiUZrcGPJQSGKhThw6dkSJI4m8u0IaCHEWNEAkVFZveijykmtm1zmbPd6V+wYru/MtzeMH2q/wIn526HIhzFc+Z2iQerVcP9X0Ez+gmkQ+BrxbLQRGdb6nfpQKYvf9hoFah7GVleP0smSVe8AsakNlDbUTnKs5KpYMhjOi/hzdWxG6Cq0sBFOUjVBS4RVzGlGhb8rrEG0MgdHraqh1yd4nZPpZT2LZAUTx9Jt1glKtGi7h/u0BD+8MuH3cYzLqEcUDMjemcGNK28fp0G9EOmLV93qEf4KhE26ck2t8ULfUcXtA6H0l9GPdbkpW/WGVC78gi3y1oYsQn/KKByHwKmANSA2mBF3hBVnOu1bHTX5QEgjDAMaxY5xbZoUPtltVhrJW67t1YJ1qFHObfk3Xfqc7hIpsbzrVDT5yuxty5V+iDyOtBSvCqlCsCh/pncaaNPYR3mnk54OiQK3D3GrrTUhXuazNTN2PGHPdpebaocCycj7ZVSqgwFohK0qWq4K7i4xbhwXTg5q+WCJTo02ODY7ADqEKEWdxGJSKEZ2AjtEmJDCa8CZQ3bJr+XlWjG7VC82Qrwi1dVgnhKFhkgbcu93no0djHj8Yc3QwII5THD0qN6CUAbWkIEFTYf53jut++yDkA+1ExE+eOOX74hIEwXvHhLe/T3uNn5HrF/f6IpdNqXtzgoNsWwXvAa6bveP2e0ztJVvUBoDa1b8dBpQma0WZBnRwSOCwMdSxUIdQK19tBAhB6OjpilHsGPctq9KyKA3zImCRaxaFYVkoikpTVELJpmdUO7UOutsVIrQvr51BUvio7zWdd0PmjetUXlpvEK12fqC1do6sVIQrRxCoxvzUV0KazQBsVdPMEjms9Q+6W62pG+iaV1mNugZ5XUNF0ij/srzi+blQliWXC8P5VcTZrGK+Eh5azS0cgWSo+hxrDqk4xMkUcT0Eg2BwKgIdoXRAYAxR6Adiw1ATGNchYKTjofbmF8APe+/7H1M6FGurnqydkBrNwSTmyYMRn350wAcPDhiPR4gaUNgepevj6OFU7B0mxO6c72prBkFu9IW7KclU3fB+35AjkxvoSnlDfu2HPubu9PKuzFOu39ejJR0kaj8nJw6FD7RzTpyx1oXD4T9tQfgehG643ZSs+lpU943ecfy8A6sdrrd7rrR0nDgPFKBw2lGHjiqxFLElM47MgSs12nr36lApQiMMA0ca1fRroV8I88jQy4VFpMlKIa/8EGtlvZN2WyF1K6J9eLl7sd00V9fduXm58/WemQcYIXvJ47/dj1621grvNdfGKQurvKasLPPMsFjBItfkVUbtDHVdU+YLBoNzdHSGC0/Q4S1Cc9j0iCwa26THGoLAEAaGMDSEgRcrOPcLi2zYeGbiZGMzZLSmlwacHCU8fjDgycMRt0/GxMmIwnoQKohxhKj1IO8bS0f+u9Q2P/L5K23Ct1PKOK3UO19+Bu/oYX0pjHS947qW6KrdnOm3CEJK3UC9dXdEfjYALLKWswoOQTW7al/rgDOOOrSUoaXQQu4MVWmwYhAUgVLERghD5RVt2puYauWdEYapo6w1pdVUtaJyLQhpautpOivdnBi1puta9VxV0/zsjgXSy6i5Dr23fjznF/vNc8k2/df8ru+hdEL31I+/JKiGgtJryyNZg2RZg7VV46UnZIXlap5z9yjk5GDBaJTRH1bEg4pI5whDatGIy9BSIEowRhMEAVHoK6Iw8FEa+97Lz7F0d33r/HXinRGM0d6aZ5py71aPB3d63LnVZzQe4vSYvByT2T5Wx4jSnaTU9yD0tkFo5wh6Bv99lMPPUUiI2pkWW8eEtsdjYyLpF742FVOUQrUTnXu9497kdJL9V3W3069kTbutn8Y5xAnKgneGsg38tBewQRFiTIzrG8yoRg8NqldAbL0tT2lY5AGrWmOdIdJCGgu9yDtp+4wd6MeOQeKhwOGNRmvn54oqJ03st2ocs9t/k83PWBpLHljlsk5mbWzEbk4A7YoZthhOWavx2uOxRa+p7b7Ny7Qn8k9e8h4cO0apLT2nBOccWVFzeuHI8orLWcGz05B7JxEf3Km4fxtu13CgahJZooIBSIS1DsjQqiYINFEUEschSeTl6a0f201Spn92WXuz1cgDvXO+d0fjD9hPPQA9ujfk0b0Rd04GjEY9wmRIZidUTChlCBJdo5zlNV/R6wSxyWu8S/nRPhV5C5+0XHsv6jpheQMndxP5eu1f3sd7/ywMglLSVRXva7ysrWtk09hW3UrobdBxa67Xvd6FpBRKAiDwmGU0EoUEcUqYDDH9AWockhxZ0sMlaW/GijnZKqNeWRaZ5TwLyOoIhSaJfLhdP3YkkZCEPqohDpt5Hy1bc0KOZrbHQWU74Xadr0XtAcg5KMrO29xquN1w+V1PJf7FDcx0KUStveNE1/w0LyxZYZktay7mNRcLx6ow5NWKqhLqumQyWhAnPTApSgeEoaADSxob0iQkjQOiUK/pSec2lKD6Gc1Wunsl2yS/xpFhNAh5cMeLER7dGzGd9AlCP4xaypCKIZYUhfEGumvjqJ+KVP1Xq5L+pcYd/2VBaPcyFaeU944Tcd5nQPZ6Re49vOoNnnXfN9s5me5OZ0vS4poKx983f24fIfDWJnGETmJUL0WlMbqXQL9PMBoRjw4wkwluEiNjR5FcUdTfUc2+JH/+DebqElcpqipgVQVULkQXjnlW0Yucr4piRxr6LKEwUOs+jW4qDNWsgI5mWJXNfFCTqoNtsoLMTQOXr7GAriucTnqgetX+UTpC35/AfqxtzPtZp9ZA1cuVSwtlZalqD9bWafLCMZsXnF+tuHWwZDrpMRz0GQ5SkiQkCoSqZxikIUnsk2Pbxb67IdrMD8lPo9zcOV7tc7c0qTGK6Tjig/t9Pn485uHdMcPhCKcGVHWfwvWpSREVNQ9WX/Ocf397nYVkrUDYzARxXb9w4zmhxBvzKSVaKQnDcF/59M7Y+PyiveMaA9OtW13XLkySGucqtKpF2jZ+x6cLtlQJsiXzVTfaHKuXecdt0W26Y8LfnExbuRHduwPqZvExgEb1e5jphOBggjk+xBxO0NMxejohODggPjpGHx7AMCFMBFWfI+d/Qb6OKPM5y6sL+lHJqk6oFLjKUJSKshJWpSMqhCgQQiPeE84ojJLGqVoIVMe1ulkI24V4cyHIun/kk1bVnoV7x57+NciJV3layl7Qf/uMfVs9b71W5T8vDyCW2UJwzrJYFjw9C/jmRcS944R7tyrunVjuYRmlPYKGfksTQz8xJLEfXlVaNe4Pfg5H4TcFbscf7+0VjLLdDGrShrVS6EAz7EfcPk548mDIhw/H3Lk1Je1NqWRMXg0oSBBC1DqsTq0tsTYbsxuN1X7U4/VjQt7bfUzvHKfUtkS/kVvTkb+tN1xrBqeDRJ0lydvNgRPlXBEEcvHNN6+7gX8PQj/G/k0p5cQ5q6DGu2m7lxa6/+wh6E59rm/25kwRrVAEfvHUegNqQeCrn34fc3BAcHxEcOuY4N5tgltHmKMD9MGU4PCA8OgIfTBB9RIiDWZ5Dt8NsG5FdvU9y6sX5PklNQWUGaCom3juvDJQbMur28gH0wgXAr2JbQjMBpDaikd1KqHWY8666zTOP3vB/9L2qbJTJQQBBGyGW61zzBaO5arm9MpwemU5mwmXy4a2FCHQQl1HVLUQBYrpKOLWQczVoqKuhWVWUzfeeFp3s4XePlvZVr8indA9EYJAk6YhJ4cp92/3eXC3z+2TAcPRCNsRI9StQ7a+yZbnfSW0/wzbz9PIa1RCe9luhXSFCTeE2r2vhN7iheQ7x0pZNK6TaLZ1iNvdhJPdZNXuvZ3T0W1jwO9W27CcbkXlWtfbVtFm1zs/wXiaTYWo2FNtKokxvR6ql6B7KWrYRw8H6OEQczAlODjAHB0S3D7CHE7RkzFmNESPh5jRCBX6Q6OBfnwX50qKxXMW51+znJ1RVjWWHNQMcTV1FWMloSoDr+5qJGleROAXPK3Fe9EpnyvkQakLQtvA1arjyh8RhN6d84z17lUrf7StkyZ0z5GV0gwDQ146lrllmVXMFxV3jhP6SQjA7aOEj4vhmmb89kXGvHKNK4GswWHjsPz2qDmlfNXjjWEd1jm0VvTTkFtHPR7fH/H4/ojbx32Ggx5B2KdwIwo34v9j77/W5Di2LF30n2buHjIjBQQhqMmlqvr0uaiLc7nPA/RtPc/Z/Sx901+vp6gH4P66e+9idy1RS1CAADIztEszm+fC3CM8EpkQJEESJIJffgCSkZERLmzYHHPMMeowQSVtj0fovcc31rz5x3n0/LH2/12bZ3cIQrIz/lEjoqpGIWgv2vuti/ZrZEqevZlCCCKiaswusq4THOyTVWOWTPC9yASJUQhqBTV9GZZpLey7BkjY8UbSUW0Sol9GTNBBaVq3N4NgEZNiphPM8RR7PMOcHkeguXWKvX0L+84t7OkJdjaLYDQZI5MJdjaNfaHhEDPIkCxDrHnmJI2PbnN85xPOHj6hyLe4oChfYkwRsRiw6QCTGDZldAJQjemr1ugeYFqE8aqob6MZ/F4ODXrI2HQmp6GnipPnaOOv2dYfDFR+Twj2Ol+zo1q7gVLfUnXS5uSoicfJNY7lOlDVDYt1xdPLkq+elLz3zoiHd8e8czbkdJbxaUt51nWgKD11HSgbv9ssCXJtgODr+OxGhCBRiOC8MrTCySzjo3dn/PaTMz5+/5Sz0yNsOqZhSqNHOJnhZdz2M2Pf82X81lSuv52/r5Ql/Q6vedPPvpbXlKsWpVwzyNhzHdcWovp0XHvzWsBYE3xQ9Srqqup5IPS2EnpdlZDElTIihVyfrBrjvbWdgehVP6b31U21xuSudtERNPQqof6iamykT9QiMoyVU5phJm2Fc3KMvX2KvX2GvX1GcvdupNzuvUPy4C727BQznSDDAWJjKSJRunYo49Y4tyHG7n59mgwZH9/n5P7vqOsSxSN4rPkaaypMUpJWKYlVEgyFYRfx0FFzBxSiRrrNf5cdws90A9yPDu9/xJgRtPflE8C5wLL2rLeO+arh6bzm6bxmsXbUjfLuOyMmo4T7t4esNhPyMoLP5bqmqn2siFS/93mom/YIoR1fCK0LxWiYcPfWkE/em/KrD495eO+E8XhGo1NcM6HUMUGG0PrDsRPa7OPp3z6+U1nEtcYv13g2ttXyrtnctiLezgm9zsd1jgntcFaQfcT3M8xpP967G1jddbhNP+OgU7NFczb1kXfRHt9tMECCpHtFmxkNkKMRZjKOPZ7ZMWY2i5XO7dMINmfH2NMzklun2Fu3sHfOsLPpS68W6twuvVyMwZiE4eSUk7sfg4ZIr+GxxmDtN4gpScyKhIKBpmzTjMqlVN7g25iHOGzaVo4qN/dI5HB/21FS3SHUF+d+/Syp/b4b+Q7I26jyEAJVHSirQNUE6ibGmG9Lx+3jAYkVHt4docBoaPnro5zHFwXrwrWg8Oye57m0nDx7/A8WLO0/R3azc9ompooIo4Hl7GTIw3dGfPTumA8eTLh9a0YyPKHUEyp3RM2IQBrFOHI1+0bejqfexAT0dmqqLfXajW90OUIHWULtBvh56ri9d1wQ0SCYICLhGnXcG6X7fjMrIdUgIk5115g5MEkXVLsbWztDs85428SBVUkMmrSmZbtez8E8PYIBm2JGQ2Q8xhwfR1Xb3Vsk9++S3DnDnJxgT44x02ns6ZzMYrUzHmFGI2Q0xIxiX+jlOROD2GRfYbf8VzoYc3R6nyRJsImNqiabYZIBxnxJZtcMbc00Tdk6YdNYlqUhr2KM93433Cv05LBFhjwb+/DSIz5yQ1rlTd9/IaHzHOXSDW9IX9NrikJo4y92328PYid9V4Wq8cxXDd4r69zxzXnBu3fHvHdvzJ3TAZ++N2U8jEFNjQs0bY8pzm7FJd3sdO16UHVxpT9wJS/hynP6MefSzgKFHf03GiacHY94//4RH797xPv3J9y5PWEynVGbM5r6FpWf4WVITzfZO34Heq0XHtsbvRZvOEmir8YpXdG/Xn9Ob0qk/b5eU/tgse/9oCFSct4jomhwbZ8gApFoz1Owz+LEboD0AAg0qCJBBR80+DRNw2QyuVoRvQWh7+sxnU6fOZjGmGBUvYIzIl541rr3oBJSfdZXUHrnSWgBJ0pPSWJ/R9I0zu8cjWMPZzbD3rpFcvcO6cP7pO8/JLl/F3t2gpkdYUYt6IxHSJY+Z7jm5YGo3wOINjOGwXCCbQFKg4IkGJuR2ITh4Cummw2bQtk0nmXZkNhAaiypNVRN56hNDMDT5+LI2x3ui+hHif591rbmn63qrHGBy1XNYtPw+KLkYlHTuMAgM9w6GfDOrSHrwrPNHc4r81VNWXmcj44SxsoLKyG95k1dvcyVtpgWdu8rqDJI90Opv/owOmTfuX3EcDRB7ZRGT6j1mEYnkQWQWHW/Dap7yU3YwfhHuwE2FkmS6LXn27RIrlRE7T3Zd7o/gMQIREFaqxVjrE+SJFy3Tr6l417jwxgTJAQfwCHiEQ1XV80OgPYy1MMbEw1oaLvyxiCDYaTWjiaxihmNMeMxZjbFnh5jjmeY4xn27DTKq+/eIb1/N8qqZ1NkNMJk6YvffAhR5HBQhUjPcLedi7kCXto1g1sVX5IOGM/eIYSAsQnpYMRofMTR8Z9ZzR+xWsxZ5TXDbUNqa6aZpWgSSmdxwe5MTLsZoH7Uw35Y8/rNo7xdfw6OhfSC/zpi2HtileOUvHBogEEW3bSrJjAeWm4dZ3z4YNLW7fCkKamagG1zlW5me/Sac3NYocgBnRf7Tc5FihBgOjLcPhnw8XtTfvvxCR++d8rJyTEkM6owpdIpTscEskhBEnbzQPAL1cJdF7R19ftdb9eYrgRtq9aA2BSTDhExqFRxk9uVz2Ev8t2Lca/ON+6OfYdRwYoEY4zevXv35b1+3oLQq28+rz6qqtKBtc4Y06DqtJUK9RVb7Y5UvVfpDDOlrfGlixO1FhkOkeMUY0ck9+5E8cCdU8zJKXY2w561PZ7ZDJlGgDLjfS9IRsMox37ZldmYHT9/Y2kvnXtB//ut42qPgM8GQ45O3iHNMkaTY45md5ndeo/Fkz8z+uYvDOffMBosGA8LqlopnVK5QOWV2lnKRiib6A1XtVLjxoPznYRXd7HfO0qq7xenB+wD14Stfmu67CVbNK+j7fMt1qW4g+3zIZ3SsCuG89LxxeOCqg5cLCreuz/h9umQh3fHKLAtHMt1TVH6vSrqCs22i4Fou9bXpI7sj3A3eNxbyILGKitLE06OBrx7b8KvP5zx649OeHj/jNH0DEfsA1VtH6hzyO5AKKpD5Rpl3E1xEi+INeHqgMX1BJ58n+DxbV7zqqJH924H+5ujB0RRlroDJN0BVDeIt3dMiGbGrb5Kd+uXRs6NZxqwIiYgeBG89z6UaRp+//vfX/fx3s4Jva6HtTYENU5UnTHSdu/1mUrIdwN54Qqvbg0yzGLv5uEMczTAjk5J339A+uF9knu3MWdn2OPjCEJ9RVuWHuYB9eiyl+hlXV9KvBDAepVSd8FKBLN0MMamA7LRjMH4jMHRXbLpHZLhCYPJn5nMv2K2OaeuSqraUzaeygmVowUhOQCj2sUv5/eu2VddBDrJdvccFw6bqL+0Skk7OW3vPEerIyFNuopEOV9UzFc183WNCzAeJpzOMm4dDzidZUxGSVTOtQc8hOudB68au95EqYoI0la5AqSJYTRIODka8sHDKZ++P+Pj92c8uDdjNjvB2TOK+pQizPBmiJK0lbe20d3hStD0L+hEH3DUV3pw1/2/ON0cmQ9V1DtUS5QCQsCXl/jlglAWkZFpS9duvOTaSkh2Ps1BTCdRlGC22zfefO6NpONU1cUqCNflaVydfO8yYoKPMYS7YLQ0wRxnpB/cJjk7IQ1n2Mkd0gd3sPdbC53jGWYyxc6mmOnkBby87i62A+Mz6e9eJDpGi9xcOrzUghcDH+JOy8a5FWMZDI8wZoBJRySDCYPRjOnJHTbn/87m8q/kqycU2yV5UVA3SuN9pOSCofGG2tN6o7WDqW1F5DonojYMLmh8XtFAUUFeKUXdcyv6hdJ1eqBEa7Uu+0VjFwfRuOiWfnoc+0JZGunV8TDheJq26jqHb6nkeL3LwUbGWiExgrGdGEIOAvI6g1RECCFSgiLCMDOczga8d2/Grz+e8duPT3n4zozpdAr2iEpPKPwphZ9FM1ZpFXE7U57QO79XTZzedI2cHvTXugr3mY1ja64oMW8ksinW7rlPH1Dn4lfdoFWFVhUhzwllDaUjVCXBLaibr/HuEpUG2usgUuHdaEnbLtptoDuqvpNn48V5n9b1TTLtt8KE7+Nx91/+5ZkDWZalpunIpVZrkeB2m4UeN6+0MzA+ToZr2JdCkiXYW0eYs/sMBh8zGL2PndzGzsbIZIiMskjTpVmsfF5ml2RaL7nrdJXXVUAvkSx53ffjTRGpuasUYJJmjKdnZFnG5OiM49vvsTp7wPzrM+bf/BvKFzj3FEKJFU+WRHpAd5s2xSst8EQA8r41N9Uo8a59BJ9VAUujrZtCO7arexXZLj77BbtKfUVrgJd69iu+/vf2mr2KMTxDVCliYsS398py0/DV04KgSmqjr9zZbNCKFKKqzvkQ1Y/S+vsRfz5LTewvJfagd6Q9ENKW1g0hysetEWZHGe/dm/K7j0/57ScnfPxedMhGBpRuQKkj6jDG6xBC0t5LLa2EiXNM7QZI5LojpzfYCL+AC9Pvad2Ub/kyfUrtIH1Ze7R4N6PQPse0/pHWQprETaZX1FVoVRO2OWGzISxX+MUaP5/jlyvCYo3PVwTZ4Kc57iTHnDYwsu3S4aOCMewVcvuomnYTLhJE1KPivLWuWC79NZ84tAvFWxB6XZWQMcGr4hTxbWl6QBT0A9R8vxIyQGox4wnJ7B6DO79mePor7OgMsbRBcyFyt8heavkqZfvrLAVkL0TXKzy1asAmKUl6xnByxmj2DiYdU9eOzWaNLBcol3jvaZzDkyA2jU1UYtSDkXhB7CpJv1fpuADWxeOa1nEjaOStTurGTXWvN2OMMGiD7ZLEkJeOR08LmiZwMs1IU8PxUUrjY47ROnd4r62duVy7GbFW2thw04YAyi4uQgRs67oRFNLEcquN6v7HX53x6QfH3DkbMxgOcVi8BzHKIA1Ydb1fKaiY3iT/9TH2PwuqrR9o9TzCURVtuWp1DgoiE1I3hM2WsG7BZ77AX8zxF3Pc+QX+ch6/iiU6rNH7BpUBcjwEm7bXje8N2R/2XWV/canE4UZvvPfDNPVXoPiN04+8cQam1tpgQvAq4hDjYjSOHrBbnVTW+RaE+smqLRDJZIQ9O8ZOz3rMmTnYQGiHZiFcoZvk+gGaH4SL2qUCHUZIoIjsT2eajRlM7pAOT7HpBDEJqoGmadiWntwlNGIJJKCKEX9g7bOz62kPgQsxcbSoYV1AUSm12yu13qrmnrNxavtEtMOuZeW5WNRRnCnCrZMB41HCpPakLTUTgsaoh/aMBwXnNO6WtbuuE+wwqu6GA7vLLspSy7DLMWp7QWfHQ95/MOWTD465f3fCeDwAk2A9GBqyJGdiLIEKxLZiBoPXqKh03uK8IWiX6hsd4eN91w5V/xTXvRdZe3T+kVEbH0c0WkWJaLcGxB2Zeoc2Dq3qWPFUNbrN0aIg5HmsdlYbwmqFny8J8yV+vsQvFvF76zUh5HAsMB0hzTEweKbS3o+WHE4utotQAPUmqG+M8Zdp2i+85U0DoJ86CMnvr9mMJEURdDptDDQi6lrN/NVKSBsf5yKc19b3qyuxPRpc/PLuuZWrtFSbGrMbKNtDwQ1lwA8ERKqdn1SPRrjysEYwSUqSDkjTFGsNQaPj83luWDUJpUtwPirdrdlHXitxns53PnyquLZ3dFXEICK7Iyi9vvULj8Trt43+SbzmjtZpm851HQjBkSZCOY3ZPGliyFLBihxE0tOajnaquaZpY8ibzolbGGSxIhoPE6aTlOk45WgaxQ7jYcJknHF8lHHrZMhknKIike5Tj1CTyRqRgJg1YlMwFsUQ1FD7lNql1CGjIqMJA4IOCWQgyW4mxnRmrM/0h+TlabTvEXkO5qj0sOdzuLJ3HpJxdkeS1k5r5+AbRzlC3aBlieY5frmJ4HK5wD8939NtqyVhk6P5lrDO0W1ByEtCVcT+UHAwUOQ4iwnP1vSMSfcMTrfxC/QTY9p7PGhAxKtRl7jgZDi8DoTeKDB64yqhxNoQQnAq0ijiMO2cUK+S7naKzscb9cBkJHjU1YSyIGy3hOEGO5j1rl+/p+N2FNhPb6svV10IVJ/pKWnwJEnGYHTEcDwjHUwQu8KFQF4FFrlnVVrySvF+fw339ROd3f8uClufjTz4pThrf6vlsK/MbIHIB6Vpos1P7QIh6MHMUXcNdyKDJDEM2oHYTrTQLVjd/E8M3gu7n4nVl2GQWYaZxRqhrDzn85LlpiFNLFlqSZP2K11GFw5ro1uHsSAJygCrAzIdIWZMImOcNjhGBDICSVsVXV0Hf1yK7TpHiRc+fEDbCGENAa0dWtdoWcW1YrMlrNa4iwX+6QXu8VPc19/gzy/wyyVhs0bLKv5M1UDjI3XXwcnAIpNBVNoOB+zlk/Rz7naVULjaDog3X1DwBHXOWv/n4+Nw01r5thJ6TY+ttWEQQoO1tag6VHaRa/sbeG+77w8UJgraltWuQZsabRrUNOyb/aGTN13LE/+0F7yw6w8pYG3K6Og2R2fvUefnbFfnDDcbhmvHKG0Y2oJCAsEbitpSe4PzXW9s34uV3qXdAY6Rw77QW0P/F2+luuC3mOfTDbSGK5Rxm8IaYiNONcZuHx+lTIYJxki0+qk8tYs0WFV7lmvFOaWsPZvcsdo0keIbJYxHKYPMkliDsTGuJE32ADXMonouSQRrDSZJSNKENMsYDEaMhiMGwynj4RSxJUEK6jCidkNqP6TWEV6zNtJEegq616Aevk7I0OvDSf8i7QZHd0FZvegW2e+y1Dm0aQhlhRZllE4XJWGbo9ucsMlbOm0TQWi+JFy2PZ+nF/jFMoJUWUBw6IG7RHTaRw2iJmanpEkM9Opbp+/HhvZ9oevTVQOqXhGXeOf4+OPAZ5/BW8eEH+4x3Gx8ODtrJIQaK/XOFbBndKatIME5bYcu94090QhEEWxaHzlre00j2kroW+ykngsQrzfOoAPRrmFsREhHR8zk3fixQk1ZbKjrEu8C1q4ZDxqmQ2GQZszzlHWZkFeRdtM2N8sasFesfHb3c29w8bt2Q+VV15837DWllQ7ubFl612jdeJrG4P2e6g1Bd84Vg4HhzumQu2dDJqMU55Xlpma+qlmsSta5Y71tWKwa0tSQJoYkMS3FZxiklqQVMSQ2Ak3a9o4GmWGQCaPMkqaxesqyhMEw4Wg84PR0wq2TMbfOck4kZzgeYc2AjCGlGYGf4fWEEKYEhpHWkt7GpUd9vcj05+V8/65OSdNLUO4BflRn7FJsOzeDnZNBbLyhwRPKmrDe4JdL3NMLwuWcsOh6OWvCahOBZrONyrd8G0GqqNC8IJQN6qIJ8l6uLlfGjQ3aVbpG9xYb7AG0Yxx8p0yNSseeOk5UhSCq3hh1jU0dv/99uHII9Ttc3m9B6GUeiyQJ0xAaC7WqujiS1y6I7UmNVO5NlVALQp1fU3eBHmxZ38D+XsvNd8IJEUOSDknSIWIMwVVUxRr1NVlimE6+5ni1ZbbyjLKG8cBwuVVWOZSN7oZQzW5BeXEC6NtK6OWAqX/JOR8oKo8xQln6qIpjT8mpKokVJqOEs+MBx9MM55UkiQaoq41Q1Z71tsH5vvXLPofGmlZNlxiyJAJQ/LKk7b+HmW1ByTIZpUwnKcdHA24va1ZnBZttTl6sOZ4NmYwHpOmIVKaMTYMxUAVoVAlkKCbOL+lr8rWQay5EuWaMtuvDOdcy7WHPMbfVj5YVfrXBzxe4p+e4rx/jnzzFX1zi5/MdCIXVmpCXaF1FwLmSDxNjV+z+bZmup9SeTyNtlHG7qzOHSLv3u+xZjoWdicruaTENXnwIuCw07ob93xs1uPXGgVCapsE554y1jVFxGAKKioh06pwusKtx7a5e5ZCO64QJ/SnLnwPfc8NjMD5hevYerimx1jAej5lMp4wuviZNLzFSkFnHKLFMU8umTskbS1nLzsan6zUIut9UPuct6FtEegaFukNielLqxgXWeUPj2sC71t/NGtnZTXXxC7ULlHXs/RSVo6w8VR1/pmoCTRMNSoMHH8KO0onmt5BYQ9pJuxPBJrFXZGysloaDhKNxyumx7sYbmsazXBU8erLieJZwdjLk7q0x79yecXLsGI2FYWqoQ6D0NZUfU7sMpymqUQSxvzT0JXYuffGAHj6lm9nZ0WotmluzU7RFVZui3qONj6BRt2q2oow9m6KM9FlLt/nlGr9Y4C7n+CexEvLLFWG9bsUFBVoVBBqgW/dNG/KXtCAkPVseItCENoVbFTUg1iCJ6QGRPHPPdK2EvttLb/6xyxFyQYJz6cBdcwe+ccz4T9077pkKfTgcBorCWZFGksQJ4mXnhh2PfwDt+Hbn445COs8mDS0Q+TZLSJ9Ddx3eJT9kKuhNz7lpWPJQKX7FwUuE0eQMc+83DEZTRtMzstEJJvs3xPwFkW8YpjmzUc0yT1nVCYvCMN8aNgUUteDbflO3mHSD4teBUHfj7A7fSwQQ6Wu6gH5ar9m5IOyzieras9o25EULJrVHJNrsKHG2rQOq88uS1aamrgOLTc18WbPJHY0L7WxQm31qO5d03XnNgbSbByG0SsegHoeg4qnrWIWNBhYxirGCC4H5suTRU0ddO2wCJ7MBHzw84nefOD750HAvEQZpIE0KUjPFMgM/JYRxdJyTtLX9iT2imxOIrtBsPW82PcCgSKdJJw9vqw6x+3BInEd9E8FmucKvVoTlCrdYtjM8S8LFHmj8ekPYbiO1ti3QIg6dah17xngXVxUESDrCO76H9rPtGjrIvqfcpTJr7+ZohXhiZdcn7NpTGjfQ2pkKx32fyl6Jqwp4VJ3FNGPnmv7d90//9E/y2WefCW+YXPuNq4SyLNNQ1z4Y46yq6wRc/eVNDyohPaTjYl+v/Qo/0y17H7Si+7ZNB4yP3yEZjLHpBMwQNUOMHZBmKUeTp2y3OcutsiiE8SaQmEAiYEUo5JDa1N6Bf6WGyVtOrnVCiAerdoGQu6h889qKDfYgFB2wlU3ugGo3Z7TOG7a5o6g8qrT9no4+lWdVVTdsXLTdeUdXEcVaYZBFWk4kmqs+vSx5ellQVo7xKOXxeUPdJCgJIXhu3yoZDEeInZJphZeGIEojUTmnLzW83wsMPKieDq+rA02ob13pnUOp4qLvQ7TK2eaE5Rp/fhmptcs5rv3Tn1/uBQXrDaHIW7DpA0bvl5qkrXpesBnt23YpveAubens1oHCyitVQr3PHYA4I2mCK7LxwaDqcrk0vIEeSsmbtrKORiNtytIX1jpV71uvP7Rnx6tKBKG2LxRCP0Z3n+Gh7ddV4vXns+hpq3Tb2/xkwyMmp+8SVDA2ZTiaMp2dkC++ZLN8wni+ZLQqGSQVKQ1jq6wyy7YSSpdQOUPtYhSE7xmYst+k7vpI10q35aaK85f02OcFOac471tJ/d7yx5q9m3oHFE2rlqvqKMt2rke3tU1v04bYaT8uvldl9KOjhcgSgPYm9bWbzwaBxillFVrFXYwtbxxYm+B9YLXOeff+mFtnY45nBYOxJ7OKSoIwJPiMoPbg/lO94kWz+2sbOmmiLxvGRBfqLvIg3tixx+PaodGmiUOjeatoywvCZhOrm+U6VjyLBX6xwi+WrX3Ocic0UFeg1K3pkLT0WoLQ9nhMzy+uTwO2B3Hn5dhnTW5SqksPgA76QnthRdfP9l652i0QI0EVH+eEjN/MZn0Qomka4Q3cCiZvEgABfPbxx+G3//f/7W3TNB5p0BBHgcz+so7KI1o6LuA1oGp7t2S7OKtrv/QgT3lPw92s57mZFpPvgBmv4MZ95fnX0nciURZ6tZocTJidPmAwGDI9vk1x6wHby7+zuvgb6fCvpINHDNJLxmnNralnXSrr0rCuLKtSWOWwLaGsuoC89tcFbbnxK5HgItdTWr0S9uBme0X662We/6O/5hVxRwQP3c1h7Qxwd30P7elmWtVaatvL3Mfr2rUmvV0fRDvw0N2vkoNjrc947PY8ymiaQFE6NtsmihkSoa6j4iuq6QxNE7hcFPzxL5ds85JHT8Z88HDKBw9nfPhew4P7wtFxgk2HEKbUfgDBtqC4d/nQa45kvExs7J8YgyRJ7Pe0rtQ7c9Ad2Gyjcm2+xJ23vZz5IoLMZoNutoT1llC0AFWUhLKM4FVWaGha8LE9ULQtzbbvwkjk83c7Kt0N0rUg9MzVLQciid0xNxJ7V6lBEjmYfejEJLEKiiGsfqeO65xLxSM4FXHG44fn5+EHYqHfgtDB4x/+QQd/+IPPvffqvW/zblV073qrvTmhrhJ6ZoHXlpYLDlWPvDl+f68I37Kj5TqXBREhG05JsiGDyRnD6R2yyT3s6DaSTEkHYybjEcezS8qyZFt6VrmyyC3zreFyrSy3yqaITtq1O+CwW447Om+/pehurvy6zc7BBqJPzbQ7YWuF4cAyHSet0s20ogGldoI4DixeAlwfC8614bAHjgxV7Vlv6qjKSwxNEwEvsdGZwftAUTZ8/cSzWJc8epLz6GnFfBlpwekk4+hozCApcL7EyhjZuSuY5zO3nRIjhKhca1VteL9TsoUOWNZRNu0vF7inF7hHj/FPzvEXl4TFgpDn0aWgbqIqLoS9Om5XWaQxUZmXtN66Lsjupdf8FnSuVkLsNyHdgLL3e3XcwSsYfFDxgA/GhCxNfxYcwk8dhJ7Vvf+f/yeT//pffRWC9yF4RL0ooV+Hhisg5Hfc6v5ZqqFVyDVocHFK/Gfb2GgHWFv/os5+3tiEzE4xyRBJRjEKIhszOTqjXH9NuXlCuV2Q52tWq5zZpma2cczGjtU2sC5gWwaKKsq6yxoqpzSujYJoQ/EO2ssKItqbN5J9xQQ32/3clB2pbzYQ7dZeOdw5d/6HQTtrHstkFEEoKGwLs1vTu+n66yrxm5jmfn/DGHaUrfOBbdFEWx8TZeBVHZV4ZR3VeGXlCcGxyRuWq4ZNHlAPx1PDu/dGvHNnxnBUklBitUIYgEmRxKJioyu46c2YdVSbjw4DWreAU1Y7q5yQx6FR35dML5f4xQp3ucCfXxAul/jVCt1sUFejOyWbHHwJFsRG2s+YXWLxwUHbiSL02dL9u2wIbTebKLuKSvUwDdr3UqHl4A6Q2A8Cb0TCaDQ6WBvTCEpvXL5Q8sasovsTqZv/+B+DD8EZrw0aPIIDsu5miz5nQRoXcC4cxHzrbmjOQ2jQUEcg0ozOkPGAPH9Fek1fc4Pj271+OyVuulzMKzEQScJockqSpAzHR9RnD6m3FxSbxxTLR2yX3zC8/IbR6oKj8YaTSUFeeopK2ZaRmtuUsNgG1oWyKSEvDZWTViK/d2GIfaq915gRPbD+kWeIDfZ93euOx565urKF+HbH6VWO86vGQPACJab0uDNts7A6JiixMcbBSOeWrVfslK6Unao399ykTxrtLYOQ2EstSkdVR6PUxkXpt+uobR/PVxJ9b6lqx2JZ8OTCcn4xZLOuaEoHE4cJNRIqRBtggEqyl1Sb3o6x/dK6iV5rq3V0I7ic4+YLwuVlq3CL5qBhHQdGw2azp9m25U7Rhm/aerAPPFeAqBMOaLi60vNMJthzGpfy7A8/8/edi4NpVXGtRFt6lVBLwx1UQap9FNKAiCNog0jjvQ/T6fTgTWVZpm8iNffmVULA6XLpa2s9og6kEVXX30WqRj7V9QZW99Ec/UqoQX0EoXgxmp8xfyQ9ulKf6S0ZmzAYn5ANp/ijezTVluH2gsHRVySjvyBmQGJhlHimg6algpSiEraVYV3A5Saw2AQWW2W1VfJKqRql8ewa3r6n+On//lfG1u+QS/RSjaAfnjHdtx16Aqudq3JL0wSJA65dlbT7OSMvv+TsilI9+JYPiq89lfqem3MHgoZxlkTqrqURm9a3Lk3jnIyKEMSiJonxBCZBkgTxbcS1D6hvoInycLxr/dmaFoDy2OO5mOO+eRK92Z6e48/Po5BgtYligqJsA+NK1Pv97rKd0xGTAunzPR+V6/s6+tpuv3iOrCBJWwm1fWwNveiZsK+Ar+yBA6pOoEHViTH+6idqK6O3ldAP8fjN7dv+0XrtjPcNqrVCc4VPPwAhF/pmgLLvCfVAKEqZfyk9Cb0yIyWIic1ZsSnGpiSDCcnwGJuOAcHXW7RZYf2aTLZ453ABqqFl6hJmteFoGjjbepbbhtWmYVM4iipQNkLj4lfloPGCC90gbM+mRPe9jb0Lkz5D1Yk8S9sJL6b1r3Nh0p+AyYlc+xlk76rgYnWy3kaPw03eViohSqqz1GAMeKcH2WxclyR/JdR3p5bTzp07Vj2qUeuTpQlHk4TT4yEnsyyKE2olLxvW24a68YxGGXdujzm7NWZyMiY9GsN0CmEGcgTVEG1Aq5JQ1VBWaFXECIRNHmdztnmMOlit8ZdL/PkF7mIeZ3oWiyg02BZRUKBxaDTSbdANjqIpIvYZNdsh6OqhRPCHkmZ2eZSJ0M497OyDujTVq1TcFaNgL6qNQIVq40Pwd+/efdsT+gGqoGuXk/nHH4fk//q/XKNai1ApUu1Wp7iDEOeRxkPjVJ1T2ZlEdhWuRjou+CruzoJ/ho37Iai2l6J9XjWB9AXP73aIfQ3PdUrqLBvA5IwmPyUfHiHJEBW7D3sUIU0tNhswHKdMp0p94ijLiu1W2RaeooRtDWVtKGrDthaK2lA0QlFC2UBZR5Gsd/vNgvTq1uf2ikQwB2ztlQVdrvmEvZOs0q/E9IX+ZnIT1aZ6PZa90rnbH1fb9kyMCE0TWG2aNn/IUDUe58Eaw6CNf6idUGvYzZcckK79oLY+/dlZjKri2ua494EQAoKQZZaTacb9u2PefzDlwb0pw9Sw3jQ8Ps85vyxxPnByMuLjD4/54L1jTk/HZKMhJEO8G+F1gK8kRhyso0w6LJZtP6fN3Fmu90Olm23b/4lDo6GsWoFB3QoVfHuezG752g+NspvL0U5a2FegPdOMe8HuQ2+g3Q6e8rz02H1hIp08OzWQGSQ1nctPdBHqObz4gOozm2b1QCWqJSKlNcb9/ve/5y0I/fAbRgX4b//tv4Xf/va3Tr1vEK00UB+0EDSqKjv/uE6cEPQFlRC/kIfcJDmPE95ibI+ui7NU3jvquiYvG6rCE9regE2ENLOMswHGxoG+4B11XVJWFUXp2JZd70hZF8q2VDalsskD2zKwLZWipe46+vS6NVxfojOjetP6oq8AAz/uzguiLLtbQxsX2BYOkF1wXZoanLOUtacoHc7r9QKEKyxd31fuAHRbl4ak9ZkbDRKOjzLu3prw/sMjPvnwhPcfzhhklvmi4otHa56c56gqt2+N+fC9GR+9e8TxaAClUrmaothSLZV64QnzBXp5jr+4iIOil/OYx3M5xy+io4Ffr6MQwTmeTQnj0J/t5XZjP41St3vDtq2CUgNJG6RHz3D5QM179TyKR7UiZrlW3nvHz8Su8U3JE+pfkWqM0U8++cSpMY2o1iJaa9wpJDsQ6qWrut7w17MgFCuh+OO/1Md+8E418mLRQsRTl2vqfE61PSdfX7Barci3UcKbpcJ0rGSJMh7ZSOGlY4xJUcB5R1U15EVNnpds8pLNtmCbV2zymm3u2ZZuJ3KoHDROqL3gg8EHofHS8uW0vaW9+lF1X8Wo9hkWPVh/tJuR0avmnq17gZFdoJ/0o61/xCnanREt+0FVgOkk5dZxRppZ8sLx5LKgLN0uEuKwx9QGFPaC8uIwZGjvh3iek1YCPh4nTEcJJ8cZt05H3Lk14p27Rzx4cMx7751w/96MNEtYritOH2yYL3JElNOjAXdOhpwNU9JcKRYr6m1gs5yzXViqRUVYzGFxGQdHLxZt5RPpt507dSiApt1N2sPB0c6ux5hnI1Z66cKHHRH96SzTbT8oRjm0c0LtQHL0ugw7h5fuvBxqWtQLdCBUJhGEnrtOvgWh7wd8diz2P4H5rN0fqSpZlrnKubo9MTUijaBJu5+XEKTdXQR1LhApOXZ2KTFXqCK4kuCrSMdd4eT3d/TPuEbqyn6JflzS6qV9cDTlmmL9hHz5Jdv539ksvmaxuGS5CQTJmIyUJHFMg8OIYpMMOziJkeLZGDEG5xzTYkuZLym28ass1hT5lrJUqlqpaqIBpxMqZykbQ+UsVWMo21mYqoGqiTNJTRfT0c1UtDdulwK7CwULrRN1gKCyA6LQqVdEuuSBdsHeK5YONiy8hEhS5HuqpoRnHQ7iv7PUcOtkwAcPJoyGCReLik0e3bOjbDr2iLprtxMsCBJpxwBqFPVRih1CS2qlhqNxyju3Rjy8O+a9BxMe3pty9+6U23eOOL415eRsxtHxGJMkHJ01nNwaUxQVRjxDIwwcyKKhfrSm+OqS4htHfu6oFzXNpoyy6aKNwu56Oz1/Ng0NfW82aePD+2EhUdEWes29frLwDfSnvJhS+y6l8LWvKYeYoHT9oBaA0o6Oi+cleKLQpwnULmjY3ZLSX4ecBi0UthpCEZLkKghRFMUbma76plRCUlxBeOecM6q1Ua0D1KLaIDLqUw67SqhvZGr7IFS3lVD9C6+EOtBt85hUCa6i3F6wnX/B+uKvrC+/ZDl/yvnlmvN1ipOEo7EgBNK0xqQ1Q+tbGfGQJDuJURJANiwYjmdMpse4akVTrWmqDa7e4lyFdw2NczROqZo4AFtUUFSQV4GyVspaKarQApFSNz3hiesri9j1RXzYq/JCrxK4Wgl110y/MfxTvAtsOzA6GSWMhgnrbawa6lal1pmYdg4UQUG97gZROxW3tcJwmJAkhtHAcnyUce/OmPfuT/no3SM+fu+Yh/eOuHU2ZXo0JskGGJthaovWMEQYZCmkRIFq0dCsKvIvtmz+bcn6jyvyL9bUT9f4TdH2deq42nZDo+FKFpbJDhN4njs02uu36Btzg7U9ofglaauQa1WIndKwcfvr78D2StUhUqhqDpSJc+458Pm2Enodp7C+skepqsqNs6zW2KzrFHIBMN20+Y5n3YGQ7Gz0o5P2WxC6frOk+Can2jxlu/iSzfxrVss5i3XJxcrzzcJShsB4aCINpA2NFhyHDVNZYG1KmgjKFJsOsWlKlp6hkxPA72jQ4AqCLwmuwrmSpq6oqpI8LymKiqKoKauasnbUtaesHLWLN2rXxHUBnDN4bV0bguB3g38czFz0qx2RPdVVVoF17llsmuhK7WMlJT2Ptx/77ARVqsaz3NQ8uSzJUsP5vGK9bXZUXWeMakQI7KtE7/ex31lmGY8SptMBZ2cjbt8acff2iPt3J9x/Z8LDd454cPeYs9mEyWCMJcWVSr1xuM2W0FQY05AmjtR68A3NoqT4esv6L2vWf1iy+cuS8ps1frlG66ql2MKVWZ2eC7U1Lc1mDs1Le55sB472Cm9cO2QnTBDIIggZKwQX58GaRqmaQO32se59EFJwghaobkWk8M+C0Nue0Gt/fPop/OlPu38eHx83TbGuUFOpailG+sM+vUyhsPeQCya6RHW2675u6bg6Rjtcc/PLz9x6ptslX914BlfSFHOqzTnFdkle1GxKwzJPOF8bNrWSJKHt5XjKpuBefYlvaqiXaPWUbDQjHR6TDk5Ihycko1PSwREmHWDEtNVoSXAl3m1pyg1VsWSSL6iLFXWxpa5zXF3hXI1zDc45QgidiqjtHZk2P0zQIO3fe9Rca9JpjJC2GTpiYqRBXgYuVzVfP61wPpC3oougYE3rMPFDbgWucz5QwXlltW344psti3WNNcI2d1wsK5wLMR/ImNjXMoLZBaS18eEhtLM+lrunQ959MOPDD094/70Z996ZcPtsyPF0yNF0zDibkOiAep3i1kp1XlM8XlM/uUTzNdZUZKknSQJa19TzkuKbLdtHW4pHOdVFTrOpwNUIviWtDNcPjtJWReEZDb1eR6+9TPXzMtGtNz1fv6fXPBDItHLshFgFteq4jvr1PlJxZR3/DK31n9nRcQoitQbNjTHr4Fw+zLLm6q/8/PPP34LQD/lYLpd+lNKoSuwJoU1UkJBoG2jYyR47N23fF8v0K6GwTwl/++gOT41vcly9pWkqGg+1TylcdEe43CoBbb3jPFVdUpYN+XbD5ijh6GjEaHzEaHqL8eweo5kytEOSYYJJj7DpGGNsjFf2DdYXmNEWO16RTZf4aoWv14RmS3Dl7jxFqyVP6ELbdrb3UVqsIar5goadJ5f2IhSsiYJuF6CsA/N1gw/K5cq1TgQcOBD8mLxGn9R3PsY55IXj0dM9jdNRj7YVVfhdfED8DNYKaRoVddNxyp2zEe89mPLJByf86qNT3n94xJ3TIdNBgsWijeDWDZu1p5xDed5QfbOl+npO8+gpulliKUlSj7VKqBqaZUU9L2lWFT5vCI1vs6cMYpLWdfplFvS9quBn56zer4TSOLBqrLRq0kgvd5VQX2DSp+NEtAiBrYgUrqrcDdD31sD0h+KNbt++7beLp41CDVQoDYjf0xfgQgSg5mrUt3Qg1KrjQhPnht4+9vWRBkJwhNbgta0l8GqomyirrtqbR0ME+zxvmC+V06lwMss5Pio4Pq45rT3Ra7bGaom4FTqcYZMxSELAgkmx2TE2nTKYvoNoDaEEX6KuglATR+1jKm7wDcHXBFfjXYlvCoIrUFfE52pkZwXtxSGAc7ApPVXu2eSOy2XDxbJmuWkoKo+/OmPzI69bHfMUdv1NT+PiMTcmpqR24OlDpJ5DK9dOE8N4lHB6PIi0290JD+5NeffelPfuTnh4OuLWKGVceZJVQ7PxlCslnyv5wlNcesqLiuY8pzlf4c/naL7FUGNtwBgl1B6XN/i8IRqX6D70zVjEth5telh9P5Mt8bJVzk9+eZLrK6QuR6gdVpUktgauglC/J2R2PUtFEadoocZsrWpRD4fuBgB6C0Kv63HVF+mzzz4Lv/71R406SqA0UEucZGubsrFn0DjVuglSu25WqDXM1ICGmuCrOLCqnmfmY1+SivsuSanf12seLF6v8B6uJZvaeSqRBGMSrLUkFjIbyGwgsXGaXhVqB+si9lC2uXK+guMJnM0Ct49rbpfrqGircpriKc367wxGU5JhpOokmSLJBJNNsemUZDAjHd0iSUfR6DG0m4Xg99HsvsY3Jb7Z4KsNrp7jywWunBMqwAdEPaIBIyaGbQaonbLOPd/MKx6dVzy6qPj6acXTRc3lqmG17ow7IwV22Je49kBf2cV/H8jTizVoc3d2svOwr+iwMXMoadVwuzwiVXwIpNYwHlru3RrFGZ4Pj/nggxMePDji1mzAzBqGpcM83rK9LHFPS6qLivK8ppzXVKuaetPgtjUhb/B5FRNHm9jfUYlApyEQXOgxCb3rqm3OXr1O9YCokmuPXz8f7GaQlivA9oLnyAuHzg7trF7lNXebWz38O22rKzGxCmrjvcXuxwG851k6TvYZUe0vaEQk16Abb20+Xa3cCw/PW++411sJAeq98YJGmbZIjUjoLp1ulqTxUU3VNbL313pod9WtPDS4fkvpF/Z41t/FJEOS4YxsfMpwdMRomDEZwmzkOZ0IpRPSaj9Lk1dCWQnLQpjnhkVhWBbKqijZ5A3zxYLjSZQCD4cDBqMJw8kp6eiEdHjMYHyLweQOw9kDkmyMDAdR5i0mUmyhjd7w0eVCzBaRDIJBXY2nwIWExgmhVtR7gnc7lVhZK6ut5+mi4YsnBV88Lvnqacnji5rV1lE1YVddmL3L/k+DFuqp+boE1f5Z62hJkei2nYwSjiYZ79we8eG7R/zmoxM+bYdNb58MGSrosqL6es36L0u2f19Tfp1TPS2oL0rcsiSU+2jrPWCYXT+nHx0RpcTtIOl1kmh9I301v9MtdPARTVcBmZ556V5+HUJbCdUtHdcXJuxfzwehEpECqPKjI8fjx288FfdG94QArHM+WFsBJSI1+2DKfaaQ6/T3EYSCxplrQfeU3A6E/Jt+SL6XO0jEYLMJg8ltRrN7jFbfMBp9w3R0ya1poDgz2FRYFib2hBqhdlA1UDqhbCCvlXXhuVw1PL5UZiPPdBiYjYXpKGE6GTI7XjI9mjGZHDGdXTCZLfBNsZPPJ6NjrE3jLFNXEfkyfjVbaDZQr9D6El+vaKotZVlSFzV13VCUNWUdKKrAauuZrx3ny4ZHFyVPLmueLmoWq4ayDnFo0whZyk82ZbfzkxOR2A9ygcaHaOcDjEcJx7OMs7MR9+9Nef/hER+8O+XDd6bcOx5ylhjGywq9LCm+XrP5y4LlX5ZsvtxQPSlolhVuVRFCjdAgOEyraovDoylIjOveVzdxMY3ev4emdDt3819MdK5wRZHQjj61LgntbBBXoix8iGtUB0KdRkp21LgSFB9EamOkDCHUDx8+9H/605/eOib8kHvBK9kZ8eQliTOqpYrkCIUITtoER1p6wrmgVRO0rL3UTewLZTsQitSOuioq5LxD7GBHXekVnlq+J+fmm6i2V6Xsvsvveu4uDjB2QDKYkY1OyEYzsmzAeCCcHcXjMp6YXcDdOo+zPaWJfThVKKpAXSvLjfJ4HhikgVGmzMbC8cRzNmu4vc45O/GcHuXUxQpXtoKEakG1fkQ2OsYmGSIaRQmuAreFUKK+QH2Ja0rKMqcqCoqiYLvNd44Mi3XDYuOYrxsuV47F2rHKHautIy8dZRVaFVxcPG07qKsvuV+/UVT1HWi6615TdtVGW6VK3G35EPAuPmswsNw6ynjvwREff3TMp5+c8v67M+6eDTm2hjR3hK/WrL7eUP5tSf7FiuLRluI8p1rUuG2D1h4JHtPGIETgsXSSau0iENjThILuLXZEr7ib9+6ffoWk8txr73nP0ZeJVBC+9XNu9It70Wv2zQpDi8BdFZRaGETPOE16ZbZ0xrGButE4mN0EXND+/9cQggo4RGqMqY5DqP/lX/7FX0O7va2EfvBKqGm8JkkpkAtSour6iZLRkDGWuWUVqJoQxQl9Os63/nFdpMPbR9tf8O1XVHNYA8NMOJ0aBpnh2BlOc8NsrCw3sK1iqF3jY6hd3VrwOK/U3uAU6iB4MQRjMIklzSBLGwbGMTAFCTniN4RqTrn6miSbYpIUCLF31xSEZhtl3b4mBEfjYshaWcWKZ5171rlnufFcrBoulg3ny4bLZcM6j7Sbbxv3QrSsSaxcx3795DhoVcBH/W5Hvw0HhtEwiU4K96f86qMTfvPpKZ9+cMw7ZyPGRtBFxfbLNZs/zSn+cEn+lwXlow1uWRNqj7ayUWl949D0uQvxM67fB0vgi/suv5hKqFPEZQYGNv5p93Eq2hOcdAPHjdfDOV5UNA7ix8QA1aaZTh280Gv3LQh9z/fftY80TV0DhRqzVdUioE0sYHS36XctCOWlp6zjSd5FGQSHuprQVUKhob9tE/llZlN7V1MXS/LlN2wXX1FuzgmuIE2E6ThlNDQcYTieCMcTwzqXdhcnbQ8u/ruqWzDy7XJmDcNMmAwNs5EwGhiyBKx4CJ7QBOoi2imZfImYISFmGuNcTV2VNFVB09TUTfRLi5lGUNaRBtyUsCmUTa4stp7lJlZDq42jrENMKgXSVEis2dFb3SKvPzIC9YsGuaKR6dRvolGSPRxYzo4HvHN3wrvvHvHxu0d89O6M9+9OuDdOGW8adFFSfrlm8+c5yz/Oyf89AlCzKlF8S7W13R6NDfPYSe8gKMqrVQ/bDgcVW9BfKOi8zE65BaFhrIQwrSOJD3j2YyTdELbzKkF3RW+bKSS1QaqA1iGEJn/wwPE//+e3WjPfgtC3vx/57LPPnjm4OnCNbbKtg1VAN6KhCoFAx7C0AWBF5dkWjm3padpqyIpDXSBIpHT8Doj8brZhH/72fKHJq4LVTwPcumCZZy0BXL0lXz5i9eTfmD/6nPX5X6i2y8hVmyxWRSKktquOYoKq8xIdDLxQ+yiHdgF8aBf5NqIgtcIghWFqGA0iIKgIjTeESinrEq8Or5s4n+SUqnbkZUNRutaZO24qoqUPrbcclDUtAEarn6qOFXDnMm13iqPO1PPbi4huBKzvAmQtDWx6jev++h98dHsfZpbbJ0M+/WDGrz4+4eOPT3nvwZTb04xpAHOes/lyTfG3FfnfVmy/XJF/s6W+LAjbpiUdZfff7vN4QMJh+FCfWnuGctSX6/voS9BfL6FMO/zRV4tg6KvjXjma4YWvKYhcyShqrXp0aGBkkYGgNp5D7wIaoG48lQvUjWrjgkRjXpV92rBgRAoDhRhb2DSt/9fRkbu+/OJtsuprroSuOaCTWm3Y4v0KZaNQaTyDO0dkH5Si8mwKT156ahcbuSqeEDyeCtNE65gY6eBb996en9oviYoLHldtKdePWZ//O4vHf2R1+YiqbgiSgiQIijHtl0CaEG16VAi9r726uVtMu+/tZbxBhaIR6mAwdTfv4ntecnvvuG0VyEtlWwY2eYx/KOvoJ9dZ+TQ9Q9P+WpAmss9//x4x43VSot17C73PMcgsw4HlztmQj9+b8R9+dcpvPznlg3dn3JpmpJXHPd2y+eMlq/91weZPc6qvNrhFiS8c6kNLuyXPB4bu97/MOqZvS6Abj2OvEtJMUNO6WDQeH6I0O3rG7fwtpXWM2qOMMQWQi6UMITT8t//2vATOt5XQazyd1x3cGtgAKxXdiEotgproX99WQocg1Lg4Xa/i444kVJimiiDk23kUu8v2PLzR5GdBwd7AZXc9tAZX5zTlkmp7Qb6+YLlcsSoNZUhwMQUOa+K8UGoFm+wtRmJlESubfRjdHoR8G8vgAgQfQQiRLoas5cg9TeNbU0dP1UDdQOnaSqeGvGzNTNsqx/ue+ajuh/1s2/Mxhr0jQp96+wnNR+6D5nQ369b5vxmJ9NtsOuDe3REfvj/j1x+d8Ov3Z7x3e8yt1JLNS+pHW7Z/nrP8twtWf5yTf7GmuSxoR+ii3EAtYq9JcOyOh76x69lPb+Gy0StOhyb+2YKQcx71sRLqQKiLm1Hth0SrA8lFJActgRpjbvKfeCtM+AGroh0IqY42ItulqqwRKY2EIGJ28ezOq+ZVYJN78jJQ14EQTAQh53BSIU2JdwXelSS+Qe1gn1ffp11u2PF9l6HRl6F6XodSDvTA8r/7fL6pcU30c9PQ4LxnUwS+WcC8DBRNwHuPEY1+ZSJYC9b032cPwCUqnbTX1+iAKEZ6R6+37ivGM/id8ab3++c7Be8F76MAIlwxKEXBiB7EgHfnT1V2xpD6PW3gv6+7fx9nHs1HIysW1Z2dk01qhZNpxocPp/z2V6f85tMTPnx3xt2jAWMX0EebOPPzpzmbf1+w/WpNdZ4TVjVC2LsYtD0ews2000uFUtykItPv5w6/6TX7ldmruvodKO54Da95QG1rrPitoLuekBBaUz/nwDdQNZ6q8Vo3qrEftE/ebLvXpYhsVXSLCTne1P0L95//+Z950xNW31R1nAB8/vnn7oMP/t95lunaGLNFaTCm9UTQVv4YIwC2pSevenScibY0XhuMq/AtJRd8jVGPXB1aVf3Jzo9835VQaEP+ouWXpXSW+RYeLZVlEWiagIjubEdMzBu7cngOF5Gdp0iceYgxA2EPPiHowd+fCcXcVVV6EPdtowkz12i2nvmE+hPtn0tvkxC6CqgFiSQxZEPD6VHG+w+m/O7TU/5fvznl0/dm3DkakJSe6tGW9R/mbP71nM0fLym/3uDWdescARYbc6J6lsxvi5wfrhLSLFZDpG0lpAqtEOEaOm5fCSlBhFJVC9A8BCklhLr/+k+ePHnjQuzeRBCSa/5u2oMe4KQMYbMVsbkQGgXt7C7aKAct6x4ItVxs9Ebz+OAwLu7+o/9Y1YoT0l8kiR06abYGOgv9aPYZlWerPDb9oydbPAmmAwe5aXXt/6E9h36hY89CiDM7Xd9oty9t52Jik1ZbPy2wdh/HgPTprGcZpv3v/+l5kxnpZTj1DFk1KGlimI4S7twa8cG7R/zqo2N+8+Exn9yfcGeQMFhWVF9t2P5hzvJ/X7D545ziqzVuU6EELIIhUm9R9daeMN9lC+lbMHrNK9ezINSa6rq9Kq6njGsd33cMiBfVCtVC0QInZVmVzQvWyDeOmvupJ6tet5xJC0Ie4G9/+//Wv/nNf9167wtjqFTVIbpLJYwSbZW8DOSl17Jx4hx469ub3uObCldtcdUa1+TYcLwznu/oMOW7CRVehyLuZV7z+ufoc6moTu2kfSdqOkpNDhf551UYevPR2mmv2icYEwchVUF7fSQxuqPW+jSbdhRfbw8o1xWrB+us7F7jujv1Zb7/qtTcdd/v2oudS7IxbYS5ahu+FyXY02HCg1sjPv3khH/49Rm/+mjGw7MRM0Cf5qz+tmb7b5ds/njJ5u9r6osc3TYtGWoObXNC+2bCFaWbfA8HQr/lgfgWr3kTpfZSJ+mm17yJTnzV1zyQse/jG8IOhNrrNYQYXd+6JFRtomp0+tfWtFSIhpZSImxFwlZVqi+//PIAhDabjbyowH4LQt8PGMnN//7PAX5TA1UIWik02pOWhoDsVVaeqvbUTnCJ7sPPXB0jC6o1ab1F/S93aFXE7GOF24rStLZXqYUsAQ1yCCTsG+rfbm/xHBpNXmFxf0P2fwcV2y6Ab0/BDTLL0Tjl/u0Rn34w4x9/fco/fHrCe7dHTBTcN1tWf5iz+tcL8v99QfnFmmZRxYvdGBKb7A2s+vZt+rbyed1MwsHDCpoImgoaHY9Qib2+nVVP06o7r7r8xw2kE9WSmKZaqOqBNdnbSuinUyXRNI0zxlSiWihSxB2E2vbmMz6wl/tWnrIWRom2AWgKvsZVG5pyias2BF//UhEogpAxGGN3X9YYrNn7LzqzPw2K7iqQoM9WIAeVkl49ibqzozG7uYhnn7PfX+7ctHq+ZLxxjLhp5z86R49+LtAos5weD3h4b8KnH834zUfH/OqDGQ+PB0wrT3hSkP/xktW/XrD835dUf1/iVnv6TTBR+WbopZK+BZ/XuhLddGyTDoQgJBGEIuUNzrVOLm0ldOCUsFcKOYECkS1BihB20/S7314Uxdue0A+4xdgtT59++il/6qWs1nXthsO0VDEbUdagZ6Cj9ryYoIgLSlUHzUtPXhgZWcWYtgnuKly1os4XZOWK4Kqeo/YVJ+AfWZvw/ajmDjVA2lvNIwAlGJtgbIqxFmsN1kTrHtNZ+OuzZ+bgNWXPTHRy6K7/0/1o53lphJ0B89U7SnrljV5ryaaH/+PlDuL1x/C7HNuXeM1ObSkiWNPFkbc9IIVBajmbZXxwf8JvfnXKP/zmlI8fTrk7SRluGoovN+R/WrD5t0u2f57TPNoS1nUvrzTm9qjX3jHUwwr1JhrqpbZ7B4Tuj/KaLzNweuMpuuFnD9Rxr3jeteOctU/TCGpa+i2DkAohISqrW8ulpgmUpaeIQ9fqnGrrjrCnn6EB2Sq6QmSbpmnzLUjItyD0Oh5Xs4WstQGoDLoBWQElqkP2WcKEthoqqmjhM8mULG138r7GVWuaYoGr1hGEfj7n+OVpBJGYH5Rk8csmbRUkJFZbGTa7gLhujTXXGy/sNnWmVQ/Yq2vyNev41Qinl3IleENOU18FFyXqewqus+D58P6E33x8zD/86oRfv3/E3UlKumko/7pk8a8XrP/XJeVflzRPc3zh2sHhVv3Wazq9LXx+5EpI2FFxIYFggba/GUJUmMa1KFDVUZSwo2pboQpCrd0cpJpNWwm90VXPmw5Cyg0xc8Ph0IdQlyhrYImSq+oUJO3OWAhQN3FwNS+FchhdFYxRtAWhumjpOPdLo+PaKkgEa1NsOsSmA2ySYK3ZmXxaEzHdt1Y6neS6m3vcgUinYGvl08ZKjFGxsqOiutA2H+JAsfPsOPG+YOGHWEQOOrj6Oo5sf3HZg09oVXCjgeX26YAP3pvxu0+P+e1Hx3z8YMqdLA6gVn9fs/7XC1b/esHm35c0T3JwrlXomL36rX29twj04+7lOs51D0JCsFFkw1UQqgJlo/tE392EUNCgVBrMWowsVFkrWl39lW26wI1r41sQ+v42jwen+mqsw3a7dYPBIBeRpQgLUV0Bx0DayVLbRqBuyyCbIjAbGbIUEhShpqk3mHJJXa5xrmxl2vZKyd6PdXi1Lfj3Fd/wqj9743OkRzOqIsZEuXWSkKRjkmxEkg5J05Q0swyy6BOXJl1lw05O2jVTu/5PNzuUWMgSIU2FQSpkrWloB1pBldpBWQW2lRKaNpytU7EJO1Xet6HFbjxuIgfUyR6IrgyzfovXfGZz3Kn9diq4uOC41s59lBnuHA/46OGU3/3qhP/w21M+vDfhzBrsecHmz0vW/zZn+8c55d9WuIscdb5Vb5qDZtzOAUL14L3Jd1ibbrp+frTXvEGd8jL3wk2v/zKxES98Td1LQbUPQBn4FEKiiFFsu/noQGhbRlsqH2LLsxt3CEE9SGHQBcqFpGZpgimvHAH97LPPwpVLVl5Qo70FoddRCc1mM19VVQmsFZaIbIiWPrszEZTWcbmthGqYOIMVBW3wqjTVFldt4+CqbzBJ9nOjX597ZwsgxraV0BibjkjSAYMsYTIMHI2UolaCV8omWu7EtM895dZuAklsjElIk8Mv23r6qUbXAxHFOYlOB9fcSa/9guo3qfT7Od16zTbqGRWcRgru9ukguiB8csw/fHLMx/cn3Mos5rxk84cF8//5lPUf5tRfb/DLChqPQSIFZ83eaifoj+4A/rYSajdjhhaAhJAKvlPFdf+FOJxaVBpzuNqcM2nvnXYjEVS1DJglqnNj7appmppnReFvK6EfsMi98Q67c+dO+PLLLytVXcdqSDaINmi3g4lN8capdjxsUQmNU1IbMOoIISBNiatzXJ3jmwKbDdvYYp7VCsvP6M65snjZZNCC0JQkGzMaZByPaxrvSW3gZASNxjhnawNpCzjWHA6X0pst6ow4o9u2tnlDcWPQ+H1V9boTuq5qBqJTg+6GWnfOxab33FfcUx5ScK0KLsQZkM7bbpgabp8M+Oj9I373qxP+4ZMTPro35lTAPM4p/rxk/a8XrP/3Jdsv1oRl1drvtKmcbUR0nCl+Cz4/Mu4cfq8dUA0DwWfgjXYUWzuUHIdTyxaEqroHQm1vT4ME0BLVVTBmbkXWf/vb3xzX3ybXGT2/MSvUTz3K4drHZ599dvDvf/mXf9FPPvmkUmvXRnWBsAbqnR26xoUggpDXTSmSV0LVBAZJADxBA64pqestTbmiKVckgwk2HbVn1PSa77uJypcv219VdfMa0lev+9ndFdyjcIxNsdmUdHTKcHKL6dElt5qaQVpzNlWcN6ixGJtgjZJZbSsdE28iostC0xqO5lVMXs3LOCVe1sq2VPIytDlEMQhPe82gFyqvb0yJlZsXiy6ZdB9r2Jl0tb0UQSUqALWnUtKrE66vAJJRgRlHAYLfB9HdPsr48MGE3316wn/4zSkf3ZtwJgKPtqz/sGD9vy7Z/HFB/fUGXcfxkIPYhc73iOudIL5TVaQvvn5+Cq+pL+Pl+KoGfy/zPuXZKldarx2VOKDqh4IfC34gBBv94oKLz/E+josUVaAo4zC989HpxRpRVSSoeEEK0KVxbuEiu+MB/n8g//n6d9rHxDfGnOmNDrXrPy+EUCfGbICVoFvQuvMC7MDD+bgAxkoo2mX4EBBayxjX4OqCplrTVGsyd4ZJBrsBzte3P/+xH4crrBiDzaYMpveYnL5PXW4RdYwHC5yLAh2TGIw1GInChK4SQqJTdtlE8BHRKGJo6dC8DKzzwLqIsQy121uVXJ0T+r43+HvnbD3ArMRGo8lD2W7vPci3P5q7aqu1yhlkhlsnGR/en/DbT4757cczPnxnzK3UYJ4UrP9tzvy/P2XzxwXN45ywbTBBgTi/JUYOej9vHz/BSigRwlDwI8EP2JmWBtfqrn287qtG2xiSds44WlFJew82QKkhbL2123//0592ooTPn39Fvo1y+LGuhRBCpaobYI2EjSqVBvVRm9XuzjsQaqOg6ybgfcBIiDSJq9sYgwhCvilIsknk3n+2j8OVP3rGCengiMnp+7imBPUYCRRLpanWoDGTBhNptKDREbuqldorlYuVz7aETamst8oqV9ZFYJWHfRVUs5Mod6q5A67hO9KefaNT6N7nXhSQWGGQGQapIU0NVgxBlca1A4Qu0mdd+upzK7PW4LafiKmqMdTP773gTo5S3ns44TefnvC7T45bCk4wTwqKPy5Yf37J+o8Lii/XaNFgIvyAbUFoZ72jPzNa+M3fxu1G465UQtpWQr4RPFGlWzVK7VQbj7SqUNMnyIFSoRCRQkOofs7H7k0FoWcSBG/fvl3neb5xVbVQWKnXPKg6REznc+k8UtQqm0LZtg3BxivGeLwPqNQ01Za6WNEUS1y5IR0cISbdrdPRRy68sob4VeMeXub738drxm+ZXe+iU1algynTs/cxSRYpttAgNBQrqKs8Jqg2nqJWisqQ17BuAWbbct1FDVXTJZyy2/3Vbfhcnz3bS1Ov+DLpi7d41zEufUAzXXROgKZ1KRARksRwNEk4maaMBhZjhLoOrAtH2Dat43pc6a3ZR0TcOLesipjo/hBa2iW0PZtBajmepLx7d8yvP5zxu1+f8MmDCbesYJ7kbP64ZPOvl2z/uMA9ztHCtZ+mjRRp1TXaqj2jqu8lUeh1JM78SK95U+zCjXEML5PJ9x1eUw+Y2nYYNRHcUHAjwQ/bGaEOhAIUMQ1YKxdHE1p1djtpHAiqDSJrVNdBJPf+0MLlyf/xfwj/8i9vbOXzc62EAJoQwlZgGUJYoRQKjcCgkxT71g26WyS7SigQm4XBNTR1QV2uqIslTbUhczU2+zlvOeUQT1sUsOmQUTsvFFxBsT4n35zDdk3jK4qyZlMELjew2FoWW2G+8RGIqgg8jd+7AlxnctqmH7+eXUovmbQDttCy5NZKG4+dcf/2gNvHGcPMUDtluWloQmBTtN6Dqq1rg7ycPqFTSHWu2ApZYjiepDy8M+KT96b86v0jPnxnzO2BJbko2f5pyfx/nLP5w4Lmm22k4Gj7kC0F94wH0vUw+PbxIxHZ0gOlkIAfgBtCyEBNNJB1IQpyiirS1Y1H40SC0Nl+RN8EClVdAisRKbMs8zf86p8FH/umRTnc9P/0P/2n/+T/y3/5L9uBMXMXwsKI5EDTDU52yqyyJtJBPRDq7NWD97i6pC5WVPmCulwzdFVbHfwyb/pseMRgckY2PiUZHGGTAYpQVY75yvHVheHxEuZbw6qI3nyVi+4UoaffMdJWJEZ2cQzmmsim70K79W1UO/Dpp62aNvZ9Mkq4fZLx4f0xH94fcTZLCaqcLxry0u97RxwWvM81aW05uF0PKESjyu733bsz4tMPjvjth0d8dHfMLSvYeUX11zWbf5uz+eOC/Ks1bOpWqhsPkrSKw7cquDeDnlETKyGfgc8gpHsQ8iFS1kUlHQj1otulZQK0QViLcqmqC0mSYjQa+R7oSOucLS/xdt6C0GsEnx3r8s///M/y+9//3v/n//yfwz/90z/leZ5fSu3mAVkDFdLeyKEFoQbJK8hL1aJSaZxGHb+CV09TF1T5gnJzQZXPcXX0Q1VNeEabIC+/MHxf6rhXfc1XUdbt4xUO/59JBqSDKYPRjMFowmAT6cmqCiw2yvnSMs+VyoVdM8cawVx5bdPFMcj1w779zLXrLoLnsTVyVRYe2nTStg9EgCQVjkaWd84GfPhwzG8/nPLRgzGToWW+btjkHucikDatkqkL7pPW+lr6qshn+gGtIW77d2thNLDcOsn44N0Jv/n0mF+/N+XeNCVb1pR/X7P5X3O2f17SPN6i26iCM60Sc9cg6E8Bf9eiXH7gO/V7fs0bB06/w5v4bq/ZN9htASgFl4FLicq4dp1oGiirKNgpKqgbNIReTzZmZpUSmCs8NcZcqOrGRTXQ/pKP6mDhZjn2VZXcWxB6nZXQv//7v++yhT777LPmH/7hH9ZgVyJsgLKvuAqKeA9V0/GysQE9TNpdblBcU1LlC4rNOeN8gWuK2IiXX0YVdPVzxirQkg6PGE5v02xvUW4uSJINYuLsXGA/axPTTvdDqXKFNHgdd8SOpdq5Ru9VbdbGwc40EWaThLunGe/fG/Gr9yf87sMp794ZAuz6P9vSsy08dRMle7bvyfY8BmQ3DxR7CdZEyu/WLOO9d8Z88sERn3x4xIOzIeOto/hyw+L/uWT9v+fUX6zx67qV+UYTWbFy8JneMnA/sRWpl2G1uw5NHFD1meAT8FZRE6PVNYBrRxY6Os75fYCj7G4UKQVdABcSwsKoFncePOjHN0hx6PXLm07Vvek9oWeszEejUdk0zVZVt6Kag9YgaXfSOlquaqKZaVEFhqmQWhN3zR0IrZ9Sbi9xTb6bx/glsdyqAUJAAWszRpPbuJP38OWCYnXOcLBkOio5OwoULoLQcgt1087fdE4K5tAWpXMp+DY+bSLX32U7N4I2FqEDwywVRgPLZJxwfJxy99aAB7cHvHt3yAfvjHhwd8jxNGFbxMpnW3o2haeoPM631ZyJm5Pnvt3+Rqcdfk0Tw2yS8O6dEZ+8O+Xj96bcvztmmgjhSUHxxZrNHxds/7JEVzXiQqy6TBenwVsK7k24U3aW8G1kwyDKsn2iuyHVqNCW3dxcFOwIzu9H4/bXttbASmAuqisjUt69e/dnfRG88eq4uq6vuv834n0hsFRjlgbdoDoTkQQRkbZPUDbKplRdFSqDFMYDgxBQX1Llc+z6KeXmgqbc4H2zG1p9Zkcq340K+L7ou+/rNaXtbcjOg8yQJCPGR/cQVUKdU67OOVovuVvXiFSMhoZhCl8Bl+vId6soaoAgr2xIKs/5xi7Su9ed6bYI3TwOQGKEcWa4NUu5d2fIwwdDHt4bce/2gFuzjNNpwnhk8UFZ5575umGxcWxLR+UioFmzy3fdod3zqEJp+1xWhMkwCh/evTvi/Xtj3rk1YjxK0NJTr2qqpwXNk5ywrMAHjLRGpAchdG8X+h99gXnORqi7T5CogAsDwY/AD8GlipewAyrvhbqBvBLy2BNSF/a0X1tvi0IFrFCde9VVBdXvf//7g7fz+Yvf7huVMZT83C6e6XSq28WiDKpLUb0QYQmM2nVphxeNU92WyLpQxgPixL9Rgq8J3sW+UL6gKdf4poTh7JfAxV1ZYQ1iLIkdkwzHmMEArzVVtcKFmnRgmUwumE5LstTRRRovc1o6a5+Vc11g3UsvBnpIDz4DCO1rJ4lgUkOaRBC4dZLy4PaA9++N+OD+iId3htw6jnLsxAq4wLoOXCxrzpc1i00UJjS73tY+Vvw6TOjmgULYZ/ZkiTBqAejO6YCTo5Tx0GJbes25QFN7Qh3AhVia716sK7r1Z2jY/ybyAS++OLs4em9jH6gZQJPFPnMgYEInkBEaJ5S1tJUQ0bSUXRSKRPpBSkVXKjJHdeVDqG9ioK/5+1vvuB9wc7I72G220O4k3L17V59+9VURQrgUkccg94ETYGRaNX4EISII5cp0KEyGkEhA1RF8oK62VMWKqojO2tnoGGOzXez1m73N02v+LvTTtCKg2IMfy4Yzpg9/jbMBM0wZHo0YPP4TNv2GwBYvkAwtT5cwXzrywlM1inPxuMcqITb7jbzkHa/01G66GzjVHriliZBlhsHIMB2nHB0lnJ1m3L2dcf90wIPjjHvThNuJ4cgpiXq8EWpgWwWWW8cyd2xKT1WHSMV1c0by/F1x58QRgpJYYTpJuXd7wL3bI27NUhBYbR2Ldc10kpEAdmBJxilmnEJm0crhg0d2O2PT0nJcr9Z4+/j+FpAXAJAc+KvvK/DWDS6KEYDGKHUqVJlQpwZnovJNQvucINTOHNBxXSaXaculoOoEclXmGsKlhUXTNNU1Z195g2x53nQQuskHaff9zz///GBX8Pvf/15/88EHuSTJE+Br4IHCPeB0H7AWQWhToqscORorx01gYAIaPBp8FCgUK4r1JeX6gmx4TDaaYZO0vTDfQEPTfs62dtWEHuyjdiArvTKg9xjYGbO7v8GqYJ3Dr9ZUyZLbkwKA8TjhaAxfmcCToAQXaFR3bTUxghWNi2w/pqFfhB0MA7ZCg+7mb+XPnQ1OamBgDccjy9lJwt1bQ+7eHXD3nSF3bw84HVpmAaZ5YLCp4vnPLDo0NImhcIFt6chrT9XEKki9oonsP/8VFaTud68AeB/wHrLUcHqc8fHDKffvjBBgmzv+9miLTw2jYcr4KCWdZWTvjEnvT7CbBvc0JzRxODWGzPbPwSEd+NKr6ctM9r7qCv0TeM0DhefNEroXbr70MLr12uHoPhm7W2CkU122Yx0oDqgtlKlQZkKVgBNBVXaVkPNtnlmDFjVSO8T71iiXLlQ1esUBF6L6tFwsFo+Wy7r3qa4zKtVvU8S9BaHXTN821uYpnKvIN6heoJJ3TgDdGXIeLWrYlEQPs0bxacC0y533jrrcUmwuyFdPGEzOSLIRNsmeyRyRN+XcH9oTPP9dtxIz7Z4fWorKK2kuDMsx1XZKuhwwWBlmNSRDYTw0DDNBiJJ2mwo2D9Ejrhf3Hdo4g+tA6PCMXnH4NoKkgsmi7c4oM5xME26fpty7nfHwzpD7dwYRgKYpI6fYixoelYRlQ20EM0vxpxnVyJL7KEgo6kDj9aVT3LUXW979zDCz3D0b8vG7U+7dHjJf1jy9rJifl5QKs0nKSTLlZJwyfHfK0abBWEM9SanPC/ymQZ3fm8sGfXaRfPt47XSbHtQ88RHafwdVvLSKUJRAoAEqE0GoygxNCt6Y2Ff14DWa+VYu0nBVIzReCNFgIxpixIuoAt2ImLkRmd9eLrd/g84527wEAL2thH4KNC1AmqZFEsK8CeFcVRcaKFEUo9EfMFo5aWeymVdK1URaJQVUBfWeutpGp4DlY0azuwwnZ6SDyTM7sx99kFVv8JzW7kbquSK8LJUogtYNWtaEoiCs1oS8QEpHyAv85RPk749I/rZkuKgwSSA9EbIjMKcGNQnZSJgulfnSs9548kIpyujd5gOx4pBnN3nau8VEdEffpYmQZobBwDAZW6aThNlRwtks5fZJwp2TjHtHKbdGlhOByaohmTeEL3KaLwvC1qHTNO5mj1Jqr2yrwCb35K1C7kX9q70vHLt+lzVClhhOpin3bg95750xd84G0acuKBfLmioot2YZt0YJ6cmAyb0xZ6nh6M6I6t0p+Vcbim9yqssct27whUdD6BKa6Obd3vaKXr6YepZaO6yfristDqk23YGP7/7sgZD2QcgK5UCoh0LIFDWKCUpwkXarmxaEXAQgH4Tu6u/9+kpVtiKsTAirz6Dp/X/DXoOj3+GQvAWhb3F9vcx08EHFPJ1Oy+3TpwuS5MKrzlHdouqMSNpx/E0QbKOS15DXUFRKPYhbE1UI3lOXa/LVEzaLrxgd32NyfJ/B9LRVxOy91lTkJXbOr5aCerjwveD5HRDuAtq0FzmxH57sjNTEPN+QVZ0j5CVhs8HPV/jzC9w3j/FPLgiLNWG1xi/muMffkFx8zchtsEcO845FNMCxwZ4ZZlPDnVPlYmG5uHCczz2XK88mF0ofiP1/aSm3fVnUvXUBrI13X2aF8cAwm1hOTzLu3Eq5dSvj7CTjdJZyPLYcZYYjhVEZGF7UyEVN+Log/C1Hn5YYI9j3JpgHI1xmcKkh3zrWW0e+9TR1NG81rUuB9N8MHUMWSUTfihGMRBn40TjlwZ0R794ZcfdswGya8rT1oyuqKPv+2xcbTscJqREeHmfMPpoxfTjFfTBj/eWG9K8rtn9dUny5Qc9L/LpTKnRVqe1d7a9Azb3q8vSGvKbebDDXiwLp35s33KfaGzZF8S3oBA24tuLxLeB4ic9R3YfUOSNUqVAPDW6k6EDBhhgN0ii1F4paol9cVMVJiItGLJdEQalQ3QgssXZFlm36b/Gf/umf5LPPPnvRFkTfRDD6uajj+rPk+tlnnzX/9PHH+QaWwBLVbZQ+atrdAUGRph1cLWso60jJJWlULwQNNFVOsTlnu/yG6fopTbWJu9ODhv1P4DyLPLN9l+cA2zPfCx51Hq1qwjYnbLaE9RY/X+CfXuAePab54ivc14/x5+f4+YKw3qCrHMqSJPGEE8FVShqEqcLgxHI8NpwNDKcDw3EmTAaG4dAwX3s2RdiFefnOabrXmzISK4xBGp2uRy0Anc0S7p5l3L+dcfcs42yWMhtYhgJpoyRrB08rwtcl9VcFfFmgX+eEjSM5TjF3h9hUkJHFpYbSK9vCUxQe5yKY26vCiauMZlsFhaBkaWvLc3vIB/fHPLw74niaYq3sWkqNC2y2jr9+tWGYxTQjp1Me3B1zcjbAHmVks4zxLMNMU+wopRisqR/n+HVNaOK+W7oy7W0p9NK0ibyAarsqMPAt4DgUT9j96XdD2b05t466t4JLDW4QPeMkbTezQfFeKJsoyy52NNwuvr7ziwuqbFVkDixQ3dImQ3ePq/OQP5cq6OfYE9qdpDvGlHkIS4G5QkxbRUfRWF0Iimm8UDWyA6GqUTIr0QVZlaYuKDYXbJePyddPqYs1wTWYzL6gcPs+abbnfP8ltM/POCAAWjeEvEDzglCWaL4l5AVhvSUslvjlmrBcRxC6nEcgevwU//QCv1jgV2u0rBDnAYMRi60T0qAYp1AF9J0Ud2wZjSyDoWF4O2E8sRydBRabmClUlIGqTZas6zibgwjGxkHTQWaZDA3jsWUythxPLSeThNuThNtjy+nAMvPKaNWQlB5ZebiscY9L/KMSfVyhT0o0r+ISUplYDGYGBgYnQumUovRUlW/dtVurHrl+YduJEYLinTLIYq/n4d0RHz2c8PDOkMkooaiiyMG1AWaLdU3TeILCtvLMtw3vbx337444HmeMJinZ+0ccTVKGxwPykyHbPy8ovtpQzytC5WKPSGKAWqRYZV/x/uJ7O/3OrF6BoT3I7AEn7Cqf0NJuexAKsRJiXwl19Ju2AZk7CDKKT/decZqCJLG60QBNYygq2FaxGnL+QJbdPUqES1UeC5yr6ubzzz/3vEHWOz9nEHreSbgacdv5yAHw//nTn5qvP/546Y25CPHEzkXkCBiDSghiAkjtYpmcV1DUyCARkgQ0BNQXlNtLtqvHbJdPKLZzmmpLkg2vLk296kJfOBT6QmruWgk1O4PMHT3T9YKMgLUv/r2AFhV+scQ9vcA/OcfPL/EXl4TFAn+5xJ/P8YsVYb0hrNeEPCdsc3SbE4oKrWq0qolOSS3hoUqSe+Q8ECoPa0+49CR3EsytFDlJSceG6cRy+1ZC3iarbnIl3wa2LSA1vpVcp3HWZjK2zMaW6dQynSZMR5ZJYpgojGtluGoYrB1m6WDRoPMG5jXMG8zKobmHsnVJR5DEQGaQgYXE4L1Su0BZBao6xORTWvPQA2W0XEv+iMAgtdw6zvjg/piPHsReUGKFbRFpvqL01HWgqDxV7fGqrAvH40XF358UPHxnzMPbY+7fGvLO6ZCzD2Yc3RqRnY2wkxRNDUHWuIsiHtsQUGT3HuNW+gUybnkJWkxfcV/9ml5Tr33Nm3yf5NrKR684nndg43a0muK1V930aLju71d7Qn01neg+BbhzzA5DIaRA5xVH1PLUjZKXwqY05LXRxsWNQ1TF7fShG1S/EeXvqvpNiCmqof9pR1EF/BaEfmIVz9V+ojx58mT3/f8M4R+Gw1yqaikwF2Pmgt4BHYJYQFSltfARzatobDpMhZEB0UBwNbUGis2cYnNBuZ1Tl2sGkxNMS8kdLvzfUyV0MDQqN6vZrlRKehW8QkCbBm0c2sTqJyzX+KcXNF99g/v6G9yTJ/gnT/GXl/iLBf5iQVhtCHmOliXq3V74IAbEtHTkvj+hCuIDyQa0CISVh4UnLDx2FRjeU9I7KUcngh8ZmqGhnMD2CDZ5YJMH8jb6QURaissyHVlmQxPBZ2gZWiF1YLceuWzgaU14WtE8rZHLGhYNsnZQBvBRBh7foo3HbWiRoUUyA4kQvLYplwG3yw56trDchZV1g6ltDy5LYx7R3bMB774z4t7tIcPMsNo2fP204KsnBfNVTdNa8gBUdeBiXrHZNpyflzz6ass3d8d89P4R5SfH8OCIszsjksQwCrobbC1rT9OUhE6s0Mrpf26VkLwS4l1Hsz1LtYUexeZaaq1PsfleRaQ39Zpa89oO9BXwBjRrAWgQw+za/g6g7WwQFHXnkiD4ELcMcigL3yI8RvnSw2PXNNur1Nv0ZzQX9HOj4w5OzBWLcwVqDWGLMUuBhQpbQY9AbEetB4XaiW4rZFMKw8yQJkoqivoG5xxlvqLYzCm3c6piyai5TZqO4oL8XQZX+wuI6hVb5h5p/BJUm6ruqpRQVWhRonmOX23QzZaw3eJXa8Jihb+Y7+g1d3lJuJjjV6uogFttCU2JUgGuhVWDELkG6bxpdnMs7UEMkSYSp7ABrRVbK1IpSamwDciZh+OEMLTUqaEwQpEJubWUCk4jxZRaYZwYxokwFhiXgWERSBvF5AFdNoTzBndeo+cVetmgqwY2DtGw3xX7VtVgO0sFE6uhxESvLwGvivN7Z4cOhPp43u8D+fbERVeGhDunAx7cGXL/9pDjSUJTB75+UvCHv6750xcbnl6WeK9MRgmD1MSeUADnPMtFTbFpWK8a1oUjV6VU4f27Y86GlvT+hNGmIWwbdFPjNw2+9M/viv5MqbabejvhSl9He2Cyr3IiveZbEAq9iij0qp3Aocr0wLSp/fbB4GLfqmcQHbOjLDv+vPdQOyib2A+qolJuR8dFBgUUKUDPJY6UPE2SJL+Kx3dfHoReR9zgLxqEblLHycsc/M8//9z97pNPcqd6icg5ygKRYyAzggTT6iI9bCqjy8LIMFNGGVgbIDi8U+piQ7G5ZLt6Qr5+yujoFmZiMcmgxYpW7BAUEb3ijHvDm9vRal1Us/JMZ74FopdR34XVBr9cxUpnsSQsFrjzeaTczi9jL2fRVjmbDWGdo3lOKMrYH6ojgKl37a1tgHQvEcbuaEcJraJHrpcMC2Abxaw82ihsPDxtYJagJwnhyGKmCXZiyYaWcRYbu6FVpSUISRPICshKT1oE7MYjaw9rh6wcsnHYjY+UW+FiH0r7pInZl2naA0vdX1XKfmixPx907VR069QQNDojjIeW26cZ794d8t7dEXePMzJrWCwrvvom59+/3PD3R1vmqxpjhNOjjFsnGceTOOy82TZcrhvW24any4oiKFuvsU+WOz56Z8zdScrg4RSWFe5xTvk4h5LekvuCLvx1N5TcMOz5XfZS3+E1b5y0lN1wwbVoe1XNFnoVj1c9+L4/oNXoyav12Z3sNaDXv//23aAYXudGgpuAHyohiRuxEPXcNF6oW2FC1RgaJ3ht3eVbik9VvcAW9CI4/8Sm6eUf//jH8upb+v2rF5Jvoxx+sEr9BXtC3zS5WnuO6jcID1W5C5wYI7t1tPGieSWyLgzTgXI88gyl3VcpOFdT5iu2y8ds5l8zmt4izSYM0uHL03FXAUmuNlOf20Taf6CwHxXQxqFlRVhtYo/n/DJWOecXUVb9+Cnu0ZMIRJdz/HKJFgXaNOD2zuC7oUhATNoDnxsowmdWEQF7pXoApNGY3LXx6LkjDBv0yKKzBI4T7EnCcJYwmFp0FPs0iGA8SBUweUDWDl063LzBLxyydkjhodmDinQBdH3p+S7NsG1fdZ3poDtg6hWcB+44V/UgB2KEsAehO6cDHt4d8c7tIUfTFLFC6QLr0rMuHNsyunEfDaOTw8cPJzy4M2KQWVabhi8eF/z16y3fnBdcLioaF7BeGRnheGg5fWfC5NYQuTuiOM6QVln3c6qEXpzMptfSbf0+Tp9eu/rVp9mue93Dzq689DKjEgPr/BDcCNxACQYkgHdRwNA4qF2sgGovuBCVcfbQnqFS2IjK3Bpz2Xi/YT8fdLW19X2uk29B6DVcx9dWS6GuNzqZPDLwd1UeIvIBbfNZW/7WedGiNmxKw7YK0Vwwaf3DRAghUOYr1pdfsXj6N0bT24yP7jKcnD4DFi+1GLwqhedb+XRZRWotL9Aix2+2UUCwWOGeXuIv54T5Enc5j8q2izn+fE5YLCPV5rYoTbv/j9WNkKAYOrPS3RzRwcrcpwyVA6sD3TeOdRf81pUO8QkxPwOkDEgekE3ArDxm4ZBZgkwtDFuqTCTOiFcezQNsHKw8umzQtUdqj0SZweGetYtr7VvdXDV+0xZw2y/ReB10kQ0vOi2dO4LdpbMOuHdryNnpgOE0jXY+Q8t4mjKbZcymKQQ4GsfnfvBwwq8/mHE8y9jkjttfbkgSQ9N4Hp+XlLljvazYbmqqJhBSg00MTFLMwLbVm74W1/Yfn2rrmv+H1UnnUtD93ff+DDsV2xUZda/P0wHW4e+VGxYPeWGltvu+aUFoAG6oURlnwLRhiiFEKq5yQu2Exhu8GtFeDa6qHtiI6sKqXooxcx9r3cCzTtj6EozQWwPTH3sj9dlnn10Nu6QZDNZW9UuFE0TeU/gdECSOgmgXdFc2hrw2rZRSqVPIbHTcDKpUxYrlxRcMxqeMj+4yu/Uh05P7h7vvzlJ3dxeFFpf08ArqFvgXDo16tKri7M5qg58vcN88xT89j5XOxUWk3pax1xM2Ucnmt9tY8RQlIa/QqkJD0x4Yu6OrpKOs+t5koXfbXq3eVG/oZ7EHn26x1075IzvSywBSK0Y91AE2Hrl0yMCgmbTgFysXbUKsdioPVYh0W62919zDkF7lVFpglHbaVPoVpdf42i5gNUryh5lhkBqsFZoQzVGf7fXpDuvS1DAbJ9w5HnD7dMDRJCVJBKfRO+/+nSGfvjfFN4HzYcVwYLl9nHHv1pAPHoy5c3tE3QRmo4TglKpsMMRI9NkkZTRMSFODaatU9QFtPe2eIQtFX6kCktcag/r83CV50bvRw3iODlC6WZ2A4jSCTzjo+fQk1D2a7iqXr9e8i2479jzQuXqIQ0vF+SxWQH4AIVUwLVHhIxVXNkLZiFbO4LxIUInwJS0EKSUhXKjqU4Fzn+eLP339dTcfpP/8z/9sOrXvtyoi34LQj9ZDOiAo/va3v5Xv/Mf/eDEriq9F9TFxgLUWGHb3QNip5Mzean0A1sSJd4LQ1AXb5ROWo78zu/UB5XaOd/WhXLvld0TMy9P1uwHEdvH2LfBstvFrmxOWK/zlAv/knObLR1HV9s1j3NOnEYQ2WzQvWxWcQ12zp+20FTiYNIoLXqYa61c+r3rke9sA7XgHPTwe4hUKhTzsc79NW0l1r9Gn2g5iFcyhqefzroCDNbsF2KaVkdeBRIRxapiMEoZDi8kdoYrzSsbSaequMHzCIDVMJymnxxkn05SBgN80NE0gawL3Zhnh/SOyxPDN0xKAB3dH3D4ZxN7QcQYiJCLklaesPaNxSuWUW8cD7r8z5miYYEpPs65pLkqadR3B8yde1bzsdlyvkGKH/RoO+jlOw07V5q6ICfoV03Xvq5fecg0B9wqfsUcABGEX4e2zCECatJueELODqkYoGkPZGBovGrTtHssOgjywAZ6i+jSILCaqRf9ttmrfn/10cvIzBB+9yqc+/h//I5/84z9e2qq6EFgobNCQdZsgBfEq0g2vFnXcxaTWkqXxyvKuptwuY19o+Q3F5oK63GBsgjF2Ly54hWtbg+4rlqKIEuq8iPM581WrWNvEAdKWXnNPzuPQ6MVl/N5mizZRyba/FUOPbrNA0gam2UMT074YohvAOxBIfB8Eac8pWzXGFYR+ycTOEetmZ5f2XjR7wLrRWfqm9G0UnKJVIBRxfigbWo4GluNpwmSUYG2N84pzgVQMfdf17jdaEbLUxLTWWcbRyJKUjvppQ7FpCMYwnaa8e3vEcJLyzr0G7wIn44SzoyxuA5oAA8vRNOXB/QmFU45ujWiA2STjzlHKSWYxy4ry6w3VlxvqyxKtw81dC/0pgdH137vqzRbYO6QfVjWHIBSptnDwfb1m7yPXXH7yPX4uhTicmsQKyA3ApxIByGjPHDmCz7a2FI2l8bFreWVAtUJZKDxW1XNVXfPoUb8XdFXt+zLr31s67kcEn2dENq3XUrfgavbb325VZK4hnAMXqjoSZSwiiIhRBReknRkybCsrqY07YsGj3tGEQLG9JF89ZbP8hu3qCdampNmInRIr6KHf7XOotrDe4M4v8ecRVNzTC/w89nPC+Tyq3VYbwmrVVjtFrIyKMgoSujmeHoW8V7NJj25rZ4Z2vZJrFvGrVRB6MCP4cpvIfSWzlznrIVBcscrXHkly84yIHL6/cEVh2Fsm9hmV15RDTiMALR1sHMPEcDKM6asns5Th3PZmbyJfa4zsZnOMCNYKw9QwmSTMZgmTzGDXDcVfV2y/KdAsYfjelNnDKaPbI965PSK4QOqVMVBfVqxyRzJKcIlhlBoe3hlzfDqEzDIaWEZBsfOK8HhL+aclxV9WVBcV2oQD8mhXB+zo0G+vdrvRt/CK2k1fsEiLPntuD8/EvmfjDsAlRIdq9s4FV9Vs2g4d98+68DyDSXllbD5UxUWhgew+u+Kt4obQjKEZgc9igrAxilHFa1TB5XXCtkooaquuleG2HaH2pMlW0Ceq+rVY+1ibZv1Z20XtPlIRveLkJbYcesMW8K1E+8ek4lqvpd3DOddY1TXwVOCxoicoA5Bk7+EEtRfNayPbyjJMlUGmpEIrvfa4pqDYXrJZPGIz/5rBaIbNBthkcGVuJ+yuB/UBmgZ1Ln4VkW7zl4tIqz16jPvmMc2jxxGQLueEy2UUHbSUnNZ1S7HpbmgUFcRkL4ULe4DRvaLteVTc93A25LqX7HnfHHD0L/cBequZvlIZIABeo5P2vIZ5zWCccDK03Lk14PbGMbusuJjXNJUnhDg/pBp2XmFddMR0nDA7Sjk6ShmmBpM7mq+2FH9a4Y2BwpEAk/sTZpMEVPGFw61qVrljDdhJijnK0FHCNLFMRwnJKCEVgVVNdV6w/euK4s8Lii/WNIsK9Yq5mmP1A5v6v2wjQq/IyK8q2rrBUafhipKNnmvBTef3UMcmP+DqEmXZ0EwUPwJNo6DF6N5JofGGorHkdULpLD5IS8ZFmwWNdMMGkW8w5kuv+tiJ5K94qN+ovs8vqSd0431TlmUzHQxWwJMAjwTuAlNi7Pdu6K3xRvPasqkSGQ+USQhkUTRGCLS03IL15VesLr9kfHqP4eltkj4AOY8vcqiaVlhQREDJux7POvqzzRe4p5Fec+eXuPOLvZJtnceB01AS6GxyDAYbh0ZJY/KpMcgzQ7PtDay6n6roOWz/uGdDDnfVcsMOdMfWtQKHb7PYXg0oC0rYRINTzkqy04zZ8ZA7gyH36sCjec3lZU1VOGqn5Op3XnJZEsULR+OE01nG6XHG0TQOoIoL+GVN9XVOXXjYNphtA49zBscxjddtHMW8pFjXBK/YSUp6OiQ7G5LNBqTTlCSzSB1onhZUf1lS/nFB+cWa5rwgFK6tMG8AoR8QfORq9AZXB0Y7BdthNdNXrvmde0G4pvLhmRpKfgKpXZ0s243BTSGMolecIVZMaJwDqr2hdJbSWWpvCBrD60w7QxhCUBU2Etehv1trHwH51Qpm9OwZ1le48uUtCP04G7PdSfr8888Pvv/ll182n3766dLAIxH5QuGuxNjv8V5ZLG0lZGVTJUxrpXEOtfEiQiF4F0Ho/EuWp39lcvshk9v3yYatQCGAv1zQXF6iy1UUDlwuItV2OSdcznHnC8IqCgr8ehNptrxsrXIqtG6gdm1mY6do66g1u6dkWvGBogeLe/9o6FXf4GvMuW5WTMlVWIvfvcnK62Ve82Ao/WZl1413nj77+tfumTvmQ9k5dKmCbh3haUFynJDcHTK5N+TWLOWBH3E+r1ldVDSlY75xlC7gvWITIbEm2vSME46nKUfjhEFqo9mtab9coJ6XsK3Ry4LqZEA6SRFj8FWg3tbUhSMo2FFCejyguT3CnQ1Ix3E2y68bmqcF5aMt5eOc5rIkFI4ulHEvO39BHaCv4fbSG9JGOTQC7aoap9e5E+xl1qFXMV3Nqz4UO7z6h9Hv8In1ynXeiSiCAZ8JfqyEiaBjkKS9voIQ1NB4S+UtlbPU3uLVSOhJc2JEvdYocxG+xpi/T8bjR/8juiQcXPafXYfHz2+fXj0tb+m4HxiM9AX9ouC930iSPAK+EHig6LuqeldEjGl7GC4Yiga2tSWvA7XzhMyQmjZHBqGpc7arxyyf/pXpyX1m4zsMb6ckmhHOL6m//pL68WPC+WUcIH16jnscZdX+/AJ3PidsNlE67fwhxdQtMMbs1GxyA0d/0J7V64DnJ7od0h9iOeHa2FYtPTqv0ccl8qQkfTBiOk25M054/86Q4uEYCcrgsma+adgWrXWRtNHkVkhsHHR2daCxBjNKyO6MGN4b4wpHWFbkf6vJ/7qOMebGtDNKoXXCFkxmcZOU5iSnOskwmUWbgFvWuHkVIxwqj7oQZ5rkkMr8vsmYF0m39Zprrj+7E64ZFHWEa4ZGD+d2boK9w+tcfjAu6qYVX4mKOD8AP4pVEIOWGQ+g3tAEQ+UiCNXeqA+dLPtgNgiNUQ1zFXliRZ789//+35ciu62doR9v9OpWPW8cTfcmgNC3PQlXN1aaZVnurH2Ec2cSwnsCK1UNbXpDm5YpVM5IXlvy2lM2gguGNEkwVglq8KEizy9YPv0rEzNjtk4YTJ6SNgnh/JLy0VfUT5+ilwv8fEmYL/CXi+jbtlzi1+uYYdX2IYWkV+HYqGKL9XuUez9DwfSHRvVnxA5/z1dNP012J10KsHVwXqFfFXB7QJpajicJ784y+GDCdGy5fVHzzUXF08uSdR7dr32IZqdF5VlvGhbLmpkI02nK+KMZp42SjCz5X1aUj3LqTbWrZk0rFu7+C6XBbx1u02DnCViJILRxhLyhkwx2P/eqUq+X1ZTIDSCkB0Bz1ZuNNuLt0H36KuC4a/zZ9MrM3PNotr3oQG4OsHuNl80BDWdbABrGLx2ApHv7RO8NtbOUTaThGm/Ua9sFopcbhFYClwLnTuQis7YPQNdwGT//h/0Z0XHXWn71dxcX/+E/hLvrdaUhWAJnSngP5Z4YydpKSH0QDQEjggySwGQQGKaB1CpWIjAEo3FxqRvsoiL9aoH5w1eE/+dPVP/zXyn/9X9R/+HPNP/+d9yXX+O+eYK/mBNWm9jnaeXU0RjU9gCoN9nQV3f1vOVkF+HwkrM8csPXdz3qP8ZrfpfvG+lZYcfJdmh5EiPY1DAYJ0yPM05PBxzPMsYDgxVwIQYeNj7+UJpGai4zljQ1pKOEbJaRzjKSaYokNo451QEpIy24P7OmXVS7XY+idUBLF2m3KlxDSkkPSK/6mumLP/sNxycqQ7tVUq6lELrqpQOYBqUmUGugIlAT/2x6X65nFNqPQtArcCdX3qgc/F2vVELyCht++fbQLFd+m0SjUj+EZgr1DJpjCGPBZC076oXaRSHCukpZl6lu6oTaWVEV6bziRKhRLhH5qxjzuTXmX4+Ojv7+aC/N5p//+Z/NlVbCD8VEvq2EvucDfr0v4r/8i/s3WH/66adPA+GJql4IslLVaetYLRCjd50XysbotopV0SBJSUzAGEVtIGhNuZ2zzj2L5QXJfES9UMyylVGXFTR+FwG+u+HEYGS4v8zlBR/nytzO24LnOwAc+50rhcc/rfCZwRhDkhpm05TJrQGnVpjNGlKBqvQs1o7LVUNexvTVNC1JE8FoG35ohHdOB4wfTsimKdNJhkxTkklK9ddWWFC7uLgag5hWBt7Sr1r7PexcGXZ+dp38fnvO1yeN7quWw3mdTtX2rE3O84ZG4XCA4GUA48e+zqW3D+gUcW4KfkLr7rF3ifIiLY1v2dYJRWNxXvZu2d08olIi8tSIfCHwtYosi6Jw/d/bj6P5OQPPmwpC37ZMvW4LtRKRC1V9InFieQYyEsFKuwkLKlTO6Laysi4ThikMrCdLPWKJoVi+oijmrBYb0qcWPVeypcMGv6NTAgbFYDCRdjMGsbYlk68OXIYD4OkPeb59fE8X0c6NVAnrBn0iME5I7wxI3p0gqSEZWcomMMwMiRFCUKo6kBeewgREomN6XSt541m5wLKacPd0wOksY/TxjNnAkmaGDbANSrMoUaetkrHHpAZ9NuSw52On1wg49AX00TUh7z3zpKuqtkOngnBFzeav+fM6mq1P4T1b7Vw3fPxmrKZKq4ibgJuBn7YgJGA0qvY8LwSh7nznauQbg/wNo19ZY5f/+I//6D///HPhujjY728NfAtC3wF0lOfLgJTny2eEq4OrQJZlpXPuUpx+JXFXcgT6DsjEiKAmSilrZ9hUiY5KlVGqjNNA1kakiUAQpUk822EgGxpsBmIFEwTbRR/0inuNZRZt+M6VXs+zejC5pt1zNUPouu/ffFvfoHb7Cb/mTZXiQdIlV8N/brhPdxO07SXTRINUWTfI5v/P3p80yXVlW5rgt8+5jbbWwtCSBEmHN48R8SIzPUJy+GoaIjl9Mc9fEhK/55XUuAY1oUhWSb4G3pDuoNMJAjQ01jfa6+3O2TU4VxszmAEGEGzgD0pRMcJM9erV2+x99tprr1UhhcM7T14Ko4ljMK4YTRxZ7oMysg/fZZI50Jys8AymFYeDkv3DjLt32nzyQZsP11M6dztYH+aSimFJVTh0VM313/SSgkZVX5SD0/MX/MWcRJGLU9B5OZ1F13vhs/Oi0+hZhYJlz52zygdc+CnyPW76q0NqcsmZvoyleZUYfg7uNOBSxXUEtwLaBonrQWYN97fzhtwZpmXEpIjIS0vl521dVUFUqYJaiz5H+U4qfZ50kv4//dM/nTGI+vzzz1+1mNZXJKr3SeiHA1KuuJR68X0vvHZra8vv7e0NjDHPge8E1kG6QNsY5nr5lRMd51bSSGknntWmo60eK6GuqSLBt4RchYkaUm9oOEujZ7BVTQs2BIl3PRdZ3oY8zvvHGwAWstD4m/2qFi11Tslyz0nu2TnK2DnMOO4X5JUnTgyddsRMfqWolHxQMJxUHJ3mHB9MGfRyvFcaqSXuxCS3WqQnHdKTnGrqqMoJflzia2q9WIPYWshWr1YSvEw2+WJywWIRcL7ycXqR4+jZKmc56VyWcuS1ezA/c7hFzzHiGoJrC74NNMPcmFHAhxmg0gVWXFZZ8iqw5GYQ/HwQHgoV6YHuOHimcFiNx9PXjGvn1xLvK6EfGYL7vu+ZrRzM9evXdX9/f4DqM7H2unp/XdFbqlyboSV+2eLBWsaNcIH5IIKLIZiw+YZQGJiqoZkbypHFTy2zRquYml6rvtZMmzmqvU9AP/mVZQQigdhAbPBWyJxyMqnYOcjYOZjSH5ZYK2yuJ7UGqlIWnnHmGE2rUC2NKvJxgPaDjUNMbNpspBZzo0XzowI/rsgmJfm0rGWAgoGfzAkT5wpi1Usu7GXnzxfTEOdgtVnFM7dTOsNqW6qE5j2eiyuey1aFV+Gk6Dt2WcwWjy4KdGzXEnxToEEgIyBIFQgJZV0BTcswG1S54BkU1Ep1VnyrwlRUTxTd9V53jDH9hw8fllxs2fA24uM7c9jtz/x6OI8g6Csguot+L7u7u7W+bXg8ePAAa61vdTo47xNgVZUbCptGJJq5o1Ye77wEplysdBpKK1ViOyvFA3VTo8BaiwqhMYFkGpxFgy1CXfEsmdG9bnat9e1eaiV++d8uZha9K9s8h0Ve+rlzyFMv+P3ykKdfgFoSGWhZ5FqKvdOGWy0mzYj9UcWTZ2MO9jOK3NFtx9y41uDmtQabKwmdZkQcCeqhrDxl4XGqYY4oNsSRIbGGJLY0YkMMSF7hTjPK04LKV8FIQ2wgIZyzaFqEoyW2mAgv55OFv5w3eyvxVBoYbSVnGW3FGTbbWQHRi6KY8LqK2W+LOnmFa++F4yKvmH1aesfMunshfIePBdcUqhWhWhfcmoGWYBOwhqCsX1nGpWWQxfQnsQ6ziLwyokEtm+BMIqKqmcIe8ECR36nq148ePRoQZFDkt7/9bbS7u8srYLaXtdnkDdcH7yuhH7F6Wu4dCcD+/n62ubl5DDx3qjs1UWGiqo3lprDTYEo1LYyOCivjwpNEYGpxUxuFLbqmULQh6yhZ22NLIcokCCDWzWW9SJhdeP/4KaC52bmwcuHTppZmOyJNDOtrKVubKZ1mhHPKYFxxeFrQaWYkseHwJCcvHL1Ryc7ehE5s6FhDJ7Z0VxPa15tw1CTvBo0XPa9eMZO0vOIi9zJGm39BefqiAdKzg6OvvpHkx9Vo+xlAtnPX1A64VcF3BJKgmmLrAlpFKL0JtOwsYlwEtWxATKB5m4D0+lKVPrAD7KjqUZZlQ2DGipN+v2+4uM32JrHunYPr/pYUE152nyw382Yn3AGjKIr287LcQ/WIMLzaQSSeLexnA6zT0uggs9KZKomFpOFJYsVEwSFPE6FoC9M1SDOwFTSdEucybzTrvwt3kHdoiaLB3oHC4wuPOMVGhu5awq0P26SpxVTKRidmayWm04jwHkbTiqO1ktVWqIiq0rN7VNEfFhhgJTbcXEsZ32rjE0O0kqArMbaxcEfVS3dLLpiG0Re0SvUFwsCy/cGMdHCe2bbQaAvyuvrCJ8oVoht/Q5fwZcOpmhDICKuC7wTTRRNc7OfJOSQhyzC3TAqrpRMPGFPf8PU838QrewKPRWTbVeZ4d3e3eI32wfdFFt8nobe4ZpUrhJVXXWfms88+Mw8ePKhqWK749a9/fYIxe3i/Y1T3QDuqrIDEMwTEK2Sl0J9abcQqaaS0EiEVTyT1KtNA1RCyNRiXSlRCVAhRoaGJeZ7x9Za1vS5rMYm849uUt3UVLTnjzY5/5ZGJQwclvlcgg5Jky7PVjUnSLvlmA5N7Wgrd2NCKLVgh8wmbK452I/hIDYcFp4Oc3tAxzRxZ7oICNxqM+myQ/JkZyuglNtKXmUsvZmz1XLUzIxcsbK5fNHxb7u/oEp6tl87tvK1kcxUZgJ+qulI5B+5L+KWiaCT4psF3Bb9i0I4gUU2bV2Fm+5JVwTNonFumpaHyovXskCoq3ntF9BTMYxX5ixjzGPJTFpYNAPrw4UO9Qr7XKy6y37nHvwcV7TMnqSiKMyfq5s2b052dnWNEngFPBF1V1RRIjBE7O7NFJTqcGkmspZ0oay3oomFlVKs9+wTyFYgcJBk0Rko6FmzxovfJ+8dPdQWwoGo7hczBoMQfZpi9KfFawnpiWGtGIWmMKmRUEU8dKYJpJlRNS7OrSCyMc8dBL+dkVGKiYHa3vtlgdS2hmVqM87hJRTWp8EUwQjKXhP+XuY2+aPRWWyGcGxy9CM85ryW4gNnkb66yeeNrY+avaAMJwbcF7RhoG0wjZBZTi1lUQdqrtmuwM/dUVOfrDFFVvPelCMci+liMeSgizxuNxviCZaj/HqfgolP+ToWYnzsx4XUP6KtOpNy8eVMODw/nJIXt7W29detWrGXZRKSJaldhHeiIWVLP9nivgaTQiJVuE9oNSGaT8xg0CmwrEbClkE4hzgRbBcMxvcg65we+XET+9rd52XvlMnRcOCcYSzhvNlBvk1Jp5J7moCTZz7A7U+xpTuwh7ibYtQSaEZUqWeHJCocXobuScOdOh3u/WOXe3S431xt0nEcOM4rtEdn2kOIoQ52viQkL9YS5GKieIxUwIxUsyAUXSeQsmG1nI5qeIxCcbeCb1yo+5S3czPxcoqWcT8pBnsc3AwRXbVr8hoFOSELW1Pe5N2SlZZhZTicRp5OIcW6ovNGacGmMYaaUMlLRv4qY30Vx/IWI7FhrJ4eHh/5//I//IZ9//rm+bvziYrPad/rxrlC0X8YUfZnfxvnX6IMHD86Tf7Sqqp7E8SOtqhX1uqlBYfu6FTEigqJauUBSiHNhlFlGmTIpIIk8sdHgNxQJPlEqFYpVyFchH4H1YDMWM6osekRyZq/1e9zYb8ek7ue3zbO6XhduRXnl71/Ksis8HBeoHeGHJdXTCazEeK9or8D1SkxiMJ90kZUEv5HiLWANzUbEjc0m1hoqI6xuNPjgVosPNlLWVTAHU/KdMdPdCdWgDLp1S+FfNUiBeg1JpFqqaip9cWZnRrM+KyaqL9wceunY58XDxW8LRruKjtbb2ubVPuHy159n+/k0JCC3YfFrBm0ZTGywNsz5eQ9Vbdvdn1oGmdWsNDh/5oLyqBogE5FjQfasMc9VdbfT6Qzu379f1XHIXjGGvSwBfd9F+fsk9AOALa/6/YXJ7ObNm9PR7u7eSPWRV73j4dfAXVQ7y8WL91BUwjgXBlPDYAqJhU6q2Dh4i6gImkDZhmytpms7SB3YvDa/gitZgL9//NBXy0LCh2GJVh53muNaE8qmDeKi/QIdV5h2jFOobjTRjZR8NcFXSiM23FhPWe/ExKlldT3l2lrKihXscU6+O2b63ZDJzphiWICp9BIAAKnOSURBVOC9n7NiQrySuVqBqz14qguZbGf7PBdHKLkkJMu7G6F+pKgxq4L8qsFvGHTFIKnBmKBxLwZKCYZ1ozwkoVFmKaoZI466+6d41QromwDxPxM4MNYO7t+/PxcqXdKI0x8h7r1PQj9B5cQVcNIZU47PP/+8+sfPPhv8Mct2gWdizC5wUh+fBILVAwSm3KQQPR0baaaQWKERe1JL0JQDvFWqlpBtQFwptgSba0hCy8IrdWvifY/oZ1BnFx4tPYxLfBSGV9UpOi4AjxlXsJ6iOxPYTFGvxMawEhtaKwl4JY0MrcjQmlTYSUXxdET2sB9sug/GVJMCpw4VxWvNUlPOkAuW+zvVJYlHL3AafRVz5z0h8yWRvJ4L0pagqwbWLNIxMJfnCdVzrSfJKLcMpoZxbrR0LBhxC7+gqcAuIg+tyGOBw62trWz5lIxGo5dxZZfbx1eB7N5p4uK7Mqx6EaR9vrJ5uaLJBdv8h3/4B7O9ve0BHhweus7aGjGsiNhNjKyKSMOIJBr4TSpGDHMreUFEJIlCb6hV94eEcLGqDRcxJvSE4inYTLEzDztTw3Hn1Ppne6iXSPGfec1FR+sNnu/KNq921SwNp75sO2euiIVIn3pFXHhSekQD4GW8IA0LrQiJgp26VYhFaAq0PDSyiqiXo3sT8scDxl+fMnzYY/x8yLQ3pSgqSvWU4ilmlgg4cg2Do1X9dHP30Ys10eQSrYTXBbrOKhvKGZVDuWCbl/3+p9rmeY245RHVF55nLgldiMgKaBxICH4zQrci2LBIy2IiQTRAp04N08LSm1gOB5aTsWFSGHUebwQ7W6Wq9yDsgfzZiPxe4c9Rs/nkX//1X0dLX8ns7u4aXk4qeNVVr5fcUe/c412haF8V6H1VJSSXrINnj0yNOTKG71TkJiItr9oEaRhDNKtcigodTEViC52GYbOAVa+kCpGCl9oGOIJClWyipH1PPBasUyQscV8kKLxfrv5EdfPS1DznfXuCz5NYg2QO3ZviEwO5QzYb2HYwpPO5oxqWVL2c4jgjP5yS7U/IDifkw5yyrEKFI9S9nxpq07NK1Od3UC5MMv/ufM/eflSZkUEiwbcMftWiqxY6FtOwSLLwn/JOyJ0wzg3DzDDOhbwUZqbIIcnp3LpblEOER4h8Y+GZMWZ0wR2uV4xjf/MI6rvirCpv6YRcxrkXQLe3t8t79+4dCzxSWFfPuqJbCOumBn29oqVHpwUMpkJ/CoPMsFYoaQSJgcgAUWDDOa8Ua0q+CfFEMQ6isWKqmiksP9UQq1ziEXaZ8vQr7pcz+3/BNi5cJujV9/WiW/Oi94u8+jtdcAxmMyAvfKqxteuGQO7xh0Gk1A1L3PoU17Z4K7jC4YYlRS8jO83JT3OKfk45DjI9M2q11tI6L1Cq5bxCgZ5ZqCwnx4uuaL0sfF24zpaXn/qL6Ib6stMic5jwxdL3Vefwpw0rmgi6atFrEboeBRmn2AQLd4JlR17BODf0JkJ/apgUIQFdEJgqRIao7is88cY8aRTF0dr6eg7wj2D/iTOqSO9XED/zJHQVyfKXrSxeterQJYrk7Kcbj8cnG+32I29MW321heqHqnpTpB5eRVGPlkBWhkR0OoZuS4gNdFOwMURRqIhIhWpFybYgKsGoR7xHRosh1lnsU9ElhbvXXAq9ltjHUkBbqgf1vIeALEVovQTikfOv13ro71yul/PJ6hzkrS/Z1/P7fKajrBe/dv4ROp9M1OXPlSUdNjlnNl1zpqWmL8osS+QePS0os4r8NCNrCnliqIzinMdljioryaeOclrh8grn3dwKe97f0bOMtrnD6XJymB+epf2V8xKmF7UN5IUi7wxeLRehPUH78IXKcPnwykXDrWcToywd8xcDtJzpwbz4HV+xWHglzPgSK5BLNquRoB0LmxFsxbBhQxIywe9YTLBryZ3QnwonI0N/IpqVgq/JCPNvoOqAkTFmRzxPveGZtXZ/7eOPR59//nn1kjbCq+5k5fVCwntn1R8wGX3f115pO7u7u5OVX//6QL3/TlU/UNVfAx+p+mvBCzrEDIDKwzhHj0dSkxRCX6iTBpFDDzgLvqWUm6GktyVEU7CZhzJ4x8jM1eHHunzm8zH68gPjr3A49VWH+Io25K+sXa9wgF5hj3FG/nMmoKBn24pnVi7LCVo96pWqUPKpMu4pI+OZiFLUfjzqFe+DYKjXFwE2PcvHX1Q5587HRQdh7sT6AnR3nnKtLz+UepUgvnx6X3EN6OUJ41IKuOorcacfJorokj6cQMvAqoWNCNmIMF2L2pCATH0onQqTUuhPhNOxMJwKZRXaudYgNScWVQqQIxEei+U7VPeNMcM6AQHowdV13fQHjJHvk9DPOLkZ6jrk66+/Hv3mN7/Z92X5BHgi8BGqiaq2BbHWiGgdGCY5ejyEJBJpxrDSNHQlwHIuCHThGkploFAhyiAZBkhOfIDn8GdRL5UzIzFvF4KbqXrrbBZ/+a8mSP/Ok5A/V2XIuUrmMkjsXHjRZYLxMrxklhgMr/jCfsEXC++29b7Oluy+dqM9L7ZeN6vrz5onovkxcGf2TOtEpLVIpddZ3yb8V+HJVRl5zxDHhEAycHPf0kAsmB1PwWIwZ+SCztQquhgv1fn5kPq9i+pH9bzJwuybnR03UV6uAHPR8XkRoJOXLFDOxlFdSJMzm4CaTUHpmbLeLO3D2d/LRdDwD7Hwmu2zFbRhYC2CjQjWIqQTYRphH40LyaqqglzXMBMGUxhOhWkOzqtaI2pCEsJ7BWWM8MyI/AXVhyJycI4R9/7xPgm9MjqfuVvb7Xb/tCh2RPWhEbkOxKrcQejWlr3qPZIV6k9HYqxB2qmw2RXWHDS8YDVoyvnIQAIOocyEcgjRpK6CprWIpspl9/nb+4rLcJZehHqeg9XO94OubD4s56A+vRhlnendzwEKffHvMyhQdClgmyWIzswhK9VzstTBPuqcStpyB2ORHHUpnHrC6FCl4GQhjTPz4ClQpnM7hEAyWBDvzyZ3mZHy1XAhz0YUncGF54jYLwr8+HNB3NRpU5ZOsbm0GtQXksZZyFQuufDkZQsaeHFxseSRdIZOLrNjIefEieQVelavRq70FeuuxaUskBpYjwMEt5kg3QhJLDaqxYZVcSVMC+hPoD8RhtPw7zLQFtWaxdckTGYcAt947/8s1n6D94dLVRAAn7/vAb1PQldIRPNHp9MphsPhMc59C6yLaluhC3SNESs1ll1WMMmhP4aTEZyOYa0jJFZIrRCZYAcsAtoSqjWhuK7YXIPNg3cYN7P9pkb8fqDC74V+j7kYt39Zn+X8718drZZ6GRestr2/fCPL8JoIYWxwqUeiGiqg2TvNMsvjfEI8m3pmNHhVc6EwaLXkOFouJSGPUgIFQoWZh/PZ0JmgC2vec4TiF6vDs99b5hMTSyRkPculCZWVPbf9c5DTpUCkLO3pi5/+ypviwnfJC6sT9f6C/dNzMNyPZBBRL27mSS8S6FjkWgzXE8xGDE2LsYaozp+VwNTBYAonQzgdwTgnsOH03BIrfKchIjvAQwMPFXb+14cPxw//xmCzf89J6HXdA7+vxJUA8v/4/HP///pP/6lXFMUj8b6lzm945I7ADREiI4KXAPw7F1ZJvTEc9JV2Q4iNsNGGJBbiqGZExYrrCsWNIOVjHNjcI5kGCGBpiPXtXa6zyscvmtxGkMiCNYiYcIN6j1Y+JAWtqw0riLGLYDh7LiUmmf1bFigXqixzHIKCtAFrkVqMTyuHFgVhuLyGZYxdCpAe9QGrFASJE6RRN9ucR/MCLYpFxWAtYqPwOUbC59RWhuoc6hzeOVRrkoA14SnBUXWegNRReUflPc47KkJfwGFQCZ5QXpaqitn3BYwx2MiGfTEmHIvS4asK7/xSIF5UM8F914QGgwlwqGp9fKowNSQ1PGfiGBPF4dwJYZi2qvBltRT4WUracqYKDp8RlMBFZH5O1deuv3o2UVx8py1eI2IwcVTvj+DLCp9neEIf3kiEiSzqNZyD2kDCmPq7zlHUt2x1f8andFH/SSTQsrAWYa4lcC2B1QhNgjZchGJEqAiMuN5YORrA6Uh1WqBeMUGgtKYoqTpVpqq6hzHbRvWxce55a22t/0+1YR0L6xi9YrzS14yHL9vme9meHzD5cAkwJG94Al9YOv9P8L9NkpEx5mk+GiVq5AbwqcItD9fmhDYREUGcg+EE9nuqSawSWWgkhrQpxFZwXkOwaytVJBQ2DLFGY8VMK8zIz2GDhfqXvPxbXYkoJEvBJSQXiRNMu4U0Gkgcg3q0KPDjCToJApsSWaTZQNKgBqB5jpZlwCCNCQHfGpAQQGfBH+/RsgqAug9uo5LESCNF0hSiGLxDJ1PccIhOpyhuoaZ27owIgI0wK13M2mpIRHmB6/fxvSFaFiFIRxGm2Qyf00gRa8E7fF7gsyluMqWsHI4Kh6AmRhsJNBqQxvjIBrWCsqTMc1ye46oyBC9riawNA8a1S27lPZF6vPc47xERoigmaTWxrSYSRWhVUY0nlONJOH51sFcJNVVILAm20SBqNLBpCpFFy5JyPKYcjcJxRLCNJkm3Q9RqY5JwzqospxqPqaZT1DlETDhvswSs9fCkC/LONomxaQOTJiE5oPjK4YsiHKeqCtfgLFnVPvdew+JE/dmFiIlj4lYT22yCATeZUvT7lNNp2Oc0xSZJSPxlFYRbI4tJYoy1eOdweYEvytDzOjOn9YbrSVmCk3VpvDcKRARZj5FrCeZagqzFSNuG1Oi1/rqG0imjTDkZw/FQ6U/Qogp7Z81Mlk/xqlNUdoFvRPUbB09tq3W8JM8jy71mLpYNu2oCeVnp+LKBfX2fhH7elZBcCI7Xl3J9MfU+u3v3uY+ixxqMqW577xsi0gExNiwsEWBaqB4NwFqlkShrHeh0hFRrBV6BKgVScGIoM4gGip0qxoNk9Y1YN0bf7gx0fTvGMXali71+Dbuxjum0g2fSaEy1f4jbP8RPpphWC7u1iel20KLEnZzg+wOoKiRJFgE/TkKSSeKwii9L/GiMjiZoVSFpgl1bwaytYtotMDYkvOEIOW3gegP8dBqohrpQVQsDolEIWCtdopvXsTe2MK0mmuW4o2OqxiGuV+9TmmLWV7Gb65iVLhiDm4zR0x6VlhQ5FFSUWlBhwMaYThO7uYFdW8E2G6E/M5kgvT7a60OWYawlarVDYjEGX1WURYGUJbgKUychYwxxo0G6ukLc7SJRRDWdkB0dh0qoTuALiE0xNiLudmhsbtLY2CBZWcUkMdVkwuRgn8nuLsUQbJLQ2Nyidf06ycoKYg0uz8h7ffKlxUbUbBK1mpgoQr3HFyUuz/FFiRghajVJVlaJV7pEzQaI4PKcYjCg6PWpphliDFGrRdRcfF9XFCFRVVUNnwomjoiaTZJul6jTRoxQTSbkp6cUwyG+cpgowsRxXfFo2HajgW2kAJTjMdnJKUU5WKoM7byP9jZguPnKLjXIWoy5niBbKWY1xrQiTGJCcq0U9UpeCcMMehM4HSuDKWRFyGexBWvEKGF+SJUhqk+Arzw85EV5nmV0RV9zYfw2Yub7Sugd7w/NL44x9Jrw1Kt+LbAF0lD4EGhaG/o9znstKvxggrFW6TaVrVVY7QhpJCQmVEQmriV71FBtQnkbbBkoofZYkYlfqMgYOVvMvy4ccSatBgjOdNtEt7aIf/EJ8d0Pia5fgzjGn5ySf/2QQgR3copdXye+9wl2axM/HFF8K2ieQa6YdjMksLUVbKuDdFqhsrIRfjJZJLMsx652ie9+QHTnFqbTDvDRaIIfjfD9PtXxKdXxCf6kjx+M0DqsCimm0yHa2iS6fYPkww9CEmo2oCipTk4on+9RPt/BnfRQazE3rhF98hH2+jWcc5S7u+TFlHyo5OJCEqIMdZdtka52SD66TeOD20Rra6gVitMe5tlz9Nlz3GBI1GzSuH6dZHMTE9mw2h+NsNMppiqJfYD3bBSTdNo0NtZJul0UyE9P8GVJMRycm4chBOR2i9aN66x88gndux/RunmLqNkk7/XoffNN4G0cxcTtNt2PP2Hl7sck3S6uzMmPj4MlUpaFBNRu0bp+nca1TWyS1MllSN7rUY0niEC80qW5dZ3mzes01taQOKIcjZns7jJ6+oyiP8CmKc0b12lsbmLjmCrLKIZDyuEIl2Woc5goJmq3SFZWQhJqNREj+LqCK4ZDytEYl+cBqrQRttEg7nRIV1awjQSX5Yx2d3FVRTEY1JNUYLA1VPg9Y6hypg8kHYu5lmBvpnAtDWy42BIZQUxgQGZVSDpHQ+VoCIMJ5GXIuzMipzEEWS6ovHIq8K3An0Xk22kcn9ZkBMML5u1vvDC+atXzvif0Dj2uAmKdoRBtb28Xv/7wwz3S9Cu87wo0VKUlhg/mM4QKlVMNsJxyPFT2e0qnCZER1lqWOBGMDT0FHyu6KlR3hFICsZVKiQqFUs9CC7Mb8vycplzQ6F1qwJ/xyzGCpCl2Y5Xoo9ukn/2K9LNfE310B2k0cLt7oXo47YFAdPsWjf/0G6IPb1MdHOLHI9zxMeodptMiurGJvXmDaH0ds76GXVtFkgTfH1A8fEzhHDIaE928TvLZL0l/fQ/TaeMGI/xpHz+Z4KdToqMTiu0nlH4bNxjidQooVhLsxirxvY9Jf3WP5MM7mLUVTBwu2XiaYe/swV/auG8fBwjw5jXsb35B9NEH+MmY0udMdp8x9SW5L6m0oqRCCTNd6Uqb9M5N2r/+JY1bNyGy5PsHASqaTCiBZG2N9r1Pad39iChNKfsDsqNjbL+PyTMS79F6hd9YXaWxuUHcbuOKApNETE9PkMjWwpaLPpCJIpKVLu0P7rDxH/+Ojc8+o/vRR0TNJuPdPTCQ905RgebmJhv/4e/Y+M1viJpNspMTBkbIewPyRkocRbRu3mT917+k+/FdolaTcjhisrvPaGeH7OgYdRXp2irtD27T/fgu7Vs3sc0mZb9Pr5lSTaagkKyusPrLe6x8fJeo2aAYjsgODpkeHVEOh3jniBpN0o11GpubxO02Jo4RazA2UDLK0Yjp3gGT3T2q6ZSo1aJ14wbtO7dpXLuGWGF6dAyxZXJ4gEogfoR3h4pa/Ev6U+dvWbngVp4pYEQgbYvdTLC3GphbDWQ1xicmJBVCP6gQZVoqx0PP3olyOPA6zlS8DzYtMwTOK6aeARtgeCbe/BXDX/JGvn0jWRk9ezGe+HNw3Ju0Dl5nuFXeJ6GffwJ61Uk77wsGgM2ygUTRd96YFqqrKnoD5RpIYxZZZK4tp/THnv1TTzP1xJEhiYXV1JAYQT2UoviWQY3BGUNVgZl4TKaYngtsOa9vVlQvm7bNkWmLNFLMWpfo5hbxR3eIP/6Q6M6tAKNNJkgzDdBas4G9tkH88YfEn95FminFN4/C3yKLNFPM2grR9U2i69exW9ew1zYxaYo7PsEPR1QHR6CK3Vgj/vAOya9+gel2cCc93OExfjpFs4yq1cRPJridg3q38/kXtqsrxHc/JPn1L4huXK8ngD2SJkRxhK52KPIM3zulGo2J1lcwH93G3ruL9Hq4Z0/ILUx9RaEVDk9Vw31WPUQW226RbKzTuHEdkyQYhWJnj2ma4pMkVA8f3KH7m18Rt9uUpz2inV04OESHA2xZgrXErTbNzQ2aW9eImk3KyZhyMgp9HhHUB1LEbI7GRAlxt0vrxnW6H99l9d6ntG/fBoViNAr7ksRErRaNa9fofvwRq7/8ReizNBKy42NMEiPGYNKIxuYGK59+wuZ//A8kqyvkvT6D1e/CMUOpJhNssxkqmLUVGtevkXRXKFpNpkdHRM0GJomJux3at2+yeu9T4k6b/LRHlCaINeRpjK8ccadD68YNmtevE7VaYXrbGuJ2C9toUo5G9OOUcjTGO0eytsrKLz5h/e9+Q3NrC5dNMfFjRs+e1YuK5RmppZ7Om1Q/y4krEkzbYjYT7PUUu5Vi1xOkFYXuqAuqJeqhqGAwVQ4HykFf6Y2UvFSdAQgoZpaFvDICnoF5JFa+FZFnT758cvqEJxfBYd/HNfWqi+n3ldCPVLVcNdO/TGvutZ0M/+Ef/sF+/vnnfuvwMOu1WgeZSCoi173qJ8AtVb1G+J1YE/rO3iOjqdf9UydRJKSxpduCTtvQqCVa1CguBokNKoLLFTdWTA5oiRlWYVDF15i2kTdDemfsNWtCEmk1MN0OptNGogg/nuAGQ4qHjykfP8GdnKBl6OWYbge7uoLr9zGNRmBBmZqJVpMNTKuJ6bQxKx1M2kCLHGk2IIkhCq8zrSZ2dQWzsY5EEQK44Qg/GCBxHFhWBFfT+QSJSHjf5jrm2gbaalCNRvjpFCMd4m4LNlbx6yuUzYQynxKlMax0MBtrGPXQSHFGqFTnnj2zJamvKnxR4LIMN53isywkwbIMrL28CD2fKCJeXaFx6ybp+jrV2hpYSwUUorjRGESwaVJDVN0QmK0JfRVbV0G1fI8QmGEmigNEtdIlWV0lajZxeU52dELv4UP6jx4xOT7C53kgAHQ7pGur2EaDcjicbxsfNN5tI6WxuUHnww9obl0j7/VR5yj6ffLT0/A9y4JqMqGcjKmyjKjRDKy1yuHLEl+WoQpuNkhWuiQrXdR7il4P22xgpyliXSBStFvEnQ5RqxX6W5ElWVkh7nQwUVT3xDLK8ZjG1mbYt48+pLG+zmRvH3Uu9KvK8u0w42ZMyCUc2jQj7LUUe7uBvdXErCfYVoRNwjkR6j5QGRLQyUg5Hnh6I2WcKd6rj6wYI1LbdSuqOlHkGSJfAQ+s9dvD1WunF0D5b5JJr0Ja+Jujeb8L2nF6wepCLmnALYtpvanU8EWv9Z+D/8ft7eGXv/nNjuT5IxXzjcBWGFXRawoNYwLHS73qNFc9Uo8YJ83Usb5i6baVJDIkVoJbYxTUd8GgW4IrAoEhUJ9BeuUS9BYgNdy5XHqRvIrUNOjlyXpjoA4OWIt6H0gBuwcUj78j//2XZH/4kmp/D9PsBt65BmptSIRhu8KCIYbzqK+fzgfW24ziPWOCOYeWLjDuxCBJYMnJeIIWJTqeoJMpWlRnlo4ewccRpAkaGapsSnZ4QNkbEK2v0UwT1ApVJOQGCnVEeJwE8ydbqyOILmDJ5UafLwuq4Yji8IhsZxdjLbbZpDg8Ij86puz3qSZTvHNIFP4Wr69hk4RyMiHq9zGnpyG5OI+WZThmdb9HrJn3gbRyeK0WKdDaUFHGERJHIBJ6MweHnH71NXv/8q8cffEHxqdHpGk7BOoZPCUSko+EaWlfVfOq2UQRcadNurGOSWLykxOS1RVsEuPLkmoyRqzBthrYOMFNc9xkQnZ6St7vU47HuCxfkChU56sq9YEJGM51SFyBQn92mFm9p8oy8l6P8d4u2ekJ6dYGRBFRu4UYQzEYMN7dY7q3TzUah32nnloTufrtuSSgOhsPm8kvmsRg12Ki203iuy3MVgrdCBOZ0AfC4NWTZUp/4jjsew56jt5ImeRh/q8WzBANtyM+2DQcquoDE0X3vcifZVLsPvvq/87fsGfzshj1Mjfpv6lq6F2C4656cl9XdVuv+tp/AveZMUOs3THef63ImnqfqEhDRBrWSCQShGCKEj/Ovdih0Dl1bKw4Ok1LbA2rLUMc17RRFJeE/lAYnDSIE3yumNwjU1cHA15obr/WYTMmQDMmTNRrluOOT6n2Dsi/+DPZ776g+PohvpwSX4+hciFJTLPwzAu0rEJSmgWhsgq/n2boZBqC1DRDizogex8m/PIi/H00xo8n4TmZ4rPA3PKVw9WOomXNj/OiJEbw1uC9p5xOme7tkx8ekWTTwLjrtHAa3EhLV1GVBT7P0SyHaQZZDkWJOIfRs5oGlBVuPKY4PiHb3UOMwbZalCenFCcnof9RFPiyCN+ZQNUOMFkSKpyagearCpemuLqyEmMoR2Oq8QQ3zfBFUVdCupSk7HxWxpclea/PeGeXoy/+xOHvfs/pX/9CRYW5ditUKUWJm2agUI7GlOMx1WSCy0Kl5MuwH+o9JoqIWi3iboe41UKsxRUFxWAQiAVpik3SsO95QX5ySjkc1dvLcFmOm2ZUcUw1nVJNs0Clrs+/LwPrzmU5Nk3rnpfBFSVajciOjhnv7jLa2SEf9Gl/eAdXhORWlBXj3T2G208Y7+5RDkegM20Ic4Y9+DoQnC7JkZvYYNcSousN4ttN4ptNzGqM2vAJEUG9IccwKR3HQ2Xv1HPU94ymnsrpTNi21ntiZtMwEuUZIg/E+y/jKHrULIrBFWPV27Ljlr+lBPQuJKHvc8BfV5L3Sp2WBw8eFP/LvXuHmbivkbjhvW+DbICuixFr6mHGWtaHae45GjieHlYksSEyhjS2NBs2DLF6pbCKa4VhTsWgFfipIoUP/q6TOqDPpH3s8mToFVtgJuD2sx6BViHB+NEYPxyFiqQMLqILuRwNlOtpNicS+CKvV8D1ze9cSERZHj6tTkAzhXAzq56KAtfrUx2dUB0ehbkZgHYL7bYpGzGZeKYqVECMI7bQjAyRAVcWFMMB+ekppDEuz5FOC6xBjeDV46sSn+chwY3HuPG43uci9JLmE/z1AGNZ4fMcN5nixhMA3DQLlOSyPDPk6mdDoVUI9r4KAXleMVBXAdMpLi/ITk/Jjo4pB4PAKuMsMy4M19ZBt6pwWUY5GlMMBpTDIa4mUIiE1wIhERUl06MjJgcHZKenVNMpsTEB3ioKXB7OT6BQN7CNRkiYzuPyArGWajKlmkyoJhN8UeHyAq2qupoNVa0vw++rLKOaV0c679eE4xKqVxNFgRbuPOV4SnZ0xPTwkOzkhHI6DvDfZEI5GqFlxeTwsN7/Hi7L6yp1IVGk9fDvK+9GWbDgtF5kmNRi1xOim03iOy2i6w2i1QTTsvPxBytB1zEvgyzPQT/0b0+Gnmmuqkq4kU1wVA75R6fAnsJjUf2r8f5RkqaH/8fubnZ/EUv9G/SA3nYyea+i/ZYTkFwBnrtKNaRvWNLq559/7s8noiJJ+lXBQ+O9EVj1qreATfV+U4OwnDeB0qlOkf7Y8fQwDLbG1tBpRXTaQioGMYpRRSPBJDVs5kCLMItgrMBBgUzckjqB1NDHZXXcBRPo9fvCTVX/3RhMu020dQ135yY6meBG40CDjqIFM887fFWGxFVDMFrzVedpbiaNsvyxM5WEyIbhyuGAfHeXav8w9IpWV+DaOv5knbydMDaOkQ+VUEpJgqMlSiTgjaAyg10CdV1MUHSQ2cH2Hl8UVNMJ5XAYnqMRLpvW+x1mB3U2zC61WkFUw2M2qn/aOnC7BZyW57jJGD+ZUo2GoY9UBrhU4jgM/Yrg8gJXFGRHxyEIj0YhgHNWxVvqHl9Q5/ZghKjZoLG5TuvWTcrxiKooiNstoiTB1JVXmWVkJydMj44o+n1cWRBJO+idZXmowCZTTJIsvp8EaFDqhcj8u9oIYxVjLWJsUDKQ4Bo7O/c6V1JYGqQWWVLDCEoOYmwY4p1OA0V7PA7VT1Xh8yJQt3t9XFGQ93oUo0D59s7N+ILMOoJyRmj1AmdeloZa6+tZCd/DrsYkH7RJPmoT3W5iNhJMGhFZM58bck6ZFkpvrBz2PAd1FTQYq5YObw3GmNp+dWHXvScif1XVr73Id6n3B3/84ovJF0v80wti1OugLW/DsuH7tCHeJ6GfMXtkeTsW8A8ePCiAw1//+texVtU1Uf1IlGvOa6KYriLG2gC2eFXy0utxH4yUtFLD+kpMt+2JrSGJDLGBKAqBSRHYEnAyn/iWslbTLGonrXoC/mrfsL4mvZ/DaDgfElC3jWk1kdjUAd1QPtkJxIJZ0kpiTLuFXVsNNGwhBLCZVE2SYNstbLcT1AymOURRkMHxDouikQ1JrR56dHmOiSySJvg0puo0yVPLRByTOkV4rWhWOUWekVYVksREG2sk3hFvrmMajfpYOERDH0gIEj0+z8NKfzSiGo9xZV7XQNGZduIMXrONBrYZJv9dlmHSFBPHIQnNkltZ4iZTquGQajjCTaeBgWJtUGuI43mvZrEQYJEIlqODLvVVanvOqNmkdesm63mGL0uMtUz2D4jbgUAyS4zLcjuYMDRqIhuCa55TjkYUgwG20aCahm2pakgycYxNU6JGg6jZIGo28fXvZ4l81nOySULUapF0u4GA0OjjppNQAarW5IoovM8r6qu59E/4W734KCQoI0xDpeeKgiqb4soC790Sm01e644MelnLaLPFrsTEN5qkH7ZJPmxjN1OkYTBWiOrEX3klrzy9sefg1LF/6jgeOEZTT1G7/hgjYq1IaH2qR7UH8tgY+ZOKfO293yvb7QmXe5bpDxSj9EeMe++T0Gsc6B/ae+PS5qBz7tiqPhL40sMKSOS9fiJG2tYIRgCnWnr1We7N6dCxc1yx2ilopoERtrUa02wYTByENAsVfLeW8hdBPFDUN1wvh6xmzHldMObmYp4XLBtnMbFyaBZgMyoXlBM21gMD7sY1JE3RLMePpnMoCgXTahHfukn6q3vo6YDq+U4IMoXDOE/UbJFc3yK6dRNJIorK4RsJBY6yzKDMaaBIo4FtN4m2NsPqN47wcUSVZ5RVQeFKCvy8JxSpo8ymFIMB1WRK3O3Q/PADks1N4k6bqNMOUjPTaVBmMDPYSgKclBcBZptM8JRhCDKEoiVYrJb7SZKQeNI0/H8UL3o2S0lFqwDf+SxHazmdGRNsFpTjViskDqAYDpns7xMdHFBWWU32AO+qeU/FFyUiEoL9zTbJajdI3dRQn7oqNMQB2wg06ubWFu2bN8lOThATEgZKTT4Ix8xMg6JCOZ6glZsnFpPEIRnVyVfqZKzq8a7Cexfo1q0WzWvXiFottKoCVNjvhSQ9+77155bjCeocUZKE4d7NDRqbG0TtFmWWLcgMy0QW9VeYA3qxmKdOwrPqJ+gKGqKVhPhmk/TDDsntFsm1BqYTgwGjSlyjB6UqoyyQEJ4fVeydVPTHXosqSJlLKLgXqtuqU4U9UfmLGv4oIt80Go3jBw8elFzNpO794x1OQq90Rn3DbX6fcnV58IyHDx/mv/rVr3bFuS+AWMEq2hLVT0yNWehCO1+K0stRv+LxbkFg0glpGtFsGhJrqLxSGkVSg4ktJqr9cqqZRD5wXELuOKMmvbz6fsGzp8bZywo/yUKSmWZB163ZJNq6hm6uoVlO8d0TpNVET2qiwTQDY7DXrtH41S9hNKUwFn9wFBhtkyxUE2urRNc3UWPQ01NKC9MyI58McKMByXRCKhCtrGDv3MJGFpdnVEVBdnxMdnhEMR7j5sOcAaYK7LVDyuOTwPq6cWNesWnlKPoDipMebjqZy8lo5XBZHqqAaRYSRn3qQmBZ0nWuWX6z3s/8Wfd5VBVc3TvKCtx0SjWe4uq+T3itzinO6vxcGSBqNSmzjPH+HvHODsV4jLqi7qcUuKmEbU0yXFFi4pjG5ibp2iq+KBnv7DB88oTs9JRyMqGaBjWYpNulc+sW+d27FIMBVKF/46sKN80oBkOygyMUmO4fkB0dU02n8wpF6qpYvQswa000qIqcKs+o6kpMrCWuqyA3nTJ6vouvKorBsE54QcKnmkzIez1cUdLcWKd5PagtNK5dI2q34bQXiCdVFfpOlasZl565t9W5Fd4LgpBLWnJiZMngL9wD0UpC+kGb9KMOyQdt4utNok6MSULPTVzYmvMwyRwnA8/zo4qnhyWHvUonWZ3k7dxuQr1XUSVXOEDlkeL/ZJz+Ke10niZJMgqXj4qI6PtE9LddCf3chrzOX2gmiqIBUfTIl2WEahtlS1XXvNeNOtqZyIjOlHdHU6c7R4WIQBobup2YTisijoKQfyKCjyT0gqwNXjNVfYsqqAOONbDOZsOshkvM5ZaWdJVDpzl+MMafDnAnPXy/j1/pLIZAZzd3XuB7fdzBIe7oGLu+hjSbYc6n08HtHaL9IXpyihn0kckInYxQhWzQZzoaMJ4MyccDXO+Y5GCf+OCANDJUeKpIKIcZ+cERk+0nTJ/vUPWHc6quEGSM/HBEvrPLdD2sqhtpmEvSsqQ4OWH67BnT3V3KwQB1ISAWJyegSnF0TDUaz+Gu5cQ8b4DXgbsaDinrJnnZ6+PGkyCqWVa46ZTy9JRsbw83GlL2+mQHhxSnp4GlVvd7fFlgo4ik28E2g4CpbQRigE1TjInwUqt4e4fmOdVoTH7aC/2j4xOSbjeoUtdVnQpzuvPk4IDp4SHoZpjlqed0bJLisoJyPCE7PmG8s4PLQjU7PTxi9Oz5fE4osNoM5WhE3utjoiCfUwyHgQWXh7me7PSU6dER6cYGNolxZYHL6wR3fIIvSvLTU7KVLuo8k/39urKbYtI0DOXaGuar+2RFf8D0OOjolcMRvigvsPF41V2r4fqvr3tDgOCS600aH3VofNwlvt7EdCJsbLBmQXSoKmU09RwPHHsnFbsnFUf9itHU4T0aRyLGiFHF1DNBOXAo8K2KfuWRryvnnvz1yy97s1367//9v5slIsL7BPTvAI67igL2myQgeYPXmOX+0D/8wz8cPXv2LAbWQW4Juqaq6lVXQSJrxIoRvFfNS+97QzXGiLQaltVOTCsN7purLUMjNhAZvIFKDH5dEF+btyF4VzeL+wpFLcHv6yrpTFXEWahuVg0NxlS7BxTffBs07w4OkTjC7R7gd/ZhOILxFH94TPnNt0gSYa9t4POc/OiQfDggGw1xkwkmMWQPO8SpYI/2wCnTZ88YfPcd49MTytEQd3SIefgNvt0g3bsR7K+HQ8qTE/K9A6ZPn5M9e041HNZUXTtPFm44Inu+g4kT1DvK4RDbaYV92T9k8t020+0nFMcnIIZ8f59xIyVrt8kPDil7/bBNieZ6fLLk2+PznKLXw+wm+CxH0oRqNKE4Og5QXlFQnvYYf/cd6h1Ro0E1GodgPxwwzXIK5/Cq2CSmGo1weU6j10OMYbK/H6jeNXQ3t6CpK7NqPGG6f0Dvm4eINWQnx9hGymT/kPFeoC9X4wnZ4RH9h99ik4Tm1gauKJns7ZP3+zXrbIyqZ7idBqJCq4kvCvJ+PwiEnvYox2PUO6o8m8/+FL0BrijCfo4ndbLoM/hum2T1C/J+H9tIyY6P6T9+HGZ7Do+oxhOiZoMqy9CqYnp4GAgH/R7lZIJEEZP9fXxeIF6pRmNGT55ia2HV8bPnlIPhfP7p8ttYlqyjtIbw6r9ElribEN9o0vyoS/PDDsnNFtFqClG4HWITiDilKtPCcdhzPD0seXZYctSvmGSBjm1EvBFMvRbDeUWhL6rfAn9UY74sjdl+8u23/eUYcHBwIFw8v/i3uoD/d5eErqqt9Dpstzexh7jo885Qxz///PPqgw8+OEqS5FtjZBOkpeHxC9CN0DQXVVV1Tr13mP6oYueooJ1aImNADdHNlLXUEkeG0itOPDQi5HqA57QeVlVq87uTEi3dzLtzzhpT5ezQ34zV5BU/nlA+3UGrgur5Dnali9gIHY6pdvZxBydBBds7/J++Ij8+RFY6eFWK4ZBi75Ds5JByOkWLMVJNMPvPMCsd8Boqhed7ZIfHuPEEp4p/8ICsd0q0uhJgkTzHjcdUp32K0z5Vf4CbTOokFM13uhpPyHb3ghjn6Qmjx4+wjRRfFJS9AcXRcV3BZBhj8VVBOehjkjqZnJ4GNQONlvTFpF4/CG6akR8e4rIp+f5BIFQUJdVgRDUaoZUjPzzEqyPbDQOtvigpplOmec7UOQrv8RqYWdNWk/HuLvFKF7E2aKjt7wdqsnfzfZjNw1TTjPHOLt47Jgf7NNbXkCSuRUX3mB4cUQ7HaFlx8uAB2ckxcbcDLBSo88NjqvGUYjigmkwY7e8GUkVZBliynvvRqiYoWEvZH5KdnGIbjfAd+71QNRYV+Wmf06//SjEYcPzgz9gkppxMme4fMH6+S9kfUiZ9XBkESMPiYIx3FeO9PUbPniHWkh2fUJ72oXKUgwG9b74hOzpCVclPT8lPTnHTbE73X9SnSw65wbtnzjvTuVOrIerGpHfaNO92adztkN5uByp2I5BijNPQk0UoneN05HhyWPBoN+f5UcFg7HBeqeXugou7x9TM0inKM4z8EfhXEflz3u8fnI8Ln3/++fmY8jaHSeWSVoK+5vt/9slI3oF9+7H38SqTyjOfELfUIzKffvrptTiO76lz/4v3/n/3qv8V1V9E1ibGGrxXKqeVKjayIu2m5eZ6wie3G/zd3TaffdLiw+sNus1gBFZ4xVkJTlqZwx/m6KMx/uEIfTRCn07R0yLYIASKUJghOqOgLfPJ9xmjijhCGtFcCy70nRyal/isCOZvVvBphDZiNIlwBirnKLOcYppRliXOCj6NoREkelDm/ROf5cHjxhpMI8E00lqiJ7DNwoxO7SVTubNiq8sH2ggSR5hGimmmc+p08AnKg3ApnKFaB268DyYwZXXGnXXWpIPaNycyZ0gIsx6POl8rQBskqanb9d8r78m9Z4pSqOJUERFs3fyXWtNtxtQLKgRVXYEs6MfBTygK/j5pHHTUZjM/WRGSR1khRjBpgm2kQS9ueVB2Wm+7rg5m3x+/IAHMz319DOaKDvXgsq+qcKxqryiTxmGflj8rL3CTbLE/SRzUHmaklzrBmVoBQsuqTn71NbD8+rKaD74GCvjyWukCeSqhPm6h7xd1Y9Lbbdr31mh+skJ6q0W0mmBTi7FB5gAXvnNeKseDikc7GV9tT3m8m3EyqCgrr9aImHptVgtj4b0vUX0K/Avw/zEi/2waje8ePHgwXooF/qLF6M+0lfHeT+gdfcgVE/VsWQ3gnHODbre7XUwmkYcOqpuIrIBeA1IjQmQwviZbjaeOfS1FDCSxodUMVRDrCe1YSCLBR2EYs2qArifIh6FvorPcIqD9AnI/H2iVeoZH5i9cfAv1Hs1zfD5G8bVp9SwsWnxtWe2cxxUON3SBjSGCE8EBVa1S4CqPzz06WBAlpD4kM9tnrRxulMFIZx6o8wWdLo1YBPaaPXNfKy6oMOQ55GPoL/9tRmIwtQeNgaKYG3XPKg6ZW5gLC9vaGc3E43NXi6a6uYimMLM+l7DMKGZMrJoVjlACFWGw1tXfw9VH0c/lkgyGpdmbucpt0JdR9VTlBF+O0JFesPKp98ErWk1gXHe0pO4bzf5WGxZqsXxMlySb5m61Z8lcfubPVu9ngHsr3CTHT9wypjvfl0BsUKoqO9NlY35GlywZ6uPvfYUrczx+6XuZc3FcuMj9fZlFZ2KD7SSkN5o0PurS+KhDerNFtJZiUosRiOokVnmY5I6jfsXTw4Lt/Zzd44LeqCIrvApBF85aEVExGhYImcCuqj5Q0S+k8l/HIs+/fPBgVO+VrZ/VzxSCeycf9h2vhN62vfdlsNtLP+O3v/2t2d3d9f1+v/pfu93iCApRNUDLirREJEUkBSIRURP0ECmdalmplE5lzjEwQmQNjcSQJAZrBXy4qZwNE+GmacNsD8HimdJDHqodWRrCDOoIwSIsSNHLUgj3taW1p0QpUAogR+unI6Mix1HgA31aPZUuwvViRa/nGv9mEcSXwtNZ5t7itfOfS4O4LzeVXQTYWQKbGb7IfMucoWOfoSbUQ7vzgdWl7cpSYlwEyhcplW6egGb7MNuOr71xlqodY+tEZM5cXIugrS9cXGeS0LkFt8zziJ4L4/ri6+bHVy69oJf/Kheeq/NbOEN0P3ecdWlxcPYaOJsYZb5AmJ+PuocpUg/zzpIQisEQr6U0P+jQ/HSV1scrpLfbxOspphHN8nowmBQoCmXGRP32ecb2fs5RryIrPBqyrzdGMMYYIVRPCruofqmq/yzK71xVfXPj2bPj7QXacZkw6Q9hp/DOWzT8e6qE9Ac4UVfRp5uDBdPpdG7j+/9++DC/e/fuURzHj63qBiItBYtqDGxZY4zUC1Pn1VVeGU0du8c5kZV5KyeyYCKhhSA+OK9aK5hOBNFsADVURV4AnaInBVrWwqEzxGCJzurrRXiodmIcSkWoaipmz5CYAs5oUexM5OaC0GUwrzhUUldYVz9BS8FP5rKWF7xKL103yBlXv1dcMjr7LrNK7HJXZVkqLM3SU9EgOLsUWJffr37mGHv+LwZDfEHiOI9GydJPuVDjU15jTSmvWJfat7w2PVuNver0L3TgdMaCkzAHlN5o0fy4S/PjFdKbLWwNwc25N15xHkqnnI4qdo4LvtvL+G4v47BXkuWhEosjEUWMiFida89pD9XHiPwBY36nVfXXkXMnny+NZHDWZ0yvAHnJ94hnb3Ob75PQ31CSu0rlpNvb29lnn32255z7CuciDzHQEtUmsCISBlmtCTq9ZaX0RxVGcqkntkmj0Au6vgrNWIhNqIrUGlwDdCOBWmjRmACXqRH0tECzMHSonlpep7aO8FpXP4IXxdXsb6/hZxADXQBas6m9EBT8PEjMGr4yqzzOR0VdGJItcP46kC47Zs50yGaJUs/ZwS5JDZ09E0vQja8huzptzgcYZyvxuhrUmQjs0vZVF+lWsBibLPV+ZkrRfskAwoDYeUUjaFASZ2bTEJh4US1w6p0POnRa1rHLvFj5GVNL67Coyupj5OfReHbswrmYy7C+IEa1OG5nZqEuo+8vXeEzyaX5fsxWSrN0r1fT1lyQMXXx2XV1s2zQeEakdNnKvr5GmWnBJTYkoJutwIL7qDuH4EjDMbQ1gcGrkuWek+EsAeU8O8w56peMp07Vo3FsxBoRrTO2V3UopwrfIfKlqH7p4JtOnu9/s78/WUKMzs8DXYWWrT9AvHkdNQZ5n4Te3aQjL1mRCGdvf//gwYPli1IA3draGhweHj4qnHOqGgt0vUgL9Ua9dGqOgIqtpUQKz8mgQkGMBCvwmT7cjfWIVrMW8nQamHOpwWylmEjQxEBs0djgLbgDT1WUuDokewlJpxIffHVUcQqz0M088ei86tGL4o5qnTfmgD3+POglLxYac005vWwIcRF4z1Q5Oo+QZ+yeBRPYgUFfv04Sbg6DzRKKaTYwcVz3fopazdrPq5/wjWvlAxsRddrYVu0BVDl8ltWCpFn9Oh8ULMwiNywfNUuw4o47HUySoGVFMRlTjcZ4Vywd3fr7W4ONY2wcL7yaZpHDLGa8fBnESX1RBgLJmaFkOXN+Fsd/CWS8KIGcMerVpdfWkJy6F997/oTp+d8vn2d9obqZe0UtH4PZ+qR+elkImIoI8WpK48MOzbsrND5shwS0kiCJDUd+yW6r8AQW3EHOo92MJ/s5h6ehAvIelbDOMoAVkXrxon2Qh6L6B696X6LoL40o2v3i0aPxJejHVcgIV9W3fJPk9DrbfK8d944noqtYP1wkqhoBrvaaP7l796631rbUmHWgpYpxTu8ItI0RawyIg8KrL0pPb1hhTU5sF30LEzW4bg1NDMYr1gfwW9oWiVKwgqnFPR0aZnGOHGVRBkM3R5146gS01L5XLnfPkktWz/ISYOx1tC1UX4Ux6AuJi3NJbd5/kBQbRZhGQtRuE6+sYNvtMCw5nlCcnFL2eng/m9kJhAZrImyrSbKxTrq1Rby6Ok8g1XBIcXpKfnJCNRyiRTEPzLMnIti4QdwM9t7Na9dobGxgGw18WZH3wzBqXs/quDzHqws+g2lKsrpKY32NuNsJCXOW7G2okGbDndlpj6IXPI60qpYOylJlh15yrC+CLvWS06JnzssZ9W99GaT5qqi3TJk4+4Yz8nE1rGmsIVpNSG+1aX6yQvPjFZIbtRJCbBbmq/WCqnBKb1yxe1LweC/j8W7GwWnJeOpQhcjKLK9ZXVTrmcJz0D8J3BfVPwM7Dx48mLyl5PHeUfVvJAm9rbLyKo2+VxEdXjWDNGsTMMOOt7e3ex/88pffpd53ABsM69QCH4iR1JhwV1gjtYWP53RQYpf6wKYGA7ZWYlqxqX12BGfBi0XXk3rGBowoGMXHnvLEkY+KwGRzvga0BTXM1+8sdVFmq+MzX/qCBoTMV7py9k0vOUrLM7RvdDvV0J+ei1ozDbOo0ybZ3CS9eZ3GjeskmxvYtIEbT8h2dhk9ekw1HuPLybxDYdtt4vVVGrdu0froQ5offEB6bRPbbIYq5uSE6c4O4+0nZM+fUx6fBIWFsqzpBw6TtEm2rtH98A6rd++y8uGHNK5dm8/gZL1Txs+eM3i8zfDpU6b7BxTjERhD3GnT+fAD1u59SufOHaJOO1Q+M2q7tVSTCePdPfqPvmP05Emwi5jUlgs6Iz/IwhL7bYYoWcoM+prvuRACvHippywU2AWDbUXEqynprVbNguuS3GhhV+NAqScQECIEVwUIrjdyIQHtZjytK6DhxKlzqtaK2KCvE9UW3SoiQ4V94C+ofmHgzzlsb3S7fc5SsPUlKIi+raP8FpPMexXtHyDjC2+mjvCqE3GZjtzruhmed3I9oy8HYIvi2CTJn8S5stYZThVtiNebEAobMWhEGG+YFp7DfonX4N5Qtz7wHm6sxbSblqienag8aGww6wliBZOATUGaim57qp2KclThZpRmsaGKOo9qX1aVXBBMljlZcskRfmHtra86gct8q8u4JoJoTbrQAAtKFBN12jRv36Lzy1/S/btf0773KY0bNxBjyPb2GHz5Z8rRmGx3Fz+pKcRiSTbW6fzqHt3PPqP7q1/S+vCDkLw6HfCe/PiYyXfbpH/5muGDDuNvH5Ef7KP9vPb78cSdFiuffsyN//Jbrv/937PyyV3i7gqmVtSuJhOGT55y/MWfiO7/PswAlQUKRJ027Tu32PiP/4GNz/6OxsYGqMc7F+aCkpRiMKD3178G99XhkHIwoppOlwQ8mffBLgbYrhqhLjqJZ+PslbapXI5kK3Oq+pkLY/naSy3JZoPmR10ad7ukd9okW01sJ0LqyVLVmgxJgLJ7IzcnIDzdz9k/LRhnDueD+l9tCWxQlVoMdiSGp6j8RUXuW/jSVNXjjWvXTu7fv19egIjohRfj9wv2r2L/vuk23zn31feV0Ns5YcvbmqFcM36v297ezoDnn92966ooSkBXEGl6VfFOr4lIYo1YYwXnlaJUnxWe40HJ8iIXAglBjNBMDEYV60GtIE0LkWAjsBFEsRIZj/WOak9hVM5veHEXFzDwc+5kSm1jEGi7s96PxBFxt0vj1i06v7zH6n/+e7qf/Ybm7VugyvjRY8rBgPjht0gcLWaBIkt6/Rrdv/sN6//lt3Tu/YJkfR3bamLbwY002bpG3O0GtWivoUc0HkG/X8+8WNL1dVZ/8SnXf/u/cfO//Jb2zRt4r/iyxKYJJopo3byBjWPK0Yjs9CRotRU5EllMEhN32jQ21mluXQsnxhjSjXXStTXK0QixhtGz5/Qffjs3wlsGueQHO2tvP37Ne0xnfoRzaxqWeKNB44MOzU9XaN7tEl9rYJtRTdleMOecQuWUXs2Ce7wbILjD05Jx5vBaE39qt/fZZ3vvC2BfVL42Rn5njPnCWPuYND2fgAxnmXA//1vkPRz3o1RFb/Ja/Z6ve5P9nCWhGavGA/zj9vbB//PevW8q6LjwN4fyG9Ab1phoJrhoDDivOktEZ0hTongfoLl2IiQ2qCR4I1RiYCXGapNEFG/Am7C6zPcnuEGBL+op+hmZWOSCNd3PdPE0h/R8DSwGQ7Wo2yHZ3CS5do1kc4NkfZ14bRWAZHODaKWLaaSBCTI7OUlCurVF+xef0vnlPRo3ruOLkrLXoxqPQyIyhnh9ndZHH1ENRuSHh0x2d8AGC/a41aJ1/TorH3/M6r1f0PngA0xkme7ukZ/2SFa6tG/doLl1je7Hd+k+ecLwyROmx8fo6Qkuz5keHTLcfkKy0iHv9ebq1VGnjbGWdHV1bouAkaAyoP7dijLnYL3zxAoTW2wnIt5s0LjVpvlRh8aHHeK6AjImWJrYGgR2FUxzR2/k5wnoyX7OwWnJcOzUaYDgTPBFtUtjClPgSOAbFfkjqn+IouibZrN5eP/+/eIV1cT75PPvPAldNVHoD7y913FI9Od+8j/B/3Y83p00mwaRwqmWEsCx2Hu/JRjxgQKt2GD9nRXKUb/Ee6RywRWyLJXqZoNbGwmdtsHWHiveK2oNphuTmhY2sdhmRNRNsK2I6ZMh5VEeeg6zNfQZFKee+FNBX/I15ZLa6TJ4R6+AOFxtm7NelF9UNNbWPkAxoLjJhLLfJ1rpYtNkTh44j33YRpNkc5PmnTukN7ZAhGxnl+nTp6j3JBvrJFtbwc5gbZXmnVukN29gO53gQeQrotU1Wjdu0L59h+bWVi3cucfR737P8MkzmlubXPtf/p7O7Tskqyu0bt6gdesmjd3d4DRaa6q5acZo9znpyiq21aRz5wMA0pVVolYzsPvKEpdlc/vuGdVbX8vM8woQ3BvcLi/d5hKdf37+ll5irCFaiUlvhv7PXIh0PcU0goKGr/nUUrMFyyowSbf3c77by3l6kLN/soDgvIiziEERVTVBuVwLEdm3xnwj8AdvzO/FmK+iKDq4f//+5AII7nzP520vWn/IRfD7ntDf2EO/xwWwvGS1gN7f3Z189tln2zhXSllaEemg2lBFKvXriETWio1NDc1VC2jO14ws77V25w7c1FYa+ijGz7TWLKwYoiTCphbbiDCxmTOPytMcLcKuqX8lRe1neCqWFakVX5VU4zH50RHRs/bcZC5eWw0OqFUVhDKXZpdso0m8ukqysUHUalGcnDJ+/Jj+H77ATac0bt2k88t7tD/9hHh1hXhtjXh9LTjRRhG2jEjaHdL1AJtFjQYuyxhuP+Xg337H6dd/pX37FlGzSdxqh/3pdkjW10hWVzBxTHl4yPTkkMn+Af3Hj0i6K6Tr62z85jc0Nzdp37pFurZKMQg25eV4TJVnqLrFnJHKO3DKFrNgs6QpYjCJIeomJDeaNO92aX66SuNOECKdKYKIX+g3VAp56TkdVjw/Kni0k/HdXs5hr2Qyg+CsiBE5A8E55yvgQEQeijF/tMb8UUT+mmXZ7l//+teZ/hD/CPafXm7N8C6w3d47q/5wIMwbH+DX4dW/ia3DVYZZ5d69e9HDhw8dtf3DvXv3dsW5po2iThVYcxXwS+CaNWJn0Jw1UFWqeRFYc1L37WealGWlXF9LWGkKSWQQK3gBZ4NsTzyzczA1P7UZkT0fUR5luFE91Dob8FSD1sOhMqMH6M/omp4lkXpgFC/4MqcY9Jju7NSDpmDiOPRy4rgWR33xhJg0xbZa2FYLE0f4LCM/OGDy+Ltgv5DlxKurNG7eJFlfC/bfrRYmTWvBUzt3J7WNBiaK8FVF3uszer7D4PF3+KoK1gfTKVGrNbfXtvU2XFmQ48myEcXOiEZ7DfU+2GBPp1TjMcYYisEg2DpMJriqWOjmiQS/qZ/bDbs0ODunsi/NSJnIBgbcWkq61SS906bxQYf0VptoPcU2gmKDcYpBwzC1U6ZZWIztHBVs7+VzCG40der9AoJTwcpcCYEMOAa+Af5gRO7bJPmq22rtfv3119nyJfFoPgF2IQR3WSx4U6juZTyPt73N90noLSSf75Mk3jRBXXVbV9rGw4cPX3BlvXfv3nP13gIlInnNso6d6ob4cBMZVKMamitKr6eD0qiqlE4pSmWaO/LS89FWwlrXBDFrD97X6s6JJVpPacSCbcRE6ynRSsz02wHZ7hjt+zOQThgeNGfIEFe9tOWNT8MVtqlLumMShf6IKlU+RY8P8FmOr0pMZEnW13GTaVCFtmY+CHpmu9bWyaS2i/YeKhfUnatq4bAqgA0ECBNHmPn7at5JrXMWVA/Mwga8KJds0vWsbegl5bJNUxrrgaCQrHQDu246pRgMKUYjqizDq2NZcWIONKrUVMq3efl/jzMmEgR03Uz/bXZ9GWwnJt1q0rhdJ5/bLeKNBqabgDWor1das8WQV6a5Z79Xsr0Xqp+do4LD05JJ7oJKFeIWpltqfIDgSkQOROShivweY/7FwZ/aabrz3/7bfxv9y7/8y5lTcf+sRtzLDtZlHkKXJa2r9Jh+iG2+E4npXauEXpX95WcMKi1f4AbQhw8fDv/+7//+cZZllaqKBmmfxHuvFbImYK0RG0uA5krnfVEpp8OK0illFfDxwAQCjNBVi5EgehqoQUDDYtIGUTMOjd7a9lgFxArVsESr2ovc12FRfoaHcR57QyXkqXDFNDz7E8QY0vV1quEIXxRzLq+YWdJY2lTdZwkOn4ptpCSbGzTv3KIar5DeuE68toptNMJ7Z3YYs6X+LHktW4KbAJGJtZioTliy9LrKLWwimFNDMEBEQmNjg87t23Tu3CZdX0eMoRyPyXt9yuEoWELU8kDz/RD5+cWZGfS2pP8mSKiAujHJ9RaNO+1Aw/6wQ1Iz4Kh9FaQKCcgDlVfGmeeoX/LkIOfbnUBCOBlUZHk4jvUg6hIE5xcQHDwUa/8g8Dvgz6PRaPubb76Z3L9/H5bMKbnYwuX7xKaf+m/vTe1+ogQl78D+zi78CvBffPHF+LPPPntWVVUi3jdFVRVyRX9RQ3Nz1pxXkcqpFqWnP9KZyZfMXLny0nNrI2G9E5MmJvSMajcCb4J3TmwW1j2SWqJuTL43peoVuHHtJUTQKQtVh1kaQPwZHcZa22xuHqB5MNGbZsGjyLklc7Tl19cq2EWwo/ZFgVhLeu0anV//KtgrjCfEa+s0P7xDvLYalAzqRIJfzu7MqyhVvwQXyvx55jUzf59liQBbQ4hEJJ0Ojc0NGpubRO0W6hzFcHgmCZ1Z7i7bt/+MAPMZ/DbXDzSCbcVEKynxVoPG7Tbp7XaA364FBpyNTKjkKpBaBSEvPcOJ46BXQ3AHOc8Oco77JZPMqyoaWZFa7cjOP1s1A46Ab0Tk98DvROSBqj7b3d1dJiEsNGjf23O/T0I/YKX0c1ofyhICM7/oHzx4ML579+6ThrXqYQJMRMSLqvVet5S5rbE3BsKiWhlNndk7LqgqJcsdw7FjcsfzyW3h+npMMw7DfZXTYAsOmEZEcr2JJIZ4JSXbbBJ1B2TPRuT7oL0cX6vOQQ3LmUUCmul6vS7IffHr5bVO4pll6hLJYK6dbevqw5wTVV1OCPWv3GRMNRgGOR7niNfW6PzyHrbTxmcZJk2JVlZIVlfnOm7zJFInmhkctrCgkGWfhaWkdC4xLWutzbYXmVCNrawG4kKS4oqCoj8Ikj+jEa4oLtQs1TcOod9//EVE6mXQ0mYcZ4z7bCuQDxo326H/c6tFdK2B7SZIwwYDDR+gYFPDyUWh9MeOnaOc7f3Q/9k7CX5AeRG6lbW9lVFVgyI+JPsSkX1j5BtEfm9F/tXCn+NWa+ePf/zj8BJ04mUsuKsMq7/qgP4U23yfhN4/XnrXvwDNbW9v939761YxbjRyr+oxJlWIgye4bohIZIxYK4INitje1zYQZaVkhWeae5xTjBWMgY0uxFGtvKyKSKArm06MbUbEnQTbiTFx8CfChv5GNSzwZS1g6fXNgIof+DDqGeM4j6WBTRuYpHZwteZSSFEBN52GuZ/vtmnevk3zgztIFJFe3wqaFlGEqYkHWlVzCO+FGR3hXAK6rHKTM2rii52p7QoiS9xqkaytkqyuYtMUN51SDAYU/ZlmnLtiQvkRz8SMeHBOUyAsCiy2E5Nca9L4MMz+pLfbpFsNbCdI8CggLgjDqkBZV0C9kWOvVsJ+tJOxc1QwGFdUTjFGSKIgUBVQheDNXUNw+0bkG2PMH4zI7yz8qTEabd9/+HByrvpxS4vBy5LCq9Zb8hpZ/afY5vsk9I4miB/qpF7UuzoDzd3f3Z389tatZ5Moip21aUDZdIrIL1V1yxiTLsVW4z1aVapl5dQ5Fe8XHYK88HywlXJtNabdsCS1NYSX2q01EcSaYAHhQWLBtiKyTkxxMKHoZbhhhXd+SXJnyRfmvHLyT1YHB6hNxM5ZazOr7nMY0ZkT6/OC6bMdevd/j5YVrU8+DooJncCYs80GIgZN/CtkiXRhQ3EZqVf1xee5R1BO6JBubJCur2MbDcp+n6J+Bqke/9M3PWXpf2ZCriyIB6hgTGC/RaspybUG6Y02jQ/aYf5no4FdiTGJxdbWDaamcBelZ5wHL6Ddk5LnBzlP9jN2akfUovAqBm+tMdYaUbDzkQUYIXKE56+i+nsj8odI5MGKyLP/3wKCk9+Cvb+w6FbeD6K+T0JvMbB/nwSiP3B5q5dAcyytyAC4v7s7ubdx76k0xOPcSEUGiDhRxau/hRo7k04QUS91pspLb3rDUrxXJoWnP6rojyrufdDkzlaDOLJ1j0hmN21wuGlFpLdb2JYhXk+JN1Oypyn22Yjs+ZjytMDX/qFBMHRJep+QwGpF09c+YG/6ennhyAar7FmloueTguo8gM/pCU7J9w/o/f6P5IeHYWj11k0aN26QXt8i2VwnXl8nvXZt3hOi9v+ZK7bNmu96LhGd8Seof+1hyUBtkYgkkOFNkhCvhJ5Qur6OjWOqaUZ2fEJ+eko1nS4BmMs+T3pGivZi3OZyLb7Xeb3MB0+XYEDPGZUNwWDbEen1mnp9p01yvU2ymWK6MaYR4SUQEOzM/BehKj3DqeegV/L8KNhx7xwVHPVLhqECUmPEi+DqQxgt5fepIM+N8LUa/aMY8zsLX6fd7s7/9/790fJ3u/8iBPd9LturMtd+im2+T0J/QxXMD5kwz1dGAujDk4eDexv3chvHU+dcJSKJEhaMzvktgYaIiBVjjSg2YERSVMrxINBWx1NHVgYTFUTwmtBt2mAXXucNA0gkyGpM1LLEKwm2G2PbMSYNApxiJ1T9Aq3hucCgW95b/ckP63wO5SWVxvlKCBF8XlCcnKBVSXFySvz0GenWFo07t2l99AGtux9hk4R4dXVhmV4TAeasL1miZ8/6TjMSgp57Tf08aygXEotNk1AJra2SdDtBQmg8Dkno5BSXZ/M6VF5px/BDXsELq4eFiGogsEhkavit1n/7uEvjTpt4o4FtRYhdEDWkUtQECKCslOHEsX9S8uxw0QM66pdMco96JaqHUOtvbP3CTiMDniPylRH5vaj+wcKDxmi0c/+bbyaygODOs+C+z6H7Abnt7yuhdyF4yxu874cYQn3b30n47DPDgwcVhDmi3/72t897vV4EJECF6kThV8AtEekaE8KRUTFeVSunWlZei8qL8wth5bxUhlPHrfWYtU5EKzWkUQiMPgqOrC42EBliE5anJrGYRkS0klDsTylPprhRGbTnqAF8NXPzsTNU4Z8yL+mMau6XnvrCrom1QXPu2iZxu433nvzoiLLXpxwMUFdhW00aW9eI2u1wIM0iyeB8EN20tiZD2Dl5wZdlMKGrqrrhLkGsNI7mag7LJ94ANkmJWi3idrvuB2UU/T7TwyPy4xNcls0N1Rf13I+QhM5ZuM2ZbywSkEkjbCvGriQkmw3SG6EKSm63iTcb2HaEieo9Vw39H69ULtCv+2PHYa/k+WHB86OcneOCw15wQ3WKRgY1YYBbvGLDHJxXhVOBXVT+guF3xto/At802/2d+98uILg6Adkl5OF1FohvI/78FNt8n4TeYqB+HTHBq7BH9Gq3249SBZ2dzg4OrQto7v798sbf//3uapY57/0Q1R4wFXDe+w8d0pk5JmsQ5fL1P21WePZOCikqpT9yHPdL7t1u8tGNlBsbMXEnmhvnuWBPgxPBtCKSGy2iVkS0lpJcb5I9GZFtD8h3AzznipmN9vLczPlvqq8skuQKp+7sSmIh1SMLF5oLExDeB4vumlZ93pNZrCG9cZ3uZ78h3dykHAyZbG9TnvQo+32qwQA3zUIyKevB03qIFReS0Lwi8gHuw9dVpqnnhCIbCBKzwda58YIGGaEaXrOYeQKKmo1ADplOyY5PmO4fkB2d4rKsPgbRWbfaS4hdV+MwyiUQHGcc1hf/FnC6xHwDE9v6OmkF/bdbLZKtJtFGDb+lAT32boZmCmKhUmWce/ZPA/z2/LBg96jgsF/SH1dMM6cKGtqW4lUR59UsekA6AHkk8Ccj/NGL+UJEvo3j+Oj+/d3zWnAv6DiyYKq+TBnhTaZ/fy7bfD8n9DOohN50m8KPMzegr0iuBvD7X3wx3ofvPv3007HxPscYCU7J6r36D1C6ErSyjDVijBFvBPGKjKeOrPCMJo7R1FGUntIFd1UFuq3AnkNB/AyeM0jXQDsiWk2I15LgZpmYoOcVTyh7Bp+5MOCKLlUa+tNeHnXPRqwN7La6lyNRhFg7j0IB0DEkW5usfPZ3tD/5mHIwJF7pMnnyFJMkwcah1njzZYmbTHCTKT4vUBe06HxR4LIsKBmUFVE7IVlbpfPBHcrxmM6d2zSvbwURUqEWIc1x0ymuLAAlAtJml+bmNRobm8SdbpD0yXPyfp/s5IRiNMS5EkM0V2X4wePMcuVzEfNNDKZRJ6CbLRp3OjQ+6NC41SJeTzFNOx8+DYkrQMNVvegZZ2H+58lBzvZezvPDUP2Mpo7KhaZnEtVK2BrQY++Da7D3fgRsi8iXWPtvxpgvTBQ93NraOqodjWcPu3Qve66uqH+Vv8lr3t8/1TbfJ6GfKCq9k+ZOywjNb8HcX6ze9NGjRwe//vDDSI2JVNWpyAR0IsiH6v2GGElMwMaMCYZ36pz6olKpqrJmzgWFhUnuGU8dNzYS1jsRndQSR4GerUZQCSw6ato2tVCqqaujAM9lVP2CalzivZ8nIZn5tM4pyz80i25JiduYkHxqXbio08EXBbbVDJ5AImdWFyZNSDY2aH/yMYqQrK3R+vQTTBzRuHWT5q1bRCsr+CynPO1Rnvbwtb2214pyPCI7PSU7PaWcTGiur9O9+xE3/vf/SvvOLZpb11j/za9J19bIej3K4ZD85JT85JRyOkHR2g7iBt2PPqRz506wBk8SfFVRjEfkwyGly1AchjiYEs4RJX27V93y2VoeOp3VVhp6WrYZEbXjQGTZapLcDFVQcqNFtJFiWzEmCgsbcX5ebVcukGYGE8fRoGTnuOTZQU1A6FUMJpWWVZgXimzwA6qTkHUBWs1V9RiRp6r6wMC/YcwXTWO+/T/+8R8P/uf//J/nRx/MW+wD/RhB/9/lwOzfimKCXghvXV66XqUq+SmrPn//RRVubLd75EejB1WSjIGewQ9VmXrVezi9aUQMc+aSeitUIljv1Q4nFU8Pgghkb+Q4HZbcvdHg7o0Gt64lrEYRkZUwq+E0hDlfa89tNmgmlng9Jb3ZIn8+IXs+In82gv0J5aBYkpNZzMPIbIbmjDj+VXgkLx9ilUt+zmZ2xJh5r0atxZh6gHUZfvIOn+W4yQT1nvTGjVDF/N2vsI0mUbcTKpLhiHzvgOnzHfK9fdxwFDTm8OSDHpO9PYbPnjHZP6B9fYvWzZtc/6//hbVf/4qk26G5tYV6pej1mezsMdnZZbp/QDkZAZC0u7Ru3qR9+zbNa1tEjWZIQJMRxWhEOR1TUS2+q9RxtYYel+dEzx5ZvfrlJkuV5Gzt4GfEg8XrRATbjkLfZ554msSbTaKVBNOOIbYE7CyQX6TWkFMNCeioH3o/z44K9o4Ljvslg7FjnHl1zqsIaowoAl41CilEUecV1QODfKXCl17kT6L6IIqiJ1M4WUpAnKt83rYa9lXmeX4O23yfhH5mfaSrnFD9mX2P88Px/sGDBwWwe+/evbGITCVA6y5Ac+o9XDfGJBiDEbE2FkTEOK9SOaU/Cnpbw2lgz00yT1kF2RvnoduyBLHU8OmiiswqoGZEvJoQbzSIVtLQbE4MRILEJujPlUtxoKaBIz8s7WfZU0edx+c51WhEcXpKfniIL0vKfh83zcAvdNfUecpej+nTZ0zv3MY2myQbGyTXNrHtNmINVX9Acdpj8uQp40ePyZ7v4kbjuseklJMJk4MDBo+/o/fNNzTXVoPuW81wM0kcgm/9msHj7xg/3yE/OaHKMywRNmkEA7s4rk3ujnBFznhnl7zXm6skyHmw+Ie4a3SRcnR5NswEKDZqx8QbDdJbLZr14GlSKx+Y2NR6eiCVn5fwDnBL+m/PDnMe7+Q8O5zL7xDyCxJFQrBhkMB5VKgroEJUD4Cvxch9Y8zvrLVfJ0nybGNjY3gOgjM/UOXzvhL6gR72Hal2zmsQvw1xwUt0jd/q/r4VaO6zzz6LDg8P5xfpyclJvvbJJ5mt8lJVSq9aAh4RJdzEDWNErDWmtjjGe/VV5bWogvp2WamUlVJUfi6G6n2A06yB2AqJEWwkmCR4v1ArK0gc/m1SG56NkJDEhqpHq5np3IzAMEPmzikHyPc4LPOqytehMgqKCY0Gtja4q0Yjsp0dRt98y+ibb8iePqfKJuHVEhQRxNhQFVUVPi9wedCgKw4PmTz+jsFXXzP48s8M//I10ydPg/NqkVOhlHiUaG4hoVWJm04px+M59DZ8+ozTB19x9McvOPnyzwwePSY7PsLhsViSTpdkZYW41QIRikGf4XfbnHz1Fadffc34+Q5VVdTqnFEgOai+8gqXlx47OSNltMg/M7KHn1dGtlkPnS6pXjc+6oT5n60W0VqCbcbYKCi4WwQb3OQoS2WcOU6GFTvHJU8OggL2s4OCg17JcFxpUQbQz4iINSLWhqFqwhrIq3IKuo3qn8WY3xmR30XGPOiurT39/R/+0N/+7ju/lHzOUgffP95XQm8JanvZ76/aqHsVBvS2JNAvElT9vmW6Pnjw4IWm6qP79wf/6T999Dhz6VRFTkT1CGOOBD5T1V96ZR0Ng6mBzYqzRsQpUjkv/VGF8yrjzHE6rNg/Kfj4VoOPbza4uZGw1hXiKIigOgVfBVtxNYLpxiS2je3GxFtN0oMJ+fNx0J/bm8BRRjUu5rtcC9PUmm7L30zDHNNLLAjkwjJ3mXln51ifm07JDw4YCpT9HtHqKuo9+eEx06fPQn+IetbXK8XxKcOvv6YcDBhvb9O4foN4dRXTaKBVSXnaI9vdY/L0GdnODuXxKW6a1UTpCMFRjccMv9tGXUW2t0fn9m3S9Q1skqLekQ8HjHf3GD15xujJU7KTEzy+nvcxuCxnenDIqTFkJ6dE7SYuL5js7zN+toMWFbamwoeraTF/9LIr69LOtywvCJZrBl1y/KkXI82IZL1Bcr1JcqNFeqNJvNkIiaebBOabEbzT+RCz1MOneaUMM8dhv2LvtOD5YcneccFhv2AwcmSFVxdIi25WcAUFhJngq6KqPYRvQb7A2i+8MV9Fqo9z1YM//fM/D+WsBNKZ6bX3of3d66n8nPftKqZ0+hNu86Iq6CJY7ftWVmeguVm/6O7du40kSTaBT1X1PwD/m6r+r8CnIrJhFm/0Ihiv1ElJJDSAhW7Lcn094eNbDX71YYtPbje4uZmw2o5IIhP0S2dGesuFiFf81FGe5uQ7Y6bfDZhuD8l3xpTHGW5c4YsleS55HVmsV6w4ZslrPjBpgt9PmmKbDWyrgSQJqOKyANG50QRflfXbA3wkkcU2G0TdDnF3pTa5i1HngrHccEg5GOLGY7QogyAsQo4yRSlFIImImg3Sbpd0dZWku4KJk6CCPRkH6Z3BkGo8DnYMWgGKIcY2UmyjQdRs1GZ3BnWOappRjSZU0yneV4t+0PdUzZZzx1kv2NZMwileT0lv1IrXt9ukN5pEqwmmYedV7+xZu1vgFYrqrPr104NQ/Rz1y2DB7RaasKqoSHAfmREsNdjenwLfqcgXovrPaswXVZpu2yzrPXz4sORiuvX7x/tK6CftA52PW2+T2v193/s25IUEMHfBbi9J/Wxvb2f/43/8j91/+qd/crn3FWVZqWouIn1V/QS4hsiKtcaY2qPFOS/OiZaVJy+9lpWXsgp9ozBA6OiNqjl7bqVhaSaGpBY4VVtr0AlBWSGue0NR6B/FaynFwYTyOKfs5bhRiZs6vLol6mJNY5jDQosZoCstAeaI1MIMUyuHcxPcdAo95qSIQPLygWhR/xd+VeKKKVUxpOyfkkfpEkTnAxW7LOsk4Op9joJWnUoQhFVPlWe4fELZ65MdHGLTYAGO81RFjssz/HwuqFYXINhl+LLClyPK0aj+Xn5uhyC1a6rBLI1f6etdMrJ0sF4ycCpisA2LaUZEK8H8MLnWIL3eJtlqEm81iNZSbDPCREItWx0GT53ivJIVQVC3N644GlTsnZTsnxTs1sOno4nTyqsaEcwS/Xp+/lQVZIDIAV4fI3xl4Au19ktrzKNvvvzyJBz0xf1wrhI6WzC/nZ6MvOX49GNt851Jyj/nnpC8wQmTK/SSXts14HvCcS/7/Wtvs39B0/Xzzz/Xjz76KM+n04n3fhCJ9EVkhIhDpAG0RUh0HnypASn1Mx1+51WmhZfx1DOYVAwnjrwIEJy1Qhpb0sRgo+Av5BWqevJdjWBig2nFgbywlhKvBqquROHFWvrwZFk+p2aymYUe2SxJXNQykheeNZtrlsjmUi4e9TN31DCsKnqWPr5MaJgTt70Pzqr1gKq6Eq+u/vtMrSBQ2F3tI+Dm8p2hOA3kiDIQJPIMVy4qn7nlhJiQPOcSQGGw1vsKVReeS6+fJ+vX/U8usJGYVRtLo74Gg23HxOsNGrdaND9eofXxSi270yHaTIPqdRJYhqIzS6XQv1GFrPScDiueH5d8t5vxeLe2XzguOB1WTHKn3qNGxBsRJ+ELmZmWng/yRX3gqRH5UuH/NiL/bOFLE8dPv7p+fcD/+X/6yyDrl6ARvIX7922iTT/WNn/OBp9/c5XQjwEx/hwqofM3nFleAd6/f78Ejm/dujVebzZ7Dk41eBOpgjqvH4Cu1IOtGCPWGIPa0JbxqjKaOLLc0xtVnA5CIppkjqLyeA+bPqbVsFhTm4zOdM9EoBkRNSJ0LcFfa1Jeq1fO7QiTWiSxmJOMalShlV+w57y+Pfm5uvKRFxbJFxzs2kgOjeaH9UV1BvMiG0bOrlNEDAaDaHRWE6H2CTJLijEiL6MLmIVC+du6kOdXyyUDp8ZgYjOH3pKtJumt4PmTbDVr0sGy5ttZ5puvFyOza2bvOAiPPjvI2T8ta++fwH4DJK6rH1WMUiegMHxaAH0r8lSM/AWR3xtr/w34a7PTOayvbVPHK3/J/fA2gu77Suh9JfRGldCbMOf4Cbb5Nlcp82G8fwR5sNRiHo1G5Wenp2N//XrmIfcilUCBaqXBgM0ImhoTekJGwsOrUlXqy0o1L5Wi9DV7Lji2zn5WAVXCCMRGSKwQWcHGBoktktQ/YxMSTyOw52wzImpFmDTCxCG4a1VXLUv/LYL9EoPrNazGF6v+QISYVRBnMojIPIEshEXPnxJzsfCoLKjHFaHRpkuvERbN/cVnWEwtcipy3qF2SU17Loa6VLUsz/C8KgGfv2rncz4LpiIEySLbjIjXksB6u9UKfj93ar+fmy3ijVD9mNQG5psRIgmzy6KBYTmeOk7q5PP0sODJfs7Tg5y9k5LTYaXjMAKgaJDria0RW19wSmBsKtoDeQZ8Y4z5IyJ/EGv/SKv116//9Ked3d1df27RLJdA7vIzvE/f5W2+T0JXTADC26db/xDb/CEumsCcu2DFsw2+s75eJM6NyzjuG+gJjDDiRTVS1VS9NmaSLPUPp9SeDYp4D2XlZZJ7huOKwbhiOHUU9UR7HBmS2JAkBlObHKmCczWLTsAkwVUzWk2I1xvEa2lQ6E5sgMMqj+YLP5oFjGRqDaFZQGbJwfSqa8RlRW1d9EfOOK6efd1cjXtZhWHZpqD+/+Uk5ObFhiJzi++zC3S5VNLoAlLAGV8ivbyMO5eAXrD8lkUVtPwpxtRGc5tN0jsdmh91aX6yQvOj7kJwtJsgcWC+zb63lbriJaheD6aOg17F04OCxzsZj3cznh3mHPaqYLRYBuPFOkeriKiAmVnR+6C31xPlESJfijH/huq/qsgXcRw/Ns4dHx4eupfUd/o9ofv3Seh9EvpBekLvEmT4VpLQucpobpZ3cnLiDnq9cePu3V6jLHtGdRKWxCrqvXXex95rEqB9wRgxkRWJI2OsEQNI5ZRpPdw6nDjGmacotbbMCZQmqZOPekWdzqE2I2ATi23HRCsJ8WpC1A3wjonNHOJBBLwgfknu54zaqL4ex3C+dHj9maQzVZRcnvQuTEKzHs68gluqfOTliMoL1t9y1Vvj3BYvtLKorQitxTZqUdqtJuntdkhAH3VpfNAJ1c9aGqrVuO5+ecLSxAfKdFUpk9zTGzv2T0ueHRZs7+Vs7wXn0+NBxSTz+FApi7UikRUb2TCACoj3inPeee9PRHmMMV8Ya39njPl9mSR/yobDJ48fPz6tE9DMguF89aM/Uqx4n4TeJ6EzP1+XLPA6Vg7ymsH/dfb/KhfNVfbnZdJC80G93/72t2Z3d5c6PjLa3S3/43/8j6NsMMi9tYWq5up9RhhwdWKMEZHEWmMiawJjqWYBO69aVkpeqhalkpcqeRkSUV4ES/GiUlwVEkVkIDFCbIUoWkB0s57QDKozaQiIthmFyqhZD7uaxbDr7D89pzspF8FOFx7d10hCsgyHzRLI5dXXhUlIFomZJUjtLBx42eebxeefSUZ6yVV17hjoMqDp58CmIEhiiDpB6SC5Phs4DUOn6e02yfUm0XqoUG1qsfEy9BaquMoF6/je0M1Vr5/UtOudo4KD04LB2GlWevVO1UjQfQvabwa7GD519fDpU1S/wpg/ijG/E5E/NxqNh3/97LOd0T//c7Ecn+7evRv1+/2rKuB/38ePLZb8c9rm+yT0ikD9MtlyfY3Af5ULWV/zb69S5H6d73bZdq46u6S7u7svaM5tb2/71c3NPI7jYeVc38OpERkJFEHulBYiLRGpoRLwqr5eC9ekukDhznIvw6mnX0N0k8xRugDRzeC5JDZYWwuhhrkkfN2fkKjuR6ymxOsp0XojzJ6kYfZEvYZeUeVfpArMk4QswXSzo7lcj7yonXqltYNccEr1xd9dXAktell6adjQq+7Ai6+VRaI7m2hn/1juqwUmn22E45zcaNG4065Zb10aH3YXVtutOFDsjSxWMyLBeluEyimjzHHcr3h6WPB4Nw/Mtxn5YFgxyb06pyrgjeAC23GR/QMjU0H1WIVHqP5e4J8F/s3Dn3yzue2y7OTk//q/yvNHrd/vc/56fsOFn179YP9NbPM9MeEHKDH1CifmZb9/VeXxpu99kxL5dffnSgnoEqw8Ajg5OakODw9HjUajb63tx8aMxZiiVvWJFCLvNVFVO6s4jBFrbYDprBHjFcmLoMA9rGnc0zzI/ugSrDSrpLwLCYVK6xkdsFFgY0UroVcUrSbYTliFSxRgupnwqWhgAZwZsFzydH4VTPdDLQ9floTkjS/1q6+BdRmqXLoyBMGIxcYRUTch3myEBPRBh+aH3Zp8EKqfeLWGRm3tfejD3A9e8bWMU1Z4+pPacO4oJKDt/SXbhUmwB1GQyIpEUbhOTJD6DnJRgQGXKxwg8siIfCEi/2KNuU8UfVVV1bPvvvlmeHJyUp2r6i+7pt8//sYeP3eKtr7D+6E/0N/liq+R34I9CsOt1Ww1ubu7OwGe/edf/rKaihTi/cSLDAQOBO6iehNYM0bSpYZ3CChOcU59Xip56SUvPWWlUpSecebojypubqZsrUVsdGPa9ZBrGhmsZW4J4a3gaz8aSS0kgU1n2hHRekqx1aA4zqhO62HXcYWbVmgZ1KvDDi1RGXQJdjt/SN7F8CUvST7nmISBkl5TrRtREJvtxESzSnMj9IDijQbRekrUiQJb0TBPPjPPH+ehLD3TwjPOQi/wZFhx1CvZOynZPS457oeh06wWq7VmDr0x9/5R8OrxykRVewoHIrKt8I2BB2rMn5uqj//w178eLQ2fUsPJdjqdyoMHD16nAnrdI/u6aMffwjbfV0I/Akz3OpXQy7Ypb3BRvAmT7k0upqtCfQCyy3y49bz2nG7evJnneT5RSU/BHRuRE2CCqgFSr9pCsTNZlxmDjsCgmxPCavKC9EcVp8OK3qhkOPWU9bo2tkISG+LYEAWKVYh5ToMMkIZIZlITVu/rKfFmMwTQWp9s0SvSc86ey2y6kOBkDlktER2QJRlVWfien/v9VS+6WSVUovNKSHhzE6szF+0s8S/1pEQWKg+L8kDn1Y+NLNFKEno+dzo07nZpfbxC46O677PVClYLzYhZ9pGaZm9D/wYlsN6GU89hr+TZUSAdPKqHTnePC04GAXqrnJ/NKPtgPxWEcxd+TYpXMlR3BP6KyB9E9Z89/FtizJ+x9tnf/ef/fPrgv/93fwGczOHhoV5yWPUtJKCL7p03ZcK+K9t8n4R+YJjubSehN92fH+PikCt+l4uS3Blxx8PDQ9fr9aant673Nr0/QaRvRTKpxzpQNc77yAd47jyDzloTujPeI1ltVDYYVwzqAdeyWgyhyiw41XI0vgownbjaMtMKtmFDA301DTBdN8G2Ymwyg+mW5n50iU2HnIPpFgCOXEGP7k0uuovguO9zAcirwNUzgJTUUKnBWkvUCkoHc9jto/r5QYf0Rotko0HUqY+jkaAcMXM7DZUt5RLr7aBOQE/2A/T29KDg4LSkP3bkZaBg1GrXJjJiTGC+ha5c6P85VQaq+gz4GvijivxORX5nrf3Ljdu3d+7fvz9+EKzszSWJ5vtI77zJvSPf87L4uW/zfRL6EZPQVQOzvGTFId9zf35IavlVvvtFSWhmCWE/++wz2d7eXtzsh4fu5ORksrGxMTHe50Y19yJTYKIiGVDWEiuJMcZYa8Ta8IswcIiWTjUw6HwYbC29BAZdcHGd5p4sd5SVR31YhUcSBl3jSIgig00MJqmhuZpJZxKLSSy2FVh04RkgJ5NYxJqgtuMCi87XBd8LygfmTWjQP0ISmpMNOGdAOyNZnGe7gcSWqB0TrSWk15qktbV2OmO83WwRbzWJ1lKidhxsNqJQhc5Yb4ZAFMhrrbfTYcXeScHzWmz06UHQezvohaHTSeapgie8zu0WZofRzIevnEJP4RnwV+BLFfmjGPOlM+YvRvXxN998c7q9vb1c/diazWlesup/Wwnnsn8v3y/6N7rNdw59/rnum74EktMr9EsuO4lXee+bJIaLGH3fh979uvt50U09U/v0F7zX3Lt3ryNFcU2i6LaHj4BfKvxKRH6p6KdGzKYxJljaeMU5Ve/xzqt3XhERE1kkicV0GpaVdsRaJ2ZrPebmRsytzZQ71xK21hLWuxGdVkQU1X0ihMpDqYt5I194dFrhJyVuVFKe5pSnGeVxRrE/pTjKKI+mFMcZrqhqOvcMpDO1NI2cUUNQFtWSvmA5fjHH5eyFo1RADmRAwcyLQK/uI7CkgvDCRT5zEq1P0+KkG2wSYLd4o4YsN9LwXE2wK4HgYRoRklpMbJHast0wGzatedKzymdYcTQoOeyVHPRKjvsVp8OS/sQxnXqmhdO8UPVBWFSNQSMjxsgCpg3yowJwhMhjUf1KVf8i1j5yzj0xxuw7544fPXrUv+K9c9H9/DrK9C9TspcroCb6N7rN95XQz7DUlbe4n28bv/0xFBuW/98CnJycZCf9fv/GrVvHCieIjMSYgiBwbOtnrKrRXL7fiLFWbBwZE0dijCDOBWbVcOoClXviGGeOrPQ4dwZZChTuGhrypUcrxXgfaMJxoHNH3SCMGq0mgVVXB1sT15YCsiAoiF8aEv3/t/dusZFlWXbY2ufcuBFBBt+PTGZWZlZVs7u6ma2SAI4wgPTB7MYYIwxGhg2bBdhjWx8Gug0/vubHf0zOvyHIEgzM/BhjYGChOB4YwsASINudtEbjbqvZanVXsbuqsrLzzUy+3xFx7z17+ePcIINMPiKCwUwGGacQYCUZseLce885++x99l7rjTAdqw74a79Fhx9w0zwh4lgh6j2iU+sNSlDwxicc7vC1Prf2M96y1zsQDuR8CDNrPSEssBd204SVpBKUI8V2KjT3atUXnFZCby+XIyyvx9jadb4omRBjRAIrplLELCKe9UCVTtUpUYRPPHgEkc8E+DcwZi4Igs+DIHh8/fr1lZ///OfFUzaXtSQYSp1zp9738Ipgto1QE89+pMmYZ+kna1joL0I/D7/M+Pi4XVhYOCBvtrS0FH3/+9/fWlhfL0uSlADsKrArwA4gJQMk4tk9Q2PEb43TDClUilydshzv8dChHFNKsaJc9iwMuyWfhVUqK+IKywI8H10uEJ/IkPHhOM9D57noZE/NNZUdSMN0QcGH6mzOnyEB4uuNoIdeh/npjrkzTTJCb9TMHgq7VUJuWs03Z43PYCtkkOnNIhz0DAe5m55cNHezE+H1DoSDVQkcudRYVcJuxns+oD/zKZb9ud3KeoyF1Rgvlnzo7cVShIVVn3K9ueNYLKehN3jFKVvhBvQepQACpSrIDQKvIPIbEZmHyC+stT835C/Lcfz1H/zBHyz8+Z//ebkq/CYAgiqVYDkhunHanKplbSBqY7GXOuZOK2C25PlKKxgh1hiOq3dxPy93tZEwwnn386hJzkP/f6CNjo5mXS7Xa0ulASdyw5J3SI4a8tsi8g2CtwDpNXuEm6xkvsWJV3QVABJYI9mMSC5rpCsfoKdgMdCdwXBfiGv9/jXSH2KgO4PeQoDOnIUJvGaRIxApkaTUMUjoU7V3HbSYwO3ESDYiJJUw3WIR0VLJh+rWI+iO1wPaz6jzZSgCA1/NUl2DtKdrU4kzYb/8df9G7YfjiAj7ecSmevWsZjao+po9L2yvgFcP9M3YlOqoK+slMfoq4bfcfk1V3qdZS2hhssYnbpi0yBREkH6XKhGVfa3P2laC1Y0Ei2sxVjZjrG0lWNt22C56qp1SpIwT+gaoFc94bQys96zEm3EllLoK8gkgX0PkoQW+VpEnCMMXgeprY8zW/Px8dMQYNIfGG+sc84cXbta4AeSJ8dbjQ/Stjsm2Ebq6Ruis33XeRginTOCD4ZHxcXs7igq5cvmaIb8B1Y8V+K6qfovkbQK9xpiwiv1ZD3+frx0hjPhkhM6cxUBPBtf7Q9wcyuL2tSxuDGYx3BuitxAgG3qdIQBw6YJdYUmoXn6ohJYckq0I8UoJ5YUdlF/uovxqF9FSCUlFVK+UgE4PxgKlvklSlxGqJRpXlYQgVjydUVcmZZPIIRzMIxzMIRzMITOwz2pdSVnHEaroFW/L6T7T9cqmNz6vViIsLHuF080dh2Iqs7CfxbgvcXQYmSQUKJFchdf8+bURmRdrfxWoPmQm8zqfz29VyS6g6tbUY2DO+p42Zou1VtITkhMi62cxrPXwyrHGUNxp/ZQG+l9LbPisNQkyPj5uCnNznN2XECfm5vQpsDZx505xLZstRyIxyS0CiyQ/FJEbAIYA9kGkyxiv4lq5OapesTVxVOeIYhlSLPtMumJZZafkGRhWNmNc6w0x0BOip9Oi0GFRyFlkg0oWXarsagAanwqgAmg+gOR9iE5yFrYQ+iLN5TKStTKSjTLijQhuJ4bbjaFFB40clHroZlTJyB3IoDv0uN54kvLmoNiTHj8o7bCfOG4gGeP52jrSDMBCiKA3LTTtyfkC054Qmd4QttuzSpjAy0UY9SnuXhTP3+PYEaVEUYp91tvWrsPmdoKVzQTLGwmW1iMsryfY2ElYKisS9SxD1p/7iDFecoHqlQ693g8VIltp5ttrks8p8htrzFcW+NoEwW8Gt7dfzH71VenQ2LQTgJk9nvOwnvASm7yO8BzWplbAbHtCZwxp8ZjBWQ+xaS3EoKdNFDbwXawTs95+NuPaa9plTdy5k3sZhgMkh1X1unHufRjzAYwZhcioALfESG+F9r9ihJzSOUdVpZAQMSJhIJLLGHTkjHR1WvQWAvQXMhjqCzHcl/FhuoEsBroDdOUtclkvG1E5y09rVz3XnVMfpism3sjsxnBbSRqqKyNaSTPplouIV0reKBXjPQYGfwPMfgGs2T8v2lM3Tb25BDyQHcfUEJnqyJ4nzTvE6FYxQgKTS5kNerMIB3y4LdObhe0JfSp6mpYue+SvAmO9d7iX7ZYaxEqtz06aCLKy6bC4FmF5I8H6VoL1nQQ7RZ8cslNSRrHSKQkKfXKJQSqpZEVSRvRU7RTAioi8APlIRb4E8NCJPDFB8DJDrhYKhfW5ubndGsZbIyTCtWwIWaN3cVUwW45DrpWM0HHeRbOFrU7K1mlUyfG8Mc/j2qsxK6+Kd2Tu3LkTZrPZgiTJCEQ+oMgYRf4GyW8BuGmAPgChpzM4gC/7zsR+KMgaIJsx6MxZ9KdhuveGs7hzPYcbA/68qKvDIhN4VmZjDlLIvXH1SmhZ4Xa8IYqWS4he76L8aseH6laKSDYiuN3Ey447HksSU8m0Oy5FuzocVx2+Opju5iXQJWN82K066eB63lPr9GZhu9KQW8akCRbYS2KoyGpXHDUASNKwWzFVN61Q7Lxc9skGG9sJdn2ygT+DYjWx+IEDK5IUklDnHMkigHUReW5EHkJkniKfAfi6FIavnufzW5ibc1VjxhyxOTxP0TmesljLFcVsuZDdZZD3PksWWiN4Z9Wvbxbm20jdFvj6IRM/fChVHHT65MmTEoDS2NhYCeVyUYGiimyp6qKI3AZ5EyIDAPsF0uPlxPfdI5eydTvHlAaG2C2p7KRZc8WSS8N0DsvrEQZ7QvR3B+jKe6+oq8MgmzEIrYGx/kwFRva5ZABojpBcAMn7ly0EsL0ZZAZyiFaLSNYiJJuRPzPaSeB2E2jJQcvuYMIATeonoTq3bi/UVjGoWjXf/arsz3l8Np/P5DMdAWynJ3DN9IbI9OaQGcwi6MkiKIQw+TTkZtMCWxJG4eXCU48violyQpRSzr7tos98W9tKsLIRY3k9wdJGjI1tr24axQQErPC8GSNi0wLeqrIpIRmT2BJgBcAigQUBHgvwSEUeisjXA8Xiwo8fPiweGjcGQDA6OioPHz5k1WblXW2k5QpjtsNxb9kTehftPChFzorJc7xPh8kCDi8u8vHHH3eUy+UeLZeH1NoREblpgA8BfEjgG0bkfQBD1njeMqaJBkrSKVQdVekz6YyIhBkj2VDQkbXSXbDo7QzQ1xVgqNcXul7vD3GtP+O9o7xFPmf26oVUUolY9S8m6mmCIgeWHbSUwG3HSLYiuM3IF8CulhEvlxAv+2LYZDPyZ0dpWrekdb4mvRVOgDIERRARCZc+PrNXYOpDbjYX+POd3kqmWxa2O0TQ7YX+fEp54DndUnqiAwZIsCepXWGGKEeKrR2H9e0Eq9sOK5sRVjf9v9e3E2zvOuwUHXbLyigmEufvsQAwFjBiYA0CSS3qHjeg32AsA3hG8iFVvxbgMYCnBngtYbjct7Oz9uPnz4vHjJFaPKHznpPN/J7LgNkSnlDbCLVbM54TAXAMyOC99wouCIZozIcw5iMAYwJ8h56BoR9AHntlOlJN2S/VRaRMQ0fWep2iQt5iMDVAN4eyuDWcxfU0rbu70+5pGBk5Ulz0wE8mhJYd3E6MeN0boGixiOhVEdFS5cyo7D2jyHniVOzpxx2ZHSdmn8oGGc8QHnSHCAfSLLeB/USDoDvjM90qQn6oMCUc6iers92AKPYyGqubyR7LweKaVzRd396X1Vb1cTuTGjJv3ytRNyGQht1UqaqxAEUYs2yMeSbAVwJ8TvILGvMkl8u9cs7tzM/Px1VhN4O2xMJFb7VqkLXDcXVa+tOMaC2H/6jhPbWquTZCw9MszNMMdKOD71Si1/HxcUnp9g+EXuaBCM+fr44B2+7OnbIzZhfAhgGWQd6ByA2QwwR6RaRbBIWKAJrAJxiQhPOJDIwTUsuEGJXdkifQLJadbBe9bMTi2n6YrrsjQGfeoDNrkMukobrAH+bD+iocNV51jRRoh8KkYTrTkYHtSrPShvK+5mi9jGQzhtuOkOwkPuGh5OBKDlTdC8WJTWUU8gGCrIXNW5iU4y7oCZHp8wkHFbaHStKByVtYe5DZQNTHxirZhF7N1mcR7pR9xtvmToLVTZ/xtroZY3UrweZOwp2SQykiqF7ZNAiMmAC+gDi19akRFfWJB9tCrkNkBeSyEC9IPjVe8+ehWvusv6/v9Y9//OPDno+MAhajo+jp6dG5uTnW6PnUo3TcjA10G7MdjnsnfTzp8L8e44AmvR9vCfM8MmJqNfBHxv5HRkY6wjDsDoE+DYJrQt4wxtyG6ocQ+QAidyC4aUQ6TLpI+oNzQpV0SnVVYTprRLJZI7mMRT5rpNBh0d0RoLcrwFBvxofqejIY6s1goCtATyFAV2eAMOMLOZWCWIlUidzLSCQKjdV7O2UH7jqfMbeTINmMkKxHiNdKPrNupYR4rYxkI0K5mKDsiHIgSPIW0ulphcKeLMK+EEFvFrYrLSztCCC5ACZr9tkfAgMJBJ5pArDcdy9UvWR6qazYTMNua1up0dlKsLGTYGPHh9xSYlhGCREnZMp0oCKA9UKE1hgREbPPnEeAdJtQvFTysQKPjOpjsfYpjVkIgSXr3EpHFG39+Pnz0jFj6XBJVC1FkyfV+Z3XWL3KmC1jiC5rqOtdFZSdRz/fJs5ZMasXpkqoTYbGxvJdpVK/iLwn5Eci8m2IfATgQwGuAegGkDtmU7FfrVNRbE21cDKBQWfeF71e6/OhuhuD/udQbwb93Rl05CwygU++rmSHHcvyp2l6cpxm1W1GiFdLKC8WEb3e9SG7lRLK2zFKsaIcCFxnANMdIuzPITuQFpb25XyWW4VCSA5NttQQHib0JoEo8azj27v7gnKLazFer6VeT6pmmzimtUj7GAcejheLk6r4pqZRxHUACyB/A+BXEPnCkF9b4Dmz2bW7d+8WZ2ZmXFXYTRscV+dRbN6Kc7MVMdtGqEnXcxrFRSOhsIu8O3pXfRMAMjk5KY8ePTJzhQIxO+uOeu+Nb397IB/H74vIByLyAVTfh8hNIa8BGAbQI0BBjGTFmAN1N5omM8SJD9M5JQQiYUbQ1Rmgrysjg90Bhvu8ARroCTDQHaKnEKC7w6IrZ5ELBdmMQSbjudCMTeuB0oy6SlxRHXwx627imRjWyt4TWikhXi17IxQ5lAy8J1TIIKxKQAi60iy3TEqkSoUhYLjvhjnHA2G3Uqpoul102Nx12Np2WNv23s/qhi/k3SoqK3LqgM92MyISWB92qyQz+EJTgVJBTzC6SXIdwIoAr6D6AiKPBXgYGPMkJ/LiZ19+uXzEw80AkLGxMeTzec7Nze0XMJ9PBKOZ61gbs22ELoT3cxFDZ+eN+S6qt4/ipDuqmbGxsd4oivqstX2qeg2qN4RyhwajQr4P4CZEBo1Ibu/QviKGp0qn0MR5dQHuZ9OZXMakYboAhTSFu7cQYqAnwHBfBiN9IYb6Mugt+HqjMGMQZvyZkaYKqYn6l6YeERIFEwdXUrASptuOEe8mKJYdSlREVsCshc0FsPkKgeq+jIKk+j3WMGWs80YoSohy5LBTUmymBmdtM8FyJclgJ8FWUVEsO6/JVFJGTunUO0AioIivk7VGApNKy1VqmtJi010Sr0X4FJTHFHkK1Wcw5qUkyaKoLoXkxrd/+7c3U++nlrBbM5SAzyNkJFd0vp8XZtsIXUW39BK3vYVsamrKPHjwwLzI523w+HFBVYfEuTsU+Q7IjwB8A8B7vsYIXWmYTo4L01Wy4PaKXlPW50wgKOQD9HcHuD4Q4vZwFjeHshju81pGnTmLfNYgzHiPK+Uq3Zccr2S6mbT79OSpruwQlx2KZYdi7FBWQlOvyliD1PE5kJ5XxckGepZxFCOf6baxnWBlw2e7La3FWNpIfIFpmohRCSHuZ8pJNecq0xTrwzxvkQAbFHlN4ImQX9KzHTwSkWfOuZVisbi9sLBQxptJBe2QUbu1jVATBnozB3mjInJ8y5iNPufzyqLDJCCLExOyvb0taSjnjd322NhYKOXytciYD6D6vpB3YMwtkNcBGQY4SKBHRAoV7+hwmM4pkSRE4mteUw8JyGa8R9TXnZHrKVv3YE8G/d0BejoCdHdaFPKBz6jLGnSExp8fGf+CEdACsAZMjUuSEHHksFt2KJYTRIkvUrUCWGNg6KXLoeqLcSsht1RIrhT5sNtWyu22se2z3da2vCe0vuOwW3SMEoVTT6tjDdIC01RSQfbtsFdVAAiUCGwS2CC5KiKvQb5U4Kn1xaZPVPWlc24xLTQ+3OxE+qwwN4e5gypHPONYOs0Incf4bHs2bSP0Tq+n0Zj1SdTxx+HWmy5ZL2YtYQw5wks4KV39bWFW/+3YCvrR0dGsqvaEZG9szKAx5pqo3iBwCz5EdxsiN0VkuGKIKjU76rPp4JROlZp6MiKAWCuSCUSyoUFH1koh70lR+woBegsB+roDDHRnMNiTwWBvBkM9GXR1BMhlfc0R4HnhaIwX30u1eaJYUSxVjJDu1TNl0mw3pOc9FXG/naJiY8fX96ylRaUbO76wdKvosFP0yQjlsrKUaKpYSxJCEdBUZBUENvX/9ouqCJBahphXAJ5A5LECT4V8psArEVkkuSwiG3Ecbx1jgI7byAGNFVLXKjHwts+NapFCuIyYbSPUwuG7Wt9zktfytjGPWjxq2Wm+bcw9Z6bq/817772XLRQKvc65ayRvGedGKTIK4EMRuQ2RwVrCdFIVRXO6H87yYTqLnj0WhpQodTDEjYEsBnr8mVE2k3LUefXY/c4SiJP0rKaUoBzrXnFoxUtxnjEc5ZTNemPbYWUjSUlFfZZbhdOtFCniJOV0S/u8V/RaLawnQpLija6vMhUgMcCWiCyJMU8U+IIiX5wQdjvqOdVb69bIc3+bmG8zJNgOXTahXfRiVanTAzkrLUYth/ryDjDZgMd3kTD3jEz6klFAQoDz+3x0AKDPPS1McXR0dAfAliE31ZhlAi8A3KZP6R4SYBBAD8AuiHQaERgRwYGUZ28MnJJxTCbqXaRi2Re/bhcT2dr1nsnKlj+TGejJ7GXUdWQtOkKDjqwvfs2GgkxgEGQEogJNBC4Byo6IUsMTp2G3YlmxU1Jf17PtsL7li0vXvdw5d3Ydyomn1SEBY0QCIwgsxOyFHKXasooSCnAXwCaADaquw9pFeI63pwI8IvCE5Is4jpeOC7uNA6Y4NiZVGW88hxD2u5Q2qDUicpkxW8aDagVRO57gxp/lzERq3HWdVvha68LcKOZRh8ZHKUtKHSGPd4V5mJW7wrhw5CZjZGQk39fX1+2c6xHn+kgOqzHXANwU8g6B2yBvKThixBQqCqDVRsiRdI4uTshUmkCs8WG6MDCSCwW50Eo+Zzw5aqdFb2E/xbu/yyc39BW8ImyhI4A1nkJnczfG+layp91TSa9e3/aGZ3PHMxxsF11FvRTlWBkliigmnTJN8vMUO2nxampP9+9qlRDEGslFgs+ofAbVZ9aYZ0ZkwZFLtHbFGLOxu7u78/xojrfKszFHPLda9K5qZcI+TSH0LJiNbJRqKWK/7JhtI3RGt7TecNBZ2tuUSKj3s60iLXFWzD3DNTU1JQ8ePDAAghcvXnQZYwYA3DTkN5T8pvqfHwAYFqBbRPJGpMIEVGWPKCd+EYDACLKhQVdHgMHezF7N0XBvJmVlyKCvK4MwI0gSxeZu6t2krAbr2/uUOutbCTZ3HLZ3HcqxVzD1YbX9TLmqUNuB+7gnqaAKkgmAojFmzYi8AvBMgK8V+Bqqj2w2+9QYsxxF0c7Dhw+jYzZvl2mM1bIQ17NuXFbMlgnltZoReht9bxuht4e55xWNARKNjgoApHIA1WG6vTY6OpoVkcEAeE+NuZ0kyS0Ad0heBzBkfI1RN0S6ABaMSGiMgYB7i32l8NUnNICp6itI0hiRfGjQUwjQ152RvkKA/u4MBrqDNLMug2woUEdslxzWtmJs7HgDtLnjw24b3viw4v0kjqxQ6QRWJKhiyZbU7fEFpl6ywTmNCGyr6ibJdS+tIIvGyIKIvBCRJyrynOSLnu9///Xcn/xJfMTcsGOAwdgYACDl+eMxi1k9fIb1GozzwmQNi/DhxVtOiZ5cRsy2EWpC304KB52HcbpoRui4MNhlwTzuOR9bmT86OpotFArdURT1lMvlHvjzoWGSIyJyxxjzHoCbnhKI/UZMhxHsKcGpsiIn4bwAqv8f7oXrjGQzItnQIhcayWctCh0GPZ3+rCgbGpA+8227oliacrl52XIyThRJQjoFlaQR0Oxr+Rj4QltvjNJCJV8TpCUqVwi8pOozAk8N+YIiC5ZchjErau2qiGyWy+XtE7LdjpJXOGkRO00OpBG5kPPAbGSs1pKNV0vYvNUw20boHIxQKxiMNub5Ye6FmSYnJ2VxcVGWlpZMJpPpKJVKvc656/BZdB8C+IDkbQGuAxgUoAdAVo5Seq36554MBKtWBPHFr7nQIJ/1tUQA0gw4RRR70lFPxYP9sBv2DZ8cuuK0yrTy2wRABHIHwBKAl/S0Ol8J8MiIPNMoem2d29wEis+fP4/QLvZsz6naPau2EXqHRui0nUWtO/ez3qu3iXmW53tRMA+E68bSf0eAPPQFr8kbX0LK3bt3+0rO3QT5XuDcTfW8dNcJXBORIQH6RKQHnjC1U0RsJUxXfVbjtJJmrXReFA6VcFrGCqz10ti6LzsBpbBCJmqNiLUQmyYbSBXDaIUBQdVFgOyC3IbIOoA1qq6KyALIBTHmuTHmiQVeFIJg8Xfn59enj6612st2qyHsVs+cqOeZvyvMo+Z+vVIpJ5UMtDImcT5Zim0jdMJArPds46SJcdLC2SxN+Fox6+knasRvBczjJupJBJrmzp073UEQdAHotmRfSv8zlGoZ3YDILYjcBnBNqX2GCPYF4DwhWxqi81ISSl8Bi73ImVTqhirnOD6VQAiAAoExECNi05/eC5JKyE0BoqzkqoCvCVkQ8ilEXoBcUJHXqrpqjFkLgmCN5Pb169eLs7OzySlzpd4MMxyzAcMJGybW+dzPA5N1vKcWSZfTzlFaFbNl2lU3QselfNdTQNqoEWr0Gmtd3E/Cb1XMo57v4QQWAYAJwCwNDeXQ0VFIgF4JghEFbgH4EMZ8k+RtktfoXC+AvPgEhowYs0/OdpiY7sTRKjgyzLcvn+0AFAFsC7BG4JV4Ke3HAL42wFM4t8AwXNna2tpdWFiI4D9z3C73Kj33ejBxSmTkbUi6XATMthFqct9qqek5qjalVuNw1A68FoNRrzzEaZg8xYuQGv4G1FZP1YqYpvp9EwC2x8dlbm6uonuTHPXcR4GsvXVrQI0ZhrU3acwdADdV9TqcG4JInxjTIyI9YqQgIp0ikjU4pPeDfdnxNx64HLwsKqBkBHKbItsANiCySnLVAMsUWTDAS0e+kDh+bsnXydOnqw+95s8bHh4Ai7ExGQNwqMj0JG/jrM+oFsaLd4lZ7yaWdXgZrYrZLlZ9Sx5RLXxnpxmiWtz503YiqHEQ1ELseNK16KEF6bTf1+sFSgthyiHMinFSnMJPF0VRRyaTKZDsVWv7jHODdO66MeaaGHNdgBGKXBdyRMFhKxKKCIwxe73Zqy493E0ecH8AogRgRYGXEHkJ8iWMeSGqr2HtMoBlY8x6FEWbzrnNzs7O3fn5+eiE8WpwQH7vyMW9XiN03DOSYzZj7wqzWd5VowtzK2G264TOsZ+toIJ6ln4efm8txuYqGqFGMasXQDMyMpLN5XJdIjJsjBkRkRskbwG4TdU7JG+BHBSRgjEmlPTM6NB3pZ4RoaqOpBMgNiJbAqwKsKAiv4HIbyjyRESeichiGIZr2Wx2q1AoRLOzs3pMWPEqPqO2EWoboZbpJ9/Bw6iVEkMa6Odph7daA77UEP46LRTZqpjVYTqZmJjA9va2AMDc3ByqvKQ3nsfI+HhHf6k0UHJuyDh3neSIJskNku+JyDBE+q0xvUakACMFEAUAWQgMCCVZJnXHKTeE3CCwaYEVEVkW1dfOmKc05oWqLgB4ba1df/jwYfmYaw0qBab5fJ4ffvihzszMHDWGTlrcGwlB44QN0EXDPO3zckwk5CzZYq2EWUsCR9sINaGvPIf34x1j8i1cb70ZNBcVU04w2EdtPk7s2+joaDafz3ckSdJdLpe7JEm61doB8Yzdw4HIDRhzE8AIPIFqN4AQQERyg+BrVT43wAtVfWWtfW1U15gkayqyijDcVNWtW7du7ZyS5VbrYn1SluVRCxoavJ8XDbPeecUG10G2MGb7TKjd2u0dj+dGPWCZmJiwS0tLZnNz0xYKhQ7nXA/JYQvcIvABRd4H8B7JPgBZ+CSCZQDPSf7GGPMYwAsRWQSwtbGxUXz9+nUZBwXiztNTb0sVtO9bS07aVu83m3yt9UgcSIOfbcT7abQOQJq0C2sFzNPqwWRiYsJUheoInwZ9JPbU1JSZmZnpTZJkmOR7AG5SOALd84TKIrJGkVd07kUQBC+ttUvzk5PrmJ7WEzwdOzo6agAgDEPOz8/rEYaq3udezyLV6P28qJj1jC+pYQ7zkmG2jdBb6DebfI21xKOlSZiH4+anFbbWiyk1/L7VMGslcDw8YQ+H7Socdce3qSnz0T/9p50AulS1x1rtJYOcqlpLG9NyB8BW5fXFxx/vYmbG1bAIHyen0MxrJ+ovI7jImM2c77VKubQ6ZvtM6AK4z/UWmb5thul2htvZMGt5FvWmBVf3p6of43Zk5GUmSfoMAGSz3a67e9MNDQ3p7OzsUePwbfTzqmGiyfO9nsW9FTDb2XFvoY+1eEFn4Wm7SEbotB3OVcasZxyfNFkPfHZ8fLwSqqt4SkmdE9nAyyhIJcutUCgwNVL1jNvzvPZWx2wbobYRuhDGp15D1Ir6PURtBI5XDbOWv58WEqyVa00bGLvmDOGkWun8zyqx0YqYzTRCJ529sYUxW84IBS3oCdXynstCcy9tzKZ9Z6Nj5EBx6hSmcH8KwN15eTC0KADw4AEwPz/LmZmaZZibPd6vEmYzWi31aZcVs+0J1dk3aXB3dRaxu9N258ftShoNSdUTfpI2ZtMm92mYB+paxsbGzO/e7bbfGMra4dvDtq8zDmItmSQu8ustJLpZLv/hP/xxjDeTIHjO/bxKmLXQYJ0233mKZ9wIg/5Fwmx7Qm9hB1OLIZJzmgi1EpRKnZOr3gPb4wbjVcGs1YCdBbOao04mh4ZwbwJhX3d/BzNRR86GWckZiZNcfGNHdrc2ggT4cbkae3JyUsZmZjB9vv28CpjVG5iTNpi1hkDrKZloVcyWiQaZFjI+NVW/v+P+tdu7GRNNf46Tkwf+rffv3dNv3ujL9vRKfyEX3Mhk9L2syPW8sX1decn/jQ/DConqgT7dP99+XkXMZm0sryJm2wid0+Je70A/6f2s0wCyCZOzFsxaGbzrua5WxqzlOZ30nE+k9yGAsbGJg7vu6fuUINPpnLsJ0VFARwVyU8T0ZQObC5T2cH8ma3v2jfTzqmI2MreaYQRbFbNthJq4EzjuXIE1LHSs4ztOM0RH4dYabpA6MY+TMOAx185TBuNlwjwtbHHS8zsJ07epKblxY3vveZH+vU6SbqXcAfEdY+QjgndIDgi0o1g2GZIHn/HkG+jN6ievGGY9xu24uV/LwlzPHL3omC3jTV30MyHW+b5aqu5rjV+/S7dbGuzHWSrJ25hVbXx8HMBcqusNcuq+cYqCGIwI8KGCeVF0iHBTVV7FYTmD+/cPMKJPtu/n28CsJZLR7PnZKphtT+gdhu/YwO/Z4ORoxC2WM3z+bYUA25hV7fO784FT00mRfgpGVHnHqd4RyHWAvaU4zuHuvLD62X6+KO372RKhpMt4xtYyKd2tboRqddlPOweqRXpXGphIR02sejB5DpOzjXkM5p5C9/Q0t7ZGDvwt3ok6aFAApDs1RNdJuUXyPQLD1trOzwF7fwr7hugecP9+UzYxV/0ZscF1QZp4Da2G2TKG6DJ6Qudl6KSJD1UaHIDnsftrY6ZN1T8LAbi0NH/g753dpkBKQSAdQuk0YrqNkUGCIwQHVVxn51CHmZ6GSoo99+W2tJ/RW8eUc5izrYrZEu2yMCac9rl66kLO0h82+FmpA68ZC1sb8wjMKq+Fn38+RgICTsnjB4/DZHO7x0C6QXSS7AxDm6WwK4qTYQD9AnbsYLd6Psmjl8VGd6TtZ9Rc3JPqi87Cbt0KmG1PqAlW/TTjUkvhlpyCV6txkBr7LCfsdOrBbO/q3hLm1BQEmNibD9PT08TUlHz1z1czebfaExgMieqgKHpE0VnIWslmJAewn8J+CLvDIvLV3/U5gOnpyhfU1f/2Mzod97i/Hze/Tvr9ad/ZqphtT+gdeEJvww1tFe61dqujzc9D+vq2q6vyU1aO1WyxZPupGBZgAEABQJixBgCN/ze7odKTFddFcllECAALC/k9L3tyHjLTvs3val6exrRwGTFbJlsuuAIDsN2u2EI0OQksLkKACdyr+uODEz74O33bUro7ZCYmJvjg3j3c9+83t2OXN4IBB7lG6CAEnQAsAFoREUqWRLeI9ItyYOWf/2fL/OkPivf/8AsuYBuTk5DFRcjv9I3L2MQcH9RxMcPD4MxM6y0qV2ydaG9KL/HFNkJ13jZGV3yRmZiYkI8+2paRtaLcHQMWViMBgJH+8MRFfAc7ZuDGoCm9LCaT94YUAL4qfmwlfnXLSPQ3hcFvKTluRL4tkFv9XSEIYmW9FAP4GYBZseavA5jP4w4szf0Vih/eeMStlwU+wJK5hyHzi9UXXK3ux/zJF7PwKM8vCnMcngVn2obotJ2+1PkZadJ6clExW6Zg9TJ4Qm3D0257E292dpazs/V/cGIC5r+512k+mZ5PMO01hD77dNgEWe0QlSEIb4jBIAR5EIgdfS42jQAsUGRYiBEReZ2PuTGJsS354UxFi8jc+MG4/au1hzrzj9uG5B0aolpLIi4j5oVt9hJ4QufhIZ2Hd9XGPP/JUk0gWtfryRPozOySq+7r0NgkPx5euA3wYxH5rjFyW0S6QQmNEVDBxClAKUKwA5ENq2Y1Y/i68MP/aav6+v9ybkHn5xvrG66ApswFWC8uG2a7WPUC7ZDqrTdoVgV4s/rTxjyhTQGYmpoCwaZ7GJN35wMKuxQYpvIGgQEjkgWAxBGOFIgYGBQgHAT1uoMOFHcPZMk1bUxPTQFTU21rc8xCyyaNsVbHbLkWXOIByiZ9Vi5Af9qYp1qj+0JM7df8AMD0NKZPMzSTkJkZcBKQsckxOz0zn6QeFTo70J/EdlCpQxT0W5EOIwIF4ZR+C2cgIHJC9IEYdnDD1rLvs08nX3z3k5mostn7R//daOYvfnHTzc7O1iwX3jY4DY0zaWO27k7iMvWtFu31k3YVrOGz9RSt1tqfZoh+neWetQLmngc0jdT9IfHg/j07dvdubrdjqWCQ6ShFLitJUdQE6spwcSZOAuQcXUSxoYhxYqwTiVWKzjB0lk7K1jLrct2WgXEB1N6Infs7SvyuAf52Pm+7AEEUK9Rnx4GggIxU3TqAh0r8BOD/U471M5vI7i52wS0EpazLuChQdZG/H6USNGOoiaFLykzojKEJIkqgdAbqXJixZRfFpUxUKmIV5Z/0P4zv3wdFDtzLZobqzuOc4bwxz1rc2qhy60XFbDlxu4vsCbGOmyp1PGjWOTlOM1Co0y0+Tu6glkHJc5jYrYJ5wADNT04KZma8doII+cc/kMcdS4VE5UZg4uFMgB4EQZgkwrCAKGAQi2pMayjG/xeIFSdGesSoS5SQjICkUMUpckKOAOYjSw5BmBXK3qGTN30ECRFBQKAb0GsKfGjAZWtoE+FWiACuU8NcohlkhAwCOk0ksSGRCGOragKr2URMAmYNJKsqQtqSS7BuTLBczOdX+vvdOoBYZH9ekAcM0lmMEeucgxcBs1HGeznBgxDULv1yUTFbLiTXCuE4nvE9chkeVLsdCqMBmKl6jp/3rUk2NhmbYYHkMMgRQHoDixAAAjFOAolJowIKfEWpGCMAhCYw+2EyEQOgg8SgALdFpA8Uow7QqtFUOYUSH5jLEugDeEuJHVC6jNEtVYFAstbYQElQARERK4ZqFQFIRyGtMaKaEaUxImVVrBnoCyNazhJbRUTmmEXpqntCV+UaL21SSqueCUkd7mg96ZAn7S6kjvfX4r0d9n54wk6RdVx/rYO4VTBrma2CtT4+vrkWK1GiMhYyMGIGIBgm2S3UECJqRRIIEhJKMZRUiI5iuPcghMZQchTpAtgnkC4IREFVL3AnhztpjZHEsRPkdQgIcIjALgQgEUJg/EkTIZXHLTBixEL9PDQGCSnbCrwWQdEYoQViIxpL0OGw7T8l53w/L8kiex5hqVbAlCb3r22ETrnJcszCfVKBF4/BOS6ExGN2no0azaMmHw/9PGnimiP6whoGH3G8+uVFwzxxUZoGwLEZVrskD+7f41jf3e3VzMpiViRnwAIF10DmDTBCkX4AGQhUhJGCCQhHEQPCyP6TEAGEQEAwIyIhhCEJqxRWZBo8uWnFIlS6akIA/VDmCI4IkVABY4xJHAVQioAUEgZEIhkRZAQSwjBWJ+si+gQOCR02FFwOaFZj5fb1/kyE7SnPQecJ6VgVmjvpvjUj/HXRMWsxZnKGhblVMRtdr9pGqMme0Fl3DscZuJMKxk7b3RyF2YgndJadZqtgHt3ug3upb/fvyz3cU3yO3YVvbC7lumKIIATYQ0EfwX4Q/QA7RST0VkZAQFMLcvCoP/1nxd2puEiKffluQerQVMJy9G4NgJxScwIBBRARkIQRgXoL5D9NKiQ1fv4rNgQoArIk4HMBnhrBS7XRyvVCfvu3fjgXT039vpmZn6/3jKHtCbU9obYRugCNV6TPvCSYp77HmEN/n75PgRDg7utPP1mOAxMmYBYWhKIEYIeQ2wSuZQPTEYYBBGKUQOIUToGUdxSkf6kSSkLVVyEd56b6DD3vF1kjCKwBIKASquI/rwaEg8AAJOLEmXLkUIqSBMAygacEvgLxGalfiA0eM3BLm6vY/nt/OBd7W3ufM598IhdoDvCSzNX2vbkgoa2r4Ak1klnXyC6jnt2N1LAzZI19qDVLkHVc17vArGVssup/DpyT8NNJ+wgo5DsyvWWXjNC623D4COR3jDXfBuSbHTlbCDMWSqAUKRKne25JxS8lCSWgqnuGSY4P9NIf96T5c1SoEs75n757/m8ColxOsLUTxyJ8HKn7tUA/U6dfONGvXVlf2XKwFpQXd+7df1I+kBGX+lFNnus8h/XjXWM2MgfPY16/bcyW84TaNCDtdqk2VZyaEpme3st0+82P/kEu2t4dCIgPhBiD4G8Zwd8i+Q2CA9ZY65QgqGmWGwgfSTt1C3Bo6fBGS6G696JzhI/BEU6dN0pOE3VuXR2fq3I+gc5R9BcJkkdaiF7/9de/KE6n/HUAMDUFc38alHZWZ7tdwha0b0G7XSYjlJ6Z7LUPvvenJQAv5v/iPy91hq6YMNkhuQXIEsgPSA6T7AGYFQMoAQgVEFSyEEgxIt4JOuUwxkft6A2ICEUEVoA0pKdlEFsAlkA8EeArQOeNMb+kKT/6q5dzr6f/a+hZrv8cdsGtgtluLdps+xa022VqY/PzmD1iYZv4jz6K7jAs0WADRtcJbJEaCSQA2QFhzlgxYuiNjs9+E4hAfEGq1LBmCpnaIMIYgbXeiiFKlCTXATyl6ucQ/H8u0X9Dg18WInk8kKys/8f//WpyFOjsLPhH9Uc1mqVy2gqY7dY2Qu3WbhfDE5qtkj2emoLMznqrMTMzz//h2/9heem2bFpubxDYgiICYQEGEA0cmXH0WnU+001E9kNzJy+WB5KlCZIGJJJEEcUuIrlC8gmIeQA/E/KnMOazKGue/d3/6l9t/o+/vcq7wzAz8w2nKjdbZruVMNutlSdtu4/tdtnG9OQkZGxxQu7eWzJD94Z0aWmYn3wy46rf9PRfTPZHRXxg4MYouAuj31HIKIFrRqTLwGTSmiEVT5Ej3DsuOsIGaVqFKiDVMU6cIbWUJMl2kuiyqj4D8QVU5hXJvGX01fh/+dcLqKr1+eMfIPOtkQkCwIMHs8A96PR0XVIOlz0cx/b6cKZ71V7gW6Rv9RCdNgO3Ucx66oh4iTGPM0QYW4TMe4lsPYzNTz+1Tzv+9+4k2h1ysO+bgH+T5G8D5q4R3BBIQSAGqSdEXwykPBCa810hfTocAWutQNWxHCWi1DV1+kydfpWom2eMz5T8ypEvu3perX33k/noUL/N1BQwP+/xZ2bq5hbEOd3Pi4J51L1odq1aLTRfrYLZEsboMsh7t1u71TqeODk5KVNjsHcxllSy6H79z/6TQStuXOj+PUD+rgDfFEivwCcViKSJBaCCcsgIpUymgCMYZAIBqdgtxVDqKzr9lSP/LcCfmtj+vFwyT37rh3+5WzE6P/jBuF1bm9Mqg9NuJy/CpxmhdjtofNqMCZf04bY1Sy425hteUd+jcYNx4O7dLVNaDez9aewlAXz73/9flx/+5X+6DOoWIBFIT1D9ZlFQWpdaNbFZIftJy1Wrtk2qIEViAXZBbopircoA4Qfj43Z8BJlSdtQtTjx0s8MgZlorlPION6dnCeNdBRYJaZVxZNpj+tjdFuv8W6O4jWLW2k9eckyegsOZGeCLuTmurc3p558v6WJ/4qaxX0v06aeTlhqHRiRjgEDBAIAwrUCFl2ygUCgiLk3BdiQUAgfsp1VXilrhAQIo81AUQOnIGxNWd+xP5uZciGX9q1cPOTsLYqam+1DPtZ/H/XyXmEclNbDOMXbqeLlEmC3R2kao/oUTTV6IG8WstZ+4pJj1GDEOe2PEtbU5/cnqU1aP/Y9zQZ8x6AfYC7BgBBljUqJSgilTghCwJAMIAgAZ8T8DAhagAQBlqrrql8ysiPQCHCY5vCtR32efTh4wRJurGS4uvnlm1cRrv4yYZ12Ma53flxGzHY47Jxf1PA5LzxJKaobeiTTxc5cds+ZnNDYGLtzIc2piwty/f8+83FoIi9Fyn2TNAKj9ALoDKxkBECegeI/HADDWilgDGCueIpyeFy5RinMwJOmUVCq8VhFC59gPyDUIh4SmKymtZwDEJPDJJ75fw8N7C4ic09yRS4bJK3rdzbg3bU+ogR3OaTeZTcCs9f3VoQAe8556+3wUZq3XdZyMwnF1GJcJs5bvqqh/cxKTAIDpaXDk5Zy7OzxMdN0QJsW8hEGvkAMC9pPsyQTGWiuVwJoj4Qiqd4kIIwojhBGft+0PiUgReBtEOBAwBhkR6RVgyMD0M2BnUmSAVIZhZgYY6Q85NnPEWVOTrv087uc7xKw1rHXUIi9nGG+tiNkyBqkVwnG1uJtE/THVs04g1NGfejHPEvK6ipinT8zJ/X9PT0Mnx8b47NGvg0hdwZADBAdJ9ImwM7BirREREUPAEhIAYhxVy1GS7JbiaHs3inaLcVQqu8Q5dem3pOE5GhGINcYakU4C/Ur0E9LtwjBf3be1hTzvn3+o6qphnrR5lCuC2TIeUaueCTX70LSWcB+bPJHOegjdLKN+GTHffM/kod/fB1xmNQtJeqkYgqCfQIGUjCeK85w7QWCCXNaKtQISRYIvCPwKxC9ImQfwhMQGSOQyBvmshTXGSko2F1iBqnZSXW/MZICq/T/9Z7/f8emnkxYA+9Y+VFPpF8/p2i8nZjs1+5K0VjRC5xGOq7UolceEEBoJKZxmUHmGa+EJfb0KmG+8Z/LQGx88eGDW491ONW4QVq9BMQCaHGCoFOecr/4x4mdJlKgC8kpEPhPIvxbI/02DfwXBzwE+AbFFEoER2FTIjqSjkqoMlex2Ca4FGXc92djq/9s7ixkAmPx0Rvc6eX9K0pqjpl77edzPd4RZmZumRm/hPAxaq2C2TGvVOiE2+fPShO/jOVwDz+H+XEnMB0OLb9RNhMZ1GGAQlBECQxB0iIhxzusIOSVip8rYRRC8IuVXAvszAl9RuEGYThh9n0RRCCmX3YdRjG7HPXE860uNmCHZJeAQiOuBySy9DoJNAKXqUqSZu/MyeXYP/zI/97agX9sIXdrWLg685K3ry+0DC9jSvWGO/LnmjWBQgRuAXDcGXUHGwFhBqjFUFGKVwgVAHkL4S6H5DFaeWBtsRU7zRuJVpeyIul2CmyBuqXIAYHcYGAiIYtmELtGCiPRDdTCG9JRKxfCNTk62n1ONc7UdimsbobYharfWbmMzsDZnOkkdInCD4DVjJS9CRIlDnGgkwCuKfCnEZzDyeQL5ygEvriX5ta14t5yEXQF0ZVODcF1glwTxggJjpH6b5AckCiaV+1ZFB6h9FDMgyp68IFdFsEAAGHqwKLjXfjY1zNGKIWrP2bYRard2a80WFsJckJR6Cbkm4HUI+kSA3XKsgGxZmkWBfGUh/5aQOcD9SvJ2Ye1R3+5HL0dcX8o796MfTRS7kuHNHpVVdbKq0E0FI5LcLSYfipGC83wKeYEMKd0Nqg5D2PVgasJ+b3r2gIbQ/fvtxbW9aWwboYvW5B0NarnAk03O+P5WwTzzjasAbr0s7O2e+emk/aq4283ADCt4A+BwPrRInCsJ5BWBZ0o+CYz5tYj9pYX86r1CxzPxKq0H2ve+N5sA2PjRj/5B+fbWboQkLifQsoDbStkQ5Xsg+kWYFXJERNcFfGaUX+B9CYA9Djv58sttSRm02+GmxgyQtMj68zYw2ynaTbqpUuN7TsqtP+nvcsLrIlx7LX0+bTDWcn0XGbOhZzE1BflkctLAK9PxAWYr/G7y73rWczYw/VAdIrWfynwpcnDEgoj8OxEzC5h/qcBskMv8PAjty6MM0EFj9KelTAlL1uZ+jcD+a2PxLwn8X1T5qZJPqYhB7RXwpjG8YUL05YN8ePh+zHgCU5k6O3lls8f2RcI8bryddfE+qYat1TBbZjNzGT0hafDvrbr7PI8d0EXDrPvZzM9Dxhb3M+Kmp6EAzP/xj/5eJlsMewAOK9BLUgCsqbIM2l9ZkZ9KYH6mRh46CV7d+J3/ZUtEyKkp8wAPzAPM6vR02qdJoEI6+unkpLn1+VjZTP/R60c/+i82dpZWVilYh8MWiJhQQ+p1gKEBu0HpScpJAcDmXqfnqvp/zszilwCTl/S6rpwndFXOhHiOD/9dGprTPAye8H6p4T3vArN5z+gegNn9yTgF4PZIV6dhNAwx1yAuB2KdkC8EWFbVz4Ig87NcZ/irrZ2+l9/6vX9cBv7Mf/juvBQX8vbv58YFUwUuLGwLxoHxPuBbIwUCi5Bpr976wff+tDQ1NfX8P7j7UxckcHBMUtG7bwFilASIAtX1/eifTGziLkrf+96se/lF4TyYAy67nEezdcfkHIxFq2C2jdAZjMtpCzYbXNAvynU14oo3cp8uGqY0YdHbx5yaQjbz604VOwiwVwSxgXlE6C4gz8TgS0ry5fXMzcWR3/uH5QMd+WRGP52cTP5n/AuMjYEjIwBeAi9HgJfTwPShL52entbJTyeXdpIdFxizLZRdQFapOgCaNaExoZFuoe02S9sOQIJ7s7pnNCchUzNv4tZ5zc28nxcB8zjuwWZu9nhFMdtGqEkLdiOMzs2I17aK99YKmE3xBMbGwLt3hw/87j6Ax4CNVQwMioC8EEFRYF874oVF9PJJ5+bKh3/nLyrJAmZqYsLMz85yBuAnMzN6wqK5V8k/NQGZnoV+95OZCMCrn/7x7286m5RgZU1gR0iKA9bVqcSBhljMWm+49nWJxsaaKqXwNjZ/bwvzvMNMVzmk1zZC72CBr4cyRM5hcsg5TV65JJin7fZOzGj8/POxN75TNCg7TVat8XoLsHbTKpZcGK584/f+t73zGRLyJz8ct7eyeXPjo3HFKRLcU1PAgwcwH+dH7fvXYpnC+8n0rE+G+K0f/uXuj/7JxONcR6ZsFC8NJZtYjVSxsY1yzMygHsa7fx+8P934tZ/H/bxAmM00TnKFMS9EuwqidkcpMTa6iLKO358FEzheObKW7CHieHr4w5i4wJgNPwufPDC9h1spDO202W2TMwtJgEeGmYdQ+yS7kV38s5/82faBmy/gy5E595NfFN3/uW+AjqXYn54GhmfBv156qP/vZ09cxQBV2vf+29mdrkLfgpKPaO3XCTLPkkhXklJxJygl7jCmCDBd33hr9F62EubhxbhZRLm4hJgtw6QtLdA3qWOg17LondVrkXeIKahP+qCW9180zKOSGuq5P0feT3p3RXAP5qviqs3kNyUudnPzWVnHX444TE9TTvbWpMaxd6RRJiCYmpIHeGCG7g6bpZ1FAwC/2HzB1dWbbnp6VlElE97g2Kinn5cJsxmb1GYu3q2C2Q7HtVu7vfWdzfemE+wXiFbPaOGnk3Zurc+M963p/c/H+EfT08oGD4AnJyFjYxPm79/4SIA54IdziXiWBT30xYL7fyDAbPvhtNuVbP8/14ih+k9k4PoAAAAASUVORK5CYII="
    c1, c2, c3 = st.columns([1, 2, 1])
    with c2:
        st.image(BytesIO(base64.b64decode(_AHLY_LOGO_B64)), width=110)
    st.markdown(
        "<h2 style='text-align:center;margin:0.25rem 0 0.1rem'>🎙️ 𝔖𝔢𝔱𝔱𝔦𝔫𝔤𝔰</h2>"
        "<p style='text-align:center;opacity:0.75;margin:0 0 0.5rem;font-size:0.9rem'>اختر القسم من القائمة</p>",
        unsafe_allow_html=True,
    )
    st.divider()
    selected_page = st.session_state.get("selected_page", DEFAULT_PAGE)
    for page_label in PAGES:
        if st.button(
            page_label,
            key=f"sidebar_page_{page_label}",
            use_container_width=True,
            type="primary" if selected_page == page_label else "secondary",
        ):
            st.session_state["selected_page"] = page_label
            st.rerun()

PAGES[selected_page]()
